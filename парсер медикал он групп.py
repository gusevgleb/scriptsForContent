from bs4 import BeautifulSoup
import requests
import re
import inspect
import time
from datetime import datetime
import os.path
from glob import glob
import pathlib
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter
import unicodedata


url = 'https://medongroup-perm.ru/prices/endocrinology/'
page = requests.get(url)
soup = BeautifulSoup(page.text, "html.parser")
price_name = soup.h2.text
print(price_name)
price_container = soup.findAll("div", class_="price__main")[1]
services_names = price_container.findAll("div", class_ ="price__line-adr")
services_prices = price_container.findAll("div", class_ ="price__line-cost")



filename_xlsx = glob('*.xlsx')
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
wb = openpyxl.load_workbook(filename_strip)
wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column

for i in range(0, len(services_names)):
	if services_names[i].text is not None:
		sheet_0.cell(row = i+1, column = 2).value = services_names[i].text

for i in range(0, len(services_prices)):
	if services_prices[i].text is not None:
		sheet_0.cell(row = i+1, column = 3).value = services_prices[i].text

print("Успех! Белиссимоs!")
wb.save(price_name + str(filename_strip))
print("end")

