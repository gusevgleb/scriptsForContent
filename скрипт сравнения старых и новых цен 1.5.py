import openpyxl
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter
import pathlib 
from glob import glob
import time
from datetime import datetime

start_time = datetime.now()
filename_xlsx = glob('*.xlsx')
print(filename_xlsx)
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
print(filename_strip)
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column


###################### эта часть конвертирует числа-как-строки в числа ######################################

comma = [","]
for i in range(2, rows + 1):
	cell_3 = str(sheet_0.cell(row = i, column = 3).value) 
	
	if comma[0] in cell_3:
		n = " "
		if n in sheet_0.cell(row = i, column = 3).value:
			b = str(sheet_0.cell(row = i, column = 3).value)
			d = b.replace(" ", "")
			c = d.replace(",", ".")
			sheet_0.cell(row = i, column = 3).value = int(float((c)))


			
		else:
			e = str(sheet_0.cell(row = i, column = 3).value)
			f = e.replace(",", ".")
			sheet_0.cell(row = i, column = 3).value = int(float((f)))


wb.active = 1
sheet_1 = wb.active
rows, columns = sheet_1.max_row, sheet_1.max_column
sheet_1.cell(row = 15, column = 7).fill = PatternFill("solid", fgColor="FFFF00")
sheet_1.cell(row = 16, column = 7).fill = PatternFill("solid", fgColor="008000")
sheet_1.cell(row = 15, column = 8).value = "ЖЕЛТЫЙ ЦВЕТ - услуга была в старом прайсе и не поменялась в новом"
sheet_1.cell(row = 16, column = 8).value = "ЗЕЛЕНЫЙ ЦВЕТ - услуга была в старом прайсе и поменялась в новом"
sheet_1.cell(row = 17, column = 8).value = "БЕЛЫЙ ЦВЕТ - услуги не было в старом прайсе, только в новом появилась"


for i in range(2, rows + 1):
	cell_3 = str(sheet_1.cell(row = i, column = 3).value) 
	if comma[0] in cell_3:
		n = " "
		if n in sheet_1.cell(row = i, column = 3).value:
			b = str(sheet_1.cell(row = i, column = 3).value)
			d = b.replace(" ", "")
			c = d.replace(",", ".")
			sheet_1.cell(row = i, column = 3).value = int(float((c)))

		else:
			e = str(sheet_1.cell(row = i, column = 3).value)
			f = e.replace(",", ".")
			print(sheet_1.cell(row = i, column = 3).value)
			sheet_1.cell(row = i, column = 3).value = int(float((f)))
##############################################################################################################
rows_0 = sheet_0.max_row + 1
rows_1 = sheet_1.max_row + 1
print(rows_0, rows_1)
##################################################### основная часть #########################################
##################################################### основная часть #########################################
print("Пошло дело, пошло! " + str(datetime.now() - start_time))
for i in range(2, rows_0):
	wb.active = 0
	sheet_0 = wb.active
	rows, columns = sheet_0.max_row, sheet_0.max_column
	cell_nazvanie_uslugi_old = sheet_0.cell(row = i, column = 2).value
	cell_price_old = sheet_0.cell(row = i, column = 3).value

	for j in range(2, rows_1):
		
		cell_nazvanie_uslugi_new = sheet_1.cell(row = j, column = 2).value
		cell_price_new = sheet_1.cell(row = j, column = 3).value

		if cell_nazvanie_uslugi_old == cell_nazvanie_uslugi_new and cell_nazvanie_uslugi_new != None:	
			if cell_price_old == cell_price_new:
				sheet_1.cell(row = j, column = 3).value = sheet_1.cell(row = j, column = 3).value
				print(cell_price_new)
				sheet_1.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="FFFF00")				
				#Заливает желтым, цены совпали все ок
								
			else:
				sheet_1.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="008000")
				#Заливает зеленым, цены не совпали 
				not_matched_value_from_old = cell_price_old
				print(cell_price_new)
				sheet_1.cell(row = j, column = 5).value = not_matched_value_from_old 
				sheet_1.cell(row = j, column = 6).value = "Старая цена"
				# print(not_matched_value_from_old)
print("Середина пути! Nel mezzo del cammin di nostra vita " + str(datetime.now() - start_time))


#################################################################################################################
print("И последний рывок! " + str(datetime.now() - start_time))
for i in range(2, rows_1):
	wb.active = 1
	sheet_1 = wb.active
	cell_nazvanie_uslugi_old = sheet_1.cell(row = i, column = 2).value
	cell_price_old = sheet_1.cell(row = i, column = 3).value

	for j in range(2, rows_0):
		
		cell_nazvanie_uslugi_new = sheet_0.cell(row = j, column = 2).value
		cell_price_new = sheet_0.cell(row = j, column = 3).value

		if cell_nazvanie_uslugi_old == cell_nazvanie_uslugi_new and cell_nazvanie_uslugi_new != None:
			#print(cell_nazvanie_uslugi_new)			
			if cell_price_old == cell_price_new:
				sheet_0.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="FFFF00")				
				#Заливает желтым, цены совпали все ок								
			else:
				sheet_0.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="008000")
				#Заливает зеленым, цены не совпали 
				not_matched_value_from_old = cell_price_old
				sheet_0.cell(row = j, column = 5).value = not_matched_value_from_old
				sheet_0.cell(row = j, column = 6).value = "Новая цена"
				print(not_matched_value_from_old)

print("Успех! Белиссимоs!")
time.sleep(2)
wb.save('__рабочий шаблон для обновления прайса клиники после скрипта__.xlsx')
