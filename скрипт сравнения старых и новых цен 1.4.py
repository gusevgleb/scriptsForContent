import openpyxl
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter
import pathlib 
from glob import glob
import time
from datetime import datetime
'''
обязательно проверять желтые цены с комментами
у кати не работал неболит королев и ленинский. ошибка в ценах
'''
#wb = openpyxl.load_workbook('рабочий шаблон для обновления прайса клиники.xlsx')
start_time = datetime.now()
filename_xlsx = glob('*.xlsx')
print(filename_xlsx)
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
print(filename_strip)
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column


###################### эта часть конвертирует числа-как-строки в числа ######################################
comma = [","]
for i in range(1, rows + 1):
	cell_3 = str(sheet_0.cell(row = i, column = 3).value) 
	
	if comma[0] in cell_3:
		n = " "
		if n in sheet_0.cell(row = i, column = 3).value:
			b = str(sheet_0.cell(row = i, column = 3).value)
			d = b.replace(" ", "")
			c = d.replace(",", ".")
			print(i)
			sheet_0.cell(row = i, column = 3).value = int(float((c)))


			
		else:
			e = str(sheet_0.cell(row = i, column = 3).value)
			f = e.replace(",", ".")
			sheet_0.cell(row = i, column = 3).value = int(float((f)))


wb.active = 2
sheet_2 = wb.active
wb.active = 1
sheet_1 = wb.active
rows, columns = sheet_1.max_row, sheet_1.max_column
sheet_1.cell(row = 15, column = 7).fill = PatternFill("solid", fgColor="FFFF00")
sheet_1.cell(row = 16, column = 7).fill = PatternFill("solid", fgColor="008000")
sheet_1.cell(row = 18, column = 7).fill = PatternFill("solid", fgColor="808080")
sheet_1.cell(row = 15, column = 8).value = "ЖЕЛТЫЙ ЦВЕТ - услуга была в старом прайсе и не поменялась в новом"
sheet_1.cell(row = 16, column = 8).value = "ЗЕЛЕНЫЙ ЦВЕТ - услуга была в старом прайсе и поменялась в новом"
sheet_1.cell(row = 17, column = 8).value = "БЕЛЫЙ ЦВЕТ - услуги не было в старом прайсе, только в новом появилась"
sheet_1.cell(row = 18, column = 8).value = "СЕРЫЙ ЦВЕТ - скрипт обновил эту услугу в нашем шаблоне самостоятельно"

for i in range(1, rows + 1):
	cell_3 = str(sheet_1.cell(row = i, column = 3).value) 
	if comma[0] in cell_3:
		n = " "
		if n in sheet_1.cell(row = i, column = 3).value:
			b = str(sheet_1.cell(row = i, column = 3).value)
			d = b.replace(" ", "")
			c = d.replace(",", ".")
			print(i)
			sheet_1.cell(row = i, column = 3).value = int(float((c)))


		else:
			e = str(sheet_0.cell(row = i, column = 3).value)
			f = e.replace(",", ".")
			sheet_1.cell(row = i, column = 3).value = int(float((f)))
			#print(sheet_1.cell(row = i, column = 3).value)
##############################################################################################################
rows_0 = sheet_0.max_row + 1
rows_1 = sheet_1.max_row + 1
rows_2 = sheet_2.max_row + 1
print(rows_0, rows_1, rows_2)
##################################################### основная часть #########################################
print("Пошло дело, пошло! " + str(datetime.now() - start_time))
for i in range(1, rows_0):
	wb.active = 0
	sheet_0 = wb.active
	rows, columns = sheet_0.max_row, sheet_0.max_column
	cell_nazvanie_uslugi_old = sheet_0.cell(row = i, column = 2).value
	cell_price_old = sheet_0.cell(row = i, column = 3).value

	for j in range(1, rows_1):
		
		cell_nazvanie_uslugi_new = sheet_1.cell(row = j, column = 2).value
		cell_price_new = sheet_1.cell(row = j, column = 3).value

		if cell_nazvanie_uslugi_old == cell_nazvanie_uslugi_new and cell_nazvanie_uslugi_new != None:
			#print(cell_nazvanie_uslugi_new)			
			if cell_price_old == cell_price_new:
				sheet_1.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="FFFF00")				
				#Заливает желтым, цены совпали все ок
				wb.active = 2
				sheet_2 = wb.active
				rows, columns = sheet_2.max_row, sheet_2.max_column
				
				for k in range(1, rows_2):
					cell_from_our_shablon = sheet_2.cell(row = k, column = 4).value
					cell_from_our_price = sheet_2.cell(row = k, column = 5).value
					cell_from_our_shablon_split = cell_from_our_shablon.split(' ')
					if len(cell_from_our_shablon_split) == 1:
							shablon_iz_nashego_price_dla_poiska = str(cell_from_our_shablon_split).rstrip('\']')
							shablon_iz_nashego_price_dla_poiska_2 = shablon_iz_nashego_price_dla_poiska.lstrip('[\'')
							if len(shablon_iz_nashego_price_dla_poiska_2) > 3:
								shablon_iz_nashego_price_dla_poiska_3 = shablon_iz_nashego_price_dla_poiska_2[1:len(shablon_iz_nashego_price_dla_poiska_2)-2]
							if len(shablon_iz_nashego_price_dla_poiska_2) <= 3:
								shablon_iz_nashego_price_dla_poiska_3 == shablon_iz_nashego_price_dla_poiska_2
							if cell_price_old == cell_from_our_price:
								if shablon_iz_nashego_price_dla_poiska_3 in cell_nazvanie_uslugi_new:									
									sheet_2.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="FFFF00")

					if len(cell_from_our_shablon_split) >= 2:
						first_word_from_template, second_word_from_template = cell_from_our_shablon_split[0], cell_from_our_shablon_split[1]
						if len(first_word_from_template) > 4:
							first_word_from_template = cell_from_our_shablon_split[0][1:len(cell_from_our_shablon_split[0])-1]	
							second_word_from_template = cell_from_our_shablon_split[1][1:len(cell_from_our_shablon_split[1])-1]
							if first_word_from_template in cell_nazvanie_uslugi_new:														
									if second_word_from_template in cell_nazvanie_uslugi_new:
										if cell_price_old == cell_from_our_price:			
											sheet_2.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="FFFF00")

							
						if len(first_word_from_template) <= 4:
							first_word_from_template = cell_from_our_shablon_split[0]
							second_word_from_template = cell_from_our_shablon_split[1]											
							if cell_price_old == cell_from_our_price:
								if first_word_from_template in cell_nazvanie_uslugi_new:
									if second_word_from_template in cell_nazvanie_uslugi_new:
										sheet_2.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="FFFF00")
								
			else:
				sheet_1.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="008000")
				#Заливает зеленым, цены не совпали 
				not_matched_value_from_old = cell_price_old
				sheet_1.cell(row = j, column = 2).value = str(sheet_1.cell(row = j, column = 2).value) + " ```"
				sheet_1.cell(row = j, column = 6).value = not_matched_value_from_old 
				sheet_1.cell(row = j, column = 7).value = "Старая цена"
				print(not_matched_value_from_old)
print("Середина пути! Nel mezzo del cammin di nostra vita " + str(datetime.now() - start_time))

############################# эта часть вносит изменившиеся цены в наш шаблон ################################################
wb.active = 1
sheet_1 = wb.active
rows, columns = sheet_1.max_row, sheet_1.max_column
not_unique_from_new = []
not_unique_from_old = []
not_unique_from_template = []
for_find_green = " ```"

for i in range(1, rows + 1):
	wb.active = 0
	sheet_0 = wb.active
	rows, columns = sheet_0.max_row, sheet_0.max_column
	cell_nazvanie_uslugi_old = sheet_0.cell(row = i, column = 2).value
	cell_price_old = sheet_0.cell(row = i, column = 3).value
	wb.active = 1
	sheet_1 = wb.active
	rows, columns = sheet_1.max_row, sheet_1.max_column
	if cell_nazvanie_uslugi_old != None:
		not_unique_from_old.append(cell_nazvanie_uslugi_old)
	counter = Counter(not_unique_from_old)
	unique_from_old = [x for x in not_unique_from_old if counter[x] == 1]


for j in range(1, rows+1):
	cell_nazvanie_uslugi_new = sheet_1.cell(row = j, column = 2).value
	cell_price_new = sheet_1.cell(row = j, column = 3).value
	if cell_nazvanie_uslugi_new != None:
		not_unique_from_new.append(cell_nazvanie_uslugi_new)
	counter = Counter(not_unique_from_new)
	unique_from_new = [x for x in not_unique_from_new if counter[x] == 1]

wb.active = 2
sheet_2 = wb.active
rows, columns = sheet_2.max_row, sheet_2.max_column
for k in range(1, rows + 1):
	cell_from_our_shablon = sheet_2.cell(row = k, column = 4).value
	if cell_from_our_shablon != None:
		not_unique_from_template.append(cell_from_our_shablon)
	counter = Counter(not_unique_from_template)
	unique_from_template = [x for x in not_unique_from_template if counter[x] == 1]


wb.active = 1
sheet_1 = wb.active
rows, columns = sheet_1.max_row, sheet_1.max_column

for i in range(1, rows + 1):
	cell_nazvanie_uslugi_new = sheet_1.cell(row = i, column = 2).value
	
	if cell_nazvanie_uslugi_new != None:
		if for_find_green in sheet_1.cell(row = i, column = 2).value:
			if sheet_1.cell(row = i, column = 2).value in unique_from_new:
				ads = sheet_1.cell(row = i, column = 2).value
				ads = str(ads).rstrip(' ```')
				sheet_1.cell(row = i, column = 2).value = str(ads).rstrip(' ```')
				if ads in unique_from_old:
					dsa = sheet_1.cell(row = i, column = 3).value
					wb.active = 2
					sheet_2 = wb.active
					rows, columns = sheet_2.max_row, sheet_2.max_column
					for i in range(1, rows + 1):
						if sheet_2.cell(row = i, column = 4).value == ads:
							if sheet_2.cell(row = i, column = 4).value in unique_from_template:
								sheet_2.cell(row = i, column = 5).value = dsa
								sheet_2.cell(row = i, column = 5).fill = PatternFill("solid", fgColor="808080")
								print(sheet_2.cell(row = i, column = 4).value, dsa)
								wb.active = 1
								sheet_1 = wb.active
								rows, columns = sheet_1.max_row, sheet_1.max_column
								for i in range(1, rows + 1):
									dark_cell = sheet_1.cell(row = i, column = 2).value
									if dark_cell == ads:
										sheet_1.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="808080")

#################################################################################################################
print("И последний рывок! " + str(datetime.now() - start_time))
for i in range(1, rows_1):
	wb.active = 1
	sheet_1 = wb.active
	cell_nazvanie_uslugi_old = sheet_1.cell(row = i, column = 2).value
	cell_price_old = sheet_1.cell(row = i, column = 3).value

	for j in range(1, rows_0):
		
		cell_nazvanie_uslugi_new = sheet_0.cell(row = j, column = 2).value
		cell_price_new = sheet_0.cell(row = j, column = 3).value

		if cell_nazvanie_uslugi_old == cell_nazvanie_uslugi_new and cell_nazvanie_uslugi_new != None:
			#print(cell_nazvanie_uslugi_new)			
			if cell_price_old == cell_price_new:
				sheet_0.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="FFFF00")				
				#Заливает желтым, цены совпали все ок								
			else:
				sheet_0.cell(row = j, column = 2).fill = PatternFill("solid", fgColor="008000")
				#Заливает зеленым, цены не совпали 
				not_matched_value_from_old = cell_price_old
				sheet_0.cell(row = j, column = 5).value = not_matched_value_from_old 
				sheet_0.cell(row = j, column = 6).value = "Новая цена"
				print(not_matched_value_from_old)

print("Успех! Белиссимоs!")
time.sleep(2)
wb.save('__рабочий шаблон для обновления прайса клиники после скрипта__.xlsx')
