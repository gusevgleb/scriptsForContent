import openpyxl
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter
import pathlib 
from glob import glob

filename_xlsx = glob('*.xlsx')
print(filename_xlsx)
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
print(filename_strip)
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column

rows_0 = sheet_0.max_row + 1
##################################################### основная часть #########################################
print("Пошло дело, пошло!")

# first_arr = []

# for i in range(2, rows_0):
# 	value_1 = sheet_0.cell(row = i, column = 2).value
# 	first_arr.append(value_1)
# 	print(i)

# 	for m in range(2, rows_0):
# 		value_2 = sheet_0.cell(row = m, column = 7).value

# 		if value_1 == value_2:
# 			sheet_0.cell(row = m, column = 6).value = 'это мы завели'



first_arr = []

for i in range(2, rows_0):
	value_2 = sheet_0.cell(row = i, column = 7).value

	if value_2 != None and value_2 != '':
		first_arr.append(value_2)


for m in range(2, rows_0):
	value_1 = sheet_0.cell(row = m, column = 5).value
	print(value_1)

	if value_1 in first_arr:
		sheet_0.cell(row = m, column = 6).value = 'это мы завели'


		

# for i in range(2, rows_0):
# 	value_from_1_list = sheet_0.cell(row = i, column = 4).value
# 	massiv_1 = value_from_1_list.split()
# 	wb.active = 1
# 	sheet_1 = wb.active
# 	rows_1 = sheet_1.max_row
# 	for m in range(2, rows_1):
# 		value_from_2_list = sheet_1.cell(row = m, column = 4).value
# 		massiv_2 = value_from_2_list.split()

# 		if massiv_1[0] == massiv_2[0]:
# 			sheet_1.cell(row = m, column = 4).fill = PatternFill("solid", fgColor="008000")

print("Успех! Белиссимоs!")
wb.save('__после скрипта__' + str(filename_strip))
print("end")
