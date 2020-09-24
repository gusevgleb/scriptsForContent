#! /usr/bin/env python
# -*- coding: utf-8 -*-
import inspect
from selenium import webdriver
import time
from datetime import datetime
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
import os.path
from glob import glob
import pathlib
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter

filename_xlsx = glob('*.xlsx')
print(filename_xlsx)
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
print(filename_strip)
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column
lk_list = []
for i in range(1, rows + 1):
	cell_3 = str(sheet_0.cell(row = i, column = 1).value) 
	if cell_3 not in lk_list:
		lk_list.append(cell_3)
print(lk_list)

wb.active = 1
sheet_1 = wb.active
rows, columns = sheet_1.max_row, sheet_1.max_column

for i in range(1, len(lk_list)):
	sheet_1.cell(row = 1, column = i+1).value = lk_list[i]
	sheet_1.cell(row = 2, column = i+1).value = 0
	sheet_1.cell(row = 3, column = i+1).value = 0
	sheet_1.cell(row = 4, column = i+1).value = 0
	sheet_1.cell(row = 5, column = i+1).value = 0
	sheet_1.cell(row = 6, column = i+1).value = 0
	sheet_1.cell(row = 7, column = i+1).value = 0
	sheet_1.cell(row = 8, column = i+1).value = 0
	sheet_1.cell(row = 9, column = i+1).value = 0
	sheet_1.cell(row = 10, column = i+1).value = 0
	sheet_1.cell(row = 11, column = i+1).value = 0
	sheet_1.cell(row = 12, column = i+1).value = 0
	sheet_1.cell(row = 13, column = i+1).value = 0
	sheet_1.cell(row = 14, column = i+1).value = 0
	sheet_1.cell(row = 15, column = i+1).value = 0
	sheet_1.cell(row = 16, column = i+1).value = 0
	sheet_1.cell(row = 17, column = i+1).value = 0
	sheet_1.cell(row = 18, column = i+1).value = 0
	sheet_1.cell(row = 19, column = i+1).value = 0
	sheet_1.cell(row = 20, column = i+1).value = 0
	sheet_1.cell(row = 21, column = i+1).value = 0
	sheet_1.cell(row = 22, column = i+1).value = 0
	sheet_1.cell(row = 23, column = i+1).value = 0
	sheet_1.cell(row = 24, column = i+1).value = 0
	sheet_1.cell(row = 25, column = i+1).value = 0
	sheet_1.cell(row = 26, column = i+1).value = 0
	sheet_1.cell(row = 27, column = i+1).value = 0
	sheet_1.cell(row = 28, column = i+1).value = 0
	sheet_1.cell(row = 29, column = i+1).value = 0
	sheet_1.cell(row = 30, column = i+1).value = 0
	sheet_1.cell(row = 31, column = i+1).value = 0
	sheet_1.cell(row = 32, column = i+1).value = 0



for i in range(1, 32):
	sheet_1.cell(row = i + 1, column = 1).value = i 

wb.save(filename_strip)
wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column

first_date = ['01.09', '02.09', '03.09', '04.09', '05.09', '06.09', '07.09', '08.09', '09.09', '10.09', '11.09', '12.09', '13.09', '14.09', '15.09', '16.09', '17.09', '18.09', '19.09', '20.09', '21.09', '22.09', '23.09', '24.09', '25.09', '26.09', '27.09', '28.09', '29.09', '30.09', '31.09']

for i in range(1, rows + 1):
	# print(str(sheet_0.cell(row = i, column = 1).value))
	if str(sheet_0.cell(row = i, column = 1).value) in lk_list:
		for k in range(0, len(first_date)):
			if first_date[k] in str(sheet_0.cell(row = i, column = 5).value):
				lk_name = str(sheet_0.cell(row = i, column = 1).value)
				wb.active = 1
				sheet_1 = wb.active
				rows, columns = sheet_1.max_row, sheet_1.max_column
				for j in range(2, len(lk_list) + 1):
					
					if sheet_1.cell(row = 1, column = j).value == lk_name:
						if sheet_1.cell(row = k+2, column = j).value != None:
							sheet_1.cell(row = k+2, column = j).value = int(sheet_1.cell(row = k+2, column = j).value) + 1
		

		


wb.save(filename_strip)
