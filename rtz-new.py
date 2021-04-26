import numpy as np
from glob import glob
import pathlib
import openpyxl

filename_xlsx = glob('*.xlsx')
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
print(filename_strip)

wb = openpyxl.load_workbook(filename_strip)
sheet = wb.active
rows = sheet.max_row
columns = sheet.max_column


def mainFunction():
	wb.create_sheet('ИТОГО', 1)
	wb.save
	doctorList = []
	for k in range(3, 7):
		for i in range(2, rows + 1):
			if sheet.cell(row = i, column = k).value != None and sheet.cell(row = i, column = k).value != "":
				
				if sheet.cell(row = i, column = k).value not in doctorList:
					doctorList.append(sheet.cell(row = i, column = k).value)


	print(str(doctorList) + ' - список врачей,', 'количество врачей - ' + str(len(doctorList)) + ' штук')

	for i in range(0, len(doctorList)):
		rtzList = []

		for g in range(3, columns):
			for k in range(2, rows + 1):
				if sheet.cell(row = k, column = g).value == doctorList[i]:
					rtzList.append(sheet.cell(row = k, column = 1).value)

					wb.active = 1
					sheet_2 = wb.active

					sheet_2.cell(row = i + 1, column = 1).value = sheet.cell(row = k, column = g).value
					sheet_2.cell(row = i + 1, column = 2).value = str(rtzList)


	wb.save('__после скрипта РТЗ__' + str(filename_strip))
mainFunction()