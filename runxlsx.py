#! /usr/bin/env python
# -*- coding: utf-8 -*-

import inspect
import time
from datetime import datetime
import os.path
from glob import glob
import pathlib
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter
import unicodedata

start_time = datetime.now()
print("заполнить прайс без лабораторки - 1")
print("заполнить лабораторку - 2")
print("заполнить прайс и лабораторку  - 3")
checker = int(input())
# checker = 1
if checker == 3:
	print("строка, где начинается лабораторка")
	where_is_lab_1 = int(input())
	# where_is_lab_1 = 960

						   
######################################################################################################################################
blood = [940399, 1917236, 1131883, 2794078, 940418, 1875930, 2007251, 1131886, 1131887, 1913343, 1131888, 1131889, 1131890, 1131891, 1131892, 1913346, 1131893, 1131894, 1913347, 1131895, 1131896, 1131897, 1131898, 1131899, 1131900, 2007252, 1131901, 2007263, 1131884, 2141263, 1875928, 2007284, 2007285, 2007286, 2007292, 2007298, 940402, 2007303, 2007304, 2007302, 2141260, 2007306, 1875952, 1917252, 1917253, 1131902, 2007432, 1131940, 1131903, 2007677, 1131904, 2008078, 1131905, 1875963, 2008631, 940407, 1917243, 940408, 1131907, 2007231, 2007258, 2007259, 2007260, 1131909, 1131908, 2007261, 2007262, 1131910, 1131911, 1131912, 2007299, 1875950, 2007311, 2007266, 2007267, 2007268, 2007269, 2007270, 2007271, 2007273, 2007277, 2007280, 2007281, 2007314, 940409, 1131913, 1131914, 1131915, 1131916, 940415, 1132736, 1131918, 1131917, 1131919, 1134803, 2007297, 2007307, 2007221, 940401, 1131920, 1131927, 1131929, 1131934, 1131933, 2007255, 2007256, 2007264, 1131971, 2007309, 2007272, 2007274, 2007275, 2007276, 2007278, 2007279, 2007254, 2008030, 2007289, 2007291, 2007312, 2007313, 2007384, 2007470, 1875945, 2007665, 1131941, 1131942, 1131945, 1131943, 1131944, 940413, 1131948, 2141267, 940416, 1131946, 940414, 1131947, 1876003, 1131979, 1131964, 1131952, 1131965, 940403, 1131953, 1131954, 1131955, 1131962, 1131956, 1131957, 1131959, 1131960, 1131961, 1131963, 1917238, 1131967, 2007282, 1131970, 1131972, 2007300, 2007301, 1131973, 1131974, 1131976, 1131977, 1131978, 2008138, 2008139, 2008467, 2007257, 2007296, 2007305, 2007308, 1132028, 2007573, 2008714, 2007616, 940400, 2009069, 2009051, 2009054]
lpg = [934658, 1131867, 1131868]
fgds = [414, 1839654, 2009151, 1929898]
vakcine = [865244, 1836640, 1836626, 2492189, 1913367, 940421, 1836613, 2492186, 1836634, 1124103, 1836620, 865253, 1134599, 2375472, 1134600, 1836631, 2245240, 1836651, 1836652, 865248, 1836615, 1836650, 1836654, 1836621, 1836628, 865250, 1836641, 1836642, 1836616, 1836622, 1132361, 1836653, 1836638, 1836618, 1836632, 1836629, 1836655, 1132362, 865257, 865254, 1134294, 1134293, 865260, 865255, 865245, 865246, 1134570, 865276, 1132363, 865256, 1828795, 865275, 2114790, 865252, 865274, 1132364, 865258, 2114794, 865249, 1134314, 1134316, 865259, 1134319, 865251, 1132365, 1836623, 1836607, 2007408, 2007409, 2007410, 1836619, 1836630, 1836645, 1836633, 1836646, 1836647, 1836635, 1836648, 1836614, 1836625, 1836624, 1836617, 1836627, 1836649, 2007407]
mazok = [1131922, 1131921, 2761762, 940412, 1131923, 1131924, 1131925, 1131926, 1132171, 2007865, 2007866, 2008175, 2008179, 2008180, 2008181, 2008182, 2007835]
vp4 = [1131878]
lab_list_1 = ["Анализ на ХГЧ", "Анализ крови на ВИЧ", "Анализ на антитела к C1q", "Анализ на антитела к ds ДНК", "Анализ на антитела к Jo-1", "Анализ на антитела к Ro/SS-A", "Анализ на антитела к нуклеосомам", "Анализ на антитела к цитоплазме нейтрофилов", "Анализ на антицентромерные антитела", "Анализ на витамин Д", "витамин д", "витамин d", "Тест (анализ) Т-Спот", "Реакция Манту", "Диаскинтест", "Анализ на сифилис", "Анализ на хламидии", "Анализ на уреаплазму", "ПЦР анализ на скрытые инфекции", "Анализ на гонорею", "Анализ на трихомониаз", "Анализ на гепатит C", "Анализ на гепатит A", "Анализ на гепатит B", "Анализ на гепатит D", "Анализ на группу крови и резус фактор", "Анализ ВПЧ высокоонкогенный", "Анализ ВПЧ", "Дыхательный тест на хеликобактер", "Гастропанель", "IgE анализ крови на Иммуноглобулин Е", "Педиатрическая панель", "Скарификационный тест", "Анализ на аллергию к деревьям", "Анализ на аллергию к насекомым", "Прик-тест", "Анализ крови на гомоцистеин", "Анализ волос на микроэлементы", "Анализ на витамин B12", "Анализ камня из почки", "Анализ на токсоплазмоз", "Анализ на гемоглобин", "Анализ АЦЦП", "Анализ на TORCH-инфекции", "Иммуноферментный анализ крови", "Спермограмма", "MAR-тест", "HLA-типирование", "Анализ кала на кальпротектин", "Анализ крови на ферритин", "Анализ на антимюллеров гормон / АМГ", "Анализ крови на кальцитонин", "Анализ крови на прогестерон", "Анализ крови на гормоны ТТГ", "Анализ крови на антитела к тиреоглобулину", "Анализ крови на гормоны Т3 / трийодтиронин", "Забор крови"]
lab_list_2 = ["Анализ крови на тиреоглобулин", "Тест на антитела к коронавирусу", "Тест на коронавирус", "Анализ мочи на наркотики", "Липидограмма", "Анализ на пролактин", "Анализ на тестостерон", "Анализ крови на ПСА / простатический специфический антиген", "Анализ крови на Са-125", "Анализ на онкомаркер HE4", "Анализ на онкомаркер Хромогранин А", "Анализ на нейронспецифическую енолазу", "Анализ Cа 72-4", "Анализ крови на бета-2-микроглобулин", "Анализ крови на Са 15-3", "Анализ на антиген плоскоклеточной карциномы", "Анализ на онкомаркер CA-242", "Анализ на онкомаркер UBC", "Анализ на онкомаркер белок S100", "Анализ на раково-эмбриональный антиген / РЭА", "Анализ Са-19-9", "ДНК тест на отцовство", "Иммунограмма", "Квантифероновый тест на туберкулез", "Копрограмма", "Общий анализ кала", "Анализ кала на скрытую кровь", "Анализ на гликозилированный гемоглобин", "Проба Зимницкого", "Проба Реберга", "Проба Сулковича", "Анализ мочи по Нечипоренко", "Анализ крови на эозинофильный катионный белок", "Анализ на соматомедин-с в крови", "Анализ крови на фибриноген", "Волчаночный антикоагулянт", "Антинуклеарный фактор", "Анализ на ФСГ", "Анализ на лютеинизирующий гормон", "Анализ на аллергию к животным", "Анализ на бактериальные и грибковые аллергены", "Анализ на бытовые аллергены", "Анализ на пищевые аллергены", "Анализ на пыльцу растений", "Взятие мазка", "Забор кров", "Анализ крови на гормоны Т4 / тироксин"]


def add_sostavnie():
	print("добавляю составные")
	wb.active = 1
	sheet = wb.active
	rows, columns = sheet.max_row, sheet.max_column
	for i in range(1, rows +1):
		cell = sheet.cell(row = i, column = 2).value
		
		if str(sheet.cell(row = i, column = 4).value) in lab_list_1 or str(sheet.cell(row = i, column = 4).value) in lab_list_2:
			sheet.cell(row = i, column = 4).fill = PatternFill("solid", fgColor="FFFF00")	

		if cell != None and cell in blood:
			cell2 = sheet.cell(row = i, column = 10)
			cell2.value = 2216897
		elif cell != None and cell in lpg:
			cell2 = sheet.cell(row = i, column = 10)
			cell2.value = 2309419
		elif cell != None and cell in fgds:
			cell2 = sheet.cell(row = i, column = 10)
			cell2.value = '546730, 2219856'
		elif cell != None and cell in vakcine:
			cell2 = sheet.cell(row = i, column = 10)
			cell2.value = 2216902
		elif cell != None and cell in mazok:
			cell2 = sheet.cell(row = i, column = 10)
			cell2.value = 2309416
		elif cell != None and cell in vp4:
			cell2 = sheet.cell(row = i, column = 10)
			cell2.value = 'ВСТАВЬ СОСТАВНУЮ ВРУЧНУЮ'
			sheet.cell(row = i, column = 8).fill = PatternFill("solid", fgColor="FFFF00")
			sheet.cell(row = i, column = 7).fill = PatternFill("solid", fgColor="FFFF00")
		




#####################################################################
def without_lab_func():
	wb = openpyxl.load_workbook('__после скрипта__' + str(filename_strip))
	wb.active = 0
	sheet_0 = wb.active
	rows = sheet_0.max_row
	if checker == 3:
		main_loop = where_is_lab_1
	else:
		main_loop = sheet_0.max_row
	######################################################################################################
	comma, ints = [","], ["Ц", "ц", "Ф", "ф" "А", "а", "Б", "б", "В", "в", "Г", "г", "Д", "д", "е", "Е", "Ж", "ж", "З", "з", "И", "и", "К", "к", "Л", "л", "М", "м", "Н", "н", "о", "О", "П", "п", "Р", "р", "С", "с", "Т", "т", "У", "у"]
	yo = "ё"
	n = " "
	counter = 0
	for i in range(1, rows+1):
		cell_value_with_comma = str(sheet_0.cell(row = i, column = 3).value)
		if cell_value_with_comma.isprintable() is False:
			cell_value_with_comma = cell_value_with_comma.replace("\xa0", "")
			sheet_0.cell(row = i, column = 3).value = cell_value_with_comma

	for i in range(1, rows+1):
		cell_value_with_comma = str(sheet_0.cell(row = i, column = 3).value) 
		cell_with_upper = str(sheet_0.cell(row = i, column = 2).value)
		if sheet_0.cell(row = i, column = 2).value != None:
			cell_without_upper = cell_with_upper.lower()
			sheet_0.cell(row = i, column = 2).value = cell_without_upper
			if yo in sheet_0.cell(row = i, column = 2).value:      
				sheet_0.cell(row = i, column = 2).value = sheet_0.cell(row = i, column = 2).value.replace('ё','е')
		if comma[0] in cell_value_with_comma:
			stop_int = re.search('|'.join(ints), cell_value_with_comma) != None
			if stop_int == False:

				q = sheet_0.cell(row = i, column = 3).value.split()
				if len(q) > 1:
					g = str(q[0] + str(q[1]))
					sheet_0.cell(row = i, column = 3).value = g
				if n in sheet_0.cell(row = i, column = 3).value:
					b = str(sheet_0.cell(row = i, column = 3).value)
					d = b.replace(" ", "")
					c = d.replace(",", ".")
					sheet_0.cell(row = i, column = 3).value = int(float((c)))                           
				elif sheet_0.cell(row = i, column = 3).value != '' or sheet_0.cell(row = i, column = 3).value != ' ':
					e = str(sheet_0.cell(row = i, column = 3).value)
					f = e.replace(",", ".")
					sheet_0.cell(row = i, column = 3).value = int(float((f)))

		elif sheet_0.cell(row = i, column = 3).value != None and sheet_0.cell(row = i, column = 3).value != " " and sheet_0.cell(row = i, column = 3).value != "  " and sheet_0.cell(row = i, column = 3).value != "   " and sheet_0.cell(row = i, column = 3).value != "":
			stop_int = re.search('|'.join(ints), cell_value_with_comma) != None
			if stop_int == False:
				try:
					sheet_0.cell(row = i, column = 3).value = int(float(sheet_0.cell(row = i, column = 3).value))
				except ValueError:
					print("в строке " + str(i) + " ошибка, " "проверь ячейку со значением " + sheet_0.cell(row = i, column = 3).value)
					time.sleep(10)
					raise ValueError

   
	dict_for_rename = {}
	dict_for_yellow_fill = {} # это словарь на случай, если скрипт найдет несколько строкв прайсе клиники, и первая строка будет больше по цене
	def new_doubler_main(name_1, name_2, stopword, what_double, what_double_service):
		wb.active = 1                                            
		sheet_1 = wb.active                                                         
		rows, columns = sheet_1.max_row, sheet_1.max_column
		for i in range(1, rows + 1):
			if sheet_1.cell(row = i, column = 4).value == what_double:                  
				sheet_1.insert_rows(i)
				sheet_1.cell(row = i, column = 2).value = what_double_service
				sheet_1.cell(row = i, column = 4).value = what_double
				break   

		wb.active = 0
		sheet_0 = wb.active
		rows, columns = sheet_0.max_row, sheet_0.max_column
		for i in range(1, main_loop + 1):
			cell_value_name_from_price = str(sheet_0.cell(row = i, column = 2).value)
			for j in range(0, len(name_1)):     
				if name_1[j] in cell_value_name_from_price:
					for m in range(0, len(name_2)):
						if name_2[m] in cell_value_name_from_price:							
							stop_word = re.search('|'.join(stopword), cell_value_name_from_price) != None
							if stop_word == False:
								if sheet_0.cell(row = i, column = 3).value != '' and sheet_0.cell(row = i, column = 3).value != None and sheet_0.cell(row = i, column = 3).value != 0:
									if str(sheet_0.cell(row = i, column = 3).value).isdigit():
										cell_value_from_price = round(int(sheet_0.cell(row = i, column = 3).value)) 
										wb.active = 1
										sheet_1 = wb.active
										rows, columns = sheet_1.max_row, sheet_1.max_column
										for k in range(1, rows):
											cell_value_name_from_template = str(sheet_1.cell(row = k, column = 4).value)

											if what_double == cell_value_name_from_template:
												print(cell_value_name_from_template + " дублируем")
												if cell_value_from_price != sheet_1.cell(row = k, column = 5).value and cell_value_from_price != "":
													if sheet_0.cell(row = i, column = 3).value != None:                             
														if sheet_1.cell(row = k, column = 5).value == None:             
															sheet_1.cell(row = k, column = 5).value = cell_value_from_price
															sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
															sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
															link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
															sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
															sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
															dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
															dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
															break
														elif int(cell_value_from_price) < int(sheet_1.cell(row = k, column = 5).value):
															sheet_1.cell(row = k, column = 5).value = cell_value_from_price
															sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
															sheet_1.cell(row = k, column = 6).value = 1 
															sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
															must_rename = dict_for_rename.get(cell_value_name_from_template)

															must_rename.value = ""
															dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
															link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
															sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
															sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
															must_refill = dict_for_yellow_fill.get(cell_value_name_from_template)
															must_refill.fill = PatternFill("solid", fgColor="FFFFFF")
															dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
															break
														else:
															sheet_1.cell(row = k, column = 6).value = 1
													break

	print("Начинаю заполнять прайс...")
	array_of_service = []
	array_out_of_service = []
	def mark_in_our_price(cell_id):
		wb.active = 1
		sheet_1 = wb.active
		rows, columns = sheet_1.max_row, sheet_1.max_column
		for i in range(1, rows+1):
			marked_cell = sheet_1.cell(row = i, column = 4).value
			if marked_cell == cell_id:
				sheet_1.cell(row = i, column = 4).fill = PatternFill("solid", fgColor="FFFF00")
	def mark_not_in_our_price(cell_id):
		wb.active = 1
		sheet_1 = wb.active
		rows, columns = sheet_1.max_row, sheet_1.max_column
		for i in range(1, rows+1):
			array_of_service.append(sheet_1.cell(row = i, column = 4).value)
		if cell_id not in array_of_service:
			array_out_of_service.append(cell_id)


		####################################################################################################

	def perfect_match(name, cell_id):
			wb.active = 0
			sheet_0 = wb.active
			rows, columns = sheet_0.max_row, sheet_0.max_column
			for i in range(1, main_loop + 1):
				cell_value_name_from_price = str(sheet_0.cell(row = i, column = 2).value)
				for j in range(0, len(name)):
					if name[j] == str(sheet_0.cell(row = i, column = 2).value):
						if sheet_0.cell(row = i, column = 3).value != None:
							cell_value_from_price = round(int(sheet_0.cell(row = i, column = 3).value))
							wb.active = 1
							sheet_1 = wb.active
							rows, columns = sheet_1.max_row, sheet_1.max_column
							for k in range(1, rows):
								cell_value_name_from_template = str(sheet_1.cell(row = k, column = 4).value)
								if cell_id == str(sheet_1.cell(row = k, column = 4).value):
									if sheet_1.cell(row = k, column = 5).value == None:             
										sheet_1.cell(row = k, column = 5).value = cell_value_from_price
										sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
										sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
										dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
										link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
										sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
										sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
										dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
										print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
										break
									elif cell_value_from_price < int(sheet_1.cell(row = k, column = 5).value):
										sheet_1.cell(row = k, column = 5).value = cell_value_from_price
										sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
										sheet_1.cell(row = k, column = 6).value = 1 
										sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
										must_rename = dict_for_rename.get(cell_value_name_from_template)
										must_rename.value = ""
										dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
										link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
										sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
										sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
										must_refill = dict_for_yellow_fill.get(cell_value_name_from_template)
										must_refill.fill = PatternFill("solid", fgColor="FFFFFF")
										dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
										print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
										break
									else:
										sheet_1.cell(row = k, column = 6).value = 1
										break

	################################################################################################################################################################
	def function_with_assistant_in_cell_or_and_min_price_main(cell_in_our_price, assistant, min_price, cell_id):
		a = []
		value_for_list = 100000
		q = None
		for i in range(1, main_loop+1):
			cell = sheet_0.cell(row = i, column = 2).value
			for x in range(0, len(cell_in_our_price)):
				try:
					if cell_in_our_price[x] in str(cell) and sheet_0.cell(row = i, column = 3).value != "" and sheet_0.cell(row = i, column = 3).value != None and int(float(sheet_0.cell(row = i, column = 3).value)) > int(min_price):
						for j in range(0, len(assistant)):
							if assistant[j] in str(cell):
								a.append(value_for_list)
								if int(float(sheet_0.cell(row = i, column = 3).value)) < value_for_list:

									value_for_list = int(float(sheet_0.cell(row = i, column = 3).value))
									q = int(float(sheet_0.cell(row = i, column = 3).value))
									p = cell
									sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
				except ValueError:
					print("в строке " + str(x) + "нечисловое значение")

		if q != None:
			print(cell_id)          
			wb.active = 1
			sheet_1 = wb.active
			for m in range(1, 3600):
				cell_sheet_1 = sheet_1.cell(row = m, column = 4).value
				if cell_id == str(cell_sheet_1):
					sheet_1.cell(row = m, column = 5).value = q
					sheet_1.cell(row = m, column = 7).value = p
					sheet_1.cell(row = m, column = 5).fill = PatternFill("solid", fgColor="008000")
					if len(set(a)) > 1:
						sheet_1.cell(row = m, column = 6).value = 1

	###################################################################################################################################################################
	def main_func(name_1, name_2, cell_id, stopword, name_3=100):
		# mark_in_our_price(cell_id)
		# mark_not_in_our_price(cell_id)
		wb.active = 0
		sheet_0 = wb.active
		rows, columns = sheet_0.max_row, sheet_0.max_column
		
		for i in range(1, main_loop + 1):
			cell_value_name_from_price = str(sheet_0.cell(row = i, column = 2).value) 
			for j in range(0, len(name_1)):		
				if name_1[j] in cell_value_name_from_price:
					for m in range(0, len(name_2)):
						if name_2[m] in cell_value_name_from_price:
							if name_3 != 100:
								for g in range(0, len(name_3)):
									if name_3[g] in cell_value_name_from_price:
										stop_word = re.search('|'.join(stopword), cell_value_name_from_price) != None

										if stop_word == False:
											if sheet_0.cell(row = i, column = 3).value != '' and sheet_0.cell(row = i, column = 3).value != None and sheet_0.cell(row = i, column = 3).value != 0:
												if str(sheet_0.cell(row = i, column = 3).value).isdigit():
													cell_value_from_price = round(int(sheet_0.cell(row = i, column = 3).value))                                  
													wb.active = 1
													sheet_1 = wb.active
													rows, columns = sheet_1.max_row, sheet_1.max_column													
													for k in range(1, rows):
														cell_value_name_from_template = str(sheet_1.cell(row = k, column = 4).value)														
														if cell_id == cell_value_name_from_template:
															if cell_value_from_price != sheet_1.cell(row = k, column = 5).value and cell_value_from_price != "": 
																if sheet_1.cell(row = k, column = 5).value == None:             
																	sheet_1.cell(row = k, column = 5).value = cell_value_from_price
																	sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
																	sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
																	dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
																	link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
																	sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
																	sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
																	dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
																	print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
																	break
																elif cell_value_from_price < int(sheet_1.cell(row = k, column = 5).value):
																	sheet_1.cell(row = k, column = 5).value = cell_value_from_price
																	sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
																	sheet_1.cell(row = k, column = 6).value = 1 
																	sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
																	must_rename = dict_for_rename.get(cell_value_name_from_template)
																	must_rename.value = ""
																	dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
																	link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
																	sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
																	sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
																	must_refill = dict_for_yellow_fill.get(cell_value_name_from_template)
																	must_refill.fill = PatternFill("solid", fgColor="FFFFFF")
																	dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
																	print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
																	break
																else:
																	sheet_1.cell(row = k, column = 6).value = 1
																	break

							if name_3 == 100:
								stop_word = re.search('|'.join(stopword), cell_value_name_from_price) != None
								if stop_word == False:
									if sheet_0.cell(row = i, column = 3).value != '' and sheet_0.cell(row = i, column = 3).value != None and sheet_0.cell(row = i, column = 3).value != 0:
										if str(sheet_0.cell(row = i, column = 3).value).isdigit():
											cell_value_from_price = round(int(sheet_0.cell(row = i, column = 3).value))
											wb.active = 1
											sheet_1 = wb.active
											rows, columns = sheet_1.max_row, sheet_1.max_column
											for k in range(1, rows):
												cell_value_name_from_template = str(sheet_1.cell(row = k, column = 4).value)
												if cell_id == cell_value_name_from_template:
													if cell_value_from_price != sheet_1.cell(row = k, column = 5).value and cell_value_from_price != "":
														if sheet_1.cell(row = k, column = 5).value == None:             
															sheet_1.cell(row = k, column = 5).value = cell_value_from_price
															sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
															sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
															link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
															sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
															sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
															dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
															dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
															print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
															break
														elif int(cell_value_from_price) < int(sheet_1.cell(row = k, column = 5).value):
															sheet_1.cell(row = k, column = 5).value = cell_value_from_price
															sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
															sheet_1.cell(row = k, column = 6).value = 1 
															sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
															must_rename = dict_for_rename.get(cell_value_name_from_template)
															must_rename.value = ""
															dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
															link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
															sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
															sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
															must_refill = dict_for_yellow_fill.get(cell_value_name_from_template)
															must_refill.fill = PatternFill("solid", fgColor="FFFFFF")
															dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
															print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
															break
														else:
															sheet_1.cell(row = k, column = 6).value = 1
															break
								
	##########################################################################################################################################


	def for_check_troubles(check_name, arg_1, arg_2, arg_3, arg_4, arg_5):
		wb.active = 1
		sheet_1 = wb.active
		rows, columns = sheet_1.max_row, sheet_1.max_column
		for i in range(1, rows):
			cell_name = str(sheet_1.cell(row = i, column = 4).value)
			cell_value = sheet_1.cell(row = i, column = 5).value

			if cell_name == check_name:
				if sheet_1.cell(row = i, column = 5).value == None:
					main_func(arg_1, arg_2, arg_3, arg_4, arg_5)

	#############################################################################################################################################
	uzi = ["узи", "ультразвук", "узд ", "ультрозвук"]
	id_uzi_vilo4kovoy_jelezi, uzi_vilo4kovoy_jelezi_2, stop_for_uzi_vilo4kovoy_jelezi = "УЗИ вилочковой железы", ["вилочко", "тимус"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_vn4s, uzi_vn4s_2, stop_for_uzi_vn4s = "УЗИ височно-нижнечелюстного сустава", ["височн", "нижнечелюстн", "нижне-челюстн"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_kisti, uzi_kisti_2, stop_for_uzi_kisti = "УЗИ кисти руки",  ["кисти", "кистей"], ["ребенк", "детя", "терапия", "воздей", "палец", "пальц", "втор", "под", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_limfouzlov, uzi_limfouzlov_2, stop_for_uzi_limfouzlov = "УЗИ лимфоузлов", ["лимф"], ["ребен", "дет", "молочн", "пункц", "щитовид", "терапия", "воздей", "под ", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_malogo_taza, uzi_malogo_taza_2, stop_for_uzi_malogo_taza = "УЗИ малого таза", ["таза", "гинекологическое"], ["ребен", "фолликул", "при наличии", "брюш", "щитовид", "плод", "триместр", "терапия", "воздей", "треместр", "беремен", "втор", "муж", "трузи", "недел", "после", "аборт", "контрол", "дет", "абдомин", "живот", "ваги", "влага", "ta", "воздей", "кавита", "беремен", "tv", "вен", "кг", "кардиограф", "фолликулометр", "под", "расшир"]
	id_uzi_malogo_taza_ta, uzi_malogo_taza__ta_2, uzi_malogo_taza__ta_3, stop_for_uzi_malogo_taza_ta = "УЗИ малого таза трансабдоминально", ["таза", "гинекологи", "матки и придатков", "женских половых органов"], ["абдомина", "живот", "ta"], ["ребен", "после", "аборт", "контрол", "втор", "терапия", "воздей", "дет", "под", "контрол", "возде", "кавита", "беремен"]
	id_uzi_malogo_taza_tv, uzi_malogo_taza__tv_2, uzi_malogo_taza__tv_3, stop_for_uzi_malogo_taza_tv = "УЗИ малого таза трансвагинально", ["таза", "гинекологи", "трансваг", "твузи", "женских половых органов"], ["ваги", "твузи", "влага", "tv", "тv"], ["ребен", "после", "терапия", "воздей", "аборт", "контрол", "втор", "дет", "под", "воздей", "кавита", "контрол", "беремен"]
	id_uzi_matki, uzi_matki_2, stop_for_uzi_matki = "УЗИ матки", ["матк"], ["ребен", "дет", "трим", "шей", "орошен", "контрол", "под", "рубц", "воздейс", "терапия", "воздей", "кавита"]
	id_uzi_molo4nih_jelez, uzi_molo4nih_jelez_2, stop_for_uzi_molo4nih_jelez, uzi_molo4nih_jelez_3 = "УЗИ молочных желез", ["молочн", "груд"], ["ребен", "дет", "лимф", "втор", "комплекс", "почк", "мочев", "муж", "пункц", "под ", "девоч", "терапия", "воздей", "возде", "кавита", "допол", "до 1", "до 2", "до 3", "прием"], ["желез", "желёз"]
	id_uzi_molo4nih_jelez_limfouzel, uzi_molo4nih_jelez_limfouzel_2, uzi_molo4nih_jelez_limfouzel_3, stop_for_uzi_molo4nih_jelez_limfouzel = "УЗИ молочных желез с лимфоузлами", ["грудных желез", "молочных желез", "молочной железы", "грудной железы"], ["лимф"], ["ребен", "дет", "прием", "девоч", "терапия", "воздей", "пункц", "под", "до 1", "до 2", "до 3"]
	id_uzi_mo4evogo_puzirya, uzi_mo4evogo_puzirya_2, stop_for_uzi_mo4evogo_puzirya = "УЗИ мочевого пузыря", ["мочево"], ["ребен", "дет", "под", "контрол", "пункц", "биопси", "остаточно", "почек", "почками", "почки", "терапия", "воздей", "мочи", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"] 
	id_uzi_mo4evogo_puzirya_mo4a, uzi_mo4evogo_puzirya_mo4a_2, uzi_mo4evogo_puzirya_mo4a_3, stop_for_uzi_mo4evogo_puzirya_mo4a = "УЗИ мочевого пузыря с определением остаточной мочи", ["мочево"], ["мочи", "мочой", "моча"], ["ребен", "дет", "терапия", "воздей", "под", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"] 
	id_uzi_mo4eto4nikov, uzi_mo4eto4nikov_2, stop_for_uzi_mo4eto4nikov = "УЗИ мочеточников", ["мочеточн"], ["ребен", "дет", "выброс", "под", "терапия", "воздей", "пункц", "контрол", "биопси", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_moshonki, uzi_moshonki_2, stop_for_uzi_moshonki = "УЗИ мошонки", ["мошонк"], ["ребен", "дет", "под", "контрол", "пункц", "терапия", "воздей", "пунктирование", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_parashitovidnih, uzi_parashitovidnih_2, uzi_parashitovidnih_3, stop_for_uzi_parashitovidnih, stop_for_uzi_parashitovidnih_2 = "УЗИ паращитовидных желез", ["пара", "около"], ["щитовид"], ["ребен", "дет", "щитовидной", "пункц", "под"], ["ребен", "терапия", "воздей", "дет", "пункц", "пунктирование", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_4lena, uzi_4lena_2, stop_for_uzi_4lena = "УЗИ полового члена", ["член"], ["ребен", "дет", "лонн", "сочлен", "сосуд", "артери", " вен", "под", "допплер", "доплер", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_pridat4nih, uzi_pridat4nih_2, stop_for_uzi_pridat4nih = "УЗИ придаточных пазух носа", ["придаточн", "пазух"], ["ребен", "дет", "под", "терапия", "воздей", "контрол"]
	id_uzi_prostati, uzi_prostati_2, stop_for_uzi_prostati = "УЗИ простаты / предстательной железы", ["простат", "предста"], ["ребен", "дет", "ректа", "под", "контрол", "терапия", "воздей", "пункц", "бипси", "трузи", "воздей", "кавита", "tr", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_truzi, uzi_truzi_2, uzi_truzi_3, stop_for_uzi_truzi = "ТРУЗИ простаты / предстательной железы", ["простат", "предста", "трузи"], ["ректа", "датчик", " tr", "трузи"], ["ребен", "экстру", "протрузи", "мат", "влаг", "придат", "терапия", "воздей", "дет", "навигация", "наведение", "абдомин", "живот", "под", "возде", "кавита"]
	id_uzi_slunnih, uzi_slunnih, stop_for_uzi_slunnih = "УЗИ слюнных желез", ["слюн"], ["ребен", "дет", "под ", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_sredosteniya, uzi_sredosteniya, stop_for_uzi_sredosteniya = "УЗИ средостения", ["средостен"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_stopi, uzi_stopi, stop_for_uzi_stopi = "УЗИ стопы", ["стопы"], ["ребен", "дет", "втор", "палец", "пальц", "под", "терапия", "воздей", "нерв", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_suhojil, uzi_suhojil, stop_for_uzi_suhojil = "УЗИ сухожилия", ["сухожили"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_uretri, uzi_uretri, stop_for_uzi_uretri = "УЗИ уретры", ["уретры"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_shitovid, uzi_shitovid, uzi_shitovid_2, stop_for_uzi_shitovid, stop_for_uzi_shitovid_2 = "УЗИ щитовидной железы", ["щито"], ["видн"], ["паращитовидн", "ребен", "дет", "пункц", "около", "школ", "брю", "дополни"], ["ребен", "дополни", "терапия", "воздей", "дет", "пункц", "под", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_yai4nikov, uzi_yai4nikov, stop_for_uzi_yai4nikov = "УЗИ яичников", ["яичник"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_follikulometriya, follikulometriya, follikulometriya_2, stop_for_uzi_follikulometriya = "Фолликулометрия", ["фолликулометрия", "фолликулогенез", "фоликулометрия", "фоликулогенез", "мониторинг", "мониторинг фолликул"], ["фолликулометрия", "фоликулометрия", "созревания фолликул", "фоликулогенез", "фолликулогенез", "мониторинг фолликул"], ["ребен", "дет", "терапия", "воздей", "под", "втор", "при", "для", "во время", "в ходе"]
	id_uzi_brushnoy, uzi_brushnoy_1, uzi_brushnoy_2, stop_for_uzi_brushnoy = "УЗИ брюшной полости", ["брюшн"], ["полост"], ["ребен", "дет", "органа", "1 орган", "один орган", "2 органа", "двух органов", "2-х органов", "под контролем", "под уз", "под ультра", "гематом", "абсцес", "плев", "жидко", "наличи", "свобод", "школ", "акц", "терапия", "воздей", "сердц", "щитовид", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_jel4nogo, uzi_jel4nogo, uzi_jel4nogo_2, stop_for_uzi_jel4nogo, stop_for_uzi_jel4nogo_2  = "УЗИ желчного пузыря", ["желчн"], ["желч"], ["ребен", "дет", "терапия", "воздей", "с исследованием функции", "с функци", "под ", "контрол", "брюшной полости комплексное", "селез"], ["ребен", "дет", "под", "контрол", "брюшной полости комплексное", "селез", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_zabrushinnogo, uzi_zabrushinnogo, stop_for_uzi_zabrushinnogo = "УЗИ забрюшинного пространства", ["забрюшинн"], ["ребен", "дет", "терапия", "воздей", "под ", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_nadpo4e4, uzi_nadpo4e4, stop_for_uzi_nadpo4e4 = "УЗИ надпочечников", ["надпочеч"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_pe4eni, uzi_pe4eni, stop_for_uzi_pe4eni = "УЗИ печени", ["печени"], ["ребен", "дет", "под", "контрол", "обеспеч", "желчного", "желчным", "терапия", "воздей", "желчный", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_podjeludo4, uzi_podjeludo4, stop_for_uzi_podjeludo4 = "УЗИ поджелудочной железы", ["поджелудочной"], ["ребен", "дет", "наведе", "пункц", "терапия", "воздей", "навигац", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_po4ek, uzi_po4ek, stop_for_uzi_po4ek = "УЗИ почек", ["почек"], ["ребен", "дет", "под", "контрол", "щитовид", "без почек", "пузыр", "мочев", "надпочеч", "терапия", "воздей", "комплексное узи органов", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_selezen, uzi_selezen, stop_for_uzi_selezen = "УЗИ селезенки", ["селезенк"], ["ребен", "дет", "под", "терапия", "воздей", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	id_uzi_funkcii_jel4nogo, uzi_funkcii_jel4nogo_1, uzi_funkcii_jel4nogo_2, stop_for_uzi_funkcii_jel4nogo = "УЗИ функции желчного пузыря", ["функци", "сократи"], ["желч", "желоч"], ["ребен", "дет", "под", "контрол", "терапия", "воздей", "без исследования функции", "без функци"]
	id_uzi_brushnoy_rebenku, uzi_brushnoy_rebenku_1, uzi_brushnoy_rebenku_2, stop_for_uzi_brushnoy_rebenku = "УЗИ брюшной полости ребенку", ["брюшной"], ["ребен", "детям", "детей"], ["диспан", "комплекс", "нейро", "терапия", "воздей", "под", "контрол"]
	id_uzi_vilo4kovoy_rebenku, uzi_vilo4kovoy_rebenku_1, uzi_vilo4kovoy_rebenku_2, stop_for_uzi_vilo4kovoy_rebenku = "УЗИ вилочковой железы ребенку", ["вилочко"], ["ребен", "дет"], ["диспан", "терапия", "воздей", "под", "контрол"]
	id_uzi_jel4nogo_rebenku, uzi_jel4nogo_rebenku_1, uzi_jel4nogo_rebenku_2, stop_for_uzi_jel4nogo_rebenku = "УЗИ желчного пузыря ребенку", ["желчного"], ["ребен", "дет"], ["функц", "диспан", "терапия", "воздей", "под", "контрол"]
	id_uzi_kolen_rebenku, uzi_kolen_rebenku_1, uzi_kolen_rebenku_2, stop_for_uzi_kolen_rebenku = "УЗИ коленных суставов ребенку", ["колен"], ["ребен", "дет"], ["диспан", "под", "терапия", "воздей", "контрол"]
	id_uzi_limfouzlov_rebenku, uzi_limfouzlov_rebenku_1, uzi_limfouzlov_rebenku_2, stop_for_uzi_limfouzlov_rebenku = "УЗИ лимфоузлов ребенку", ["лимфоузл", "лимфаузл", "лимфатических узлов", "лимфатическими узлами", "лимфотических узлов", "лимфотическими узлами"], ["ребен", "дет"], ["щитовид", "брюшн", "комплекс", "диспан", "под", "контрол"]
	id_uzi_kolen_rebenku, uzi_kolen_rebenku_1, uzi_kolen_rebenku_2, stop_for_uzi_kolen_rebenku = "УЗИ коленных суставов ребенку", ["колен"], ["ребен", "дет"], ["диспан", "терапия", "воздей", "под", "контрол"]
	id_uzi_taza_rebenku, uzi_taza_rebenku_1, uzi_taza_rebenku_2, stop_for_uzi_taza_rebenku = "УЗИ малого таза ребенку", ["таза", "тазов", "гинекологическое"], ["ребен", "дет"], ["фолликул", "втор", "преры", "терапия", "воздей", "бере", "контрол", "под", "под", "контрол"]
	uzi_molo4noy_rebenku, id_uzi_molo4noy_rebenku, stop_for_uzi_molo4noy_rebenku, uzi_molo4noy_rebenku_2 = ["молочн"], "УЗИ молочной железы ребенку", ["диспан", "под", "терапия", "воздей", "контрол"], ["ребен", "дет"]
	uzi_mo4evogo_rebenku, id_uzi_mo4evogo_rebenku, stop_for_uzi_mo4evogo_rebenku, uzi_mo4evogo_rebenku_2 = ["мочево"], "УЗИ мочевого пузыря ребенку", ["диспан", "под", "комплекс", "терапия", "воздей", "контрол"], ["ребен", "дет"]
	uzi_moshonki_rebenku, id_uzi_moshonki_rebenku, stop_for_uzi_moshonki_rebenku, uzi_moshonki_rebenku_2 = ["мошонк"], "УЗИ мошонки ребенку", ["диспан", "под", "контрол"], ["ребен", "дет"]
	uzi_m9gkih_rebenku, id_uzi_m9gkih_rebenku, stop_for_uzi_m9gkih_rebenku, uzi_m9gkih_rebenku_2 = ["мягк"], "УЗИ мягких тканей ребенку", ["под", "терапия", "воздей", "контрол"], ["ребен", "дет"]
	uzi_nadpo4e4_rebenku, id_uzi_nadpo4e4_rebenku, stop_for_uzi_nadpo4e4_rebenku, uzi_nadpo4e4_rebenku_2 = ["надпочечник"], "УЗИ надпочечников ребенку", ["диспан", "терапия", "воздей", "под", "контрол"], ["ребен", "дет"]
	uzi_pe4eni_rebenku, id_uzi_pe4eni_rebenku, stop_for_uzi_pe4eni_rebenku, uzi_pe4eni_rebenku_2 = ["печен"], "УЗИ печени ребенку", ["диспан", "под", "контрол"], ["ребен", "дет",]
	uzi_podjelud_rebenku, id_uzi_podjelud_rebenku, stop_for_uzi_podjelud_rebenku, uzi_podjelud_rebenku_2 = ["поджелудочн"], "УЗИ поджелудочной железы ребенку", ["диспан", "под", "терапия", "воздей", "контрол"], ["ребен", "дет"]
	uzi_pozvono4_rebenku, id_uzi_pozvono4_rebenku, stop_for_uzi_pozvono4_rebenku, uzi_pozvono4_rebenku_2 = ["позвоночник"], "УЗИ позвоночника ребенку", ["шейн", "под", "терапия", "воздей", "контрол"], ["ребен", "дет"]
	uzi_po4ek_rebenku, id_uzi_po4ek_rebenku, stop_for_uzi_po4ek_rebenku, uzi_po4ek_rebenku_2 = ["почек", "почк"], "УЗИ почек ребенку", ["диспан", "под", "комплекс", "терапия", "воздей", "контрол"], ["ребен", "дет"]
	uzi_pazuh_rebenku, id_uzi_pazuh_rebenku, stop_for_uzi_pazuh_rebenku, uzi_pazuh_rebenku_2 = ["пазух"], "УЗИ придаточных пазух носа ребенку", ["диспан", "терапия", "воздей", "под", "контрол"], ["ребен", "дет"]
	uzi_selezen_rebenku, id_uzi_selezen_rebenku, stop_for_uzi_selezen_rebenku, uzi_selezen_rebenku_2 = ["селезен"], "УЗИ селезенки ребенку", ["диспан", "под", "терапия", "воздей", "контрол"], ["ребен", "дет"]
	uzi_serdca_rebenku, uzi_serdca_rebenku_2, id_uzi_serdca_rebenku, stop_for_uzi_serdca_rebenku, uzi_serdca_rebenku_3 = ["эхо", "узи", "ультразвук"], ["кг", "кардиограф", "серд"], "УЗИ сердца (ЭхоКГ) ребенку", ["диспан", "комплекс", "нейро", "под", "контрол"], ["ребен", "дет"]
	uzi_tazobedr_rebenku, id_uzi_tazobedr_rebenku, stop_for_uzi_tazobedr_rebenku, uzi_tazobedr_rebenku_2 = ["тазобедр"], "УЗИ тазобедренного сустава ребенку", ["втор", "терапия", "воздей", "под", "контрол"], ["ребен", "дет", "новорожд", "младе"]
	uzi_shenyogo_rebenku, id_uzi_shenyogo_rebenku, stop_for_uzi_shenyogo_rebenku, uzi_shenyogo_rebenku_2 = ["шейн"], "УЗИ шейного отдела позвоночника ребенку", ["диспан", "терапия", "воздей", "под", "контрол"], ["ребен", "дет"]
	uzi_shitovid_rebenku, id_uzi_shitovid_rebenku, stop_for_uzi_shitovid_rebenku, uzi_shitovid_rebenku_2 = ["щитовидн"], "УЗИ щитовидной железы ребенку", ["диспан", "под", "терапия", "воздей", "комплекс", "контрол"], ["ребен", "дет"]
	uzi_m9gkih_tkaney, id_uzi_m9gkih_tkaney, stop_for_uzi_m9gkih_tkaney, uzi_m9gkih_tkaney_2, stop_for_uzi_m9gkih_tkaney_2  = ["мягк"], "УЗИ мягких тканей", ["ребен", "дет", "терапия", "мошо", "чле", "воздей", "лиц", "образ", "кератом", "липо", "фибро", "папил", "шов", "швов", "швы", "шеи", "шея", "шейн", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["ткан", "област", "зон"], ["ребен", "дет"]
	uzi_m9gkih_tkaney_lica, id_uzi_m9gkih_tkaney_lica, stop_for_uzi_m9gkih_tkaney_lica, id_uzi_m9gkih_tkaney_lica_2 = ["мягк"], "УЗИ мягких тканей лица", ["ребен", "дет", "терапия", "воздей", "шеи", "шея", "шейн", "втор", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["лиц"]
	uzi_m9gkih_tkaney_shei, id_uzi_m9gkih_tkaney_shei, stop_for_uzi_m9gkih_tkaney_shei, id_uzi_m9gkih_tkaney_shei_2 = ["мягк"], "УЗИ мягких тканей шеи", ["ребен", "дет", "терапия", "воздей", "лица", "втор", "комплекс", "узл", "лимф", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["шеи", "шейн", "шея"]
	uzi_m9gkih_plevri, id_uzi_plevri, stop_for_uzi_plevri = ["плевр"], "УЗИ плевральной полости", ["ребен", "терапия", "воздей", "дет", "дому", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	uzi_pozvono4, id_uzi_pozvono4, stop_for_uzi_pozvono4 = ["позвоночн"], "УЗИ позвоночника", ["ребен", "дет", "транс", "карни", "крани", "шей", "протру", "груд", "терапия", "воздей", "поясн", "отдела", "втор", "артерий", "сосудов", "вен", "втор", "навигац", "блок", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"] 
	uzi_grudnogo_pozvono4, id_uzi_grudnogo_pozvono4, stop_for_uzi_grudnogo_pozvono4, uzi_grudnogo_pozvono4_2 = ["позвоноч"], "УЗИ грудного отдела позвоночника", ["ребен", "дет", "терапия", "экстру", "протру", "воздей", "втор", "навигац", "блок", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["грудн", "одного отдела", "один отдел", "1 отдел"]
	uzi_kop4ik, id_uzi_kop4ik, stop_for_uzi_kop4ik, uzi_kop4ik_2 = ["копч"], "УЗИ копчика", ["ребен", "дет", "терапия", "воздей", "навигац", "блок", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["копчи", "одного отдела", "один отдел", "1 отдел"]
	uzi_po9snic, id_uzi_po9snic, stop_for_uzi_po9snic, uzi_po9snic_2 = ["позвоноч"], "УЗИ поясничного отдела позвоночника", ["ребен", "дет", "терапия", "воздей", "втор", "навигац", "экстру", "протру", "блок", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["поясни", "одного отдела", "1 отдел", "один отдел"]
	uzi_sheynogo_pozvon, id_uzi_sheynogo_pozvon, stop_for_uzi_psheynogo_pozvon, uzi_sheynogo_pozvon_2 = ["позвоноч"], "УЗИ шейного отдела позвоночника", ["ребен", "дет", "втор", "терапия", "воздей", "навигац", "экстру", "протру", "блок"], ["шейн", "1 отдел", "одного отдела", "один отдел"]
	uzi_serdca, uzi_serdca_2, id_uzi_serdca, stop_for_uzi_serdca = ["эхо", "узи", "ультразвук"], ["кг", "кардиограф", "серд"], "УЗИ сердца / ЭхоКГ", ["ребен", "кроме", "биени", "по результату", "дом", "перед", "терапия", "воздей", "для", "дет", "нагруз", "школ", "стресс", "пище", "плод", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	stress_eho, stress_eho_2, id_stress_eho, stop_for_stress_eho, stress_eho_3 = ["эхо", "узи", "ультразвук"], ["стресс", "нагрузк"], "Стресс-эхокардиография / Стресс ЭхоКГ", ["ребен", "дет", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["кг", "кардио"]
	uzi_perikarda, id_uzi_perikarda, stop_for_uzi_perikarda = ["перикард"], "УЗИ перикарда", ["ребен", "дет", "под", "терапия", "воздей", "контрол", "пункц", "пунктирование", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	chrespishevod_eho, chrespishevod_eho_2, id_chrespishevod_eho, stop_for_chrespishevod_eho, chrespishevod_eho_3 = ["пищево"], ["эхо"], "Чреспищеводная эхокардиография", ["ребен", "терапия", "воздей", "дет", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"], ["кг", "кардио"]
	uzi_sustavov, id_uzi_sustavov, stop_for_uzi_sustavov = ["сустав"], "УЗИ суставов", ["ребен", "дет", "височ", "челюст", "нерв", "колен", "локт", "голеност", "терапия", "воздей", "лучеза", "плече", "тазо", "стоп", "пункц", "наведен", "навигац", "контрол", "под", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	uzi_golenostop, id_uzi_golenostop, stop_for_uzi_golenostop = ["голеностоп", "парных суставов", "крупных суставов", "крупного сустава", "крупные суставы"], "УЗИ голеностопного сустава", ["ребен", "нерв", "дет", "терапия", "воздей", "втор", "под", "контрол", "пункц", "пунктирование", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	uzi_kolen, id_uzi_kolen, stop_for_uzi_kolen = ["колен", "парных суставов", "крупных суставов", "крупного сустава", "крупные суставы"], "УЗИ коленного сустава", ["ребен", "дет", "втор", "под", "терапия", "воздей", "нерв", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	uzi_loktevogo, id_uzi_loktevogo, stop_for_uzi_loktevogo = ["локт", "локот", "парных суставов", "крупных суставов", "крупного сустава", "крупные суставы"], "УЗИ локтевого сустава", ["ребен", "терапия", "воздей", "дет", "нерв", "втор", "под", "контрол", "пункц", "пунктирование", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	uzi_lu4ezap, id_uzi_lu4ezap, stop_for_uzi_lu4ezap = ["лучезап", "запяст", "парных суставов", "крупных суставов", "крупного сустава", "крупные суставы"], "УЗИ лучезапястного сустава", ["ребен", "терапия", "воздей", "дет", "нерв", "втор", "под", "контрол", "пункц", "пунктирование", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	uzi_ple4, id_uzi_ple4, stop_for_uzi_ple4 = ["плечевого", "сустава плеча", "плечевых", "парных суставов", "плечевой", "крупных суставов", "крупного сустава", "крупные суставы"], "УЗИ плечевого сустава", ["ребен", "терапия", "воздей", "нерв", "дет", "втор", "вен", "до плеча", "дуплекс", "под", "пункц", "контрол", "сплетен", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	uzi_tazobedr, id_uzi_tazobedr, stop_for_uzi_tazobedr = ["тазобедр"], "УЗИ тазобедренного сустава", ["ребен", "терапия", "воздей", "дет", "втор", "нерв", "рожд", "млад", "пункц", "под", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	dupleks, tripleks, uzdg = ["дуплекс", "уздс", "дс ", "уздc"], ["триплекс"], ["допплерограф", "доплерограф", "уздг", "доплер", "допплер"]
	dupleks_aorti, id_dupleks_aorti, stop_for_dupleks_aorti = ["аорт"], "Дуплексное сканирование аорты", ["ребен", "дет", "под", "терапия", "воздей", "контрол"]
	dupleks_arteriy_verh, id_dupleks_arteriy_verh, stop_for_dupleks_arteriy_verh, dupleks_arteriy_verh_2 = ["артери"], "Дуплексное сканирование артерий верхних конечностей", ["ребен", "перед", "терапия", "воздей", "дет", "вен", "одного", "бры", "чрев"], ["верх", "рук"]
	dupleks_arteriy_niz, id_dupleks_arteriy_niz, stop_for_dupleks_arteriy_niz, dupleks_arteriy_niz_2 = ["артери"], "Дуплексное сканирование артерий нижних конечностей", ["ребен", "дет", "вен", "терапия", "воздей", "перед", "брюшного", "ного", "голов", "чрев", "одного", "бры", "чрев"], ["низ", "ниж", " ног"]
	dupleks_brahiocefal, id_dupleks_brahiocefal, stop_for_brahiocefal = ["брахиоцефал", "брахицефал"], "Дуплексное сканирование брахиоцефальных артерий", ["ребен", "терапия", "воздей", "дет", "под", "контрол"]
	dupleks_ven_verh, id_dupleks_ven_verh, stop_for_dupleks_ven_verh, dupleks_ven_verh_2 = ["вен"], "Дуплексное сканирование вен верхних конечностей", ["ребен", "дет", "терапия", "воздей", "перед", "артери", "одного", "бры"], ["верх", "рук"]
	dupleks_ven_niz, id_dupleks_ven_niz, stop_for_dupleks_ven_niz, dupleks_ven_niz_2 = ["вен"], "Дуплексное сканирование вен нижних конечностей", ["ребен", "дет", "терапия", "воздей", "перед", "нижней полой вены", "портальной", "брюшного", "артери", "ного", "чрев", "головн", "грозд", "тестикул", "нижняя", "нижней", "одного"], ["низ", "ниж", " ног"]
	dupleks_uzds, id_dupleks_uzds, stop_for_dupleks_uzds = ["голов"], "Дуплексное сканирование сосудов головы / УЗДС", ["ребен", "терапия", "воздей", "дет", "под", "контрол", "мозг", "ткдг", "транскарни", "транскрани"]
	dupleks_po4ek, id_dupleks_po4ek, stop_for_dupleks_po4ek = ["почек", "почеч"], "Дуплексное сканирование сосудов почек", ["ребен", "терапия", "воздей", "дет", "под", "контрол"]
	dupleks_shei, id_dupleks_shei, stop_for_dupleks_shei = ["шеи"], "Дуплексное сканирование сосудов шеи / УЗДС", ["ребен", "терапия", "воздей", "дет", "под", "контрол", "до 1", "до 2", "до 3", "до 18", "до 12", "до 10", "до 4", "до 5", "до 6"]
	tripleks_arteriy_verh, id_tripleks_arteriy_verh, stop_for_tripleks_arteriy_verh, tripleks_arteriy_verh_2 = ["артери", "сосуд"], "Триплексное сканирование артерий верхних конечностей", ["ребен", "дет", "вис"], ["верх", "рук"]
	tripleks_arteriy_niz, id_tripleks_arteriy_niz, stop_for_tripleks_arteriy_niz, tripleks_arteriy_niz_2 = ["артери", "сосуд"], "Триплексное сканирование артерий нижних конечностей", ["ребен", "дет", "под", "голов", "чрев", "аорт", "контрол"], ["низ", "ниж", " ног"]
	tripleks_ven_verh, id_tripleks_ven_verh, stop_for_tripleks_ven_verh, tripleks_ven_verh_2 = ["вен", "сосуд"], "Триплексное сканирование вен верхних конечностей", ["ребен", "дет", "под", "контрол"], ["верх", "рук"]
	tripleks_ven_niz, id_tripleks_ven_niz, stop_for_tripleks_ven_niz, tripleks_ven_niz_2 = ["вен", "сосуд"], "Триплексное сканирование вен нижних конечностей", ["ребен", "дет", "голов", "чрев", "аорт", "полой", "поч"], ["низ", "ниж", " ног"]
	tripleks_sosudov_golov_i_shei, id_tripleks_sosudov_golov_i_shei, stop_for_tripleks_sosudov_golov_i_shei, tripleks_sosudov_golov_i_shei_2 = ["голов"], "Триплексное сканирование сосудов головы и шеи", ["ребен", "дет", "под", "контрол"], ["шеи", "шейн", "шея", "брахиоцефал"]
	uzdg_shitovid, id_uzdg_shitovid, stop_for_uzdg_shitovid = ["щитовидн"], "Допплерография щитовидной железы", ["ребен", "дет", "узи", "ультразвук"]
	tkdg, id_tkdg, stop_for_tkdg, tkdg_2 = ["транскрани", "ткдг", "транскарни", "ткдс"], "ТКДГ сосудов головного мозга", ["ребен", "дет", "под", "контрол", "стимуляц"], ["мозг", "допплеро", "артерий", "вен", "ткдс", "дуплекс"]
	uzdg_sosudov_golovi, id_uzdg_sosudov_golovi, stop_for_uzdg_sosudov_golovi = ["голов", "мозг"], "УЗДГ / Допплерография сосудов головы", ["ребен", "дет", "под", "контрол", "ткдг", "транскарн", "транскрани", "мозг"]
	uzdg_sosudov_arteriy_niz, id_uzdg_arteriy_niz, stop_for_uzdg_arteriy_niz, uzdg_arteriy_niz_2 = ["артери", "сосуд"], "УЗДГ артерий нижних конечностей", ["ребен", "дет", "семен", "голов", "брюшного", "под", "контрол", "голов"], ["низ", "ниж", " ног"]
	uzdg_brushn_aort, id_uzdg_brushn_aort, stop_for_uzdg_brushn_aort, uzdg_brushn_aort_2 = ["брюш"], "УЗДГ брюшной аорты", ["ребен", "дет", "полост", "втор", "курс"], ["аорт"]
	uzdg_sosudov_ven_verh, id_uzdg_ven_verh, stop_for_ven_verh, uzdg_ven_verh_2 = ["вен", "сосуд"], "УЗДГ вен верхних конечностей", ["ребен", "дет", "под", "контрол", "полост", "анастомо", "портал"], ["верх", "рук"]
	uzdg_sosudov_ven_niz, id_uzdg_ven_niz, stop_for_ven_niz, uzdg_ven_niz_2 = ["вен", "сосуд"], "УЗДГ вен нижних конечностей", ["ребен", "семен", "голов", "дет", "под", "нижней", "контрол", "голов"], ["низ", "ниж", " ног"]
	uzdg_glaz_sosud, id_uzdg_glaz_sosud, stop_for_glaz_sosud, uzdg_glaz_sosud_2 = ["глаз"], "УЗДГ глазных сосудов", ["ребен", "дет", "под", "контрол"], ["сосуд"]
	uzdg_limf_sosud, uzdg_limf_sosud_2, id_uzdg_limf_sosud, stop_for_limf_sosud = ["лимф"], ["сосуд"], "УЗДГ лимфатических сосудов", ["ребен", "дет", "молоч", "щитовид"]
	uzdg_sosud_brush_polo, id_uzdg_sosud_brush_polo, stop_for_sosud_brush_polo = ["брюшн"], "УЗДГ сосудов брюшной полости", ["аорт"]
	uzdg_sosud_molo4, id_uzdg_molo4, stop_for_molo4 = ["молочн"], "УЗДГ сосудов молочной железы", ["заглушка"]
	uzdg_sosud_po4ek, id_uzdg_po4ek, stop_for_po4ek = ["почек", "почечн"], "УЗДГ сосудов почек", ["надпочеч"]
	uzdg_ekstrakranial, id_uzdg_uzdg_ekstrakranial, stop_for_uzdg_ekstrakranial = ["экстракрани"], "УЗДГ экстракраниальных сосудов", ["заглушка"]
	sonoelastograf, sonoelastograf_2, id_sonoelastograf, stop_for_sonoelastograf = ["соно", "сэг"], ["ластограф"], "Соноэластография", ["молоч", "печен", "допол"]
	uzgss_mato4nih, mato4nih_2, id_mato4nih, stop_for_mato4nih = ["узгс", "узи", "ультразвук", "эхо", "соногистеросальпингоскопия", "соногистеросальпингография", "соносальпингоскопия", "проходимости", "соносальпингографи", "эхогистеросальпингоскопия", "эхо-гсг", "эхографическая гсг", "гидросонография"], [" маточн", "эхо-гсг", "эхографическая гсг", "труб", "гсс", "соносальпингографи", "соногистеросальпингография", "эхогистеросальпингоскопия", "гистеро", "соногистеросальпингоскопия", "соносальпингоскопия", "гидросонография"], "УЗГСС / УЗИ проходимости маточных труб", ["кровоток", "операц", "подгото", "евста", "пуп", "двух", "подт", "яйц", "артери", "сосуд", "слух", "плацент", "восстанов", "опера", "турбац", "тубация", "тубации", "обследование для", "осмотр для", "исследование для", "стерили", "создание", "лапаротом", "лапароскоп"]
	elastograf_molo4, elastograf_molo4_2, id_elastograf_molo4, stop_for_elastograf_molo4 = ["эластограф", "сэг", "эластометрия"], ["молоч"], "Эластография молочных желез", ["ребе", "дет", "мрт", "магнит", "мр", "мр-", "мр -", "дополни"]
	elastograf_pe4en, elastograf_pe4en_2, id_elastograf_pe4en, stop_for_elastograf_pe4en = ["эластограф", "сэг", "эластометрия"], ["печен"], "Эластография печени", ["ребе", "втор", "дет"]
	uzi_glaza, id_uzi_glaza, stop_for_uzi_glaza = ["глаз", "орбит", "а-в сканирование"], "УЗИ глаза", ["биометр", "груз", "давле", "лифт", "пери", "дейс", "лин", "после", "пред"]
	neirosonografia, neirosonografia_2, id_neirosonografia, stop_for_neirosonograf = ["нейро", "узи", "ультразвук"], ["соно", "мозга"], "Нейросонография / УЗИ головного мозга ребенку", ["перфуз", "лизис", "под контролем", "яич", "овари", "спасти", "инсул", "склеро", "щитовид", "эластограф", "допол", "черн", "субстанц", "паркинс", "транскрани", "магнит", "терапия", "воздей", "мрт", "терапи", "сонопульс", "аппарат", "зона", "мр", "эндосонограф", "эндоузи"]
	fetometria, fetometria_2, id_fetometria, stop_for_fetometria = ["фето"], ["метр"], "Фетометрия", ["кровоток" ,"плацент", "втор", "контр", "пола плода", "без фетометрии", "не включая фетометри", "прием", "консул"]
	kolposkopia, kolposkopia_2, id_kolposkopia, stop_for_kolposkopia = ["кольпо"], ["скоп"], "Кольпоскопия", ["видео", "реб", "дет", "регистр", "контр", "втор," "запись", "данн", "вульво", "прием", "консул"]
	kolposkopia_video, kolposkopia_video_2, id_kolposkopia_video, stop_for_kolposkopia_video, kolposkopia_video_3 = ["кольпо"], ["скоп"], "Видеокольпоскопия", ["реб", "дет", "в рамках", "в ходе кон", "прием", "без вид", "без кольп"], ["видео"]
	anoskop, anoskop_2, id_anoskop, stop_for_anoskop = ["аноскопи"], ["аноскопи"], "Аноскопия", ["ректор", "реб", "дет", "диафан", "контро", "втор", "готов", "регистр", "запись", "данн", "просвет", "восстанов", "прием", "консул"]
	rektoromanoskop, rektoromanoskop_2, id_rektoromanoskop, stop_for_rektoromanoskop = ["ректор"], ["скоп"], "Ректороманоскопия", ["реб", "дет", "втор", "конт", "готов", "диафан", "регистр", "запись", "данн", "прием", "консул"]
	rektoromanoskop_det, rektoromanoskop_det_2, id_rektoromanoskop_det, stop_for_rektoromanoskop_det, rektoromanoskop_det_3 = ["ректор"], ["скоп"], "Ректороманоскопия ребенку", ["диафан", "втор", "контр", "прием", "консул"], ["реб", "дет"]
	bronhoskop, bronhoskop_2, id_bronhoskop, stop_for_bronhoskop = ["бронх"], ["скоп"], "Бронхоскопия", ["реб", "дет", "биоп", "мокрот", "просвет", "восстан", "забор", "взят", "томограф", "кт ", "кт-", "кт - ", "кт- "]
	vulvoskop, vulvoskop_2, id_vulvoskop, stop_for_vulvoskop = ["вульв"], ["скоп"], "Вульвоскопия", ["реб", "дет", "регистр", "запись", "данн"]
	diagnoz_gisteroskop, diagnoz_gisterosko_2, id_diagnoz_gisterosko, stop_for_diagnoz_gisterosko = ["диагно"], ["гистеро"], "Диагностическая гистероскопия", ["реб", "дет", "рдв", "выскабл", "раздел", "сальпи", "прием", "консул"]
	laringoskop, laringoskop_2, id_laringoskop, stop_for_laringoskop = ["ларингоскопия", "эндоскоп"], ["ларингоскопия", "гортан", "глотк", "глодк", "лор-органов", "лор органов"], "Ларингоскопия", ["микро", "инород", "биопси", "мини", "реб", "дет", "регистр", "запись", "данн", "прием", "консул"]
	mediastinoskop, mediastinoskop_2, id_mediastinoskop, stop_for_mediastinoskop = ["медиастиноскоп", "эндоскоп"], ["медиа", "средосте"], "Медиастиноскопия", ["реб", "дет", "регистр", "запись", "данн", "прием", "консул"]
	bioimpedansometria, bioimpedansometria_2, id_bioimpedansometria, stop_for_bioimpedansometria = ["биоимпеданс", "биоимпенданс"], ["анализ", "метр", "графи", "состав", "биоимпеданс", "биоимпенданс"], "Биоимпедансометрия", ["реб", "ух", "лор", "уш", "втор", "дет", "прием", "консул"]
	audiogramma, audiogramma_2, id_audiogramma, stop_for_audiogramma = ["аудио"], ["грамм", "метр"], "Аудиометрия / Аудиограмма", ["реб", "дет", "вебер", "лмк", "порог", "прием", "консул"]
	audiogramma_weber, audiogramma_weber_2, id_audiogramma_weber, stop_for_audiogramma_weber, audiogramma_weber_3 = ["аудио"], ["грамм", "метр"], "Аудиометрия с тестом Вебера", ["реб", "дет", "порог", "прием", "консул"], ["вебер"]
	audiogramma_porog, audiogramma_porog_2, id_audiogramma_porog, stop_for_audiogramma_porog, audiogramma_porog_3 = ["аудио"], ["грамм", "метр"], "Пороговая аудиометрия", ["реб", "дет", "вебер", "прием", "консул"], ["порог", "тон"]
	tamponada_nosa, tamponada_nosa_2, id_tamponada_nosa, stop_for_tamponada_nosa = ["тампон"], ["нос"], "Тампонада носа", ["реб", "удал", "претам", "ринит", "хронич", "репози", "уретр", "перетамп", "извлеч", "лун", "зуб", "стома", "дес", "влага", "гине", "вуль", "мат", "дет", "удал", "извлеч"]
	zaush_blok, zaush_blok_2, id_zaush_blok, stop_for_zaush_blok = ["зауш", "ушн", "парамеатал"], ["блокад"], "Заушная блокада", ["реб", "дет"]
	mass_pereponki, mass_pereponki_2, id_mass_pereponki, stop_for_mass_pereponki = ["массаж", "массирова", "пневм"], ["перепон", "зигле"], "Пневмомассаж барабанной перепонки", ["реб", "дет", "в рамках", "консул", "прием"]
	otoskopia, otoskopia_2, id_otoskopia, stop_for_otoskopia = ["отоскопи", "эндоскопи"], ["отоскопи", "уха", "ушных", "ушей", "слухо", "лор-органов", "лор органов"], "Отоскопия", ["микро", "регистр", "запись", "данн", "оториноларинголог", "лор"]
	kamerton, kamerton_2, id_kamerton, stop_for_kamerton = ["камертон"], ["слух", "исследова", "пробы", "проба"], "Камертональное исследование слуха", ["реб", "дет", "регистр", "запись", "данн"]
	elektrokohleo, elektrokohleo_2, id_elektrokohleo, stop_for_elektrokohleo = ["электро"], ["кохлео"], "Электрокохлеография", ["реб", "дет", "регистр", "запись", "данн"]
	otoakustik, otoakustik_2, id_otoakustik, stop_for_otoakustik = ["акусти", "аккусти"], ["эмисс"], "Отоакустическая эмиссия", ["реб", "дет"]
	timpanometr, timpanometr_2, id_timpanometr, stop_for_timpanometr = ["тимпан", "акусти"], ["метр", "импеданс", "импенданс"], "Тимпанометрия", ["реб", "дет", "рино", "втор"]
	perimetr, perimetr_2, id_perimetr, stop_for_perimetr = ["периметрия", "периметрии"], ["периметрия", "периметрии"], "Периметрия", ["период", "контрол"]
	skiaskop, skiaskop_2, id_skiaskop, stop_for_skiaskop = ["скиа"], ["скопия", "скопии"], "Скиаскопия глаза", ["заглушка"]
	maklakov, maklakov_2, id_maklakov, stop_for_maklakov = ["маклаков"], ["маклаков"], "Тонометрия по Маклакову", ["в рамках"]
	gonioskop, gonioskop_2, id_gonioskop, stop_for_gonioskop = ["гонио"], ["скопия", "скопии"], "Гониоскопия", ["заглушка"]
	refraktometr, refraktometr_2, id_refraktometr, stop_for_refraktometr = ["рефрактометрия", "авторефрактометрия"], ["рефрактометрия", "авторефрактометрия"], "Рефрактометрия", ["дет", "реб", "повтор", "направле", "после", "преодплат", "проф"] 
	shirmer, shirmer_2, id_shirmer, stop_for_shirmer = ["тест", "проб", "исслед"], ["ширмер"], "Тест / проба Ширмера", ["комплекс", "с проведением"]
	norn, norn_2, id_norn, stop_for_norn = ["тест", "проб", "исслед"], ["норн", "проба норма"], "Проба Норна", ["комплекс", "с проведением"]
	ekzoftalmometr, ekzoftalmometr_2, id_ekzoftalmometr, stop_for_ekzoftalmometr = ["экз"], ["офтальмометр"], "Экзофтальмометрия", ["комплекс", "с проведением"]
	mass_prostati, mass_prostati_2, id_mass_prostati, stop_for_mass_prostati = ["массаж"], ["простат", "предстательн"], "Массаж простаты", ["комплекс", "с проведением", "с помощью", "получение секрета"]
	ust_pessar, ust_pessar_2, id_ust_pessar, stop_for_ust_pessar = ["введе", "установ", "постанов"], ["пессар", "маточного кольца"], "Введение акушерского пессария", ["заглушка"]
	udal_pessar, udal_pessar_2, id_udal_pessar, stop_for_udal_pessar = ["удал", "извле"], ["пессар", "маточного кольца"], "Извлечение акушерского пессария", ["заглушка"]
	inorod_vlag, inorod_vlag_2, id_inorod_vlag, stop_for_inorod_vlag = ["инород", "посторонн"], ["влагал"], "Удаление инородного тела из влагалища", ["дет", "реб", "втор"]
	inorod_glaz, inorod_glaz_2, id_inorod_glaz, stop_for_inorod_glaz = ["инород", "посторонн"], ["глаз", "конъюнктив", "конъюктив", "коньюктив", "коньюнк", "роговиц", "роговичн"], "Удаление инородного тела из глаза", ["дет", "реб", "втор"]
	inorod_glotki, inorod_glotki_2, id_inorod_glotki, stop_for_inorod_glotki = ["инород", "посторонн"], ["глотки", "глодки"], "Удаление инородного тела из глотки", ["дет", "реб", "втор"]
	inorod_gortani, inorod_gortani_2, id_inorod_gortani, stop_for_inorod_gortani = ["инород", "посторонн"], ["гортани"], "Удаление инородного тела из гортани", ["дет", "реб", "втор"]
	inorod_nosa, inorod_nosa_2, id_inorod_nosa, stop_for_inorod_nosa = ["инород", "посторонн"], ["носа", "носовой"], "Удаление инородного тела из носа", ["дет", "реб", "втор"]
	inorod_uha, inorod_uha_2, id_inorod_uha, stop_for_inorod_uha = ["инород", "посторонн"], ["уха", "ушной", "ушей", "ушных", "ухо"], "Удаление инородного тела из уха", ["дет", "реб", "втор", "чел", "пазухах"]
	politser, politser_2, id_politser, stop_for_politser = ["проду"], ["политцер", "слуховых труб", "слуховой трубы"], "Продувание слуховых труб по Политцеру", ["дет", "реб", "втор", "в рамках", "прием", "консул"]
	kukushka, kukushka_2, id_kukushka, stop_for_kukushka = ["процедур", "промыва", "санац", "чистк", "кукушк", "проетц", "проэтц"], ["кукушк", "проетц", "проэтц", "перемещения"], "Процедура Кукушка", ["дет", "реб", "втор", "в рамках кон", "в ходе кон"]
	bujir_cervik, bujir_cervik_2, id_bujir_cervik, stop_for_bujir_cervik = ["бужирован"], ["цервика"], "Бужирование цервикального канала", ["дет", "реб", "втор"]
	bujir_uretri, bujir_uretri_2, id_bujir_uretri, stop_for_bujir_uretri = ["бужирован"], ["уретр"], "Бужирование уретры", ["дет", "реб", "втор"]
	bujir_uretri_jen, bujir_uretri_jen_2, id_bujir_uretri_jen, stop_for_bujir_uretri_jen, bujir_uretri_jen_3 = ["бужирован"], ["уретр"], "Бужирование уретры у женщин", ["дет", "реб", "втор"], ["женщин"]
	bujir_uretri_muj, bujir_uretri_muj_2, id_bujir_uretri_muj, stop_for_bujir_uretri_muj, bujir_uretri_muj_3 = ["бужирован"], ["уретр"], "Бужирование уретры у мужчин", ["дет", "реб", "втор"], ["мужчин"]
	bujir_anal, bujir_anal_2, id_bujir_anal, stop_for_bujir_anal = ["бужирован"], ["анальн", "анус"], "Бужирование анального отверстия", ["дет", "реб", "втор"]
	bujir_kolostom, bujir_kolostom_2, id_bujir_kolostom, stop_for_bujir_kolostom = ["бужирован"], ["колостом"], "Бужирование колостомы", ["дет", "реб", "втор"]
	inorod_kishe4, inorod_kishe4_2, id_inorod_kishe4, stop_for_inorod_kishe4 = ["инородн"], ["кишечник", "прямой кишки", "толстой кишки", "толстого кишечник"], "Удаление инородного тела прямой кишки", ["дет", "реб", "втор", "тонк"]
	blokad_grushevid, blokad_grushevid_2, id_blokad_grushevid, stop_for_blokad_grushevid = ["блокад"], ["грушевидн"], "Блокада грушевидной мышцы", ["дет", "реб", "втор"]
	block_krest_povz, block_krest_povz_2, id_block_krest_povz, stop_for_block_krest_povz, block_krest_povz_3 = ["блокад"], ["крест"], "Блокада крестцово-подвздошного сочленения", ["дет", "реб", "втор", "курс"], ["подвздош", "повздош", "позвз"]
	block_per_nerva, block_per_nerva_2, id_block_per_nerva, stop_for_block_per_nerva = ["блокад"], ["нерва", "нервов", "тройничн", "нервы"], "Блокада периферического нерва", ["дет", "реб", "ствол", "втор"]
	block_pozvon, block_pozvon_2, id_block_pozvon, stop_for_block_pozvon = ["блокад"], ["позвоночник", "спинальн", "вертебральн", "вертербральн"], "Блокада позвоночника", ["дет", "реб", "втор", "корешков"]
	block_p9t_shpor, block_p9t_shpor_2, id_block_p9t_shpor, stop_for_block_p9t_shpor, block_p9t_shpor_2 = ["блокад"], ["пяточн"], "Блокада при пяточной шпоре", ["дет", "реб", "втор"], ["шпор", "кости"]
	block_semen_kanat, block_semen_kanat_2, id_block_semen_kanat, stop_for_block_semen_kanat, block_block_semen_kanat_2 = ["блокад"], ["семен", "сменног", "лорин"], "Блокада семенного канатика", ["дет", "реб", "втор"], ["канат", "эпштейн"]
	block_sustav, block_sustav_2, id_block_sustav, stop_for_block_sustav = ["блокад", "периартикулярное введение алфлутоп"], ["сустав", "периартикулярн", "параартикулярн", "интраартикулярн", "периартикулярное введение алфлутоп"], "Блокада сустава", ["дет", "реб", "снятие бло", "устранение бло", "позвон", "лечен", "исправлен"]
	block_to4ek, block_to4ek_2, id_block_to4ek, stop_for_block_to4ek = ["блокад"], ["триггер", "миофасциальн"], "Блокада триггерных точек", ["дет", "реб", "втор", "деся"]
	block_nosa, block_nosa_2, id_block_nosa, stop_for_block_nosa = ["блокад"], ["носа", "носов", "эндоназальн", "интраназаль"], "Внутриносовая блокада", ["дет", "реб", "втор"]
	block_zaushn, block_zaushn_2, id_block_zaushn, stop_for_block_zaushn = ["блокад"], ["ушей", "уха", "заушная"], "Заушная блокада", ["дет", "реб", "втор"]
	block_intratonz, block_intratonz_2, id_block_intratonz, stop_for_block_intratonz = ["блокад"], ["тонзилляр", "глотки", "глоточн"], "Интратонзиллярная блокада", ["дет", "реб", "втор"]
	block_koreshkov, block_koreshkov_2, id_block_koreshkov, stop_for_block_koreshkov = ["блокад"], ["корешков"], "Корешковая блокада", ["дет", "реб", "втор"]
	block_mezhreber, block_mezhreber_2, id_block_mezhreber, stop_for_block_mezhreber = ["блокад"], ["ребер"], "Межреберная блокада", ["заглушка"]
	block_paraprost, block_paraprost_2, id_block_paraprost, stop_for_block_paraprost = ["блокад"], ["простат", "предстат"], "Парапростатическая блокада", ["дет", "реб", "втор"]
	block_epidural, block_epidural_2, id_block_epidural, stop_for_block_epidural = ["блокад"], ["эпидурал"], "Эпидуральная блокада", ["дет", "реб", "втор", "курс"]
	block_presakral, block_presakral_2, id_block_presakral, stop_for_block_presakral = ["блокад"], ["сакрал"], "Пресакральная блокада", ["дет", "реб", "втор"]
	piling_karbon, piling_karbon_2, id_piling_karbon, stop_for_piling_karbon = ["пилинг"], ["карбонов"], "Карбоновый пилинг", ["дет", "реб", "втор", "втор", "акци", "допол"]
	piling_gazojid, piling_gazojid_2, id_piling_gazojid, stop_for_piling_gazojid = ["пилинг"], ["газожидк", "газо-жид"], "Газожидкостный пилинг лица", ["дет", "акци", "реб", "втор", "втор", "допол"]
	piling_korall, piling_korall_2, id_piling_korall, stop_for_piling_korall = ["пилинг"], ["кораллов"], "Коралловый пилинг лица", ["дет", "реб", "втор", "акци", "втор", "допол"]
	piling_mehani4, piling_mehani4_2, id_piling_mehani4, stop_for_piling_mehani4 = ["пилинг"], ["механическ"], "Механический пилинг лица", ["дет", "реб", "акци", "втор", "химич", "втор", "допол"]
	piling_abr, piling_abr_2, id_piling_abr, stop_for_piling_abr = ["пилинг"], ["abr", "абр ", " абр", "абр-"], "ABR-пилинг", ["дет", "акци", "реб", "втор", "допол"]
	piling_spa, piling_spa_2, id_piling_spa, stop_for_piling_spa = ["пилинг"], ["spa", " спа"], "SPA-пилинг", ["дет", "реб", "акци", "втор", "испа", "допол"]
	piling_almaz, piling_almaz_2, id_piling_almaz, stop_for_piling_almaz = ["пилинг", "дермабраз"], ["алмазн"], "Алмазный пилинг", ["дет", "реб", "втор", "акци", "допол"]
	piling_glikol, piling_glikol_2, id_piling_glikol, stop_for_piling_glikol = ["пилинг", "%"], ["гликоли", "гликоле"], "Гликолиевый химический пилинг", ["дет", "акци", "реб", "втор", "допол"]
	piling_jeltiy, piling_jeltiy_2, id_piling_jeltiy, stop_for_piling_jeltiy = ["пилинг"], ["желт", "ретино", "yellow"], "Желтый химический пилинг", ["дет", "реб", "закры", "втор", "акци", "допол"]
	piling_mindal, piling_mindal_2, id_piling_mindal, stop_for_piling_mindal = ["пилинг", "%"], ["миндаль"], "Миндальный химический пилинг", ["дет", "реб", "акци", "втор", "допол"]
	piling_molo4, piling_molo4_2, id_piling_molo4, stop_for_piling_molo4 = ["пилинг"], ["молоч", "молок", "лакто"], "Молочный пилинг", ["дет", "реб", "акци", "втор", "допол"]
	piling_jessner, piling_jessner_2, id_piling_jessner, stop_for_piling_jessner = ["джесснер", "джеснер"], ["джесснер", "джеснер"], "Пилинг Джесснера", ["дет", "реб", "втор", "допол"]
	piling_tsa, piling_tsa_2, id_piling_tsa, stop_for_piling_tsa = ["пилинг", "tca 15%"], ["тса", "трихлор", "tca", "tca 15%"], "Пилинг ТСА", ["дет", "реб", "втор", "допол", "веки", "век"]
	piling_pirovino, piling_pirovino_2, id_piling_pirovino, stop_for_piling_pirovino = ["пилинг"], ["пировино"], "Пировиноградный пилинг", ["дет", "акци", "реб", "втор", "допол"]
	piling_salicil, piling_salicil_2, id_piling_salicil, stop_for_piling_salicil = ["пилинг"], ["салицил"], "Салициловый химический пилинг", ["дет", "реб", "акци", "ше", "рук", "кист", "декол", "втор", "допол"]
	piling_fenol, piling_fenol_2, id_piling_fenol, stop_for_piling_fenol = ["пилинг"], ["фенол"], "Феноловый пилинг лица", ["дет", "реб", "втор", "акци", "допол"]
	piling_ferul, piling_ferul_2, id_piling_ferul, stop_for_piling_ferul = ["пилинг"], ["ферул"], "Феруловый пилинг", ["дет", "реб", "втор", "акци", "допол"]
	piling_frukt, piling_frukt_2, id_piling_frukt, stop_for_piling_frukt = ["пилинг"], ["фрукт"], "Фруктовый пилинг", ["дет", "реб", "втор", "акци", "допол"]
	amnioskopia, amnioskopia_2, id_amnioskopia, stop_for_amnioskopia = ["амниоскоп"], ["амниоскоп"], "Амниоскопия", ["дет", "реб", "втор"]
	amniocentez, amniocentez_2, id_amniocentez, stop_for_amniocentez = ["амниоценте"], ["амниоценте"], "Амниоцентез", ["дет", "реб", "втор"]
	kordocentez, kordocentez_2, id_kordocentez, stop_for_kordocentez = ["кордоцент"], ["кордоцент"], "Кордоцентез", ["дет", "реб", "втор"]
	placentocentez, placentocentez_2, id_placentocentez, stop_for_placentocentez = ["плацентоценте"], ["плацентоценте"], "Плацентоцентез", ["дет", "реб", "втор"]
	prenatal_trisom, prenatal_trisom_2, id_prenatal_trisom, stop_for_prenatal_trisom, prenatal_trisom_3 = ["пренатальн"], ["скрини", "исследован"], "Пренатальный скрининг трисомий", ["заглушка"], ["трисом"]
	uzi_vne_skrining, uzi_vne_skrining_2, id_uzi_vne_skrining, stop_for_uzi_vne_skrining = ["узи", "ультразвук", "ранние сроки беременности", "ранних сроках беременности", "ранних сроков беременности"], ["не скрининг"], "УЗИ беременным вне скрининга", ["заглушка"]
	uzi_serdca_ploda,uzi_serdca_ploda_2, id_uzi_serdca_ploda, stop_for_uzi_serdca_ploda, uzi_serdca_ploda_3 = ["узи", "ультразвук"], ["сердц"], "УЗИ сердца плода", ["заглушка"], ["плод"]
	cervikometr, cervikometr_2, id_cervikometr, stop_for_cervikometr = ["цервикометр", "узи", "ультразвук"], ["цервикометр", "шейки матки"], "Цервикометрия при беременности", ["руб", "узи плода", "контрольн", "плода", "аборт", "готов", "прерыв"]
	diafanoskop, diafanoskop_2, id_diafanoskop, stop_for_diafanoskop = ["диафаноскоп"], ["мошонк"], "Диафаноскопия мошонки", ["дет", "реб", "втор"]
	kolonoskop, kolonoskop_2, id_kolonoskop, stop_for_kolonoskop = ["колоноскоп", "фкс"], ["колоноскоп", "фкс"], "Колоноскопия кишечника / ФКС", ["дет", "реб", "прерванн", "досрочно", "полип", "в ходе", "образова", "готов", "втор", "без колоно", "без фкс", "без эндо", "акци", "анестез", "седац", "томограф", "кт ", "кт-", "кт - ", "кт- ", "наркоз"]
	fks_and_fgds, fks_and_fgds_2, id_fks_and_fgds, stop_for_fks_and_fgds = ["колоноскоп", "фкс", "вкс"], ["фгдс", "гастро", "эндоскопия желудка", "вгдс", "эгдс"], "Колоноскопия и ФГДС под наркозом одновременно", ["дет", "готов", "полип", "образо", "в ходе", "акци", "реб", "втор", "при", "для"]
	fks_narkoz, fks_narkoz_2, id_fks_narkoz, stop_for_fks_narkoz = ["колоноскоп", "фкс"], ["под наркозом", "с наркозом", "седаци"], "Колоноскопия под наркозом", ["дет", "реб", "без седации", "без наркоза", "втор", "полип", "образ", "акци", "в ходе", "фгдс", "гастро", "готов", "эгдс", "эзофаго", "гастроскоп"]
	angiograf_set4atk, angiograf_set4atk_2, id_angiograf_set4atk, stop_for_angiograf_set4atk = ["ангиограф"], ["сетчатк"], "Ангиография сетчатки", ["дет", "реб", "втор"]
	beskon_tonometr, beskon_tonometr_2, id_beskon_tonometr, stop_for_beskon_tonometr = ["бесконтакт", "пневмо"], ["тонометр"], "Бесконтактная тонометрия", ["дет", "реб", "втор", "проф"]
	retinometria, retinometria_2, id_retinometria, stop_for_retinometria = ["ретинометрия"], ["ретинометрия"], "Ретинометрия", ["дет", "реб", "втор"]
	fluor_angio, fluor_angio_2, id_fluor_angio, stop_for_fluor_angio = ["флуоресце", "флюоресце"], ["ангиограф"], "Флуоресцентная ангиография", ["дет", "реб", "втор"]
	fluor_proba, fluor_proba_2, id_fluor_proba, stop_for_fluor_proba = ["флюоресце", "тест флюоресциииновый", "тест флюоресциин", "флуоресце", "флюоресцииин", "флюоресцин", "флюоресциииновый тест", "флюоресцеиновый тест", "флюоресцииновый тест", "флюоресцеиновая проба", "флюоресцеиновая инстилляционная проба", "зайдел"], ["зайдел", "инстилл", "флюоресцеиновый тест", "тест флюоресциииновый", "флюоресцеиновая проба", "тест флюоресциин", "флюоресциииновый тест", "флюоресцииновый тест", "люоресцеиновая инстилляционная проба"], "Флюоресцеиновая инстилляционная проба", ["дет", "реб", "втор"]
	cvet_slez_nos, cvet_slez_nos_2, id_cvet_slez_nos, stop_for_cvet_slez_nos, cvet_slez_nos_3 = ["носов"], ["слезн"], "Цветная слезно-носовая проба", ["дет", "реб", "втор"], ["проб"]
	elastotonometr, elastotonometr_2, id_elastotonometr, stop_for_elastotonometr = ["эластотонометрия"], ["эластотонометрия"], "Эластотонометрия", ["дет", "реб", "втор"]
	elektrookulograf, elektrookulograf_2, id_elektrookulograf, stop_for_elektrookulograf = ["электроокулограф"], ["электроокулограф"], "Электроокулография", ["дет", "реб", "втор"]
	elektroretinograf, elektroretinograf_2, id_elektroretinograf, stop_for_elektroretinograf = ["электроретинограф"], ["электроретинограф"], "Электроретинография", ["дет", "реб", "втор"]
	biopsia_bronh, biopsia_bronh_2, id_biopsia_bronh, stop_for_biopsia_bronh = ["биопсия", "биопсии"], ["бронхов"], "Биопсия бронхов", ["дет", "реб", "втор"]
	biopsia_vulvi, biopsia_vulvi_2, id_biopsia_vulvi, stop_for_biopsia_vulvi = ["биопсия", "биопсии"], ["вульвы"], "Биопсия вульвы", ["дет", "реб", "втор"]
	biopsia_mozg, biopsia_mozg_2, id_biopsia_mozg, stop_for_biopsia_mozg = ["биопсия", "биопсии"], ["мозг"], "Биопсия головного мозга", ["дет", "реб", "костн", "опухол", "образовани"]
	biopsia_gortan, biopsia_gortan_2, id_biopsia_gortan, stop_for_biopsia_gortan = ["биопсия", "биопсии"], ["гортан", "лор-органов", "лор органов"], "Биопсия гортани", ["дет", "реб", "втор"]
	biopsia_jeludka, biopsia_jeludka_2, id_biopsia_jeludka, stop_for_biopsia_jeludka = ["биопсия", "биопсии"], ["желудк"], "Биопсия желудка", ["дет", "реб", "втор"]
	biopsia_kavern_4lena, biopsia_kavern_4lena_2, id_biopsia_kavern_4lena, stop_for_biopsia_kavern_4lena, biopsia_kavern_4lena_3 = ["биопсия", "биопсии"], ["член"], "Биопсия кавернозной ткани полового члена", ["дет", "реб", "втор", "опухол", "образован"], ["каверн", "член"]
	biopsia_kishe4, biopsia_kishe4_2, id_biopsia_kishe4, stop_for_biopsia_kishe4 = ["биопсия", "биопсии"], ["кишки", "кишечник"], "Биопсия кишечника", ["толст", "прям", "тонк"]
	biopsia_koji, biopsia_koji_2, id_biopsia_koji, stop_for_biopsia_koji = ["биопсия", "биопсии"], ["кожи"], "Биопсия кожи", ["дет", "реб", "втор", "накожн", "образова", "элемент"]
	biopsia_kost_mozg, biopsia_kost_mozg_2, id_biopsia_kost_mozg, stop_for_biopsia_kost_mozg, biopsia_kost_mozg_3 = ["биопсия", "биопсии"], ["мозг"], "Биопсия костного мозга", ["дет", "реб", "головн"], ["кост"]
	biopsia_legkih, biopsia_legkih_2, id_biopsia_legkih, stop_for_biopsia_legkih = ["биопсия", "биопсии"], ["легких", "легкого"], "Биопсия легких", ["дет", "реб", "втор"]
	biopsia_limfouzl, biopsia_limfouzl_2, id_biopsia_limfouzl, stop_for_biopsia_limfouzl = ["биопсия", "биопсии"], ["лимфоузла", "лимфоузлов", "лимфатических узлов", "лимфатического узла"], "Биопсия лимфоузла", ["дет", "реб", "подмыш"]
	biopsia_molo4, biopsia_molo4_2, id_biopsia_molo4, stop_forbiopsia_molo4, biopsia_molo4_3 = ["биопсия", "биопсии"], ["молочн"], "Биопсия молочных желез", ["дет", "реб", "навигация", "наведение", "втор", "дополнит"], ["желез"]
	biopsia_mo4evo, biopsia_mo4evo_2, id_biopsia_mo4evo, stop_for_biopsia_mo4evo = ["биопсия", "биопсии"], ["мочевого пузыря", "мочевой пузырь"], "Биопсия мочевого пузыря", ["дет", "реб", "втор"]
	biopsia_mishc, biopsia_mishc_2, id_biopsia_mishc, stop_for_biopsia_mishc = ["биопсия", "биопсии"], ["мышц", "мышечн", "мускул", "фасц"], "Биопсия мышцы и мышечной фасции", ["дет", "реб", "втор", "лимф"]
	biopsia_m9gkih, biopsia_m9gkih_2, id_biopsia_m9gkih, stop_for_ebiopsia_m9gkih = ["биопсия", "биопсии"], ["мягких", "мягкой"], "Биопсия мягких тканей", ["дет", "реб", "втор"]
	biopsia_nakojnih, biopsia_nakojnih_2, id_biopsia_nakojnih, stop_biopsia_nakojnih, biopsia_nakojnih_3 = ["биопсия", "биопсии"], ["накожн"], "Биопсия накожных элементов", ["дет", "реб", "втор"], ["элементов", "образован"]
	biopsia_obraz_4len, biopsia_obraz_4len_2, id_biopsia_obraz_4len, stop_for_biopsia_obraz_4len, biopsia_obraz_4len_3 = ["биопсия", "биопсии"], ["образован"], "Биопсия образования полового члена", ["дет", "реб", "втор"], ["член"]
	biopsia_pe4en, biopsia_pe4en_2, id_biopsia_pe4en, stop_for_biopsia_pe4en = ["биопсия", "биопсии"], ["печен"], "Биопсия печени", ["дет", "реб", "втор"]
	biopsia_pishevod, biopsia_pishevod_2, id_biopsia_pishevod, stop_for_biopsia_pishevod = ["биопсия", "биопсии"], ["пищевод"], "Биопсия пищевода", ["дет", "реб", "втор"]
	biopsia_podjelud, biopsia_podjelud_2, id_biopsia_podjelud, stop_for_biopsia_podjelud = ["биопсия", "биопсии"], ["поджелудочн"], "Биопсия поджелудочной железы", ["дет", "реб", "втор"]
	biopsia_polip, biopsia_polip_2, id_biopsia_polip, stop_for_biopsia_polip = ["биопсия", "биопсии"], ["полип"], "Биопсия полипа", ["дет", "реб", "втор", "эктом"]
	biopsia_po4ki, biopsia_po4ki_2, id_biopsia_po4ki, stop_for_biopsia_po4ki = ["биопсия", "биопсии"], ["почки", "почек"], "Биопсия почки", ["дет", "реб", "втор"]
	biopsia_prostati, biopsia_prostati_2, id_biopsia_prostati, stop_for_biopsia_prostati = ["биопсия", "биопсии"], ["предстател", "простаты"], "Биопсия предстательной железы / простаты", ["дет", "реб", "втор", "навигация", "наведение"]
	biopsia_pr9m_kishki, biopsia_pr9m_kishki_2, id_biopsia_pr9m_kishki, stop_for_biopsia_pr9m_kishki, biopsia_pr9m_kishki_3 = ["биопсия", "биопсии"], ["прямой", "толст"], "Биопсия прямой кишки", ["дет", "реб", "втор", "полип"], ["кишк"]
	biopsia_sustav, biopsia_sustav_2, id_biopsia_sustav, stop_for_biopsia_sustav, biopsia_sustav_3 = ["биопсия", "биопсии"], ["оболоч", "ткан", "жидк"], "Биопсия синовиальной оболочки сустава", ["дет", "реб", "втор"], ["сустав", "синовиал"]
	biopsia_slunnih, biopsia_slunnih_2, id_biopsia_slunnih, stop_for_biopsia_slunnih = ["биопсия", "биопсии"], ["слюнн"], "Биопсия слюнных желез", ["дет", "реб", "втор"]
	biopsia_ton_kishe4, biopsia_ton_kishe4_2, id_biopsia_ton_kishe4, stop_for_biopsia_ton_kishe4, biopsia_ton_kishe4_3 = ["биопсия", "биопсии"], ["тонк"], "Биопсия тонкого кишечника", ["дет", "реб", "втор"], ["кишечник", "кишк"]
	biopsia_sheiki_matki, biopsia_sheiki_matki_2, id_biopsia_sheiki_matki, stop_for_biopsia_sheiki_matki = ["биопсия", "биопсии"], ["шейки матки"], "Биопсия шейки матки", ["дет", "реб", "после", "обработ", "втор", "прицель", "ножева"]
	biopsia_shitovid, biopsia_shitovid_2, id_biopsia_shitovid, stop_for_biopsia_shitovid = ["биопсия", "биопсии"], ["щитовидн"], "Биопсия щитовидной железы", ["дет", "реб", "навигация", "наведение"]
	laparos_biops_9i4nikov, laparos_biops_9i4nikov_2, id_laparos_biops_9i4nikov, stop_for_laparos_biops_9i4nikov, laparos_biops_9i4nikov_3 = ["биопсия", "биопсии"], ["лапароскоп"], "Лапароскопическая биопсия яичников", ["дет", "реб", "втор"], ["яичник"]
	nojeva9a_biops_sheiki_matki, nojeva9a_biops_sheiki_matki_2, id_nojeva9a_biops_sheiki_matki, stop_for_nojeva9a_biops_sheiki_matki = ["биопсия", "биопсии"], ["ножева", "ножевой"], "Ножевая биопсия шейки матки", ["дет", "реб", "кожи", "вуль", "влаг", "наружных поло", "губ"]
	otkrita_biopsi_9i4ka, otkrita_biopsi_9i4ka_2, id_otkrita_biopsi_9i4ka, stop_for_otkrita_biopsi_9i4ka, otkrita_biopsi_9i4ka_3 = ["биопсия", "биопсии"], ["открыт"], "Открытая биопсия яичка", ["дет", "реб", "втор"], ["яичка", "яичек"]
	preskalen_biopsia, preskalen_biopsia_2, id_preskalen_biopsia, stop_for_preskalen_biopsia = ["биопсия", "биопсии"], ["прескален"], "Прескаленная биопсия", ["дет", "реб", "втор"]
	pricel_biopsia_sheiki_matki, pricel_biopsia_sheiki_matki_2, id_pricel_biopsia_sheiki_matki, stop_for_pricel_biopsia_sheiki_matki, pricel_biopsia_sheiki_matki_3 = ["биопсия", "биопсии"], ["прицель"], "Прицельная биопсия шейки матки", ["дет", "реб", "втор"], ["шейки матки"]
	biopsia_plevri, biopsia_plevri_2, id_biopsia_plevri, stop_for_biopsia_plevri = ["биопсия", "биопсии"], ["плевры", "плевральной"], "Пункционная биопсия плевры", ["дет", "реб", "втор"]
	biopsia_podmish_limfo, biopsia_podmish_limfo_2, id_biopsia_podmish_limfo, stop_for_biopsia_podmish_limfo, biopsia_podmish_limfo_3 = ["биопсия", "биопсии"], ["подмыш"], "Пункционная биопсия подмышечных лимфоузлов", ["дет", "реб", "втор"], ["лимфоузл"]
	biopsia_pozvonka, biopsia_pozvonka_2, id_biopsia_pozvonka, stop_for_biopsia_pozvonka = ["биопсия", "биопсии"], ["позвонка", "позвонков"], "Пункционная биопсия позвонка", ["дет", "реб", "втор"]
	biopsi_endometri, biopsi_endometri_2, id_biopsi_endometri, stop_for_biopsi_endometri = ["биопсия", "биопсии"], ["эндометр"], "Биопсия эндометрия", ["дет", "реб", "пайпел", "цуг", "pipel", "штрих", "аспирац"]
	aspirac_biop_endometr, aspirac_biop_endometr_2, id_aspirac_biop_endometr, stop_for_aspirac_biop_endometr = ["биопсия", "биопсии"], ["аспирац"], "Аспирационная биопсия эндометрия", ["дет", "молоч", "яич", "желез", "груд", "реб", "втор"]
	paipel, paipel_2, id_paipel, stop_for_paipel, paipel_3 = ["биопсия", "биопсии", "диагности", "аспирата из полости матки"], ["пайпел", "pipel"], "Пайпель биопсия эндометрия", ["дет", "реб", "втор"], ["пайпел", "pipel"]
	tsug_biopsi, tsug_biopsi_2, id_tsug_biopsi, stop_for_tsug_biopsi, tsug_biopsi_3 = ["биопсия", "биопсии"], ["цуг", "штрих"], "ЦУГ биопсия эндометрия", ["дет", "реб", "втор"], ["эндометри", "полости матки"]
	fgds_1, fgds_2_1, id_fgds_1, stop_for_fgds_1 = ["гастроскопия"], ["желудка"], "ФГДС / Гастроскопия желудка", ["дет", "реб", "при", "кислот", "хромоскоп," "полип", "в ходе", "образова", "акци", "эгдс", "эндоназальн", "наркоз", "эзофагогастродуоденоскоп", "анестез", "седац", "хелико", "helico"]
	fgds_2, fgds_2_2, id_fgds_2, stop_for_fgds_2 = ["фгдс"], ["фгдс"], "ФГДС / Гастроскопия желудка", ["дет", "реб", "эгдс", "эндоназальн", "наркоз", "при", "хромоскоп", "кислот", "эзофагогастродуоденоскоп", "полип", "в ходе", "акц", "образ", "анестез", "седац", "хелико", "helico"]
	fgds_3, fgds_2_3, id_fgds_3, stop_for_fgds_3 = ["гастроскопия"], ["гастроскопия"], "ФГДС / Гастроскопия желудка", ["дет", "реб", "эгдс", "эндоназальн", "наркоз", "хромоскоп" ,"при", "кислот", "полип", "в ходе", "удал", "образ", "акци", "эзофагогастродуоденоскоп", "анестез", "седац", "хелико", "helico"]
	fgds_narko, fgds_narko_2, id_fgds_narko_2, stop_for_fgds_narko_2 = ["фгдс", "гастроскоп", "эгдс", "гастродуоденоскопия"], ["наркоз", "седац"], "ФГДС под наркозом", ["дет", "реб", "втор", "колонос", "хромоскоп", "полип", "в ходе", "удал", "образ", "акци", "без фгд", "без эндо", "без опер", "без гастро", "без эгдс"]
	ezofagogastroduodenoskop, ezofagogastroduodenoskop_2, id_ezofagogastroduodenoskop, stop_for_ezofagogastroduodenoskop = ["эзофагогастродуоденоскоп", "эгдс", "гастродуоденоскопия"], ["эзофагогастродуоденоскопия", "эгдс", "гастродуоденоскопия"], "ЭГДС", ["дет", "полип", "прерва", "досроч", "в ходе", "удал", "для", "образ", "акци", "при", "кислот", "реб", "втор", "удал", "хелико", "helico"]
	endonazal_fgds, endonazal_fgds_2, id_endonazal_fgds, stop_for_endonazal_fgds = ["эндоназальн"], ["фгдс", "гастроскоп", "эзофагогастродуоденоскоп"], "Эндоназальная ФГДС", ["дет", "полип", "хромоскоп", "в ходе", "удал", "образ", "акци", "реб", "втор", "удал"]
	gisteroskop, gisteroskop_2, id_gisteroskop, stop_for_gisteroskop = ["гистероскопия", "гистероскопии"], ["гистероскопия", "гистероскопии"], "Гистероскопия", ["дет", "соногис", "узгсс", "проходимо", "реб", "втор", "рдв", "выскаблива", "диагности", "гистероскопическ"]
	gisteroskop_rdv, gisteroskop_rdv_2, id_gisteroskop_rdv, stop_for_gisteroskop_rdv = ["гистероскопия", "гистероскопии"], ["рдв", "выскабливани"], "Гистероскопия с РДВ", ["дет", "реб", "втор", "без рдв", "без выскабливан"]
	razdel_viskabl, razdel_viskabl_2, id_razdel_viskabl, stop_for_razdel_viskabl = ["раздельн", "рдв", "диагностическ"], ["выскабл", "рдв"], "Раздельное диагностическое выскабливание", ["дет", "реб", "гистероскоп"]
	biopsi_9i4ka, biopsi_9i4ka_2, id_biopsi_9i4ka, stop_for_biopsi_9i4ka = ["биопсия", "биопсии"], ["яичка", "яичек"], "Биопсия яичка", ["дет", "реб", "открыт", "придатк"]
	aromaterapi, aromaterapi_2, id_aromaterapi, stop_for_aromaterapi = ["арома"], ["терапи"], "Ароматерапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	barokameri, barokameri_2, id_barokameri, stop_for_barokameri = ["барокамер"], ["барокамер"], "Барокамеры", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	biorezonans, biorezonans_2, id_biorezonans, stop_for_biorezonans = ["биорезонанс"], ["биорезонанс"], "Биорезонансная терапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	biotok, biotok_2, id_biotok, stop_for_biotok = ["биоток", "биологическ"], ["биоток", "электричеств"], "Биотоковая терапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	bos_terapi, bos_terapi_2, id_bos_terapi, stop_for_bos_terapi = ["бос-терапи", "биологическ", "бос "], ["обратн", "бос-терапи", "терапи"], "Бос-терапия", ["дет", "пияв", "гируд", "перед", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "прогрев"]
	galvanization, galvanization_2, id_galvanization, stop_for_galvanization = ["гальванизац", "гальванич", "гальван"], ["гальванизац", "током", "токами", "терапи", "токи"], "Гальванизация", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	vlok, vlok_2, id_vlok, stop_for_vlok, vlok_3 = ["влок", "внутривенн", "облуч"], ["влок", "лазер"], "ВЛОК / Внутривенное лазерное облучение крови", ["дет", "реб", "над", "наруж", "втор", "члок", "чрезкож", "чрескож", "вторич", "абонемент", "занятий", "посещени", "внеполостн", "при проведении"], ["влок", "кров"]
	gidrokolonoterap, gidrokolonoterap_2, id_gidrokolonoterap, stop_for_gidrokolonoterap = ["гидроколонотерапи", "промывание"], ["гидроколонотерапи", "кишечник", "кишки"], "Гидроколонотерапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	gr9zele4en, gr9zele4en_2, id_gr9zele4en, stop_for_gr9zele4en = ["грязелеч", "лечени", "терапи"], ["грязелеч", "гряз"], "Грязелечение", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "оберт"]
	diadinamo, diadinamo_2, id_diadinamo, stop_for_diadinamo = ["диадинамотерапи", "диадинам"], ["диадинамотерапи", "ток", "терапи", "лечени"], "Диадинамотерапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "оберт"]
	infrakras_terapi, infrakras_terapi_2, id_infrakras_terapi, stop_for_infrakras_terapi = ["инфракрас"], ["терапи", "лечени", "саун"], "Инфракрасная терапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "внеполостн"]
	kv4, kv4_2, id_kv4, stop_for_kv4 = ["квч"], ["квч", "терапи", "лечени"], "КВЧ-терапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "внеполостн"]
	lazeroterapi, lazeroterapi_2, id_lazeroterapi, stop_for_lazeroterapi = ["лазеротерапи", "лазерная терапия", "лазерное облучение кожи", "воздействие"], ["лазеротерапи", "лазер", "лазерная терапия", "лазерное облучение кожи"], "Лазеротерапия", ["дет", "реб", "влок", "доп", "зве", "ангиоэ", "ангиэ", "крови", "папил", "образ", "кондил", "образо", "добро", "фибр", "кровь", "втор", "вторич", "абонемент", "занятий", "посещени", "магнит", "десен", "стоматол"]
	elektromagnit, elektromagnit_2, id_elektromagnit, stop_for_elektromagnit, elektromagnit_3 = ["электр", "текар"], ["магнит", "текар"], "Лечение электромагнитными полями", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "внеполостн"], ["лечени", "воздейств", "терапи", "применени"]
	magnitolazer, magnitolazer_2, id_magnitolazer, stop_for_magnitolazer, magnitolazer_3 = ["магнит"], ["лазер"], "Магнитолазеротерапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "внеполостн"], ["лечени", "воздейств", "терапи", "применени"]
	magnitoterap, magnitoterap_2, id_magnitoterap, stop_for_magnitoterap = ["магнитотерап", "воздействие"], ["магнитотерап", "магнитным полем", "магнитное поле", "магнитными полями"], "Магнитотерапия", ["дет", "реб", "от 3", "более", "втор", "вторич", "абонемент", "занятий", "посещени", "внеполостн"]
	mikrovolni, mikrovolni_2, id_mikrovolni, stop_for_mikrovolni = ["микроволнов"], ["терапи", "лечен", "воздейств"], "Микроволновая терапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "внеполостн"]
	okuf, okuf_2, id_okuf, stop_for_okuf = ["куф", "ультрафиолет"], ["куф", "терапи", "лечен", "воздейств", "облучени"], "ОКУФ-терапия", ["дет", "длинн", "уфо", "реб", "втор", "кров", "вторич", "абонемент", "занятий", "посещени", "внеполостн"]
	smt_terapi, smt_terapi_2, id_smt_terapi, stop_for_smt_terapi = ["смт", "амплипульс", "синусоидальн", "импульсн"], ["смт", "амплипульс", "терапи", "лечен", "воздейств", "ток"], "СМТ-терапия", ["дет", "доп", "реб", "телеанги", "звезд", "сетч", "хром", "глаз", "втор", "внеполос", "наружн", "вторич", "абонемент", "занятий", "посещени"]
	smt_terapi_vne, smt_terapi_vne_2, id_smt_terapi_vne, stop_for_smt_terapi_vne, smt_terapi_3 = ["смт", "амплипульс", "синусоидальн"], ["смт", "амплипульс", "терапи", "лечен", "воздейств", "ток"], "СМТ-терапия внеполостная", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["внеполост", "наружн"]
	transkrani_elektrostim, transkrani_elektrostim_2, id_transkrani_elektrostim, stop_transkrani_elektrostim = ["тэс", "транскраниальн", "ткэс"], ["тэс", "стимул", "ткэс"], "Транскраниальная электростимуляция", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	transrektal_elektrostim_prostati, transrektal_elektrostim_prostati_2, id_transrektal_elektrostim_prostati, stop_transrektal_elektrostim_prostati, transrektal_elektrostim_prostati_3 = ["ректал"], ["электр"], "Трансректальная электростимуляция простаты", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["простат", "предста", "стимул"]
	transuretral_elektrostim_prostati, transuretral_elektrostim_prostati_2, id_transuretral_elektrostim, stop_transuretral_elektrostim_prostati, transuretral_elektrostim_prostati_3 = ["уретр"], ["электр"], "Трансуретральная электростимуляция простаты", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["простат", "предста"]
	uv4_terapi, uv4_terapi_2, id_uv4_terapi, stop_uv4_terapi = ["увч", "ультравысокочасто", "ультразвуковое лечение кожи"], ["терапи", "ультразвуковое лечение кожи", "воздействие", "1 поле"], "УВЧ-терапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	uz_terapi, uz_terapi_2, id_uz_terapi, stop_uz_terapi = ["узи", "ультразвуко", "уз-", "уз - "], ["терапия", "ультразвуковой терапии", "сеанс"], "Ультразвуковая терапия", ["дет", "реб", "инфу", "втор", "вторич", "протруз", "абонемент", "занятий", "посещени"]
	ufo_terapi, ufo_terapi_2, id_ufo_terapi, stop_ufo_terapi = ["уфо", "ультрафиолет"], ["терапи", "лечен", "воздейств", "уфо", "облучени"], "УФО-терапия", ["дет", "куфо", "форм", " куф", "куф", "окуф", "кров", "вен", "коротк", "сосуд", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	fdt_terapi, fdt_terapi_2, id_fdt_terapi, stop_fdt_terapi = ["фдт", "фотодинамичес"], ["фдт", "терапи", "лечен", "воздейств"], "ФДТ / Фотодинамическая терапия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "курс", "абонемент"]
	fonoforez, fonoforez_2, id_fonoforez, stop_fonoforez = ["фонофорез"], ["фонофорез"], "Фонофорез", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "лиц"]
	fotoforez, fotoforez_2, id_fotoforez, stop_fotoforez = ["фотофорез", "фотоферез"], ["фотофорез", "фотоферез"], "Фотоферез", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	chens, chens_2, id_chens, stop_chens = ["чэнс", "чрескожная", "чрезскожная", "tens-терапия", "чрезкожная электронейростимуляция", "черескожная электонейростимуляция", "черескожная электронейростимуляция", "динамическая электронейростимуляци"], ["динамическая электронейростимуляци", "tens-терапия", "чэнс", "нейростимуляция", "электростимуляция", "электронейростимуляция", "чрезкожная электронейростимуляция", "черескожная электонейростимуляция", "черескожная электронейростимуляция"], "Чрескожная электронейростимуляция", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	elektromiostimul, elektromiostimul_2, id_elektromiostimul, stop_elektromiostimul = ["электро"], ["миостимуляци"], "Электромиостимуляция", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	elektroson, eelektroson_2, id_elektroson, stop_elektroson = ["электросон"], ["электросон"], "Электросон", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	elektrostimul_mo4evogo, elektrostimul_mo4evogo_2, id_elektrostimul_mo4evogo, stop_elektrostimul_mo4evogo, elektrostimul_mo4evogo_3 = ["электр"], ["стимул", "терапи"], "Электростимуляция мочевого пузыря", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["мочевого"]
	elektrostimul_uretri, elektrostimul_uretri_2, id_elektrostimul_uretri, stop_elektrostimul_uretri, elektrostimul_uretri_3 = ["электр"], ["стимул", "терапи"], "Электростимуляция уретры", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "проста", "предста"], ["уретры", "уретра"]
	elektrostimul_cervik, elektrostimul_cervik_2, id_elektrostimul_cervik, stop_elektrostimul_cervik, elektrostimul_cervik_3 = ["электр"], ["стимул", "терапи"], "Электростимуляция цервикального канала", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["цервикал"]
	uretroskopia, uretroskopiak_2, id_uretroskopia, stop_uretroskopia = ["уретроскопия", "уретроскопии"], ["уретроскопия", "уретроскопии"], "Уретроскопия", ["втор", "вторич", "абонемент", "занятий", "посещени"]
	cistoskopia, cistoskopia_2, id_cistoskopia, stop_cistoskopia = ["цистоскопия", "цистоскопии"], ["цистоскопия", "цистоскопии"], "Цистоскопия", ["дет", "реб", "дополните", "при цисто", "для цисто", "втор", "вторич", "абонемент", "занятий", "посещени", "хромо"]
	cirkumsicizo, cirkumsicizo_2, id_cirkumsicizo, stop_cirkumsicizo = ["циркумцизи", "обрезание"], ["циркумцизи", "крайней плоти"], "Обрезание / Циркумцизия", ["дет", "радио", "снятие", "после цир", "после обре", "сурги", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	viskab_cerviko, viskab_cerviko_2, id_viskab_cerviko, stop_viskab_cerviko, viskab_cerviko_3 = ["выскабливани", "кюретаж"], ["цервикальн"], "Выскабливание цервикального канала", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "наркоз при", "наркоз для", "посещени"], ["канал"]
	defloracia, defloracia_2, id_defloracia, stop_defloracia = ["дефлораци", "сечен"], ["дефлораци", "девственной плевы"], "Дефлорация", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	gineko_massaj, gineko_massaj_2, id_gineko_massaj, stop_gineko_massaj = ["гинекологическ", "при заболеваниях женских половых органов"], ["массаж"], "Гинекологический массаж", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	gsg, gsg_2, id_gsg, stop_gsg = ["гсг", "гистеросальпингография", "гистеросальпинография"], ["гсг", "гистеросальпингография", "гистеросальпинография"], "ГСГ / Гистеросальпингография", ["дет", "пневмо", "операц", "реб", "втор", "эхо", "уз", "ультразвук", "подгото", "узд", "соно", "вторич", "абонемент", "занятий", "посещени"]
	punkc_shitovid, punkc_shitovid_2, id_punkc_shitovid, stop_punkc_shitovid = ["пункция щитовидной", "пункция узла щитовидной"], ["пункция"], "Пункция щитовидной железы", ["дет", "реб", "при", "втор", "вторич", "абонемент", "занятий", "посещени"]
	massaj_veka, massaj_veka_2, id_massaj_veka, stop_massaj_veka = ["массаж"], ["века", " век"], "Массаж века", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	rinoskopia, rinoskopia_2, id_rinoskopia, stop_rinoskopia = ["рино", "носа", "носо", "лор-органов", "лор органов"], ["скопия", "скопии", "эндоскопи"], "Риноскопия", ["дет", "реб", "блок", "втор", "вторич", "абонемент", "биопси", "занятий", "посещени"]
	reposicia_nosa, reposicia_nosa_2, id_reposicia_nosa, stop_reposicia_nosa = ["репозици"], ["носа"], "Репозиция костей носа", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	udalenie_probok, udalenie_probok_2, id_udalenie_probok, stop_udalenie_probok = ["удаление", "промывание серных пробок", "промывание серной пробки"], ["пробок", "пробки", "ушной серы", "серы из ушей", "серы из слуховых", "серы из слухового", "промывание серных пробок", "промывание серной пробки"], "Удаление серных пробок", ["дет", "реб", "миндал", "втор", "вторич", "в рамках кон", "в ходе кон", "абонемент", "занятий", "посещени"]
	promiv_mndalin, promiv_mndalin_2, id_promiv_mndalin, stop_promiv_mndalin = ["промывани", "очищени", "санация"], ["миндалин"], "Промывание лакун небных миндалин", ["дет", "реб", "консультация", "в рамках консуль", "в ходе консульта", "прием", "в рамках", "втор", "вторич", "абонемент", "занятий", "посещени"]
	tonzillor, tonzillor_2, id_tonzillor, stop_tonzillor = ["тонзиллор", "тонзилор"], ["тонзиллор", "тонзилор"], "Лечение на аппарате Тонзиллор", ["дет", "реб", "дезинтеграц", "втор", "в рамках", "консул", "прием", "вторич", "абонемент", "занятий", "посещени"]
	vnutriven_ozon, vnutriven_ozon_2, id_vnutriven_ozon, stop_vnutriven_ozon = ["внутривенн", "в/в"], ["озон", "введение офр"], "Внутривенная озонотерапия", ["дет", "ауто", "курс", "гемо", "агт", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	kateter_sluh, kateter_sluh_2, id_kateter_sluh, stop_kateter_sluh = ["катетер"], ["слухово", "слуховых", "евстахи"], "Катетеризация слуховой трубы", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	trihogramm, trihogramm_2, id_trihogramm, stop_trihogramm = ["трихограмм"], ["трихограмм"], "Трихограмма", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "фото"]
	trihoskop, trihoskop_2, id_trihoskop, stop_trihoskop = ["трихоскопи"], ["трихоскопи"], "Трихоскопия", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	fototrihogramm, fototrihogramm_2, id_fototrihogramm, stop_fototrihogramm = ["фото"], ["трихограмм"], "Фототрихограмма", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	lampa_vuda, lampa_vuda_2, id_lampa_vuda, stop_lampa_vuda = ["ламп"], ["вуда"], "Люминесцентная диагностика под лампой Вуда", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	ustanovka_vms, ustanovka_vms_2, id_ustanovka_vms, stop_ustanovka_vms = ["становк", "введени"], ["спирали", "вмс", "мирена", "вмк", "контрацептив в/м"], "Установка внутриматочной спирали / ВМС", ["дет", "реб", "втор", "после", "нит", "лифти", "косич", "фил", "напол", "мезо", "вторич", "абонемент", "занятий", "посещени"]
	udalenie_vms, udalenie_vms_2, id_udalenie_vms, stop_udalenie_vms = ["удалени", "извлечени"], ["спирали", "вмс", "вмк", "мирена"], "Удаление внутриматочной спирали / ВМС", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "врос", "сложное", "сложного"]
	udal_vros_vms, udal_vros_vms_2, id_udal_vros_vms, stop_udal_vros_vms, udal_vros_vms_3 = ["удалени", "извлечен"], ["спирали", "вмс", "вмк"], "Удаление вросшей спирали", ["дет", "реб", "втор", "первой катего", "не осложне", "первая катег", "вторич", "абонемент", "неосложненн", "несложн", "занятий", "посещени"], ["сложн", "вросш"]
	zondir_polosti_matki, zondir_polosti_matki_2, id_zondir_polosti_matki, stop_zondir_polosti_matki, zondir_polosti_matki_3 = ["зондировани"], ["полости"], "Зондирование полости матки", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["матки"]
	kateter_mo4evogo_jen, kateter_mo4evogo_jen_2, id_kateter_mo4evogo_jen, stop_kateter_mo4evogo_jen, kateter_mo4evogo_jen_3 = ["катетер"], ["мочевого"], "Катетеризация мочевого пузыря у женщин", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["жен"]
	kateter_mo4evogo_muj, kateter_mo4evogo_muj_2, id_kateter_mo4evogo_muj, stop_kateter_mo4evogo_muj, kateter_mo4evogo_muj_3 = ["катетер"], ["мочевого"], "Катетеризация мочевого пузыря у мужчин", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"], ["муж"]
	vnutriven_inekc, vnutriven_inekc_2, id_vnutriven_inekc, stop_vnutriven_inekc = ["внутривенн"], ["инъекци", "введение лекарственных", "вливание"], "Внутривенная инъекция", ["дет", "дом", "капель", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "курс"]
	vnutrimish_inekc, vnutrimish_inekc_2, id_vnutrimish_inekc, stop_vnutrimish_inekc = ["внутримышеч"], ["инъекц", "введение лекарственных "], "Внутримышечная инъекция", ["дет", "диспорт", "гипер", "пото", "dispo", "ксео", "наличии", "направлен", "xeo", "botox", "ботокс", "сустанон", "реб", "хгч", "гонадо",  "втор", "вторич", "абонемент", "занятий", "посещени", "курс"]
	lumbal_punkc, lumbal_punkc_2, id_lumbal_punkc, stop_lumbal_punkc = ["люмбальн"], ["пункци"], "Люмбальная пункция", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени"]
	kt = ["кт ", "кт-", "кт -", "кт- ", "компьютерная томограф", "компьютерной томограф", "компьютерно-томографи", "компьютеро-томографи"]
	kt_brushn_aorti_2, id_kt_brushn_aorti, stop_kt_brushn_aorti, kt_brushn_aorti_3 = ["брюшной"], "КТ брюшной аорты", ["дет", "описание ", "дефект", "подго", "приго", "перед", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраст", "введением контраст"], ["аорты"]
	kt_gortani_2, id_kt_gortani, stop_kt_gortani = ["гортани", "гортань"], "КТ гортани", ["дет", "реб", "втор", "вторич", "абонемент", "описание ", "занятий", "дефект", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_grud_aorti_2, id_kt_grud_aorti, stop_kt_grud_aorti, kt_grud_aorti_3 = ["грудной"], "КТ грудной аорты", ["дет", "реб", "описание ", "втор", "дефект", "подго", "приго", "перед", "вторич", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["аорты"]
	kt_kisti_ruki_2, id_kt_kisti_ruki, stop_kt_kisti_ruki = ["кисти", "кисть",], "КТ кисти руки", ["дет", "реб", "втор", "вторич", "описание ", "абонемент", "дефект", "подго", "приго", "перед", "занятий", "посещени", "топометр", "скт", "мультиспиральн", "спирал"]
	kt_korona_calc_2, id_kt_korona_calc, stop_kt_korona_calc, kt_korona_calc_3 = ["коронар"], "КТ коронарного кальция", ["дет", "описание ", "реб", "дефект", "подго", "приго", "перед", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["кальци"]
	kt_kostey_taza_2, id_kt_kostey_taza, stop_kt_kostey_taza, kt_kostey_taza_3 = ["костей"], "КТ костей таза", ["дет", "реб", "описание ", "втор", "дефект", "подго", "приго", "перед", "вторич", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["таза", "тазовых"]
	kt_limfouzlov_2, id_kt_limfouzlov, stop_kt_limfouzlov, kt_limfouzlov_3 = ["лимфатическ", "лимфоузл"], "КТ лимфоузлов", ["дет", "описание ", "реб", "втор", "вторич", "подго", "приго", "перед", "дефект", "абонемент", "занятий", "посещени", "топометр", "скт", "мультиспиральн", "спирал"], ["узлов", "лимфоузл", "систем"]
	kt_licevogo_skelet_2, id_kt_licevogo_skelet, stop_kt_licevogo_skelet, kt_llicevogo_skelet_3 = ["лицев", "лица"], "КТ лицевого скелета", ["дет", "описание ", "реб", "подго", "приго", "перед", "втор", "дефект", "вторич", "абонемент", "занятий", "посещени", "топометр", "скт", "мультиспиральн", "спирал"], ["кост", "скелет"]
	kt_molo4nih_2, id_kt_molo4nih, stop_kt_molo4nih, kt_molo4nih_3 = ["молочн"], "КТ молочных желез", ["дет", "реб", "втор", "вторич", "абонемент", "описание ", "занятий", "подго", "приго", "перед", "посещени", "дефект", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["желез"]
	kt_m9gk_shei_2, id_kt_m9gk_shei, stop_kt_m9gk_shei, kt_m9gk_shei_3 = ["мягк"], "КТ мягких тканей шеи", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "описание ", "подго", "приго", "перед", "посещени", "дефект", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["шеи"]
	kt_nosoglotki_2, id_kt_nosoglotki, stop_kt_nosoglotki = ["носоглотк"], "КТ носоглотки", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "подго", "описание ", "приго", "перед", "топометр", "дефект", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_slun_jele_2, id_kt_slun_jele, stop_kt_slun_jele, kt_slun_jele_3 = ["слюн"], "КТ слюнной железы", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "подго", "описание ", "приго", "перед", "посещени", "дефект", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["желе"]
	kt_sredosten_2, id_kt_sredosten, stop_kt_sredosten = ["средостени"], "КТ средостения", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "подго", "описание ", "приго", "перед", "топометр", "дефект", "скт", "мультиспиральн", "спирал"]
	kt_stopi_2, id_kt_stopi, stop_kt_stopi = ["стопы", "стопа"], "КТ стопы", ["дет", "реб", "втор", "пальц", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "описание ", "приго", "перед", "контрастом", "дефект", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_shitovid_2, id_kt_shitovid, stop_kt_shitovid = ["щитовидн"], "КТ щитовидной железы", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "описание ", "подго", "приго", "перед", "топометр", "дефект", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_kolonoskop_2, id_kt_kolonoskop, stop_kt_kolonoskop = ["колоноскопи"], "КТ-виртуальная колоноскопия кишечника", ["дет", "реб", "втор", "вторич", "абонемент", "описание ", "занятий", "подго", "приго", "перед", "дефект", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "прием контраста"]
	kt_brush_polo_2, id_kt_brush_polo, stop_kt_brush_polo, kt_brush_polo_3 = ["брюшной"], "КТ брюшной полости", ["дет", "реб", "втор", "вторич", "абонемент", "описание ", "занятий", "подго", "приго", "перед", "дефект", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["полост"]
	kt_jeludok_2, id_kt_jeludok, stop_kt_jeludok = ["желудк"], "КТ желудка", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "полип", "топометр", "описание ", "контрастом", "подго", "приго", "перед", "дефект", "скт", "мультиспиральн", "спирал"]
	kt_jel4_puz_2, id_kt_kt_jel4_puz, stop_kt_jel4_puz, kt_jel4_puz_3 = ["желчн", "желоч"], "КТ желчного пузыря", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "описание ", "посещени", "подго", "приго", "перед", "дефект", "топометр", "скт", "мультиспиральн", "спирал"], ["пузыр"]
	kt_zabrush_2, id_kt_zabrush, stop_kt_zabrush = ["забрюшинн"], "КТ забрюшинного пространства", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "описание ", "топометр", "подго", "приго", "перед", "дефект", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_kishe4_2, id_kt_kishe4, stop_kt_kishe4 = ["кишечник", "кишки"], "КТ кишечника", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "описание ", "подго", "приго", "перед", "колоноскоп", "дефект", "виртуал", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_nadpo4e4_2, id_kt_nadpo4e4, stop_kt_nadpo4e4 = ["надпочечник"], "КТ надпочечников", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "подго", "описание ", "приго", "перед", "топометр", "контрастом", "дефект", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_pe4eni_2, id_kt_pe4eni, stop_kt_pe4eni = ["печен"], "КТ печени", ["офэ", "дет", "обеспеч", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "описание ", "приго", "перед", "контрастом", "дефект", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_podjeludo4_2, id_kt_podjeludo4, stop_kt_podjeludo4 = ["поджелудочн"], "КТ поджелудочной железы", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "подго", "описание ", "приго", "перед", "посещени", "топометр", "дефект", "скт", "мультиспиральн", "спирал"]
	kt_po4ek_2, id_kt_po4ek, stop_kt_po4ek = ["почек", "почки"], "КТ почек", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "контрастом", "подго", "описание ", "приго", "перед", "контрастированием", "дефект", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_selezen_2, id_kt_selezen, stop_kt_selezen = ["селезенк"], "КТ селезенки", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "описание ", "приго", "перед", "скт", "мультиспиральн", "спирал"]
	kt_golovi_2, id_kt_golovi, stop_kt_golovi = ["головы"], "КТ головы", ["дет", "реб", "ангиограф", "сосуд", "вен", "артери", "втор", "вторич", "абонемент", "подго", "приго", "описание ", "перед", "занятий", "дефект", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_viso4nih_2, id_kt_viso4nih, stop_kt_viso4nih, kt_viso4nih_3 = ["височн"], "КТ височных костей", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "подго", "приго", "описание ", "перед", "посещени", "дефект", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["костей", "кости"]
	kt_gipofiz_2, id_kt_gipofiz, stop_kt_gipofiz = ["гипофиз"], "КТ гипофиза", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "описание ", "перед", "контрастом", "дефект", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_glaza_2, id_kt_glaza, stop_kt_glaza = ["глаз", "орбит"], "КТ глаза / орбит", ["окт", "оптическая когерентная", "описание ", "дет", "реб", "втор", "вторич", "подго", "приго", "описание ", "перед", "абонемент", "занятий", "дефект", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_golov_mozga_2, id_kt_golov_mozga, stop_kt_golov_mozga = ["мозг"], "КТ головного мозга", ["дет", "пэт-кт", "ангиогр", "сосуд", "реб", "втор", "подго", "приго", "перед", "вторич", "описание ", "абонемент", "дефект", "занятий", "посещени", "топометр", "спинн", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_pazuh_2, id_kt_pazuh, stop_kt_pazuh = ["пазух носа", "носовых пазух", "придаточных пазух"], "КТ придаточных пазух носа", ["дет", "реб", "втор", "подго", "приго", "перед", "вторич", "описание ", "абонемент", "дефект", "занятий", "посещени", "топометр", "спинн", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_sedla_2, id_kt_sedla, stop_kt_sedla, kt_sedla_3 = ["турецк"], "КТ турецкого седла", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "подго", "приго", "перед", "посещени", "описание ", "топометр", "дефект", "спинн", "скт", "мультиспиральн", "спирал"], ["седл"]
	kt_4elust_2, id_kt_4elust, stop_kt_4elust = ["челюст", "дентальная", "дентальной"], "КТ челюсти", ["дет", "реб", "втор", "1-3", "1 з", "несколько зу", "до 3", "2-3", "одного зуба", "один зуб", "для", "консуль", "план", "лечен", "терапи", "ортопед", "прием", "анализ", "вторич", "абонемент", "занятий", "посещени", "подго", "клкт", "приго", "перед", "сустав", "височ", "описание ", "топометр", "дефект", "спинн", "скт", "мультиспиральн", "сустав", "внчс", "спирал"]
	kt_4erep_2, id_kt_4erep, stop_kt_4erep = ["череп"], "КТ черепа", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "перед", "спинн", "описание ", "контрастом", "дефект", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_grudnoy_kletki_2, id_kt_grudnoy_kletki, stop_kt_grudnoy_kletki, kt_grudnoy_kletki_3 = ["грудной"], "КТ грудной клетки", ["дет", "реб", "втор", "подго", "приго", "перед", "вторич", "описание ", "абонемент", "дефект", "занятий", "посещени", "топометр", "спинн", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["клетк", "полости"]
	kt_legkih_2, id_kt_legkih_2, stop_kt_legkih = ["легких"], "КТ легких", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "перед", "спинн", "описание ", "контрастом", "дефект", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_serdca_2, id_kt_serdca_2, stop_kt_serdca = ["сердечной", "сердца"], "КТ сердца", ["дет", "сосуд", "пэт-кт", "ангиограф", "реб", "втор", "вторич", "подго", "приго", "перед", "абонемент", "описание ", "занятий", "дефект", "посещени", "топометр", "спинн", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_bronhoskop_2, id_kt_bronhoskop_2, stop_kt_bronhoskop = ["бронхоскоп"], "КТ-виртуальная бронхоскопия", ["дет", "удаление", "инород", "реб", "втор", "вторич", "абонемент", "подго", "приго", "перед", "занятий", "описание ", "посещени", "дефект", "топометр", "спинн", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_malogo_taza_2, id_kt_malogo_taza_2, stop_kt_malogo_taza, kt_malogo_taza_3 = ["органов", "малого"], "КТ органов малого таза", ["дет", "реб", "втор", "подго", "приго", "перед", "вторич", "описание ", "абонемент", "дефект", "занятий", "посещени", "топометр", "кост", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["таза"]
	kt_matki_2, id_kt_matki_2, stop_kt_matki = ["матки"], "КТ матки", ["дет", "реб", "втор", "дефект", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "перед", "описание ", "дефект", "скт", "мультиспиральн", "спирал"]
	kt_mo4evogo_2, id_kt_mo4evogo_2, stop_kt_mo4evogo = ["мочевого", "мочевым", "мочевой пузырь"], "КТ мочевого пузыря", ["дет", "реб", "втор", "вторич", "подго", "приго", "перед", "абонемент", "описание ", "занятий", "дефект", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_prostati_2, id_kt_prostati_2, stop_kt_prostati = ["простат", "предстательн"], "КТ предстательной железы", ["дет", "реб", "пэт", "втор", "вторич", "абонемент", "подго", "приго", "перед", "занятий", "описание ", "посещени", "дефект", "топометр", "скт", "мультиспиральн", "спирал"]
	kt_9i4nik_2, id_kt_9i4nik_2, stop_kt_9i4nik = ["яичник"], "КТ яичников", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "перед", "скт", "мультиспиральн", "спирал"]
	kt_pozvono4_2, id_kt_pozvono4, stop_kt_pozvono4 = ["позвоночн"], "КТ позвоночника", ["дет", "реб", "один отдел", "одного отдела", "1 отдел", "2 отдел", "подго", "приго", "перед", "2 отдел", "два отдела", "описание ", "дефект", "двух отделов", "втор", "вторич", "абонемент", "занятий", "спирал", "прием контраста", "посещени", "топометр", "грудн", "копч", "поясни", "шейн", "шеи", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн"]
	kt_grud_pozvono4_2, id_kt_grud_pozvono4_2, stop_kt_grud_pozvono4, kt_grud_pozvono4_3 = ["позвоночн"], "КТ грудного отдела позвоночника", ["дет", "реб", "подго", "приго", "перед", "втор", "вторич", "дефект", "описание ", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["груд", "одного отдела", "1 отдел", "один отдел"]
	kt_kop4ik_2, id_kt_kop4ik_2, stop_kt_kop4ik = ["копчик"], "КТ копчика", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "перед", "контрастом", "описание ", "дефект", "скт", "мультиспиральн", "спирал"]
	kt_po9s_pozvon_2, id_kt_po9s_pozvon_2, stop_kt_po9s_pozvon, kt_po9s_pozvon_3 = ["позвоночн"], "КТ поясничного отдела позвоночника", ["дет", "реб", "втор", "подго", "приго", "перед", "вторич", "описание ", "дефект", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["поясни", "одного отдела", "1 отдел", "один отдел"]
	kt_shei_pozvon_2, id_kt_shei_pozvon_2, stop_kt_shei_pozvon, kt_shei_pozvon_3 = ["позвоночн"], "КТ шейного отдела позвоночника", ["дет", "реб", "втор", "вторич", "подго", "приго", "перед", "описание ", "дефект", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["шейн", "одного отдела", "1 отдел", "один отдел"]
	kt_vn4s_2, id_kt_vn4s_2, stop_kt_vn4s, kt_vn4s_3 = ["челюст", "внчс"], "КТ височно-нижнечелюстных суставов", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "подго", "приго", "перед", "описание ", "дефект", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"], ["сустав"]
	kt_golenostop_2, id_kt_golenostop, stop_kt_golenostop = ["голеностопн"], "КТ голеностопного сустава", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "подго", "приго", "перед", "дефект", "описание ", "посещени", "топометр", "скт", "мультиспиральн", "спирал"]
	kt_kolena_2, id_kt_kolena, stop_kt_kolena = ["колен"], "КТ коленного сустава", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "перед", "дефект", "описание ", "скт", "мультиспиральн", "спирал"]
	kt_lokt9_2, id_kt_lokt9, stop_kt_lokt9 = ["локтевого", "локтевых"], "КТ локтевого сустава", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "подго", "приго", "перед", "дефект", "описание ", "топометр", "скт", "мультиспиральн", "спирал"]
	kt_ple4_2, id_kt_ple4, stop_kt_ple4 = ["плечев"], "КТ плечевого сустава", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "подго", "приго", "перед", "описание ", "дефект", "скт", "мультиспиральн", "предплеч", "спирал"]
	kt_tazobedr_2, id_kt_tazobedr, stop_kt_tazobedr = ["тазобедр"], "КТ тазобедренного сустава", ["дет", "реб", "втор", "вторич", "абонемент", "занятий", "посещени", "подго", "приго", "перед", "топометр", "описание ", "дефект", "скт", "мультиспиральн", "спирал"]
	duktrograf, duktrograf_2, id_kt_druktograf, stop_kt_duktograf = ["дуктографи", "галактограф"],["дуктографи", "галактограф"], "Дуктография / Галактография", ["дет", "подго", "приго", "перед", "реб", "описание ", "втор", "вторич", "абонемент", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "спирал", "прием контраста", "введением контраст", "введение контраст", "скт", "мультиспиральн"]
	obzor_mammograf, obzor_mammograf_2, id_kt_obzor_mammograf, stop_obzor_mammograf = ["маммографи"], ["маммографи"], "Обзорная маммография", ["дет", "реб", "подго", "приго", "перед", "втор", "описани", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "спирал", "прием контраста", "мультиспиральн", "дукто", "галакто", "прицель"]
	pricel_mammograf, pricel_mammograf_2, id_pricel_mammograf, stop_pricel_mammograf = ["прицель"], ["маммограф"], "Прицельная маммография", ["дет", "реб", "втор", "подго", "приго", "перед", "описани", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн", "спирал", "прием контраста"]
	kt_lu4ezap_2, id_kt_lu4ezap, stop_kt_lu4ezap = ["лучезапястн"], "КТ лучезапястного сустава", ["дет", "реб", "втор", "вторич", "абонемент", "описани", "подго", "приго", "перед", "расшифров", "топометр", "описание ", "скт", "мультиспиральн", "спирал"]
	kt_kontrast_stop = ["дет", "реб", "втор", "вторич", "абонемент", "описани", "расшифров", "топометр", "спинн", "артери", "пэт", "венами", "вены", "вену", "дефект", "сосуд", "спирал", "скт", "подго", "приго", "описание ", "перед"]
	kt_for_kontrast = ["контрастом", "контрастным", "контрастированием", "введением контраст", "использованием контраст"]
	kt_brush_polo_2_k, id_kt_brush_polo_k = ["брюшной полости"], "КТ брюшной полости с контрастом"
	kt_gortani_2_k, id_kt_gortani_k = ["гортани"], "КТ гортани с контрастом"
	kt_viskov_2_k, id_kt_viskov_k = ["височных"], "КТ височных костей с контрастом"
	kt_gipofiz_2_k, id_kt_gipofiz_k = ["гипофиз"], "КТ гипофиза с контрастом"
	kt_glaza_2_k, id_kt_glaza_k = ["глаз", "орбит"], "КТ глаза / орбит с контрастом"
	kt_mozg_2_k, id_kt_mozg_k = ["мозга"], "КТ головного мозга с контрастом"
	kt_golova_2_k, id_kt_golova_k = ["головы"], "КТ головы с контрастом"
	kt_grud_pozvon2_k, id_kt_grud_pozvon_k = ["грудного отдела позвоноч"], "КТ грудного отдела позвоночника с контрастом"
	kt_grud_aorti_k, id_kt_grud_aorti_k = ["грудной аорты", "грудного отдела аорты"], "КТ грудной аорты с контрастом"
	kt_grud_kletki_k, id_kt_grud_kletki_k = ["грудной клетки"], "КТ грудной клетки с контрастом"
	kt_zabrush_k, id_kt_zabrush_k = ["забрюшинн"], "КТ забрюшинного пространства с контрастом"
	kt_kishe4_k, id_kt_kishe4_k = ["кишеч", "кишки"], "КТ кишечника с контрастом"
	kt_kostey_taza_k, id_kt_kostey_taza_k = ["костей таза", "тазовых костей"], "КТ костей таза с контрастом"
	kt_legkih_k, id_kt_legkih_k = ["легких"], "КТ легких с контрастом"
	kt_molo4nih_2_k, id_kt_molo4nih_k = ["молочных желез"], "КТ молочных желез с контрастом"
	kt_mo4evoy_2_k, id_kt_mo4evoy_k = ["мочевого"], "КТ мочевого пузыря с контрастом"
	kt_shei_2_k, id_kt_shei_k = ["тканей шеи"], "КТ мягких тканей шеи с контрастом"
	kt_nadpo4e4_2_k, id_kt_nadpo4e4_k = ["надпочечни"], "КТ надпочечников с контрастом"
	kt_nosoglot_2_k, id_kt_nosoglot_k = ["носоглотки", "носа и глотки", "глотки и носа"], "КТ носоглотки с контрастом"
	kt_organov_malogo_taza_2_k, id_kt_organov_malogo_taza_k = ["органов малого таза"], "КТ органов малого таза с контрастом"
	kt_pe4eni_2_k, id_kt_pe4eni_k = ["печени"], "КТ печени с контрастом"
	kt_po4ek_2_k, id_kt_po4ek_k = ["почек", "почки"], "КТ почек с контрастом"
	kt_po9s_pozvon_2_k, id_kt_po9s_pozvon_k = ["поясничного отдела позвоночника", "позвоночника в поясничном отделе"], "КТ поясничного отдела позвоночника с контрастом"
	kt_pridat_pazuh_2_k, id_kt_pridat_pazuh_k = ["придаточных пазух", "пазух носа", "носовых пазух", "придаточные пазухи"], "КТ придаточных пазух носа с контрастом"
	kt_serdca_2_k, id_kt_serdca_k = ["сердца"], "КТ сердца с контрастом"
	kt_slunnih_2_k, id_kt_slunnih_k = ["слюнн"], "КТ слюнной железы с контрастом"
	kt_stopi_2_k, id_kt_stopi_k = ["стопы"], "КТ стопы с контрастом"
	kt_4erepa_2_k, id_kt_4erepa_k = ["черепа"], "КТ черепа с контрастом"
	kt_sheynogo_pozvon_2_k, id_kt_sheynogo_pozvon_k = ["шейного отдела позвоночн" , "позвоночника в шейном отделе"], "КТ шейного отдела позвоночника с контрастом"
	kt_shitovid_2_k, id_kt_shitovid_k = ["щитовидной"], "КТ щитовидной железы с контрастом"
	ren = ["рентген", "rg-графи"]
	ren_bedra, id_ren_bedra, stop_ren_bedra = ["бедра", "бедрен"], "Рентген бедра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "блокада", "курс", "тазобедр"]
	ren_viso4, id_ren_viso4, stop_ren_viso4 = ["височны", "височной" , "висков"], "Рентген височных костей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_gortan, id_ren_gortan, stop_ren_gortan = ["гортан"], "Рентген гортани", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_grudini, id_ren_grudini, stop_ren_grudini = ["грудины"], "Рентген грудины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_kisti, id_ren_kisti, stop_ren_kisti = ["кисти", "кисть", "кистей"], "Рентген кисти", ["дет", "пальц", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_klu4ic, id_ren_kLu4ic, stop_ren_klu4ic = ["ключиц"], "Рентген ключицы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_kostey_nosa, id_ren_kostey_nosa, stop_ren_koste_nosa, ren_kostey_nosa_2 = ["носа", "носовых"], "Рентген костей носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["кост"]
	ren_kost_taz, id_ren_kost_taz, stop_ren_kost_taz, ren_kost_taza_2 = ["таза", "тазовых"], "Рентген костей таза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["кост"]
	ren_legkih, id_ren_legkih, stop_ren_legkih, ren_legkih_2 = ["легких", "грудной"], "Рентген легких / грудной клетки", ["дет", "реб", "втор", "доп", "описани", "расшифров", "занятий", "посещени", "топометр"], ["легких", "клетки"]
	ren_lopatki, id_ren_lopatki, stop_ren_lopatki = ["лопатки", "лопаток"], "Рентген лопатки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_nadkolennik, id_ren_nadkolennik, stop_ren_nadkolennik = ["надколенник"], "Рентген надколенника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "коленного", "коленей", "колена", "коленных"]
	ren_nosoglotki, id_ren_nosoglotki, stop_ren_nosoglotki = ["носоглотк"], "Рентген носоглотки", ["дет", "реб", "втор", "описани", "расшифров", "направлению", "направления", "занятий", "посещени", "топометр"]
	ren_palca, id_ren_palca, stop_ren_palca = ["пальц"], "Рентген пальца", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_predple4, id_ren_predple4, stop_ren_predple4 = ["предплеч"], "Рентген предплечья", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_pazuh_nos, id_ren_pazuh_nos, stop_ren_pazuh_nos, ren_pazuh_nos_2 = ["пазух"], "Рентген придаточных пазух носа", ["дет", "реб", "втор", "для", "описани", "расшифров", "занятий", "посещени", "топометр"], ["придаточ", "носа", "носовых"]
	ren_p9t_kost, id_ren_p9t_kost, stop_ren_p9t_kost, ren_p9t_kost_2 = ["пят"], "Рентген пяточной кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["кост"]
	ren_reber, id_ren_reber, stop_ren_reber = ["ребер", "ребра"], "Рентген ребер", ["дет", "ребен", "втор", "запись", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_skul_kost, id_ren_skul_kost, stop_ren_skul, ren_skul_kost_2 = ["скул"], "Рентген скуловой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["кост"]
	ren_stopi, id_ren_stopi, stop_ren_stopi = ["стопы"], "Рентген стопы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "нагруз", "пальц"]
	ren_stopi_nagruz, id_ren_stopi_nagruz, stop_ren_stopi_nagruz, ren_stopi_nagruz_2 = ["стоп"], "Рентген стопы с нагрузкой", ["дет", "реб", "втор", "голе", "описани", "расшифров", "занятий", "посещени", "топометр"], ["нагруз"]
	ren_taza, id_ren_taza, stop_ren_taza = ["таза"], "Рентген таза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "кост", "суста"]
	ren_tur_sed, id_ren_tur_sed, stop_ren_tur_sed, ren_tur_sed_2 = ["туре"], "Рентген турецкого седла", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["седл"]
	ren_4erep, id_ren_4erep, stop_ren_4erep = ["череп"], "Рентген черепа", ["дет", "реб", "втор", "доп", "теле", "трг", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_brushnoy, id_ren_brushnoy, stop_ren_brushnoy, ren_brushnoy_2 = ["брюшной"], "Рентген брюшной полости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["полост"]
	ren_jeludka, id_ren_jeludka, stop_ren_jeludka = ["желудк"], "Рентген желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_pishevod, id_ren_pishevod, stop_ren_pishevod = ["пищевод"], "Рентген пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_po4ek, id_ren_po4ek, stop_ren_po4ek = ["почек"], "Рентген почек", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_pozvono4_grud, id_ren_pozvono4_grud, stop_ren_pozvono4_grud, ren_pozvono4_grud_2 = ["грудн"], "Рентген грудного отдела позвоночника", ["дет", "реб", "втор", "блокада", "курс", "описани", "расшифров", "доп", "занятий", "посещени", "топометр"], ["позвоночн", "поз-ка"]
	ren_pozvono4_po9s_krest, id_ren_pozvono4_po9s_krest, stop_ren_pozvono4_po9s_krest, ren_pozvono4_po9s_krest_2 = ["поясни"], "Рентген пояснично-крестцового отдела позвоночника", ["дет", "реб", "доп", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["крест"]
	ren_pozvono4_po9s, id_ren_pozvono4_po9s, stop_ren_pozvono4_po9s = ["поясни"], "Рентген поясничного отдела позвоночника", ["дет", "реб", "втор", "блокада", "с функц", "с проб", "курс", "описани", "расшифров", "занятий", "посещени", "доп", "крест"]
	ren_pozvono4_shei, id_ren_pozvono4_shei, stop_ren_pozvono4_shei = ["шейн", "шеи"], "Рентген шейного отдела позвоночника", ["дет", "реб", "втор", "с функц", "с проб", "блокада", "курс", "описани", "позвонка", "расшифров", "занятий", "доп", "посещени"]
	main_func(["рентген", "rg-графи"], ["шейн", "шеи"], "Рентген шейного отдела позвоночника с функциональными пробами", ["дет", "реб", "втор", "блокада", "курс", "описани", "позвонка", "расшифров", "занятий", "доп", "посещени"], ["с функц", "с проб"])
	main_func(["рентген", "rg-графи"], ["поясни"], "Рентген поясничного отдела позвоночника с функциональными пробами", ["дет", "реб", "втор", "блокада", "курс", "описани", "расшифров", "занятий", "посещени", "доп", "крест"], ["с функц", "с проб"])
	ren_kop4ik, id_ren_kop4ik, stop_ren_kop4ik = ["копчика"], "Рентген копчика", ["дет", "реб", "втор", "описани", "расшифров", "доп", "занятий", "посещени"]
	ureteropiel, ureteropiel_2, id_ureteropiel, stop_ureteropiel = ["уретеропиелограф"], ["уретеропиелограф"], "Ретроградная уретеропиелография", ["дет", "антероград", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "восходящ"]
	fistulograf, fistulograf_2, id_fistulograf, stop_fistulograf = ["фистулографи"], ["фистулографи"], "Фистулография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	holangiograf, holangiograf_2, id_holangiograf, stop_holangiograf = ["холангиографи"], ["холангиографи"], "Холангиография", ["дет", "магнит", "мрт", "мр-", "мр ", "мр -", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	laterograf_jel, laterograf_jel_2, id_laterograf_jel, stop_laterograf_jel = ["латерографи"], ["желудк"], "Латерография желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	sialograf, sialograf_2, id_sialograf, stop_sialograf = ["сиалография", "сиалографии", "рентген"], ["сиалография", "сиалографии", "слюнн"], "Сиалография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "восходящ"]
	cistograf, cistograf_2, id_cistograf, stop_cistograf = ["цистография", "цистографии", "рентген"], ["цистография", "цистографии", "мочево"], "Цистография", ["дакрио", "слез" ,"дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "восходящ"]
	rebenok = ["ребенку", "ребенок", "ребенка", "дети", "детям"]
	ren_brush_reb, id_ren_brush_reb, stop_ren_brush_reb = ["брюшной полости"], "Рентген брюшной полости ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_viso4_kost_reb, id_ren_viso4_kost_reb, stop_ren_viso4_kost_reb = ["височных костей", "костей виска"], "Рентген височных костей ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_grud_pozvon_reb, id_ren_grud_pozvon_reb, stop_ren_grud_pozvon_reb = ["грудного отдела позвоночн", "грудной отдел позвоночн"], "Рентген грудного отдела позвоночника ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_po9s_pozvon_reb, id_ren_po9s_pozvon_reb, stop_ren_po9s_pozvon_reb = ["поясничного отдела позвоночн", "поясничный отдел позвоночн"], "Рентген поясничного отдела позвоночника ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_shei_pozvon_reb, id_ren_shei_pozvon_reb, stop_ren_shei_pozvon_reb = ["шейного отдела позвоночн", "шейный отдел позвоночн"], "Рентген шейного отдела позвоночника ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_jelud_reb, id_ren_jelud_reb, stop_ren_jelud_reb = ["желудк"], "Рентген желудка ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_kisti_reb, id_ren_kisti_reb, stop_ren_kisti_reb = ["кисти"], "Рентген кисти ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_kost_taz_reb, id_ren_kost_taz_reb, stop_ren_kost_taz_reb = ["костей таза", "тазовых костей"], "Рентген костей таза ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_legkih_reb, id_ren_legkih_reb, stop_ren_legkih_reb = ["легких", "грудной клетки"], "Рентген легких / грудной клетки ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_nosoglot_reb, id_ren_nosoglot_reb, stop_ren_nosoglot_reb = ["носоглотки"], "Рентген носоглотки ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_orbit_reb, id_ren_orbit_reb, stop_ren_orbit_reb = ["орбит"], "Рентген орбит ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_pishevod_reb, id_ren_pishevod_reb, stop_ren_pishevod_reb = ["пищевод"], "Рентген пищевода ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_pridat_reb, id_ren_pridat_reb, stop_ren_pridat_reb = ["придаточных пазух", "пазух носа", "носовых пазух"], "Рентген придаточных пазух носа ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_stopi_reb, id_ren_stopi_reb, stop_ren_stopi_reb = ["стопы"], "Рентген стопы ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_tazobed_reb, id_ren_tazobed_reb, stop_ren_tazobed_reb = ["тазобедренн"], "Рентген тазобедренного сустава ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_ton_kish_reb, id_ren_ton_kish_reb, stop_ren_ton_kish_reb = ["тонкого кишечника", "тонкой кишки"], "Рентген тонкого кишечника ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_trub_kost_reb, id_ren_trub_kost_reb, stop_ren_trub_kost = ["трубчаты"], "Рентген трубчатых костей ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_turec_reb, id_ren_turec_reb, stop_ren_turec_kost = ["турецк"], "Рентген турецкого седла ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_4elust_reb, id_ren_4elust_reb, stop_ren_4elust_kost = ["челюсти"], "Рентген челюсти ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_4erepa_reb, id_ren_4erepa_reb, stop_ren_4erepa_kost = ["черепа"], "Рентген черепа ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_golenostop, id_ren_ren_golenostop, stop_ren_ren_golenostop = ["голеностоп"], "Рентген голеностопного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_kolena, id_ren_ren_kolena, stop_ren_ren_kolena = ["колен"], "Рентген коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "надколен"]
	ren_krest_podvzdo, id_ren_krest_podvzdo, stop_ren_krest_podvzdo, ren_krest_podvzdo_2 = ["крестцов"], "Рентген крестцово-подвздошных суставов", ["дет", "реб", "блокада", "курс", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"], ["повздошн", "подвздошн"]
	ren_lokt9, id_ren_lokt9, stop_ren_lokt9 = ["локт"], "Рентген локтевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_lu4ezap, id_ren_lu4ezap, stop_ren_lu4ezap = ["лучезапяст"], "Рентген лучезапястного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_ple4a, id_ren_ple4a, stop_ren_ple4a = ["плечевого", "плечевой"], "Рентген плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	ren_tazobedr, id_ren_tazobedr, stop_ren_tazobedr = ["тазобедренн"], "Рентген тазобедренного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	duodenograf, duodenograf_2, id_duodenograf, stop_duodenograf = ["дуоденография"], ["дуоденография"], "Релаксационная дуоденография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр"]
	mrt = ["мрт", "мр-", "мр -", "мр ", "магнито-резонанс", "магниторезонанс", "магнитно-резонанс", "магнитнорезонанс", "магнито резонанс", "магнитно резонанс"]
	mrt_vse_telo, id_mrt_vse_telo, stop_mrt_vse_telo = ["всего тела", "тела полностью", "онко-поиск", "поиск метаста", "онкопоиск", "онкоскрининг"], "МРТ всего тела", ["дет", "реб", "втор", "описание ", "подго", "приго", "перед", "расшифров", "занятий", "посещени", "топометр"] 
	mrt_gortan, id_mrt_gortan, stop_mrt_gortan = ["гортани"], "МРТ гортани", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр"] 
	mrt_kist, id_mrt_kist, stop_mrt_kist = ["кисти", "кисть"], "МРТ кисти", ["дет", "реб", "пальц", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_kraniover, id_mrt_kraniover, stop_mrt_kraniover = ["краниовертебра"], "МРТ краниовертебрального перехода", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_limfouzl, id_mrt_limfouzl, stop_mrt_limfouzl = ["лимфоузлов"], "МРТ лимфоузлов", ["дет", "реб", "втор", "описание ", "расшифров", "подго", "приго", "перед", "занятий", "посещени", "топометр"] 
	mrt_molo4, id_mrt_molo4, stop_mrt_molo4 = ["молочных желез", "молочной железы", "грудных желез", "грудной железы"], "МРТ молочных желез", ["дет", "подго", "приго", "перед", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_moshonki, id_mrt_moshonki, stop_mrt_moshonki = ["мошонки"], "МРТ мошонки", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_m9gk_tkan, id_mrt_m9gk_tkan, stop_mrt_m9gk_tkan = ["мягких тканей"], "МРТ мягких тканей", ["дет", "реб", "втор", "шеи", "описание ", "расшифров", "подго", "приго", "перед", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_shei, id_mrt_shei, stop_mrt_shei = ["мягких тканей шеи", "мягкие ткани шеи"], "МРТ мягких тканей шеи", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_ple4sple, id_mrt_ple4sple, stop_mrt_ple4sple = ["плечевого сплетения", "плечевых сплетений"], "МРТ плечевого сплетения", ["дет", "реб", "втор", "подго", "приго", "перед", "описание ", "расшифров", "занятий", "посещени", "топометр"] 
	mrt_plod, id_mrt_plod, stop_mrt_plod = ["плода"], "МРТ плода", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "подго", "приго", "перед", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_4len, id_mrt_4len, stop_mrt_4len = ["члена"], "МРТ полового члена", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_slun, id_mrt_slun, stop_mrt_slun = ["слюнн"], "МРТ слюнных желез", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр"] 
	mrt_spinnogo, id_mrt_spinnogo, stop_mrt_spinnogo = ["спинного мозга", "миелографи"], "МРТ спинного мозга", ["дет", "реб", "втор", "описание ", "подго", "приго", "перед", "рентген", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_sredosten, id_mrt_sredosten, stop_mrt_sredoten = ["средостения"], "МРТ средостения", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_stopi, id_mrt_stopi, stop_mrt_stopi = ["стопы", "стопа"], "МРТ стопы", ["дет", "реб", "пальц", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_shitovid, id_mrt_shitovid, stop_mrt_shitovid = ["щитовидной"], "МРТ щитовидной железы", ["дет", "реб", "втор", "описание ", "расшифров", "подго", "приго", "перед", "занятий", "посещени", "топометр"] 
	mrt_brush, id_mrt_brush, stop_mrt_brush = ["брюшной полости", "брюшная полость"], "МРТ брюшной полости", ["дет", "реб", "втор", "описание ", "онко", "метастаз", "всего тела", "все тело", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_jelud, id_mrt_jelud, stop_mrt_jelud = ["желудка"], "МРТ желудка", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр"] 
	mrt_zabrush, id_mrt_zabrush, stop_mrt_zabrush = ["забрюшинн"], "МРТ забрюшинного пространства", ["дет", "реб", "втор", "описание ", "расшифров", "подго", "приго", "перед", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_nadpo4e4, id_mrt_nadpo4e4, stop_mrt_nadpo4e4 = ["надпочечник"], "МРТ надпочечников", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_pe4en, id_mrt_pe4en, stop_mrt_pe4en = ["печени", "печень"], "МРТ печени", ["дет", "реб", "втор", "эластограф", "обеспеч", "описание ", "расшифров", "занятий", "онко", "поиск", "всего тела", "все тело", "метастаз", "посещени", "подго", "приго", "перед", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_pishevod, id_mrt_pishevod, stop_mrt_pishevod = ["пищевода"], "МРТ пищевода", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр"] 
	mrt_podjel, id_mrt_podjel, stop_mrt_podjel = ["поджелудочн"], "МРТ поджелудочной железы", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр"] 
	mrt_po4ek, id_mrt_po4ek, stop_mrt_po4ek = ["почек", "почки"], "МРТ почек", ["дет", "реб", "втор", "описание ", "предыдущ", "наличии", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_golova, id_mrt_golova, stop_mrt_golova = ["головы"], "МРТ головы", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_spektroskop, id_mrt_spektroskop, stop_mrt_spektroskop = ["спектроскоп"], "МР-спектроскопия головного мозга", ["дет", "подго", "приго", "без спектро", "перед", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mrt_bn4s, id_mrt_vn4s, stop_mrt_vn4s = ["внчс", "нижнечелюстн"], "МРТ височно-нижнечелюстных суставов", ["дет", "реб", "втор", "описание ", "подго", "приго", "перед", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_gipofiz, id_mrt_gipofiz, stop_mrt_gipofiz = ["гипофиз"], "МРТ гипофиза", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_gippokmap, id_mrt_gippo, stop_mrt_gippo = ["гиппокамп"], "МРТ гиппокампа", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_glaz, id_mrt_glaz, stop_mrt_glaz = ["глаз", "орбит"], "МРТ глаза / орбит", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_golovnogo, id_mrt_golovnogo, stop_mrt_golovnogo = ["головного мозга"], "МРТ головного мозга", ["дет", "ангиограф", "перфузи", "сосуд", "артери", "подго", "приго", "перед", "вен", "реб", "втор", "открыт", "спектроскоп", "цистерн", "каверн", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_licnerv, id_mrt_licnerv, stop_mrt_licnerv = ["лицевого нерва", "лицевых нервов"], "МРТ лицевого нерва", ["дет", "реб", "втор", "описание ", "подго", "приго", "перед", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_pridatpazuh, id_mrt_pridatpazuh, stop_mrt_pridatpazuh = ["пазух носа", "носовых пазух", "придаточных пазух", "пазух в носу", "придаточные пазухи"], "МРТ придаточных пазух носа", ["дет", "подго", "приго", "перед", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_uha, id_mrt_uha, stop_mrt_uha = [" уха", "улит"], "МРТ структуры внутреннего уха", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_troynerv, id_mrt_troynerv, stop_mrt_troynerv = ["тройничн"], "МРТ тройничного нерва", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_sedlo, id_mrt_sedlo, stop_mrt_sedlo = ["седла"], "МРТ турецкого седла", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "подго", "приго", "перед"] 
	mrt_4erep, id_mrt_4erep, stop_mrt_4erep = ["черепа"], "МРТ черепа", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "подго", "приго", "перед", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_4erepnerv, id_mrt_4erepnerv, stop_mrt_4erepnerv = ["черепных нервов"], "МРТ черепных нервов", ["дет", "реб", "втор", "описание ", "расшифров", "подго", "приго", "перед", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_grudkletka, id_mrt_grudkletka, stop_mrt_grudkletka = ["грудной клетки", "грудной полости"], "МРТ грудной клетки", ["дет", "реб", "втор", "подго", "приго", "перед", "описание ", "расшифров", "занятий", "посещени", "топометр"] 
	mrt_legkih, id_mrt_legkih, stop_mrt_legkih = ["легких"], "МРТ легких", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "подго", "приго", "перед", "топометр"] 
	mrt_serdca, id_mrt_serdca, stop_mrt_serdca = ["сердца", "сердечной мышцы"], "МРТ сердца", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_holangio, id_mrt_holangio, stop_mrt_holangio = ["холангио", "МРХПГ"], "МР-холангиография", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "подго", "приго", "онко", "всего тела", "все тело", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_enterograf, id_mrt_enterograf, stop_mrt_enterograf = ["энтерография", "энтерографии"], "МР-энтерография", ["дет", "реб", "втор", "описание ", "подго", "приго", "перед", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_bedrkost, id_mrt_bedrkost, stop_mrt_bedrkost = ["бедра", "бедренной кости", "кости бедра", "костей бедра", "бедренных костей"], "МРТ бедренной кости", ["дет", "мягких тканей", "мягкие ткани", "подго", "приго", "перед", "реб", "втор", "описание ", "сустав", "расшифров", "занятий", "посещени", "топометр"] 
	mrt_viskost, id_mrt_viskost, stop_mrt_viskost = ["височных костей", "костей виска", "височной кости"], "МРТ височных костей", ["дет", "реб", "втор", "подго", "приго", "перед", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_kosttaza, id_mrt_kosttaza, stop_mrt_kosttaza = ["костей таза", "тазовых костей", "костей малого таза"], "МРТ костей таза", ["дет", "реб", "втор", "подго", "приго", "перед", "описание ", "расшифров", "занятий", "посещени", "топометр"] 
	mrt_taz, id_mrt_taz, stop_mrt_taz = ["малого таза"], "МРТ малого таза", ["дет", "реб", "втор", "мужчин", "без органов", "не включая органы", "без малого та", "женщин", "описание ", "расшифров", "занятий", "подго", "приго", "перед", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_tazwomen, id_mrt_tazwomen, stop_mrt_tazwomen, mrt_tazwomen_2 = ["малого таза"], "МРТ малого таза у женщин", ["дет", "реб", "втор", "описание ", "подго", "приго", "перед", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["женщин"]
	mrt_tazmen, id_mrt_tazmen, stop_mrt_tazmen, mrt_tazmen_2 = ["малого таза"], "МРТ малого таза у мужчин", ["дет", "реб", "втор", "описание ", "расшифров", "подго", "приго", "перед", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["мужчин"] 
	mrt_matki, id_mrt_matki, stop_mrt_matki = ["матки"], "МРТ матки", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "подго", "приго", "перед", "онко", "метастаз", "всего тела", "все тело"] 
	mrt_mo4evoy, id_mrt_mo4evoy, stop_mrt_mo4evoy = ["мочевого", "мочевым пузырем"], "МРТ мочевого пузыря", ["дет", "реб", "втор", "описание ", "расшифров", "подго", "приго", "перед", "занятий", "онко", "метастаз", "всего тела", "все тело", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_prostati, id_mrt_prostati, stop_mrt_prostati = ["простаты", "предстательн"], "МРТ простаты / предстательной железы", ["дет", "реб", "втор", "описание ", "расшифров", "онко", "метастаз", "всего тела", "все тело", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_9i4nik, id_mrt_9i4nik, stop_mrt_9i4nik = ["яичник"], "МРТ яичников", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "подго", "приго", "перед", "онко", "метастаз", "всего тела", "все тело"] 
	mrt_golovnogo_open, id_mrt_golovnogo_open, stop_mrt_golovnogo_open, mrt_golovnogo_open_2 = ["головного мозга", "мозга"], "МРТ головного мозга на аппарате открытого типа", ["дет", "реб", "спектроск", "цистерн",  "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["открыт"]
	mrt_kolen_open, id_mrt_kolen_open, stop_mrt_kolen_open, mrt_kolen_open_2 = ["колен"], "МРТ коленного сустава на аппарате открытого типа", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["открыт"]
	mrt_pozvon_open, id_mrt_pozvon_open, stop_mrt_pozvon_open, mrt_pozvon_open_2 = ["позвоночник"], "МРТ позвоночника на аппарате открытого типа", ["дет", "груд", "поясни", "копч", "крестц", "шейн", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["открыт"]
	mrt_grudpozvon, id_mrt_grudpozvon, stop_mrt_grudpozvon, mrt_grudpozvon_2 = ["грудн"], "МРТ грудного отдела позвоночника", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["позвоноч", "отдел"]
	mrt_kop4ik, id_mrt_kop4ik, stop_mrt_kop4ik = ["копчик"], "МРТ копчика", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед"] 
	mrt_krestpodvz, id_mrt_krestpodzv, stop_mrt_krestpodzv, mrt_krestpodzv_2 = ["крестц", "кресц"], "МРТ крестцово-подвздошных сочленений", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["подвз", "повздо"]
	mrt_po9spozvon, id_mrt_po9spozvon, stop_mrt_po9spozvon, mrt_po9spozvon_2 = ["пояс"], "МРТ поясничного отдела позвоночника", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["позвон", "отдел"]
	mrt_sheinogo, id_mrt_sheinogo, stop_mrt_sheinogo, mrt_sheinogo_2 = ["шейн"], "МРТ шейного отдела позвоночника", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "артери", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"], ["позвон", "отдел"]
	mrt_golenostop, id_mrt_golenostop, stop_mrt_golenostop = ["голеностоп"], "МРТ голеностопного сустава", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_kolennogo, id_mrt_kolennogo, stop_mrt_kolennogo = ["колен"], "МРТ коленного сустава", ["дет", "реб", "втор", "открыт", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_lokt9, id_mrt_lokt9, stop_mrt_lokt9 = ["локт"], "МРТ локтевого сустава", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "подго", "приго", "перед"] 
	mrt_lu4ezap, id_mrt_lu4ezap, stop_mrt_lu4ezap = ["лучезап"], "МРТ лучезапястного сустава", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "подго", "приго", "перед"] 
	mrt_ple4evogo, id_mrt_ple4evogo, stop_mrt_ple4evogo = ["плечевого сустава", "суставов плеч", "плечевых суставов", "плечевой сустав"], "МРТ плечевого сустава", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_tazobedr, id_mrt_tazobedr, stop_mrt_razobedr = ["тазобедренн"], "МРТ тазобедренного сустава", ["дет", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"] 
	mrt_pozvon, id_mrt_pozvon, stop_mrt_pozvon = ["позвоночник", "шейного, грудного, пояснично-крестцового"], "МРТ позвоночника", ["дет", "отдела", "двух отделов", "2 отделов", "2-х отделов", "груд", "поясни", "копч", "крестц", "шейн", "открыт", "реб", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_angiograf_mozga_reb, id_mr_angiograf_mozga_reb, stop_mr_angiograf_mozga_reb = ["ангиография сосудов головного мозга", "ангиография сосудов мозга"], "МР-ангиография сосудов головного мозга ребенку", ["артери", "наличии", "проведенного", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_brushpol_reb, id_mrt_brupshpol_reb, stop_mrt_brushpol_reb = ["брюшной полости"], "МРТ брюшной полости ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "онко", "всего тела", "метастаз", "всё тело", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_gipofiz_reb, id_mr_gipofiz_reb, stop_mr_gipofiz_reb = ["гипофиза"], "МРТ гипофиза ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_golovnogo_reb, id_mr_golovnogo_reb, stop_mr_golovnogo_reb = ["головного мозга"], "МРТ головного мозга ребенку", ["втор", "сосуд", "артер", "вен", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_grudklet_reb, id_mr_grudklet_reb, stop_mr_grudklet_reb = ["грудной клетки"], "МРТ грудной клетки ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_kolena_reb, id_mr_kolena_reb, stop_mr_kolena_reb = ["колен"], "МРТ коленного сустава ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_malogo_taza_reb, id_mr_malogotaza_reb, stop_mr_maliyyaz_reb = ["малого таза"], "МРТ малого таза ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_nadpo4e4nik_reb, id_mr_nadpo4e4nik_reb, stop_mr_nadpo4e4nik_reb = ["надпочечник"], "МРТ надпочечников ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_pazuh_reb, id_mr_pazuh_reb, stop_mr_pazuh_reb = ["пазух носа", "придаточных пазух", "носовых пазух", "пазух в носу"], "МРТ пазух носа ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_po4ek_reb, id_mr_po4ek_reb, stop_mr_po4ek_reb = ["почек"], "МРТ почек ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_serdca_reb, id_mr_serdca_reb, stop_mr_serdca_reb = ["сердца"], "МРТ сердца ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mr_tazobedr_reb, id_mr_tazobedr_reb, stop_mr_tazobedr_reb = ["тазобедрен"], "МРТ тазобедренного сустава ребенку", ["втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "контрастом", "подго", "приго", "перед", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст"]
	mrt_for_kontrast = ["контрастом", "контрастным", "контрастированием", "введением контраст", "введение контраст", "использованием контраст", "прием контраст", "введением контраст", "примовист"]
	mr_brushpolk, id_brushpolk, stop_mr_brushpolk = ["брюшной полости"], "МРТ брюшной полости с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_vn4sk, id_vn4sk, stop_mr_vn4sk = ["височно-нижнечелюстных", "внчс", "височно-нижнечелюстного"], "МРТ височно-нижнечелюстных суставов с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_gipofizk, id_gipofizk, stop_mr_gipofizk = ["гипофиз"], "МРТ гипофиза с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_glazk, id_glazk, stop_mr_glazk = ["глаз", "орбит"], "МРТ глаза с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_golenostopk, id_golenostopk, stop_mr_golenostopk = ["голеностопн"], "МРТ голеностопного сустава с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_golovnoyk, id_golovnoyk, stop_mr_golovnoyk = ["головного мозга", "головным мозгом", "мозг"], "МРТ головного мозга с контрастом", ["дет", "спинн", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_golovik, id_golovik, stop_mr_golovik = ["головы"], "МРТ головы с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_grudpozvon, id_mrt_grudpozvonk, stop_mrt_grudpozvonk = ["грудного отдела позвоночни", "позвоночника в грудной отдел", "грудной отдел позвоно", "1 отдел"], "МРТ грудного отдела позвоночника с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_zabrushk, id_zabrushk, stop_mr_zabrushk = ["забрюшинного про"], "МРТ забрюшинного пространства с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_kistkik, id_kistkik, stop_mr_kistkik = ["кисти", "кисть"], "МРТ кисти с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_kolenak, id_kolenak, stop_mr_kolenak = ["колен"], "МРТ коленного сустава с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_krestpodvk, id_krestpodvk, stop_mr_krestpodvk = ["крестцово-подвздошн", "крестцово подвздошн"], "МРТ крестцово-подвздошных сочленений с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_maliytazk, id_maliytazk, stop_mr_maliytazk = ["малого таза"], "МРТ малого таза с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_molo4k, id_molo4k, stop_mr_molo4k = ["молочных желез", "грудных желез", "молочной железы", "грудной железы"], "МРТ молочных желез с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_mo4evoyk, id_mo4evoyk, stop_mr_mo4evoyk = ["мочевого пузыря", "мочевым пузырем"], "МРТ мочевого пузыря с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_moshonkik, id_moshonkik, stop_mr_moshonkik = ["мошонки"], "МРТ мошонки с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_m9gkihk, id_m9gkihk, stop_mr_m9gkihk = ["мягких тканей"], "МРТ мягких тканей с контрастом", ["дет", "шеи", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_m9gksheik, id_m9gksheik, stop_mr_m9gksheik = ["мягких тканей шеи"], "МРТ мягких тканей шеи с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_nadpo4e4k, id_nadpo4e4k, stop_mr_nadpo4e4k = ["надпочечник"], "МРТ надпочечников с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_pe4enik, id_pe4enik, stop_mr_pe4enik = ["печени", "печень"], "МРТ печени с контрастом", ["дет", "реб", "обеспеч", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_ple4esusk, id_ple4esusk, stop_mr_ple4esusk = ["плечевого сустава", "плечевых суставов", "суставов плеч", "плечевой сустав"], "МРТ плечевого сустава с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_pozvonkon, id_pozvonkon, stop_mr_pozvonkon = ["позвоночник"], "МРТ позвоночника с контрастом", ["дет", "реб", "подго", "приго", "перед", "грудн", "поясн", "шейн", "1 отдел", "один отдел" "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_4lenk, id_4lenk, stop_mr_4lenk = ["члена"], "МРТ полового члена с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_po4ekk, id_po4ekk, stop_mr_po4ekk = ["почек"], "МРТ почек с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_po9skk, id_po9skk, stop_mr_po9skk = ["поясничн", "1 отдел"], "МРТ поясничного отдела позвоночника с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_pridatkk, id_pridatkk, stop_mr_pridatkk = ["придаточных пазух", "пазух носа", "носовых пазух", "придаточные пазухи"], "МРТ придаточных пазух носа с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_prostatik, id_prostatik, stop_mr_prostatik = ["простаты", "предстательн"], "МРТ простаты с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_serdcak, id_sercak, stop_mr_sercak = ["сердца"], "МРТ сердца с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_spinnoyk, id_spinnoyk, stop_mr_spinnoyk = ["спинного мозга"], "МРТ спинного мозга с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_stopik, id_stopik, stop_mr_stopik = ["стопы", "стопа"], "МРТ стопы с контрастом", ["дет", "реб", "подго", "приго", "перед", "пальц", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_sustavk, id_sustavk, stop_mr_sustavk = ["суставов"], "МРТ суставов с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "челюст", "внчс", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_tazobedk, id_tazobedk, stop_mr_tazobedk = ["тазобедренн"], "МРТ тазобедренного сустава с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_4erepk, id_4erepk, stop_mr_4erepk = ["черепа"], "МРТ черепа с контрастом", ["дет", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	mr_sheynogok, id_sheynogok, stop_mr_sheynogok = ["шейного отдела позвоночника", "1 отдел"], "МРТ шейного отдела позвоночника с контрастом", ["дет", "ангиограф", "реб", "подго", "приго", "перед", "втор", "описание ", "расшифров", "занятий", "посещени", "топометр", "без учета контраст", "без констраст"]
	anorek_manomentr, anorek_manomentr_2, id_anorek_manonemtr, stop_anorek_manometr = ["аноректальн"], ["манометри"], "Аноректальная манометрия", ["втор", "описани", "расшифров", "занятий", "посещени"]
	keratometr, keratometr_2, id_keratometr, stop_keratometr = ["кератометри", "кривизны роговицы", "кривизны роговой оболочки", "офтальмометрия"], ["офтальмометрия", "кривизны роговицы", "кератометри", "кривизны роговой оболочки"], "Кератометрия", ["втор", "экзо", "описани", "расшифров", "занятий", "посещени", "комп"]
	komp_kerato, kompkerato_2, id_kompkerato, stop_kompkerato =  ["кератометри", "кривизны роговицы"], ["кривизны роговицы", "комп"], "Компьютерная кератометрия", ["втор", "описани", "расшифров", "занятий", "посещени"]
	pahimetr, pahimetr_2, id_pahimetr, stop_pahimetr =  ["пахиметри", "толщины роговицы", "толщины роговой оболочки"], ["толщины роговицы", "толщины роговой оболочки", "пахиметри"], "Пахиметрия роговицы", ["втор", "направлени", "описани", "расшифров", "занятий", "посещени"]
	uz_biometr_glaz, uz_biometr_glaz_2, id_uz_biometr_glaz, stop_uz_biomtr_glaz =  ["узи", "ультразвук"], ["биометрия глаз", "биометрии глаз"], "Ультразвуковая биометрия глаза", ["втор", "описани", "расшифров", "занятий", "посещени"]
	bio_trofekto, biotrofekto_2, id_biotrofekto, stop_biotrofekto =  ["биопси"], ["трофэктодерм"], "Биопсия трофэктодермы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	bio_horion, biohorion_2, id_biohorion, stop_biohorion =  ["биопси"], ["хорион"], "Биопсия хориона", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	bio_9zik, bio9zik_2, id_bio9zik, stop_bio9zik =  ["биопси"], ["языка"], "Биопсия языка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	bio_shl9pa, bioshl9pa_2, id_bioshl9pa, stop_bioshl9pa =  ["биопси"], ["опухолей головного мозга"], "Стереотаксическая биопсия опухолей головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	vidurodin, vidurodin_2, id_vidurodin, stop_vidurodin =  ["уродинамическо"], ["исследовани"], "Видеоуродинамическое исследование", ["дет", "реб", "втор", "исследо", "орган", "описани", "расшифров", "занятий", "посещени"]
	viz_pot, viz_pot_2, id_vizpot, stop_vizpot =  ["вызванн", "стволовые вп", "соматосенсорные вп", "вп при стимуляции"], ["потенциал", "стволовые вп", "соматосенсорные вп", "вп при стимуляции"], "Вызванные потенциалы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "зритель", "глаз", "офтальмо"]
	densitometr, densitometr_2, id_densitometr, stop_densitometr =  ["денситометри"], ["денситометри"], "Денситометрия костей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "зритель", "концентрац", "клеток", "бактери", "микроорганизм"]
	dermatoskop, dermatoskop_2, id_dermatoskop, stop_dermatoskop =  ["дерматоскоп"], ["дерматоскопи"], "Дерматоскопия", ["дет", "реб", "втор", "более", "до 10", "описани", "расшифров", "занятий", "документ", "дерматоскопических", "изображени"]
	dopplerometr, dopplerometr_2, id_dopplerometr, stop_dopplerometr =  ["допплерометрия", "доплерометрия", "допплерография плод", "кровотока", "допплеровская оценка показателей кровотока "], ["доплерометрия", "допплерометрия", "допплеровская оценка показателей кровотока", "доплерография плода", "допплерография плод", "маточно-плацентарн", "маточного и плацентарн", "маточного", "плацентарного"], "Допплерометрия при беременности", ["втор", "допол", "с допплерометри", "при узи органов", "нейросон", "член"]
	invazprenatal, invazprenatal_2, id_invazprenatal, stop_invazprenatal =  ["инвазивн"], ["пренатальная диагностика"], "Инвазивная пренатальная диагностика", ["дет", "неинваз", "втор", "не инваз"]
	amniotik, amniotik_2, id_amniotik, stop_amniotik =  ["амниотической жидкости"], ["амниотической жидкости"], "Индекс амниотической жидкости", ["дет", "неинваз", "втор", "не инваз"]
	fono_plod, fonoplod_2, id_fonoplod, stop_fonoplod =  ["фонокардиограф"], ["плод", "реб", "беременн"], "Фонокардиография плода", ["втор", "вторич"]
	ktg, ktg_2, id_ktg, stop_ktg =  ["кардиотокограф", "ктг", "кардиотография"], ["плод", "беременн", "кардиотокограф", "кардиотография"], "КТГ при беременности", ["втор", "вторич", "актг"]
	ninvazprenatal, ninvazprenatal_2, id_ninvazprenatal, stop_ninvazprenatal =  ["неинвазивн"], ["пренатальная диагноастика"], "Неинвазивная пренатальная диагностика", ["дет", "втор", "не инваз"]
	kt_krest_2, id_kt_krest, stop_kt_krest = ["крестцово-подвз", "крестцово подвзд", "крестцово-повз", "крестцово повз"], "КТ крестцово-подвздошных сочленений", ["дет", "реб", "втор", "вторич", "абонемент", "описани", "расшифров", "топометр", "скт", "мультиспиральн"]
	kt_brushpolr_2, id_kt_brushpolr, stop_kt_brushpolr = ["брюшной полости"], "КТ брюшной полости ребенку", ["втор", "вторич", "абонемент", "описани", "расшифров", "топометр", "скт", "мультиспиральн"]
	kt_viskosr_2, id_kt_viskosr, stop_kt_viskosr = ["височных костей", "височной кости", "костей височных"], "КТ височных костей ребенку", ["втор", "вторич", "абонемент", "описани", "расшифров", "топометр", "скт", "мультиспиральн"]
	kt_golovnoyr_2, id_kt_golovnoyr, stop_kt_golovnoyr = ["головного мозга", "мозга"], "КТ головного мозга ребенку", ["втор", "спинн", "вторич", "абонемент", "описани", "расшифров", "топометр", "скт", "мультиспиральн"]
	kt_grudkletr_2, id_kt_grudkletr, stop_kt_grudkletr = ["грудной клетки"], "КТ органов грудной клетки ребенку", ["втор", "спинн", "вторич", "абонемент", "описани", "расшифров", "топометр", "скт", "мультиспиральн"]
	kt_po4ekr_2, id_kt_po4ekr, stop_kt_po4ekr = ["почек"], "КТ почек ребенку", ["втор", "спинн", "вторич", "абонемент", "описани", "расшифров", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "скт", "мультиспиральн"]
	kt_pazuhnosr_2, id_kt_pazuhnos, stop_kt_pazuhnosr = ["пазух носа ребенку", "придаточных пазух носа ребенку", "придаточных пазух ребенку", "носовых пазух ребенку"], "КТ пазух носа ребенку", ["втор", "спинн", "вторич", "абонемент", "описани", "расшифров", "топометр", "скт", "мультиспиральн"]
	mr_pozvon_reb, id_mr_pozvon_reb, stop_mr_pozvon_reb = ["позвоночник"], "МРТ позвоночника ребенку", ["груд", "поясни", "ше", "отдел" "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст"]
	mr_shei_reb, id_mr_shei_reb, stop_mr_shei_reb = ["шейного отдела позвоночник"], "МРТ шейного отдела позвоночника ребенку", ["груд", "поясни", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст"]
	kt_po4ekr_2_k, id_kt_po4ekr_k = ["почек ребенку"], "КТ почек ребенку с контрастом"
	kt_brushaor_2_k, id_kt_brushaor_k = ["брюшной аорты", "брюшного отдела аорты", "аорты (брюшной"], "КТ брюшной аорты с контрастом"
	kt_vn4s_2_k, id_kt_vn4s_k = ["височно-нижнечелюстн", "внчс"], "КТ височно-нижнечелюстных суставов с контрастом"
	mrt_cisterno, id_mrt_cisterno, stop_mrt_cisterno = ["цистернография мозга", "цистернография головного мозга"], "МР-цистернография головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "топометр", "контрастом", "контрастированием", "введением контраст", "введение контраст", "использованием контраст"] 
	kt_kontrast_stop_reb = ["втор", "вторич", "абонемент", "описани", "расшифров", "топометр", "спинн", "артери", "вен", "подго", "приго", "перед", "сосуд"]
	mskt = ["скт", "мультиспиральная томография", "мультисрезовая томография", "спиральная компьютерная томографи", "спиральной компьютерной томограф", "мультисрезовая компьютерная томографи", "мультиспиральная кт", "мультисрезовая кт", "мультидетекторная кт", "мультидетекторная томограф", "мультидетекторная компьютерная томо"]
	msktbrushpol_2, id_msktrbrushpol, stop_msktbrushpol =  ["брюшной полости"], "МСКТ брюшной полости", ["дет", "введение контраст", "реб", "подго", "приго", "перед", "для", "вторич", "втор", "абонеме", "занят", "посещен"]
	msktgolovnoy_2, id_golovnoy, stop_golovnoy =  ["головного мозга", "мозга"], "МСКТ головного мозга", ["ангиогра", "подго", "приго", "введение контраст", "перед", "для", "сосуд", "артери", "вен", "дет", "реб", "спинн", "вторич", "втор", "абонеме", "занят", "посещен", "перфуз"]
	msktgrudklet_2, id_msktgrudklet, stop_msktgrudklet =  ["грудной клетки"], "МСКТ грудной клетки", ["дет", "реб", "подго", "приго", "перед", "введение контраст", "для", "вторич", "втор", "абонеме", "занят", "посещен"]
	msktkorcalc_2, id_msktcorcalc, stop_msktcorcalc =  ["коронарного кальция"], "МСКТ индексация коронарного кальция", ["дет", "подго", "приго", "перед", "для", "реб", "вторич", "втор", "абонеме", "занят", "посещен"]
	msktlegkie_2, id_msktlegkie, stop_msktlegkie =  ["легких"], "МСКТ легких", ["дет", "реб", "вторич", "втор", "абонеме", "введение контраст", "подго", "приго", "перед", "для", "занят", "посещен"]
	msktnadpo4e4_2, id_msktnadpo4e4, stop_msktnadpo4e4 =  ["надпочечников"], "МСКТ надпочечников", ["дет", "реб", "вторич", "втор", "абонеме", "занят", "подго", "приго", "перед", "для", "посещен"]
	msktnosoglot_2, id_msktnosoglot, stop_msktnosoglot =  ["носоглотки"], "МСКТ носоглотки", ["дет", "реб", "вторич", "введение контраст", "втор", "абонеме", "занят", "подго", "приго", "перед", "для", "посещен"]
	msktperfuz_2, id_msktperfuz, stop_msktperfuz =  ["перфузия мозга", "перфузия головного мозга"], "МСКТ перфузия головного мозга", ["дет", "введение контраст", "реб", "вторич", "втор", "подго", "приго", "перед", "для", "абонеме", "занят", "посещен"]
	msktpozvon_2, id_msktpozvon, stop_pozvon =  ["позвоночник"], "МСКТ позвоночника", ["дет", "реб", "вторич", "втор", "абонеме", "подго", "введение контраст", "приго", "перед", "для", "занят", "посещен"]
	msktpo4ek_2, id_msktpo4ek, stop_msktpo4ek =  ["почек", "почк", "мочевой си"], "МСКТ почек", ["дет", "реб", "вторич", "сосуд", "артери", "вен", "подго", "введение контраст", "приго", "перед", "для", "втор", "абонеме", "занят", "посещен"]
	msktpazuh_2, id_msktpazuh, stop_msktpazuh =  ["пазух носа", "придаточных пазух", "носовых пазух", "придаточные пазухи"], "МСКТ придаточных пазух носа", ["дет", "введение контраст", "подго", "приго", "перед", "для", "реб", "вторич", "втор", "абонеме", "занят", "посещен"]
	msktserca_2, id_msktserdca, stop_msktserdca =  ["сердца"], "МСКТ сердца", ["артери", "сосуд", "вен", "ангиограф", "дет", "реб", "вторич", "введение контраст", "втор", "подго", "приго", "перед", "для", "абонеме", "занят", "посещен"]
	msktsustav_2, id_msktsustav, stop_msktsustav =  ["сустав"], "МСКТ суставов", ["дет", "реб", "вторич", "втор", "подго", "приго", "введение контраст", "перед", "для", "абонеме", "занят", "посещен"]
	msktcistouretrograf_2, id_msktcistouretrograf, stop_msktcistouretrograf =  ["цистоуретрограф"], "МСКТ цистоуретрография", ["дет", "реб", "введение контраст", "вторич", "подго", "приго", "перед", "для", "втор", "абонеме", "занят", "посещен"]
	msktcherep_2, id_msktcherep, stop_msktcherep =  ["черепа"], "МСКТ черепа", ["дет", "реб", "вторич", "втор", "подго", "введение контраст", "приго", "перед", "для", "абонеме", "занят", "посещен"]
	polplodamater, polplodamater_2, id_polplodamater, stop_polplodamater =  ["пола плода", "пола ребенка"], ["крови"], "Определение пола плода по крови матери", ["втор", "описани", "расшифров", "занятий", "посещени"]
	rezusmater, resuzmater_2, id_resuzmater, stop_resuzmater, rezusmater_3 =  ["резус-фактор", "резус фактор", "резуса"], ["крови"], "Определение резус-фактора плода по крови матери", ["втор", "описани", "расшифров", "занятий", "посещени"], ["ребенка", "плод"]
	oktdiska, oktdiska_2, id_oktdiska, stop_oktdiska =  ["окт", "оптическая когерентная томография"], ["нерв", "дзн"], "ОКТ диска зрительного нерва", ["дет", "реб", "локт", "втор", "описани", "расшифров", "занятий", "посещени", "макул"]
	oktmakuli, oktmakuli_2, id_oktmakuli, stop_oktmakuli =  ["окт ", "оптическая когерентная томография"], ["макул"], "ОКТ макулы и ДЗН", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	oktrogovici, oktrogovici_2, id_oktrogovoci, stop_oktrogovici =  ["окт", "оптическая когерентная томография"], ["рогови"], "ОКТ роговицы глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	oktset, oktset_2, id_oktset, stop_oktset =  ["окт", "оптическая когерентная томография"], ["сетчатк"], "ОКТ сетчатки глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	diafanglaz, diafanglaz_2, id_diafaglaz, stop_diafalglaz =  ["диафаноскопи"], ["глаз"], "Диафаноскопия глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	potencizri, potenczri_2, id_potenczri, stop_potenczri =  ["зритель", "глаз", "офтальм", "зрительные вп", "зрительных вп"], ["потенциал", "зрительные вп", "зрительных вп"], "Исследование зрительных вызванных потенциалов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	kompkeratotop, kompkeratotop_2, id_kompkeratotop, stop_kompkeratotop =  ["корнеотопография", "фотокератоскопия", "видеокератография", "кератотопограф"], ["корнеотопография", "фотокератоскопия", "видеокератография", "кератотопограф"], "Компьютерная кератотопография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	mikrorog, mikrorog_2, id_mikrorog, stop_mikrorog =  ["микроскоп"], ["рогови"], "Микроскопия роговицы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	k4sm, k4sm_2, id_k4sm, stop_k4sm =  ["кчсм", "критической частоты"], ["кчсм", "слияния мельканий"], "Определение КЧСМ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	oftalmoskopop, oftalmoskop_2, id_oftalmoskop, stop_oftalmoskop =  ["офтальмоскоп", "исследование", "линза гольдман", "линзой гольдман", "линзы голдмана", "линзой голдмана", "осмотр сетчатки"], ["офтальмоскоп", "линза гольдман", "линзой гольдман", "линзы голдмана", "линзой голдмана", "глазного дна"], "Офтальмоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	suttono, suttono_2, id_suttono, stop_suttono =  ["суточная"], ["тонометрия"], "Суточная тонометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	fotoglaz, fotoglaz_2, id_fotoglaz, stop_fotoglaz =  ["фотографи"], ["глазного дна"], "Фотография глазного дна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	tonoglaz, tonoglaz_2, id_tonoglaz, stop_tonoglaz =  ["тонограф"], ["глаза"], "Электронная тонография глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	polisomno, polisomno_2, id_polisomno, stop_polisomno =  ["полисомнографи"], ["полисомнографи"], "Полисомнография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	ventri, ventri_2, id_ventri, stop_ventri =  ["вентрикулография"], ["вентрикулография"], "Вентрикулография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	dakrio, dakrio_2, id_dakrio, stop_dakrio =  ["дактриоцистография", "цистография"], ["дакриоцистография", "слез"], "Дакриоцистография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	irrigo, irrigo_2, id_irrigo, stop_irrigo =  ["ирригоскопия", "иригоскопия", "рентген"], ["ирригоскопия", "иригоскопия", "кишечника", "толстой кишки"], "Ирригоскопия кишечника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	kavegraf, kavergraf_2, id_kavergraf, stop_kavergraf =  ["кавернозография", "рентген"], ["кавернозография", "члена"], "Кавернозография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	mielograf, mielograf_2, id_mielograf, stop_mieograf =  ["миелография", "рентген"], ["миелография", "спинного мозга", "ликворопроводящ"], "Миелография", ["дет", "реб", "мрт", "мр", "магнито", "втор", "описани", "расшифров", "занятий", "посещени"]
	rinomano, rinomano_2, id_rinomano, stop_rinomano =  ["риноманометрия", "манометри"], ["риноманометрия", "нос"], "Риноманометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	tride, tride_2, id_tride, stop_tride =  ["3д", "3 д", "3-д", "3-d", "3 d", "3d"], ["узи", "ультразвук", "плода", "беременн"], "3Д УЗИ", ["3 до", "4 до", "област", "1 орган", "одного органа", "1 органа", "органа", "кт", "томограф", "втор", "описани", "расшифров", "занятий", "посещени"]
	chde, chde_2, id_chde, stop_chde =  ["4д", "4 д", "4-д", "4-d", "4 d", "4d"], ["узи", "ультразвук", "плода", "беременн"], "4Д УЗИ", ["дет", "4 до", "3 до", "втор", "кт", "томограф", "описани", "расшифров", "занятий", "посещени"]
	uzipola, uzipola_2, id_uzipola, stop_uzipola =  ["узи", "ультразвук"], ["пола плода", "пола ребенка"], "УЗИ определение пола ребенка", ["дет", "втор", "описани", "расшифров", "занятий", "посещени"]
	yzrubc, uzrubc_2, id_uzrubc, stop_uzrubc, uzrubc_3 =  ["узи", "ультразвук"], ["рубц"], "УЗИ рубца на матке при беременности", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"], ["матке", "маточного", "беременн"]
	termografi, termografi_2, id_termografi, stop_termografi =  ["термография"], ["термография"], "Термография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scintrigraf = ["сцинтиграф"]
	osteoscin, id_osteoscin, stop_osteoscin =  ["остео", "костей скелета"], "Остеосцинтиграфия / Сцинтиграфия костей скелета", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	renosci_2, id_renosci, stop_renosci =  ["рено", "нефро", "почек"], "Реносцинтиграфия / Нефросцинтиграфия почек", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	statisci_2, id_statisci, stop_statisci =  ["статическ"], "Статическая сцинтиграфия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scitelo_2, id_scitelo, stop_scitelo =  ["всего тела", "тела полностью"], "Сцинтиграфия всего тела", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scikost_2, id_scikost, stop_scikost =  ["костного мозга", "спинного мозга"], "Сцинтиграфия костного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scileg_2, id_scileg, stop_scileg =  ["легких"], "Сцинтиграфия легких", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scimiokard_2, id_scimiokard, stop_scimiokard =  ["миокард"], "Сцинтиграфия миокарда", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	sciparashit_2, id_sciparashit, stop_sciparashit =  ["паращитовидн"], "Сцинтиграфия паращитовидных желез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scipe4_2, id_scipe4, stop_scipe4 =  ["печени"], "Сцинтиграфия печени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scishit_2, id_scishit, stop_scishit =  ["щитовидной"], "Сцинтиграфия щитовидной железы", ["дет", "пара", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	flebosci_2, id_flebosci, stop_flebosci =  ["флебо"], "Флебосцинтиграфия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	holesci_2, id_holesci, stop_holesci =  ["холе"], "Холесцинтиграфия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scipo4reb_2, id_scipo4reb, stop_scipo4reb =  ["почек ребенку", "почек детям"], "Сцинтиграфия почек ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"]
	dinscin_2, id_discin, stop_discin =  ["пищевода и желудка", "желудка и пищевода", "функции желудка", "эвакуатор"], "Динамическая сцинтиграфия пищевода и желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	scilim_2, id_scilim, stop_scilim =  ["лимфатическ", "лимф"], "Лимфосцинтиграфия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	uzperner, uzner_2, id_uzner, stop_uzner =  ["узи", "ультразвук"], ["нерв"], "УЗИ периферических нервов", ["дет", "реб", "локт", "втор", "блокад", "под ", "описани", "расшифров", "занятий", "посещени"]
	voseretro, vosuretro_2, id_vosuretro, stop_vosuretro =  ["восходящ"], ["уретрограф", "уретерограф"], "Восходящая уретрография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	retrouretro, retrouretro_2, id_retruoretro, stop_retrouretro =  ["ретроград"], ["уретрограф"], "Ретроградная уретрография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	urograf, urograf_2, id_urograf, stop_ureograf =  ["урография"], ["урография"], "Урография", ["дет", "мрт", "мр", "мскт", "магнит", "эксре", "экскре", "эскре", "обзор", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	mrurograf, mrurograf_2, id_mrurograf, stop_mrurograf =  ["мрт", "мр", "мскт"], ["урография"], "МРТ-урография", ["дет", "реб", "втор", "описани", "предыду", "наличии", "расшифров", "занятий", "посещени"]
	obzouro, obzouro_2, id_obzouro, stop_obzouro =  ["обзорн", "рентген"], ["урография", "мочевыдел"], "Обзорная урография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	uroreb, ureoreb_2, id_uroreb, stop_uroreb =  ["ребен", "дет"], ["урография"], "Урография ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"]
	eskuro, eskuro_2, id_eksyro, stop_eksuro =  ["экскре", "эксре", "эскре"], ["урография"], "Экскреторная урография", ["дет", "реб", "предыду", "наличии", "втор", "описани", "расшифров", "занятий", "посещени"]
	fertiloskop, fertiloskop_2, id_fertiloskop, stop_fertiskolop =  ["фертилоскопи"], ["фертилоскопи"], "Фертилоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	fleboskon, flebokon_2, id_flebokon, stop_flebokon =  ["восходящ"], ["флебографи"], "Восходящая флебография конечностей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	tazozfle, tazofle_2, id_tazofle, stop_tazofle =  ["тазов"], ["флебографи"], "Тазовая флебография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	cheroksi, cheroksi_2, id_cheroksi, stop_cheroksi =  ["чрезкож", "чрескож", "черезкож", "черескож"], ["оксиметри"], "Чрескожная оксиметрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	fluoro, fluoro_2, id_fluoro, stop_fluoro =  ["флюорография", "флюрография"], ["флюорография", "флюрография"], "Флюорография", ["дет", "книжк", "доп", "реб", "втор", "на ул", "описани", "расшифров", "занятий", "посещени"]
	aktis, aktis_2, id_aktis, stop_aktis =  ["актиграфия"], ["актиграфия"], "Актиграфия сна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	akurino, akurino_2, id_akurino, stop_akurino =  ["ринометрия"], ["ринометрия"], "Акустическая ринометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	vesti, vesti_2, id_vesti, stop_vesti =  ["вестибулометрия"], ["вестибулометрия"], "Вестибулометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	phjel, phjel_2, id_phjel, stop_phjel, phjel_3 =  ["желуд"], ["ph", "рн", "pн"], "Внутрижелудочная pH-метрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"], ["метри"]
	phpi, phpi_2, id_phpi, stop_phpi, phpi_3 =  ["пищевод"], ["ph", "рн", "pн"], "Внутрипищеводная рН-метрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"], ["метри"]
	ishetes, ishetest_2, id_ishetest, stop_ishetest =  ["ишемическ"], ["тест"], "Ишемический тест", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	karesmon, karesmon_2, id_karesmon, stop_karesmon =  ["кардиореспираторн", "кардио-респираторн", "кардио респираторн"], ["мониторинг"], "Кардио-респираторный мониторинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	limfograf, limfograf_2, id_limfograf, stop_limfograf =  ["лимфография"], ["лимфография"], "Лимфография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	otolito, otolito_2, id_otolito, stop_otolito =  ["отолитометрия"], ["отолитометрия"], "Непрямая отолитометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	olfaktome, olfaktome_2, id_olfaktome, stop_olfaktome =  ["ольфактометрия", "ольфакометрия"], ["ольфакометрия", "ольфактометрия"], "Ольфактометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	prohosluh, proosluh_2, id_prohosluh, stop_prohosluh =  ["проходимост"], ["слуховой", "евстахиев"], "Определение проходимости слуховой трубы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	ortoporb, ortoprob_2, id_ortoprob, stop_ortoprob =  ["ортостатическ"], ["проба", "пробы"], "Ортостатическая проба", ["дет", "экг", "электрокардиогр", "регистр", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	pikflow, pikflow_2, id_pikflow, stop_pikflow =  ["пикфлоуметрия", "пикфлуометрия"], ["пикфлоуметрия", "пикфлуометрия"], "Пикфлоуметрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	plantograf, plantograf_2, id_plantograf, stop_plantograf =  ["плантографи"], ["плантографи"], "Плантография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	platismograf, platismograf_2, id_platismograf, stop_platismograf =  ["плетизмографи", "платизмографи"], ["плетизмографи", "платизмографи"], "Плетизмография", ["боди",  "дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	profilouretri, profilouretri_2, id_profilouretri, stop_profilouretri =  ["профилометри"], ["уретры"], "Профилометрия уретры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	pulsoksi, pulsoksi_2, id_pulsoksi, stop_pulsoksi =  ["пульсоксиметрия"], ["пульсоксиметрия"], "Пульсоксиметрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	rvg, rvg_2, id_rvg, stop_rvg =  ["реовазографи", "рвг"], ["реовазографи", "рвг"], "Реовазография / РВГ", ["дет", "реб", "втор", "описани", "расшифров", "при", "для", "занятий", "посещени"]
	stabilograf, stabilograf_2, id_stabilograf, stop_stabilograf =  ["стабилография"], ["стабилография"], "Стабилография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]
	stroboskop, stroboskoop_2, id_stroboskop, stop_stroboskop =  ["стробоскопия", "стробоскопии"], ["стробоскопии", "стробоскопия"], "Стробоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"]

	main_func(["сфинктерометрия", "сфинктерометрии", "тонуса"], ["сфинктерометрия", "сфинктерометрии", "сфинктера"], "Сфинктерометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["урофлоуметрия", "урофлоуметрии", "урофлоурометрия", "урофлуометрия"], ["урофлуометрия", "урофлоуметрия", "урофлоуметрии", "урофлоурометрия"], "Урофлоуметрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["цистостометри", "цистоманометри"], ["цистометри", "цистоманометри"], "Цистометрия / Цистоманометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["чреспищевод", "чрезпищевод", "черезпищевод", "череспищевод"], ["кардиостимуляци"], "Чреспищеводная электрокардиостимуляция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эгг", "электрогастрографи"], ["эгг", "электрогастрографи"], "Электрогастрография / ЭГГ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["электронистагмографи"], ["электронистагмографи"], "Электронистагмография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эхосинусоскопи", "эхо- синусоскопи", "эхо-синусоскопи", "эхо синусоскопи"], ["эхосинусоскопи", "эхо- синусоскопи", "эхо-синусоскопи", "эхо синусоскопи"], "Эхосинусоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["фвд", "функции внешнего дыхания", "функция внешнего дыхания", "исследование дыхательных объемов"], ["исследование дыхательных объемов", "фвд", "функции внешнего дыхания", "функция внешнего дыхания"], "Функция внешнего дыхания / ФВД", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["бодиплетизмографи"], ["бодиплетизмографи"], "Бодиплетизмография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["пневмотахометри", "пневмотахиметри"], ["пневмотахиметр", "пневмотахометри"], "Пневмотахометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["диффузионн"], ["способности легких"], "Проверка диффузионной способности лёгких", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["спирометрия", "спирография", "спирометрии", "спирографии"], ["спирометрия", "спирография", "спирометрии", "спирографии"], "Спирометрия / Спирография", ["дет", "эрго", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эргоспирометри", "эргоспирографи"], ["эргоспирометри", "эргоспирографи"], "Эргоспирометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["велоэргометри", "велоэгрометри"], ["велоэргометри", "велоэгрометри"], "Велоэргометрия", ["дет", "реб", "диспансер", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["стептест", "степ тест", "степ-тест"], ["стептест", "степ тест", "степ-тест"], "Степ-тест", ["дет", "реб", "втор", "описани", "расшифровка", "занятий", "посещени"])
	main_func(["тредмил-тест", "тредмилтест", "тредмил тест", "тредмил"], ["тредмил-тест", "тредмилтест", "тредмил тест", "тредмил"], "Тредмил-тест", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["электронейрография"], ["электронейрография"], "Электронейрография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["энмг", "электронейромиографи", "декремент", "электромиография"], ["энмг", "электромиография", "электронейромиографи", "тест"], "ЭНМГ / Электронейромиография", ["дет", "при проведе", "при электро", "лиц", "верх", "осмотр", "объема", "консульт", "прием", "ниж", "рук", "ног", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["энмг", "электронейромиографи", "электромиография"], ["рук", "верх"], "ЭНМГ верхних конечностей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["энмг", "электронейромиографи", "электромиография"], ["нижн", "ног"], "ЭНМГ нижних конечностей", ["дет", "реб", "одного", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["энмг", "электронейромиографи", "электромиография"], ["лиц"], "ЭНМГ лицевого нерва", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["энмг", "электронейромиографи", "электромиография"], ["реб", "дет"], "ЭНМГ ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["бронхоскопи"], ["биопси"], "Бронхоскопия с биопсией", ["дет", "реб", "удалени", "при бронхоскопии", "биопсия трахеи", "биопсия бронхов", "инород", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["бронхоскопи"], ["забор мокроты", "забором мокроты"], "Бронхоскопия с забором мокроты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["дуоденальн"], ["зондир"], "Дуоденальное зондирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["инстил", "инстал"], ["мочевого", "мочевой"], "Инстилляция мочевого пузыря", ["втор"])
	main_func(["биомикроскопия", "осмотр"], ["биомикроскопия", "щелев"], "Биомикроскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"], ["ламп", "биомикроскоп"])
	main_func(["холеходоскопия", "холедохоскопия"], ["холеходоскопия", "холедохоскопия"], "Холедохоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["фарингоскопия"], ["фарингоскопия"], "Фарингоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["уретероскопия"], ["уретероскопия"], "Уретероскопия", ["дет", "пиело", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["торакоскопия"], ["торакоскопия"], "Торакоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["сигмоскопия", "сигмоидоскопия"], ["сигмоскопия", "сигмоидоскопия"], "Сигмоскопия", ["дет", "ректо", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["ректоскопия"], ["ректоскопия"], "Ректоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["ректосигмоскопия"], ["ректосигмоскопия"], "Ректосигмоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["нефроскопия"], ["нефроскопия"], "Нефроскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["микроотоскоп", "миниотоскоп", "отомикроскопи", "отоэндоскопия", "уха", "отоскоп"], ["микроотоскоп", "миниотоскоп", "отомикроскопи", "отоэндоскопия", "микроскоп"], "Микроотоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["микроларингоскопи", "миниларингоскопи"], ["микроларингоскопи", "миниларингоскопи"], "Микроларингоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["капсуль"], ["эндоскоп"], "Капсульная эндоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["лифтинг", "термолифтинг", "радиочастотный лифтинг"], "RF лифтинг / Термолифтинг", ["дет", "фрак", "лиц", "орбит", "rfe", "инфек", "антител", "глаз", "шея", "декольте", "периорбит", "шеи", "шей", "тел", "спин", "век", "подбород", "lumenis", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "при"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["глаз", "периорбит"], "RF лифтинг вокруг глаз", ["дет", "реб", "втор", "описани", "расшифров", "инфек", "фракц", "антител", "инфек", "занятий", "посещени", "под", "для", "во время", "rfe"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["лица", "лицево", "лицо"], "RF лифтинг лица", ["дет", "реб", "втор", "описани", "расшифров", "инфек", "антител", "фракц", "инфек", "занятий", "посещени", "под", "для", "rfe", "во время"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["тела"], "RF лифтинг тела", ["дет", "томограф", "живот", "реб", "антител", "втор", "описани", "расшифров", "инфек", "антител", "инфек", "фракц", "занятий", "посещени", "под", "rfe", "для", "во время"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["шеи", "шейн", "шея"], "RF лифтинг шеи", ["дет", "реб", "втор", "описани", "расшифров", "инфек", "антител", "занятий", "посещени", "под", "для", "rfe", "во время"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["игольчат", "микроигл"], "Игольчатый RF-лифтинг", ["дет", "реб", "втор", "описани","инфек", "антител", "расшифров", "занятий", "посещени", "rfe", "под", "для", "во время"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["век"], "Термолифтинг век", ["дет", "реб", "rfe", "втор", "описани", "без обработки области век", "расшифров", "инфек", "занятий", "антител", "посещени", "rfe", "под", "для", "во время"])
	main_func(["rf", "термолифтинг", "рф", "радиоволновой лифтинг", "радиочастотный лифтинг", "радиочастотная термоабляция"], ["подбород"], "Термолифтинг подбородка", ["дет", "реб", "втор", "описани", "инфек", "расшифров", "занятий", "антител", "rfe", "посещени", "для", "во время"])
	main_func(["sprs"], ["терапи", "лечени"], "SPRS-терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["абдоминопластик"], ["живот"], "Абдоминопластика живота", ["дет", "реб", "пупк", "мини", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["абдоминопластик"], ["без перемещени", "без транслокац"], "Абдоминопластика без перемещения пупка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["абдоминопластик"], ["с перемещ", "с транслокац"], "Абдоминопластика с перемещением пупка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["мини"], ["абдоминопластик"], "Миниабдоминопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["коррекци", "реконструкци"], ["передней брюшной стенки"], "Операция по коррекции передней брюшной стенки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["пластика пупка"], ["пластика пупка"], "Пластика пупка", ["дет", "реб", "эндоскоп", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["эндоскопическая абдоминопластика"], ["эндоскопическая абдоминопластика"], "Эндоскопическая абдоминопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ангиопластик"], ["артерий нижних конечностей", "артерий ног", "артерий на ног", "сосудов нижних конечностей", "сосудов ног", "сосудов на ногах"], "Ангиопластика артерий нижних конечностей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ангиопластик"], ["почечных артерий", "артерий почек", "почечных сосудов", "сосудов почек"], "Ангиопластика почечных артерий", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["стентировани"], ["коронарных артерий", "коронарной артерии"], "Стентирование коронарных артерий", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["стентировани"], ["сонных артерий", "сонной артерии"], "Стентирование сонных артерий", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["стентировани"], ["трахеи"], "Стентирование трахеи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["стентировани"], ["сосудов шеи", "шейных сосудов"], "Стентирование шейных сосудов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["антицеллюлитн", "антициллюлитный", "цилюлит", "целюлит"], ["массаж"], "Антицеллюлитный массаж", ["дет", "ягоди", "реб", "курс", "допол", "живот", "ног", "нижней конечности", "нижних конечностей", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["антицеллюлитн", "антициллюлитный", "цилюлит", "целюлит"], ["массаж"], "Антицеллюлитный массаж ног", ["дет", "реб", "втор", "курс", "допол", "описани", "расшифров", "занятий", "посещени", "под", "для", "дополнит", "во время"], ["ног", "нижних конечностей", "бедер", "бедра"])
	main_func(["антицеллюлитн", "антициллюлитный", "цилюлит", "целюлит"], ["массаж"], "Антицеллюлитный массаж ягодиц", ["дет", "реб", "втор", "курс", "допол", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"], ["ягоди", "таз", "мышц таза"])
	main_func(["аппаратное лечение зрения"], ["аппаратное лечение зрения"], "Аппаратное лечение зрения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["амблиокор"], ["амблиокор"], "Амблиокор", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["офтальмологическ"], ["массаж"], "Вакуумный офтальмологический массаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["визотроник"], ["визотроник"], "Визотроник", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["лазер"], ["стимуляция глаз", "стимуляции глаз", "окулостимуляц"], "Лазерстимуляция глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["магнит"], ["стимуляция глаз", "стимуляции глаз", "окулостимуляц"], "Магнитостимуляция глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["синоптофор"], ["синоптофор"], "Синоптофор", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["тренировк", "аккомодац"], ["ручеек"], "Тренировка аккомодации Ручеек", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["тренировк", "аккомодац"], ["аветисов"], "Тренировки аккомодации по Аветисову", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["аппендэктомия", "апендэктомия", "удаление", "резекция"], ["аппендэктомия", "аппенди", "апендэктомия"], "Аппендэктомия", ["дет", "лапарос", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["лапароскопи"], ["аппендэктоми", "удаление аппендикса", "аппендицит", "апендэктомия"], "Лапароскопическая аппендэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["артроскопия"], ["сустава"], "Артроскопия сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["артроскопия"], ["голеностоп"], "Артроскопия голеностопного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["артроскопия"], ["колен"], "Артроскопия коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["артроскопия"], ["локт"], "Артроскопия локтевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["артроскопия"], ["лучезап"], "Артроскопия лучезапястного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["артроскопия"], ["плеч"], "Артроскопия плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["артроскопия"], ["тазобедр"], "Артроскопия тазобедренного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["аутогемотерапия"], ["аутогемотерапия"], "Аутогемотерапия / АГТ", ["дет", "озон", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["аутогемотерапия", "аутогемоозонотерапия"], ["озон"], "Аутогемотерапия с озоном", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["аутодермопластика"], ["аутодермопластика"], "Аутодермопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["формирова"], ["филатов"], "Формирование Филатовского стебля", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["бандаж"], ["желудк"], "Бандажирование желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["гастропласти"], ["гастропласти"], "Гастропластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["безыгольн", "безигольн", "безинъекцио", "неинваз"], ["мезотерапи"], "Безыгольная мезотерапия", ["дет", "шеи", "шейн", "реб", "втор", "лиц", "кист", "деко", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["безыгольн", "безигольн", "безинъекцио", "неинваз"], ["декольте"], "Безыгольная мезотерапия декольте", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["безыгольн", "безигольн", "безинъекцио", "неинваз"], ["лица", "лицо"], "Безыгольная мезотерапия лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["мезотерапи"], ["кислород"], "Кислородная мезотерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["мезотерапи"], ["лазер"], "Лазерная мезотерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["электропораци", "гидромезотерапи", "электр"], ["электропорац", "гидромезотерапи", "мезотерапи"], "Электропорация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ревитализац"], ["гиалурон", "belotero", "reneal"], "Биоревитализация гиалуроновой кислотой", ["дет", "лазер", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["aquashine", "аквашайн"], ["aquashine", "аквашайн"], "Биоревитализация Aquashine", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["hyalax", "hyaluh"], ["hyalax", "hyaluh"], "Биоревитализация Hyalax", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ial-system", "ial system", "иал систем", "иал-систем", "ial – system", "yal- system", "иал - систем", "lаl system", "ial-sistem"], ["ial-system", "ial system", "иал систем", "ial-sistem", "иал-систем", "ial – system", "yal- system", "иал - систем", "lаl system"], "Биоревитализация IAL System", ["дет", "реб", "facial", "втор", "описани", "расшифров", "занятий", "посещени", "акци", "под", "для", "во время"])
	main_func(["биоревитализац", "dermaheal"], ["возраст", "age", "уставше"], "Биоревитализация возрастной кожи", ["дет", "лазер", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["биоревитализац", "биорепарац"], ["декольте"], "Биоревитализация декольте", ["дет", "реб", "лазер", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["биоревитализац", "биорепарац"], ["глаз", "орбит"], "Биоревитализация зоны глаз", ["дет", "лазер", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["биоревитализац", "биорепарац"], ["лиц"], "Биоревитализация лица", ["дет", "реб", "втор", "лазер", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["биоревитализац", "биорепарац"], ["шеи", "шейн"], "Биоревитализация шеи", ["дет", "реб", "втор", "описани", "расшифров", "лазер", "занятий", "посещени", "под", "для", "во время"])
	main_func(["биоревитализац", "биорепарац"], ["инъекционн", "препарат"], "Инъекционная биоревитализация", ["дет", "реб", "втор", "безинъекцио", "безын", "описани", "лазер", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["биоревитализац", "биорепарац"], ["лазер"], "Лазерная биоревитализация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["пиляция", "биоэпиляция"], ["воск", "полимер"], "Биоэпиляция воском", ["дет", "реб", "втор", "описани", "бикини", "расшифров", "занятий", "посещени", "под", "во время"])
	main_func(["пиляция", "биоэпиляция"], ["воск", "полимер"], "Биоэпиляция бикини воском", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время"], ["бикини"])
	main_func(["блефаропластика", "пластика век"], ["блефаропластика", "пластика век"], "Блефаропластика", ["дет", "массаж", "транскон", "безоперационная", "лазер", "азиат", "кругов", "нижних и верхних", "ниж", "верх", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["блефаропластик"], ["кругов", "верх"], "Блефаропластика круговая", ["дет", "реб", "втор", "описани", "лазер", "массаж", "расшифров", "занятий", "посещени", "для", "во время"], ["ниж", "кругова"])
	main_func(["блефаропласти"], ["верх"], "Верхняя блефаропластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "лазер", "массаж", "посещени", "для", "во время"])
	main_func(["блефаропласти"], ["ниж"], "Нижняя блефаропластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "лазер", "массаж", "для", "во время"])
	main_func(["пластик", "сангапури"], ["азиатски", "восточны", "сангапури"], "Пластика азиатских век", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["коньюктив", "конюктив", "конъюктив", "конъюн", "коньюн"], ["блефаропластик"], "Трансконъюнктивальная блефаропластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["малярн"], ["мешк"], "Удаление малярных мешков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["эпикантопластик", "сангапури", "сингапури", "пластика"], ["эпикантопластик", "сангапури", "сингапури", "эпикантус"], "Эпикантопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["брахиопластика"], ["брахиопластика"], "Брахиопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["липоскульптур", "липомоделирован"], ["рук", "верхних конечностей"], "Липоскульптура рук", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вакуумный массаж", "вакуумная терапия", "вакуум массаж"], ["вакуумный массаж", "вакуумная терапия", "вакуум массаж"], "Вакуумный массаж", ["дет", "допол", "ран", "реб", "лод", "живот", "лиц", "тел", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вакуумный массаж", "вакуумная терапия"], ["лица"], "Вакуумный массаж лица", ["дет", "реб", "втор", "лод", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вакуумный массаж", "вакуумная терапия"], ["тела"], "Вакуумный массаж тела", ["дет", "реб", "втор", "лод", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вальвулопластика"], ["вальвулопластика"], "Вальвулопластика", ["дет", "реб", "бал", "стен", "лег", "втор", "описани",  "допол","расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вестибулопластика", "операция вестибулопластик", "операции вестибулопластики"], ["вестибулопластика", "операция вестибулопластик", "операции вестибулопластики"], "Вестибулопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вибр"], ["массаж", "воздействие"], "Вибромассаж", ["дет", "реб", "механич", "электростатич", "втор", "допол", "описани", "расшифров", "занятий", "барабан", "перепон", "уш", "б/п", "б / п", "посещени", "под", "для", "во время"])
	main_func(["вибромассаж", "вибрационный массаж"], ["механич"], "Механический вибрационный массаж", ["дет", "реб", "втор", "допол", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вибромассаж", "вибрационный массаж"], ["электростати"], "Электростатический вибрационный массаж", ["дет", "допол", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["абляци"], ["эндометри"], "Абляция эндометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["вправлени", "устранение"], ["вывих"], "Вправление вывиха", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["вскрыти", "дренаж", "дренир"], ["абсцесса полости рта", "абсцесса рта", "абсцесса глотки", "внутриротовой абсцес", "внутриротовой абцес", "внутриротового абсце", "внутриротового абце", "абсцесса ротогло", "глоточного абсцесса", "пародонтального абсце", "пародонтального абцес"], "Вскрытие абсцесса полости рта", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["вскрыти", "дренаж", "дренир", "лечение"], ["гидраденит", "гидроденит", "гидроаденит"], "Вскрытие гидраденита", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["вскрыти", "дренаж", "дренир"], ["лимфаденит", "лимфоденит"], "Вскрытие гнойного лимфаденита", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["вскрыти", "дренаж", "дренир"], ["панарици", "понариц"], "Вскрытие панариция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время"])
	main_func(["вскрыти", "дренаж", "дренир"], ["флегмон", "флигмон"], "Вскрытие флегмоны", ["дет", "слез", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["вскрыти", "дренаж", "дренир"], ["ячмен"], "Вскрытие ячменя", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["вульвэктомия"], ["вульвэктомия"], "Вульвэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["капельн"], ["алкоголь"], "Капельница при алкогольной интоксикации", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время"])
	main_func(["вытяжени"], ["позвоночник"], "Вытяжение позвоночника", ["дет", "подвод", "сухо", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["подвод", "мокрое"], ["вытяжение позвоночник"], "Подводное вытяжение позвоночника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["сухое"], ["вытяжение позвоночника"], "Сухое вытяжение позвоночника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["миллиган", "милиган"], ["милиган", "миллиган"], "Геморроидэктомия по Миллигану Моргану", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["геморроидэктомия", "геммороидэктмия", "гемороидэктомия", "иссеч", "удал", "вапоризаци", "оперативное лечение", "деструкци"], ["геморроидэктомия", "геммороидэктомия", "гемороидэктомия", "геморр"], "Геморроидэктомия", ["дет", "блок", "милиган", "неоперативн", "бахром", "миллиган", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["пластик", "сечение", "удал", "хирургич", "опера"], ["спигел", "спеге"], "Герниопластика грыжи спигелиевой линии", ["дет", "реб", "втор", "без опе", "описани", "снятие", "после", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["пластик", "сечение", "удал", "хирургич", "опера"], ["пах"], "Герниопластика пахово-бедренной грыжи", ["дет", "реб", "втор", "дюкен", "лимфаденэктоми", "описани", "расшифров", "без опе", "снятие", "после", "занятий", "посещени", "для", "во время"], ["бедр"])
	main_func(["пластик", "сечение", "удал", "хирургич", "опера"], ["вентрал", "послеоп"], "Герниопластика послеоперационной вентральной грыжи", ["дет", "реб", "втор", "описани", "снятие", "после ", "без опе", "расшифров", "занятий", "посещени", "для", "во время"], ["грыж"])
	main_func(["пластик", "сечение", "удал", "хирургич", "опера"], ["прямых", "диастаз", "передней брюшной стенки"], "Герниопластика при диастазе прямых мышц живота", ["дет", "реб", "снятие", "после", "втор", "снятие", "после", "описани", "расшифров", "без опе", "занятий", "посещени", "для", "во время"])
	main_func(["пластик", "сечение", "удал", "хирургич", "опера"], ["рецидив"], "Герниопластика рецидивной грыжи", ["дет", "нерец", "не рец", "реб", "втор", "описани", "расшифров", "занятий", "без опе", "посещени", "для", "во время"], ["грыж"])
	main_func(["пластик", "сечение", "удал", "хирургич", "опера"], ["щемл"], "Герниопластика ущемленной грыжи", ["дет", "реб", "кольц", "фимоз", "втор", "описани", "расшифров", "занятий", "снятие", "после",  "посещени", "без опе", "для", "во время"], ["грыж"])
	main_func(["ампутация бедра"], ["ампутация бедра"], "Ампутация бедра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ампутация голени"], ["ампутация голени"], "Ампутация голени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ампутация кисти"], ["ампутация кисти"], "Ампутация кисти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ампутация клитора"], ["ампутация клитора"], "Ампутация клитора", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ампутация пальц"], ["кист"], "Ампутация пальцев кисти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["ампутация пальц"], ["стоп"], "Ампутация пальцев стопы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["ампутация плеча"], ["ампутация плеча"], "Ампутация плеча", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["ампутация предплечья"], ["ампутация предплечья"], "Ампутация предплечья", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["ампутация стопы"], ["шарп"], "Ампутация стопы по Шарпу", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["экзартикуляция пальц"], ["кист"], "Экзартикуляция пальцев кисти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["экзартикуляц"], ["плечевого сустава", "сустава плеча"], "Экзартикуляция плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["гирудотерапи", "пиявк", "пиявок"], ["гирудотерапи", "пиявк", "пиявок"], "Гирудотерапия", ["дет", "реб", "консультац", "1 пиявки", "одной пиявки", "проб", "перед", "перевя", "втор", "рефлексо", "внутриполост", "описани", "расшифров", "занятий", "допол", "след", "посещени", "под", "для", "во время"])
	main_func(["внутри", "полостн", "внутрен"], ["гирудотерапи"], "Внутриполостная гирудотерапия", ["дет", "реб", "втор", "описани", "внеполост", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["гистерэктомия", "удаление матки"], ["гистерэктомия", "удаление матки"], "Гистерэктомия / Операция по удалению матки", ["дет", "пан", "лапароскоп", "реб", "втор", "влагалищ", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["влагалищ"], ["гистерэктоми", "удаление матки"], "Влагалищная гистерэктомия", ["дет", "реб", "надвлагалищ", "над влагалищ", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["лапароскоп"], ["гистерэктом", "удаление матки"], "Лапароскопическая гистерэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ампутация матки"], ["надвлагалищ", "над влагалищ"], "Надвлагалищная ампутация матки", ["дет", "реб", "с придатками", "включая придатки", "и придатков", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["ампутация матки"], ["с придатками", " и придатков", "включая придатки"], "Надвлагалищная ампутация матки с придатками", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["пангистерэктомия"], ["пангистерэктомия"], "Пангистерэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["гомеомезотерапия"], ["гомеомезотерапия"], "Гомеомезотерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["дакриоцисториностоми", "дакриоцисторинотом"], ["дакриоцисториностоми", "дакриоцисторинотоми"], "Дакриоцисториностомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["дарсонвал", "д-арсонвал", "д-арсонваль", "дарсонал", "дарсанвал", "дарсенвал"], ["дарсонвал", "дарсанвал", "д-арсонвал", "д-арсонваль", "дарсонал", "дарсенвал"], "Дарсонвализация", ["дет", "лиц", "голов", "спин", "сустав", "курс", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["дарсонвал", "дарсанвал", "д-арсонвал", "д-арсонваль", "дарсонал", "дарсенвал"], ["волос"], "Дарсонвализация волосистой части головы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["дарсонвал", "д-арсонвал", "дарсанвал", "д-арсонваль", "дарсонал", "дарсенвал"], ["лиц"], "Дарсонвализация лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["дарсонвал", "д-арсонвал", "д-арсонваль", "дарсанвал", "дарсонал", "дарсенвал"], ["полост"], "Дарсонвализация полостная", ["дет", "реб", "втор", "вне", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["дарсонвал", "д-арсонвал", "д-арсонваль", "дарсанвал", "дарсонал", "дарсенвал"], ["спины"], "Дарсонвализация спины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["дренирован", "дренаж"], ["плеврал"], "Дренирование плевральной полости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["зондир"], ["слез"], "Зондирование слезных каналов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["иглоукалывани", "рефлексотерапи", "иглотерапия", "акупунктура корпоральная", "иглоук"], ["иглотерапия", "иглоукалывани", "рефлексотерапи", "акупунктура корпоральная"], "Иглоукалывание / Рефлексотерапия", ["дет", "допол", "реб", "други", "гиру", "электр", "аури", "ушн", "микросистем", "ушей", " уха", "пиявк", "втор", "описани", "лазер", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	new_doubler_main(["иглоукалывани", "рефлексотерапи", "аурикулотерапия", "акупунктур"], ["аурик", "ушей", "ушн", "микросистем"], ["дет", "реб", "други", "гиру", "электр", "пиявк", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], "Иглоукалывание / Рефлексотерапия", 961113)
	main_func(["лазеропунктур", "лазероапунктур", "лазер"], ["лазеропунктур", "лазероапунктур", "пунктура"], "Лазеропунктура", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["магнитопунктур", "магнитоапунктур"], ["магнитопунктур", "магнитоапунктур"], "Магнитопунктура", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электропунктур", "электроакупунктур"], ["электропунктур", "электроакупунктур"], "Электроакупунктура", ["дет", "матки", "кист", "набот", "образ", "эроз", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["вагинопластик", "кольпорафи", "кольпоррафи"], ["вагинопластик", "кольпорафия", "кольпоррафи"], "Вагинопластика / Кольпорафия", ["дет", "реб", "втор", "при", "описани", "расшифров", "занятий", "посещени", "для", "во время"])
	main_func(["вульв", "вагин"], ["омоложен"], "Вульво-вагинальное омоложение", ["дет", "лазер", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["гименопластик", "восстановление девственности", "восстановления девственности", "пластик", "восстановление", "восстановления", "восстановлением"], ["гименопластик", "плевы", "восстановление девственности", "восстановления девственности"], "Гименопластика / Восстановление девственности", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время"])
	main_func(["интимн", "цервикал", "влагалищ", "вульв", "гинекологи", "матки"], ["плазмолифтинг"], "Интимный плазмолифтинг", ["дет", "реб", "втор", "дополни", "десен", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["клитор"], ["пластик", "коррекц"], "Клиторопластика / Пластика клитора", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кольпоперинеора"], ["леваторо"], "Кольпоперинеорафия с леваторопластикой", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["половых губ", "половой губы"], ["гелем", "гель"], "Коррекция малых половых губ гелем", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лабиопластик", "коррекция", "увели", "восстановл", "пластика"], ["лабиопластик", "половых губ"], "Лабиопластика / Коррекция половых губ", ["дет", "реб", "гел", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер"], ["омолож", "шлифовк", "моделирование"], "Лазерное омоложение влагалища", ["дет", "реб", "втор", "описани", "рубц", "рубец", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["влаг", "вагин"])
	main_func(["лазер"], ["омоложе", "шлифовк", "моделирование"], "Лазерное омоложение вульвы", ["дет", "реб", "втор", "описани", "расшифров", "рубц", "рубец", "занятий", "посещени", "для", "во время", "курс"], ["вульв"])
	main_func(["перинеопластик"], ["перинеопластик"], "Перинеопластика", ["дет", "кольпо", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс", "рубц", "рубец", "коррекц", "удал", "шлиф", "лазер", "со2", "углекислот", "аппарат"])
	main_func(["пластик"], ["лобк"], "Пластика лобка у женщин", ["дет", "мужч", "контур", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж"], ["влагал"], "Подтяжка влагалища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж"], ["лобк"], "Подтяжка лобка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["увеличени", "аугментац", "увелече"], ["клитор"], "Увеличение клитора", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["увеличени", "аугментац", "увелече"], ["графенберг", "грефенберг", "g"], "Увеличение точки G", ["дет", "реб", "втор", "описани", "графи", "рентген", "молоч", "груд", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["уменьше", "редукц", "резекция"], ["больших половых губ", "большой половой губы"], "Уменьшение больших половых губ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["уменьше", "редукц", "резекция"], ["малых половых", "малой половой", "половых губ"], "Уменьшение малых половых губ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["уменьше"], ["влаг"], "Уменьшение объема влагалища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["инъекц", "укол", "шприц"], ["ботулотоксин"], "Инъекции ботулотоксина", ["дет", "реб", "мигрен", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ботокс", "botox"], ["ботокс", "botox"], "Инъекции ботокса", ["дет", "реб", "гипергидроз", "акци", "лоб", "лиц", "трещи", "лба", "губ", "втор", "шеи", "шея", "шейн", "описани", "расшифров", "занятий", "пептид", "эффектом", "посещени", "под", "для", "во время", "курс"])
	main_func(["ботокс", "botox"], ["губ"], "Инъекции ботокса в губы", ["дет", "реб", "втор", "описани", "гипергидроз", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ботокс"])
	main_func(["ботокс", "botox"], [ "лиц"], "Инъекции ботокса в лицо", ["дет", "реб", "втор", "описани", "расшифров", "гипергидроз", "занятий", "посещени", "под", "для", "во время", "курс"], ["ботокс"])
	main_func(["ботокс", "botox"], ["лоб"], "Инъекции ботокса в лоб", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "гипергидроз", "акци", "посещени", "под", "для", "во время", "курс"], ["ботокс"])
	main_func(["ботокс", "botox"], ["шею", "шеи", "шейн"], "Инъекции ботокса в шею", ["дет", "реб", "втор", "описани", "расшифров", "акци", "занятий", "посещени", "под", "для", "во время", "курс"], ["ботокс"])
	main_func(["диспорт", "disport", "dysport"], ["диспорт", "disport", "dysport"], "Инъекции диспорта", ["дет", "реб", "акци", "втор", "описани", "гипергидроз", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ксеомин"], ["ксеомин"], "Инъекции ксеомина", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "гипергидроз", "во время", "курс"])
	main_func(["мигрен", "головной боли"], ["ботокс", "ботулин", "ботулотокс"], "Лечение мигрени ботоксом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["инъекц", "укол", "шприц", "шриц"], ["гиалурон"], "Инъекции гиалуроновой кислоты", ["дет", "реб", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ионофорез"], ["ионофорез"], "Ионофорез лица", ["дет", "реб", "тел", "рук", "ног", "конечно", "спин", "шея", "шеи", "деколь", "кров", "инфек", "анали", "иссле", "лабора", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["капельн", "непрерывное внутривенное введение"], ["капельн", "непрерывное внутривенное введение"], "Капельницы", ["дет", "реб", "допол", "алкоголь", "выезд", "дом", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["капельн"], ["дом", "выезд"], "Капельница на дому", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["цистостомическ", "удаление цистостомы", "удаление цистостомического дрен", "мена цистомы", "мена цистоми", "мена цистостом", "мена эпицистомического дренаж", "мена цистомического дренаж"], ["мена цистоы", "мена цистоми", "мена цистостом", "дрен", "замен", "удаление цистостомы", "удаление цистостомического дрен", "мена эпицистомического дренаж", "мена цистомического дренаж"], "Замена цистостомического дренажа", ["дет", "реб", "фоле", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["катетер"], ["мочеточник", "мочеточечник"], "Катетеризация мочеточника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фолея", "катетера фоли", "фолей"], ["фолея", "катетера фоли", "фолей"], "Постановка постоянного катетера Фолея", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кератопластик", "пересад", "трансплан"], ["кератопластик", "роговиц"], "Кератопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "после"])
	main_func(["керато", "рогови"], ["протез"], "Кератопротезирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кросс"], ["линкин"], "Роговичный кросс-линкинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кинезио", "медицин", "лечеб", "наложение", "тейпировани"], ["тейп"], "Кинезиотейпирование", ["дет", "реб", "втор", "кросс", "апликац", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["вазорезекц", "перевяз"], ["вазорезекц", "семявывод", "семяновыно"], "Вазорезекция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вазэктоми", "стерилизация"], ["вазэктоми", "муж"], "Вазэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["подбор"], ["контрацеп"], "Подбор контрацептива", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["кожн", "импланон"], ["контрацеп", "импланон"], "Подкожный контрацептив", ["дет", "реб", "втор", "удал", "извлеч", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["контрацептив", "влагалищного", "вагинального", "нова-ринг", "новаринг", "нова ринг"], ["кольц", "новаринг", "нова-ринг"], "Постановка контрацептивного кольца", ["дет", "реб", "втор", "обучение устано", "обучение поста", "обучение по", "описани", "расшифров", "занятий", "пессар", "посещени", "под", "для", "во время", "курс"], ["станов", "введе"])
	main_func(["посткоитал", "постокоитал", "экстренн"], ["контрацепц"], "Посткоитальная контрацепция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "извлеч"], ["контрацеп", "гормональной системы", "нова-ринг", "новаринг", "нова ринг", "nuvaring"], "Удаление гормонального контрацептива", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["становк", "введе"], ["гормон", "мат"], "Установка внутриматочной гормональной системы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["систем", "контрацептив"])
	main_func(["интимн", "полов", "влагалищ", "вульв", "половых губ", "контурная интимная пластика"], ["контурная пластика", "контурной пластики", "контурная интимная пластика"], "Интимная контурная пластика", ["дет", "акци", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["контурная пластика", "контурной пластики"], ["гиалурон"], "Контурная пластика гиалуроновой кислотой", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["контурная пластика", "контурной пластики", "коррекция контура", "juvederm", "juviderm", "рестилайн", "restyline"], ["губ", "smile", "lips", "smail", "kiss", "липп"], "Контурная пластика губ", ["дет", "нос", "реб", "вестибуло", "полов", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["контурная пластика", "контурной пластики"], ["лица"], "Контурная пластика лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["контурная пластика", "контурной пластики", "коррекция"], ["носогуб"], "Контурная пластика носогубных складок", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["контурная пластика", "контурной пластики"], ["слез"], "Контурная пластика носослезной борозды", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["контурная пластика", "контурной пластики"], ["скул"], "Контурная пластика скул", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["контурная пластика", "контурной пластики"], ["шеи", "шейн"], "Контурная пластика шеи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["коррекц"], ["мимич"], "Коррекция мимических морщин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["коррекц"], ["морщ"], "Коррекция морщин", ["дет", "реб", "мимич", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["коррекц"], ["ногт"], "Коррекция ногтевой пластины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["маск"], ["космет", "коллагеновая"], "Косметические маски", ["дет", "рук", "реб", "акци", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["маск"], ["альгинат", "альгинан", "альгена", "альгинал"], "Альгинатная маска", ["дет", "реб", "рук", "акци", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["биоматрикс"], ["биоматрикс"], "Биоматрикс", ["дет", "реб", "втор", "акци", "описани", "губча", "мембран", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["маск"], ["кремо"], "Кремообразная маска", ["дет", "реб", "рук", "втор", "описани", "пилинг", "расшифров", "занятий", "посещени", "акци", "под", "для", "во время", "курс"])
	main_func(["маск"], ["термопла"], "Термопластическая маска", ["дет", "реб", "рук", "втор", "описани", "расшифров", "занятий", "посещени", "под", "акци", "для", "во время", "курс"])
	main_func(["криотерап", "криомассаж", "криоорошение"], ["криотерапи", "криомассаж", "криоорошение"], "Криотерапия", ["дет", "лиц", "маск", "точеч", "1 ед", "тел", "волос", "акне", "ороше", "груд", "спин", "алопец", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криомассаж", "криотерапи", "криоорошение"], ["волос"], "Криомассаж волосистой части головы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криомассаж", "криотерапи", "криоорошение"], ["лиц"], "Криомассаж лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криомассаж", "криотерапи", "криоорошение"], ["тел", "ног", "рук", "спин"], "Криомассаж тела", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криомассаж", "криотерапи", "криоорошение"], ["акне"], "Криотерапия акне", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["блефаропластик"], "Лазерная блефаропластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["коагуляция сосудов"], "Лазерная коагуляция сосудов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["удаление капилляр"], "Лазерное удаление капилляра", ["дет", "сет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удаление капиллярной сет"], ["удаление капиллярной сет"], "Удаление капиллярной сети", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["коррекция зрения", "коррекции зрения"], "Лазерная коррекция зрения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "после", "предоплат"])
	main_func(["барраж"], ["макул", "сетчатки"], "Барраж макулярной области", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func([" ласик", "lasik", "lasek"], ["lasik", " ласик", "lasek"], "Коррекция зрения ЛАСИК", ["дет", "реб", "классик", "втор", "femto", "фемто", "описани", "перел", "расшифров", "занятий", "посещени", "под", "после", "предоплат", "для", "во время", "курс"])
	main_func(["коррекция", "зрения"], ["смайл", "smile"], "Коррекция зрения Смайл", ["дет", "реб", "втор", "губ", "филлер", "контур", "объем", "juve", "описани", "расшифров", "занятий", "перед", "посещени", "под", "для", "во время", "курс"])
	main_func(["фемтоласик", "femtolasik"], ["femtolasik", "femto lasik", "фемтоласик", "фемто ласик"], "Коррекция зрения ФЕМТОЛасик", ["дет", "реб", "перед", "втор", "описани", "после", "предоплат", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["неоваскул", "субретин"], "Лазеркоагуляция субретинальной неоваскулярной мембраны", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["швартотом"], "Лазерная швартотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["липосакци"], "Лазерная липосакция", ["дет", "голен", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["липосакция голеней", "липосакции голеней"], "Лазерная липосакция голеней", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция бикини", ["дет", "реб", "втор", "процедур ", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["бикин"])
	new_doubler_main(["лазерная эпиляция бикини", "лазерная эпиляция зоны бикини", "лазерная эпиляция области бикини", "лазерная эпиляция в области бикини", "процедур ", "лазерная эпиляция в области глубокого би", "лазерная эпиляция (бикини)", "лазерная эпиляция глубокое бикини", "удаление волос в области бикини лазер", "удаление волос в зоне бикини лазер", "эпиляция бикини лазером", "лазерная эпиляция глубокого бикини", "эпиляция глубокого бикини лазер", "лазерная эпиляция"], ["глубок"], ["втор", "дет", "реб", "посеще", "во время", "курс"], "Лазерная эпиляция бикини", 1839930)
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция верхней губы", ["дет", "реб", "втор", "описани", "расшифров", "процедур ", "занятий", "посещени", "во время", "курс"], ["губ", "усы", "усов"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция голеней", ["дет", "реб", "втор", "описани", "расшифров", "процедур ", "занятий", "посещени", "под", "во время", "курс"], ["голен"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция живота", ["дет", "реб", "втор", "описани", "расшифров", "процедур ", "занятий", "посещени", "под", "во время", "курс"], ["живот"])
	new_doubler_main(["лазерная эпиляция живот", "лазерная эпиляция зоны живот", "лазерная эпиляция (живот)", "удаление волос в области живота лазе", "удаление волос в зоне живота лазер", "эпиляция живота лазер", "лазерная эпиляция живота"], ["полност"], ["втор", "дет", "реб", "процедур ", "посеще", "во время", "курс"], "Лазерная эпиляция живота", 1978638)
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция лба", ["дет", "реб", "процедур ", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "лобка", "лобок"], ["лба", "лоб"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция лица", ["дет", "реб", "процедур ", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["лиц"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция ног", ["дет", "реб", "процедур ", "втор", "пальц", "описани", "расшифров", "мног", "нного", "сложн", "голен", "занятий", "посещени", "под", "во время", "курс"], ["ног", "нижних конечностей", "нижняя конечность", "нижней конечности"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция носа", ["дет", "реб", "процедур ", "втор", "полнос", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["нос"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция подбородка", ["дет", "процедур ", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["подбород"])
	main_func(["лазер", "motus ax", "duetto"], ["пиляци", "удаление волос"], "Лазерная эпиляция подмышек", ["дет", "реб", "процедур ", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["подмыше"])
	main_func(["лазер", "smooth"], ["омолож", "смус", "smooth"], "Лазерное омоложение", ["дет", "реб", "aft", "шея", "вагин", " акци", "процедур ", "растя", "влаг", "вульв", "пол", "инти", "шеи", "шей", "декольт", "кист", "рук", "локт", "орбит", "глаз", "лиц", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер", "smooth"], ["омолож", "смус", "smooth"], "Лазерное омоложение кистей рук", ["дет", "реб", "aft", "втор", " акци", "процедур ", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["кист"])
	main_func(["лазер", "smooth"], ["омолож", "смус", "smooth"], "Лазерное омоложение лица", ["дет", "реб", "втор", "aft", "описани", " акци", "процедур ", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["лазер", "smooth"], ["омолож", "смус", "smooth"], "Лазерное омоложение локтей", ["дет", "реб", "втор", "aft", "описани", " акци", "процедур ", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["локт"])
	main_func(["лазер", "smooth"], ["омолож", "смус", "smooth"], "Лазерное омоложение периорбитальной зоны", ["дет", "aft", "реб", "втор", " акци", "процедур ", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["орбит", "глаз"])
	main_func(["лазер", "smooth"], ["омолож", "смус", "smooth"], "Лазерное омоложение шеи", ["дет", "реб", "втор", "aft", "описани", "расшифров", " акци", "процедур ", "занятий", "посещени", "под", "для", "во время", "курс"], ["шейн", "шеи", "шея"])
	main_func(["омолож", "шлифовк"], ["фраксел"], "Омоложение Фраксель", ["дет", "реб", "втор", "описани", "расшифров", " акци", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["липолиз"], "Лазерный липолиз", ["дет", "реб", "втор", "лиц", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["липолиз"], "Лазерный липолиз лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["лазер", "карбон"], ["пилинг"], "Лазерный пилинг", ["дет", "реб", "втор", "после", "программ", "реабил", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лакунотоми", "уменьше", "редукц"], ["лакунотоми", "миндалин"], "Лакунотомия", ["дет", "реб", "лазер", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лакунотоми", "уменьше", "редукц"], ["миндалин", "лакунотоми"], "Лазерная лакунотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лазер"])
	main_func(["лапароцентез"], ["лапароцентез"], "Лапароцентез", ["дет", "плод", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["акне"], ["лазер", "дот"], "Лечение акне лазером", ["дет", "реб", "щек", "лбу", "лоб", "подборо", "спин", "лиц", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["акне"], ["лоб", "лбу", "лба"], "Лазерное лечение акне на лбу", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лазер"])
	main_func(["акне"], ["лиц", "щек"], "Лазерное лечение акне на лице", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лазер"])
	main_func(["акне"], ["подбород"], "Лазерное лечение акне на подбородке", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лазер"])
	main_func(["акне"], ["спин"], "Лазерное лечение акне на спине", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лазер"])
	main_func(["асит", "аллерген-специфическ"], ["терапи", "асит"], "АСИТ-терапия", ["дет", "реб", "втор", "описани", "расшифров", "прием", "консул", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["инъекционн"], ["иммунокоррекц"], "Инъекционная иммунокоррекция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["неспецифическ"], ["терапи", "лечен"], "Неспецифическая терапия аллергических заболеваний", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["аллерг"])
	main_func(["небулайзер"], ["небулайзер"], "Небулайзерная терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вскрыти", "дренаж", "дренир", "удален"], ["бартолин", "бактолин", "женских половых органов", "бортолин"], "Вскрытие абсцесса бартолиновой железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["абсцес", "абцес"])
	main_func(["склеропластик"], ["склеропластик"], "Склеропластика", ["дет", "реб", "колла", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кроссэктоми"], ["кроссэктоми"], "Кроссэктомия", ["дет", "реб", "втор", "тромбоз ", "остр", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["склеротерап"], ["ретикул"], "Микросклеротерапия ретикулярных вен", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время"])
	main_func(["мини", "мало", "микро"], ["флебэктом"], "Минифлебэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["линтон"], ["линтон"], "Операция Линтона", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["тренделен"], ["тренделен"], "Операция Троянова-Тренделенбурга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["перевяз"], ["перфора"], "Перевязка перфорантной вены", ["дет", "реб", "флебэктомия", "комбин", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["абляц", "облитерац"], ["вен", "сосуд"], "Радиочастотная абляция вен", ["дет", "реб", "втор", "описани", "фоам", "foam", "расшифров", "занятий", "нос", "киссель", "посещени" "во время", "курс"], ["радио", "сурги"])
	main_func(["склерози", "склерозац", "склеротизац"], ["вен"], "Склерозирование вен нижних конечностей", ["дет", "реб", "втор", "описани", "расшифров", "канат", "семен", "занятий", "посещени", "во время", "курс"], ["ног", "нижних конечностей", "нижняя конечность"])
	main_func(["склеротерапия", "склерозирован", "склеротизац", "сеанс склеротерапии"], ["сеанс склеротерапии", "склеротерапия", "склерозирован", "склеротизац"], "Склеротерапия", ["дет", "молоч", "кист", "образова", "гемор", "семен", "канат", "микро", "звезд", "эхо", "ретикул", "поч", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["вен", "сосуд", "варикоз", "конечност", "склерозант", "категори"])
	main_func(["стриппинг", "стрипинг"], ["вен"], "Стриппинг варикозных вен", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["флебэктоми"], ["флебэктоми"], "Флебэктомия", ["дет", "мини", "микро", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["эндовазальн", "эндовеноз", "эвлк"], ["коагул", "облитерац", "эвлк"], "Эндовазальная (эндовенозная) коагуляция", ["дет", "консультац", "после эвлк" , "после эндо", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["эхо"], ["склеротерап", "облитерац"], "Эхо-склеротерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["паломо"], ["паломо"], "Операция Паломо", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эмболизац"], ["варикоцел"], "Эмболизация варикоцеле", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фототерап", "фотолече"], ["витилиго"], "Фототерапия витилиго", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазерный душ"], ["голов", "волос"], "Лазерный душ кожи головы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["иониз"], ["волос"], "Микроионизация волосистой части головы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стимул"], ["волос"], "Электростимуляция волосистой части головы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["электр"])
	main_func(["лигирован", "легирован"], ["геморр", "узла", "узлов"], "Лигирование геморроя", ["дет", "реб", "втор", "описани", "следую", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["гинеколог", "половых", "интимн"], ["кавитац"], "Гинекологическая ультразвуковая кавитация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лечение"], ["гипергидроз"], "Лечение гипергидроза", ["дет", "реб", "ботокс", "botox", "xeomin", "dysport", "ботулотоксин", "диспорт", "лазер", "ксеомин", "инъекц", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["гипергидроз"], "Лазерное лечение гипергидроза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["гипергидроза"], ["ботокс", "botox", "диспорт", "ксеомин", "ботулотоксин", "ботулакс", "dysport", "xeomin"], "Лечение гипергидроза ботулотоксином", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["симпатэктоми"], ["симпатэктоми"], "Симпатэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["трепанац", "трепонац"], ["склер"], "Задняя трепанация склеры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["иридопласт"], "Лазерная иридопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["иридотом"], "Лазерная иридотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["иридэктом"], "Лазерная периферическая иридэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["трабекулопласти"], "Лазерная трабекулопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер", "диод"], ["циклокоагул"], "Лазерная циклокоагуляция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер", "операция"], ["глауком"], "Лазерное лечение глаукомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["непроник"], ["склерэктоми"], "Непроникающая глубокая склерэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["трабекулэктоми", "трабекулэто"], ["трабекулэктоми", "трабекулэто"], "Трабекулэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func([" дом", "по москве"], ["консульт", "прием", "вызов", "выезд"], "Вызов терапевта на дом", ["дет", "реб", "втор", "2-го", "к основному", "описани", "ортопедом", "мануал", "пси", "рефлекс", "расшифров", "дмс", "занятий", "посещени", "под", "для", "во время", "курс"], ["терапевт"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов аллерголога на дом", ["дет", "реб", "втор", "описани", "расшифров", "к основному", "занятий", "посещени", "для", "во время", "курс"], ["аллерголог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов гастроэнтеролога на дом", ["дет", "реб", "втор", "описани", "расшифров", "к основному", "занятий", "посещени", "для", "во время", "курс"], ["гастроэнтеролог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов гинеколога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["гинеколог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов дерматолога на дом", ["дет", "реб", "втор", "обработ", "описани", "расшифров", "к основному", "занятий", "посещени", "для", "во время", "курс"], ["дерматолог", "дерматовенеролог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов детского врача / педиатра на дом", ["втор", "2-го", "к основному", "описани", "расшифров", "занятий", "посещени", "для", "во время", "дмс", "от ", "крови", "курс"], ["педиатр", "детского врача"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов детского ЛОРа на дом", ["втор", "описани", "расшифров", "занятий", "посещени", "к основному", "для", "во время", "курс"], ["детского лор", "детского оторинолар"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов инфекциониста на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["инфекционист"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов кардиолога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["кардиолог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов ЛОРа на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["лор", "оторино"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов мануального терапевта на дом", ["дет", "реб", "втор", "описани", "расшифров", "к основному", "занятий", "посещени", "для", "во время", "курс"], ["мануальн"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов массажиста на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["массажист"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов нарколога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["нарколог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов невролога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["невролог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов онколога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["онколог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов ортопеда на дом", ["дет", "реб", "втор", "ортопедом", "описани", "расшифров", "к основному", "занятий", "посещени", "для", "во время", "курс"], ["ортопед"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов остеопата на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["остеопат"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов офтальмолога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["офтальмолог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов проктолога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["проктолог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов психиатра на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["психиатр"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов психотерапевта на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["психотерапевт"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов пульмонолога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["пульмонолог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов стоматолога на дом", ["дет", "реб", "ортопедом", "втор", "описани", "расшифров", "к основному", "занятий", "посещени", "для", "во время", "курс"], ["стоматолог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов травматолога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "к основному", "посещени", "для", "во время", "курс"], ["травматолог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов уролога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "к основному", "для", "во время", "курс"], ["уролог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов флеболога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "к основному", "к основному", "для", "во время", "курс"], ["флеболог"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов хирурга на дом", ["дет", "реб", "абдоминал", "втор", "флебол", "ортопедом", "описани", "расшифров", "к основному", "занятий", "посещени", "для", "во время", "курс"], ["хирург"])
	main_func([" дом"], ["консульт", "прием", "вызов", "выезд"], "Вызов эндокринолога на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "к основному", "для", "во время", "курс"], ["эндокринолог"])
	main_func([" дом", "в пределах мкад", "выезд", "вызов"], ["медсестр", "сестра", "м/с", "сестры", "медицинская сестра", "медицинской сестрой", "медицинской сестры", "среднего мед"], "Медсестра на дом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["лазер"], ["гриб"], "Лазерное лечение грибковых поражений", ["дет", "ногт", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["гриб"], "Лазерное лечение грибковых поражений ногтей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ногт"])
	main_func(["пластик", "сечение", "удал", "хирургич", "лапароскопическ", "опера"], ["бедр"], "Герниопластика бедренной грыжи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "пах", "курс"], ["грыж"])
	main_func(["пластик", "сечение", "удал", "хирургич", "лапароскопическ", "опера"], ["пах"], "Герниопластика паховой грыжи", ["дет", "реб", "бедр", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["грыж"])
	main_func(["пластик", "сечение", "удал", "хирургич", "лапароскопическ", "опера"], ["пупоч"], "Герниопластика пупочной грыжи", ["дет", "реб", "около", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["грыж"])
	main_func(["грыж"], ["белой линии"], "Грыжесечение при грыже белой линии живота", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["деструк"], ["фасеточ"], "Деструкция фасеточных нервов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дискэктомия"], ["дискэктомия"], "Дискэктомия", ["дет", "микро", "мини", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вапоризац"], ["межпозвонковой грыж"], "Лазерная вапоризация межпозвонковой грыжи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["реконструкция диск", "реконй межпозвоночн"], "Лазерная реконструкция дисков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["микродискэктомия", "удаление"], ["микродискэктомия", "позвонковой грыжи", "позвоночной грыжи"], "Микродискэктомия / Удаление межпозвоночной грыжи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["опера", "хирург", "коррекци"], ["вальгус", "валгус" "hallux valgus", "halux valgus", "halluxvalgus"], "Операция при Халюс вальгус", ["дет", "без опер", "безопер", "неинва", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["детокс"], ["наркоти"], "Детоксикация от наркотиков", ["реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кодир"], ["двойной", "двойным"], "Кодирование Двойной блок", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кодир"], ["алкогол"], "Кодирование от алкоголизма", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["коди"], ["довжен"], "Кодирование по методу Довженко", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ксено"], ["терап", "лечени"], "Ксенонотерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кристостоми", "кристотоми"], ["кристотоми", "кристотоми"], "Кристотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["перфорации нос"], "Пластика перфорации носовой перегородки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекци", "отслое", "отслойк"], ["носовой перегородки", "перегородки носа", "перегородки в носу"], "Подслизистая резекция носовой перегородки", ["дет", "реб", "опухо", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["иссечен", "дренирован"], ["кисты почки", "почечной кисты", "кист почек", "кист почки", "почечных кист"], "Иссечение кисты почки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["аспирац"], ["почечной кисты", "кисты почки"], "Пункционная аспирация почечной кисты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ренефростоми"], ["ренефростоми"], "Ренефростомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["склеротерап"], ["кист почки", "кисты почки", "почечной кист", "почечных кист", "кист в почке"], "Склеротерапия кист почки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["операция", "хирургическ", "коррекц"], ["при косоглазии", "косоглазия"], "Операция при косоглазии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "без", "под", "для", "во время", "курс"])
	main_func(["пластика мышц"], ["косоглаз"], "Пластика мышц при косоглазии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миомэктомия", "удал"], ["миомэктомия", "миомы матки"], "Удаление миомы матки / Миомэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эмболиза"], ["маточных артерий", "артерий матки"], "Эмболизация маточных артерий", ["дет", "реб", "втор", "описани", "расшифров", "сопрово", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["имплан"], ["сфинктера мочевого пузыря"], "Имплантация сфинктера мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["берча"], ["берча"], "Операция Берча", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["слинг"], ["недержании мочи", "недержания мочи"], "Слинговая операция при недержании мочи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радиоволн", "радиволн", "сургитрон", "радиочастот", "радионож", "радиохирурги", "коагуляци"], ["кровоточ", "кровотеч", "киссель", "сосудов полости носа", "сосудов в полости носа", "сосудов перегородки", "сосудов носовой"], "Радиоволновая коагуляция кровоточащих сосудов", ["дет", "реб", "втор", "описани", "ног", "гинеко", "флебо", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["витрэктомия"], ["витрэктомия"], "Витрэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фибротест"], ["фибротест"], "Фибротест", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гипертерми", "прогрев"], ["простат", "предстате"], "Гипертермия простаты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лод", "яровит", "ллод"], ["ллод", "терапи", "простат", "предста", "вакуумный массаж"], "ЛОД-терапия", ["дет", "реб", "втор", "описани", "бесплод", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фототерап", "фотолече"], ["псиориаз", "псориаз"], "Узковолновая УФБ-фототерапия псориаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удалени", "резекц"], ["шпор"], "Удаление пяточной шпоры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увт", "ударно"], ["шпоры", "шпоре", "пяточ"], "Ударно-волновое лечение пяточной шпоры", ["дет", "реб", "осмотр", "консул", "прием", "перед", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["розацеа"], "Лазерное лечение розацеа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пройд", "зондерман", "зандерман"], ["пройд", "зондерман", "зандерман"], "Аспирация слизи из носа по Пройду / Зондерману", ["дет", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дрен"], ["носовой пазухи", "носовых пазух", "придаточных пазух", "пазух носа"], "Дренирование околоносовой пазухи", ["дет", "реб", "втор", "адренал", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["промывани"], ["соусть", "соустие"], "Промывание придаточных пазух через соустье", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ямик"], ["ямик"], "Санация околоносовых пазух ЯМИК-катетером", ["дет", "реб", "втор", "описани", "расшифров", "следую", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["абсцесстонзиллэктомия"], ["абсцесстонзиллэктомия"], "Абсцесстонзиллэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["функции"], ["миндалин"], "Анализ функции миндалин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вскрыти", "дренаж", "дренир"], ["паратонзиллярно", "околоминдалинн", "паратонзилярн", "пара(пери)тонзиллярного"], "Вскрытие паратонзиллярного абсцесса", ["дет", "разведе", "после", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["абсцесс"])
	main_func(["тонзиллотомия", "тонзилотомия"], ["тонзиллотоми", "тонзилотомия"], "Тонзиллотомия", ["дет", "аден", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["палатопластика"], ["палатопластика"], "Палатопластика", ["дет", "уволо", "увуло", "увулу", "увол", "реб", "увул", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увулопалатопластика", "увулупалатопластика", "уволопалатопластика"], ["увулопалатопластика", "увулупалатопластика", "уволопалатопластика"], "Увулопалатопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увулопалатофарингопластика", "увулупалатофарингопластика", "уволопалатофарингопластика", "уволупалатофарингопластика"], ["увулопалатофарингопластика", "уволопалатофарингопластика", "увулупалатофарингопластика", "уволупалатофарингопластика"], "Увулопалатофарингопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увулопластика", "уволопластика", "увулупластика"], ["увулопластика", "уволопластика", "увулупластика"], "Увулопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гистерорезектоскопия", "гистерорезектоскопии", "лазерное удаление", "лазерная вапоризация"], ["гистерорезектоскопия", "гистерорезектоскопии", "эндометриоз"], "Гистерорезектоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["иссеч"], ["ретроцервикальн"], "Иссечение ретроцервикального эндометриоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["эндометр"])
	main_func(["лапароскоп"], ["коагул"], "Лапароскопическая эндокоагуляция эндометриоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["эндометри"])
	main_func(["удалени", "полипэктоми", "деструкц"], ["полип"], "Удаление полипа цервикального канала", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["церви", "церкви", "женских половых органов"])
	main_func(["удалени", "полипэктоми"], ["полип"], "Удаление полипа эндометрия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["эндометри", "женских половых органов"])
	main_func(["лечени", "терапи"], ["эндометрита"], "Лечение эндометрита", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кавердже"], ["тест"], "Каверджект-тест", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лигирован", "перевязка"], ["дорсаль"], "Лигирование дорсальной вены", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лигирован"], ["член"], "Лигирование сосудов полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["операци"], ["приапизм"], "Операция при приапизме", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["реваскуляризаци"], ["каверн"], "Реваскуляризация кавернозных тел", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тест", "диагно", "проба", "пробы"], ["эрек"], "Тесты на эректильную дисфункцию", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["функц"])
	main_func(["фалло", "член"], ["протез"], "Фаллопротезирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лимфодренаж", "лимфадренаж"], ["массаж", "аппарат"], "Лимфодренажный массаж", ["дет", "конеч", "живот", "рук", "скуль", "букка", "допол", "ног", "лиц", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лимфодренаж", "лимфадренаж"], ["массаж", "аппарат"], "Лимфодренажный массаж конечностей", ["дет", "реб", "втор", "допол", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["рук", "ног", "конечн"])
	main_func(["лимфодренаж", "лимфадренаж"], ["массаж", "аппарат"], "Лимфодренажный массаж лица", ["дет", "реб", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["липосакция"], ["липосакция"], "Липосакция", ["дет", "спин", "шеи", "следу", "боле", "шейн", "радио", "лазер", "сурги", "мини", "ягоди", "щек", "щеч", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криолиполиз"], ["криолиполиз"], "Криолиполиз", ["дет", "zeltiq", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криолиполиз"], ["zeltiq"], "Криолиполиз Zeltiq", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["бедер", "бедра"], "Липосакция бедер", ["дет", "реб", "втор", "описани", "безопер", "без опера", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци", "интралипотерапия", "липолити", "липолитическая терапия"], ["липолитическая терапия", "без операции", "интралипотерапия", "безопер", "без опера", "мезотерапи", "коктейл"], "Липосакция без операции", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["бок", "тали", "фланк"], "Липосакция бока", ["дет", "реб", "втор", "безопер", "без опера", "описани", "без тали", "без бок", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["галиф"], "Липосакция галифе", ["дет", "реб", "втор", "описани", "безопер", "без опера", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["голен"], "Липосакция голеней", ["дет", "реб", "лазер", "втор", "безопер", "без опера", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["живот"], "Липосакция живота", ["дет", "реб", "втор", "описани", "безопер", "без опера", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["лиц"], "Липосакция лица", ["дет", "реб", "втор", "описани", "безопер", "без опера", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], [" ног", "нижних кон", "нижней кон"], "Липосакция ног", ["дет", "безопер", "без опера", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["подбород"], "Липосакция подбородка", ["дет", "реб", "втор", "безопер", "без опера", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["липосакци"], ["рук", "верних кон", "верхней кон"], "Липосакция рук", ["дет", "безопер", "без опера", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["спин"], "Липосакция спины", ["дет", "реб", "втор", "описани", "безопер", "без опера", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["шеи", "шейн", "шея"], "Липосакция шеи", ["дет", "реб", "втор", "безопер", "без опера", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["щек"], "Липосакция щек", ["дет", "реб", "мини", "втор", "описани", "безопер", "без опера", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["ягоди"], "Липосакция ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "безопер", "без опера", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липосакци"], ["радиочастот"], "Радиочастотная Липосакция Body Tite", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["бед"], "Липофилинг бёдер", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["груд", "декольт"], "Липофилинг груди", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["век"], "Липофилинг век", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["губ"], "Липофилинг губ", ["дет", "реб", "втор", "полов", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["лоб", "лба"], "Липофилинг лба", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["лиц"], "Липофилинг лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["подборо"], "Липофилинг подбородка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["тела", "тело"], "Липофилинг тела", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["щек", "щечн"], "Липофилинг щёк", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липофилинг", "липофиллинг"], ["ягоди"], "Липофилинг ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["логопед"], ["массаж"], "Логопедический массаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["аденотонзиллотоми", "аденотонзилотоми"], ["аденотонзиллотоми", "аденотонзилотом"], "Аденотонзиллотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["балон", "баллон"], ["синусопластик"], "Балонная синусопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вазотом"], ["раковин"], "Вазотомия носовых раковин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["вскрыти", "дренир", "дренаж", "удал"], ["атером", "адером"], "Вскрытие атеромы уха", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "после", "посещени", "под", "для", "во время", "курс"], ["уха", "ушей", "ушах", "ушной", "лор орган", "лор-орган"])
	main_func(["вскрыти", "дренир", "дренаж", "удал"], ["кист"], "Вскрытие кисты глотки", ["дет", "реб", "миндал", "втор", "описани", "расшифров", "занятий", "после", "посещени", "под", "для", "во время", "курс"], ["глот", "лор орган", "лор-орган"])
	main_func(["вскрыти", "дренир", "дренаж", "удал"], ["кист"], "Вскрытие кисты миндалины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "после", "под", "для", "во время", "курс"], ["миндалин", "лор орган", "лор-орган"])
	main_func(["вскрыти", "дренир", "дренаж", "удал"], ["фурункул"], "Вскрытие фурункула в слуховом проходе / ухе", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "после", "посещени", "под", "для", "во время", "курс"], ["слух", "ушн", "уха", "ухе", "уша", "лор-органов", "лор органов", "лор органа", "лор-органа"])
	main_func(["гайморотоми", "гаймаротоми"], ["гайморотоми", "гаймаротоми"], "Гайморотомия", ["дет", "реб", "микро", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["крио", "азот"], ["аденоид"], "Криодеструкция аденоидов", ["дет", "реб", "терапи", "лечен", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["гранул"], "Лазерная коагуляция гранул глотки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["глот", "фаринги"])
	main_func(["микрогайморотом"], ["микрогайморотом"], "Микрогайморотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мирингопластик"], ["мирингопластик"], "Мирингопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование", "вскрыти"], ["гаймор"], "Пункция при гайморите", ["дет", "реб", "втор", "без пунк", "безпунк", "беспун", "безпун", "описани", "расшифров", "контрол", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["тушировани"], ["глотк"], "Радиоволновое туширование задней стенки глотки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["радио", "сурги"])
	main_func(["увулотомия", "уволотомия", "увулутомия", "увулотамия", "увулатамия", "увулатомия"], ["увулотомия", "увулотамия", "уволотомия", "увулутомия", "увулатамия", "увулатомия"], "Увулотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удалени", "аденотоми", "аденоиэктомия", "аденоидэктомия", "дестру"], ["аденоидэктомия", "аденоиктомия", "аденотоми", "аденоид"], "Удаление аденоидов / Аденотомия", ["дет", "при", "пособи", "реб", "образо", "вегет", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удалени"], ["кист"], "Удаление кисты шеи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["шеи", "шейн"])
	main_func(["эндоскоп"], ["труб"], "Эндоскопическая операция на трубных валиках", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["валик"])
	main_func(["арома"], ["массаж"], "Ароматерапевтический массаж", ["дет", "реб", "втор", "допол", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["висцера"], ["массаж"], "Висцеральный массаж", ["дет", "реб", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["восстано"], ["массаж"], "Восстановительный массаж", ["дет", "реб", "втор", "допол", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["индий"], ["массаж"], "Индийский массаж", ["дет", "реб", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["китай"], "Китайский массаж", ["дет", "реб", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["грудно"], "Классический массаж грудной клетки", ["дет", "перкуатор", "допол", "отдел", "перкутор", "перкус", "пневмони", "бронхит", "сегментар", "позвон", "шей", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["позвоночник"], "Классический массаж позвоночника", ["дет", "реб", "отдела", "допол", "сегмен", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "вибро"], ["класс", "общ", "медицин", "лечеб"])
	main_func(["массаж"], ["лечеб", "медицинский"], "Лечебный массаж", ["дет", "реб", "предста", "перепо", "бараба", "простат", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["банк", "баноч"], "Массаж банками", ["дет", "реб", "втор", "описани", "расшифров", "допол", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["4 руки", "четыре руки", "четырьмя руками", "4 рук", "4 руками"], "Массаж в 4 руки", ["дет", "допол", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["воротник"], "Массаж воротниковой зоны", ["дет", "реб", "сегмен", "акц", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["головы"], "Массаж головы", ["дет", "реб", "крио", "азот", "холод", "втор", "описани", "расшифров", "допол", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["гуаш", "гуа-ша", "гуа ша"], "Массаж гуаша", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "допол", "для", "во время", "курс"])
	main_func(["массаж"], ["беремен"], "Массаж для беременных", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "во время", "курс"])
	main_func(["массаж"], ["коррекц"], "Массаж для коррекции бюста", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "во время", "курс"], ["бюст", "груд", "декольт", "молоч"])
	main_func(["массаж", "косметический массаж"], ["лиц", "косметический массаж", "жакэ", "жаке"], "Массаж лица", ["дет", "реб", "доп", "ультра", "втор", "крио", "lpg", "буккаль", "вакуум", "лимф", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["дом"], "Массаж на дому", ["дет", "реб", "втор", "доплата", "кист", "сустав", "ног", "рук", "плеч", "голо", "ворот", "массажиста", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["класс", "общ"])
	main_func(["массаж"], [" ног", "нижних конечностей", "нижних  конечностей", "нижней конечности", "нижняя конечность", "н/конечност"], "Массаж ног", ["дет", "рук", "цел", "допол", "сегмент", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["рук", "верхней конечност", "верхняя конечност", "верхних конечност", "в/конечност", "верних конечностей", "кисти и пред"], "Массаж рук", ["дет", "допол", " ног", "реб", "04", " 4", "четыр", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["спины"], "Массаж спины", ["дет", "воротник", "шеи", "крио", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["стоп"], "Массаж стоп", ["дет", "реб", "втор", "описани", "без стоп", "не включая стоп", "расшифров", "занятий", "допол", "посещени", "под", "для", "во время", "голено", "курс"])
	main_func(["массаж"], ["ягодиц", "ягодичной"], "Массаж ягодиц", ["дет", "реб", "целлю", "до", "втор", "описани", "расшифров", "занятий", "допол", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["медов", "меда"], "Медовый массаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "допол", "во время", "курс"])
	main_func(["массаж"], ["миофасци"], "Миофасциальный массаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["нейроседа", "расслабля", "успокаива", "релакс", "стресс", "relax"], "Нейроседативный массаж", ["дет", "реб", "допол", "втор", "голов", "описани", "расшифров", "занятий", "посещени", "допол", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["общий", "общего"], "Общий классический массаж", ["дет", "бан", "цел", "реб", "втор", "спины", "допол", "позвон", "описани", "лимфодренаж", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["перкутор", "пневмонии", "бронхите", "перкус"], "Перкуторный массаж грудной клетки", ["дет", "реб", "втор", "допол", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["груд"])
	main_func(["массаж"], ["скульптур", "моделирующ"], "Скульптурный массаж", ["дет", "реб", "втор", "описани", "макия", "лиц", "маск", "пита", "увлаж", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["спа", "spa"], "СПА массаж", ["дет", "реб", "втор", "пневмо", "бара", "евста", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["спорт"], "Спортивный массаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["тайск"], "Тайский массаж", ["дет", "реб", "китай", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж", "акупрессура", "акупресура"], ["точечн", "акупрессура", "акупресура"], "Точечный массаж / Акупрессура", ["дет", "на фоне", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["хиро"], "Хиромассаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж", "кобидо"], ["япон", "кобидо"], "Японский массаж / Кобидо", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["живот"], "Массаж живота", ["дет", "реб", "вакуум", "лимф", "жиро", "целлю", "похуд", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["живот"], "Антицеллюлитный массаж живота", ["дет", "реб", "без живот", "не включая жи", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["целлюлит", "циллюлит", "цилюлит", "целюлит"])
	main_func(["массаж"], ["живот"], "Вакуумный массаж живота", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "во время", "курс"], ["вакуум"])
	main_func(["массаж"], ["живот"], "Лимфодренажный массаж живота", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "для", "во время", "курс"], ["дренаж"])
	main_func(["массаж"], ["живот"], "Массаж живота для похудения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "во время", "курс"], ["похуде"])
	main_func(["массаж"], ["брюшной стенки"], "Массаж мышц брюшной стенки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "под", "для", "во время", "курс"])
	main_func(["мастэктомия"], ["мастэктомия"], "Мастэктомия", ["дет", "подкож", "радика", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["подкож"], ["мастэктоми"], "Подкожная мастэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радикал"], ["мастэктоми"], "Радикальная мастэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["медикамент", "лекарств"], ["акне"], "Медикаментозное лечение акне", ["дет", "аппарат", "тек", "дот", "фракц", "ток", "фото", "радио", "сурги", "азот", "крио", "реб", "лазер", "без", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["после"], ["рейс"], "Послерейсовый медицинский осмотр водителей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["осмотр"])
	main_func(["пред"], ["рейс"], "Предрейсовый медицинский осмотр водителей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["осмотр"])
	main_func(["мезо"], ["ботокс"], "Мезоботокс", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапия", "мезококтейль"], ["мезотерапия", "мезококтейль"], "Мезотерапия", ["дет", "крио", "голов", "рубц", "рубец", "dermahe", "гомео", "безыголь", "лазер", "безигол", "кислород", "волос", "meso", "век", "глаз", "орбит", "кокте", "липоли", "лиц", "тел", "шеи", "дермахи", "мезолай", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["dermaheal", "дермахил"], ["dermaheal", "дермахил"], "Мезотерапия Dermaheal", ["дет", "крем", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["mesoline", "мезолайн"], ["mesoline", "мезолайн"], "Мезотерапия Mesoline", ["дет", "реб", "крем", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапи", "мезококтейль"], ["век"], "Мезотерапия век", ["дет", "реб", "безинъекцион", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапи", "мезококтейль", "mesoline"], ["волос", "голов"], "Мезотерапия волос", ["дет", "реб", "втор", "безинъекцион", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапи", "mesoeye", "meso eye", "мезококтейль", "мезоай"], ["глаз", "орбит", "mesoeye", "meso eye", "мезоай"], "Мезотерапия зоны вокруг глаз", ["дет", "реб", "безинъекцион", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапи", "липолитик", "интралинотерапи", "интралипотерапи", "nctf-135", "акваликс", "aqualyx"], ["коктейл", "интралинотерапи", "липолитик", "жиросжиг", "интралипотерапи", "nctf-135", "акваликс", "aqualyx"], "Мезотерапия коктейлями", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапи"], ["лиц"], "Мезотерапия лица", ["дет", "реб", "безыголь", "безинъекцион", "угревой", "безиголь", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапи", "мезококтейль"], ["тела", "тело"], "Мезотерапия тела", ["дет", "реб", "втор", "описани", "безинъекцион", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мезотерапи", "мезококтейль"], ["шеи", "шея", "шейн"], "Мезотерапия шеи", ["дет", "безыголь", "безиголь", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миелограмм"], ["миелограмм"], "Миелограмма костного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["микронидлинг"], ["микронидлинг"], "Микронидлинг", ["дет", "реб", "втор", "описани", "2-ая", "вторая", "2 ая", "2-я", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миостимуляция"], ["миостимуляци"], "Миостимуляция", ["дет", "электр", "живот", "лиц", "тел", "ягоди", "транзио", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миостимуляци"], ["живот"], "Миостимуляция живота", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миостимуляци"], ["лиц"], "Миостимуляция лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миостимуляци"], ["тела", "тело"], "Миостимуляция тела", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миостимуляци"], ["транзион", "transion"], "Миостимуляция Транзион", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["миостимуляци"], ["ягоди"], "Миостимуляция ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эпиляц", "удаление волос"], ["спин"], "Мужская эпиляция спины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["муж"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевая подтяжка лица", ["дет", "золот", "реб", "дополнит", "мезо", "синус", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лиц"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевая подтяжка тела", ["дет", "реб", "втор", "дополнит", "синус", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["тел"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевой лифтинг бровей", ["дет", "реб", "втор", "синус", "дополнит", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["бров"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевой лифтинг живота", ["дет", "реб", "втор", "синус", "дополнит","описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["живот"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевой лифтинг молочных желез", ["дет", "реб", "втор", "дополнит", "синус", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["молочн", "груд", "декольт"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевой лифтинг носогубных складок", ["дет", "реб", "втор", "дополнит", "синус", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["носогуб"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевой лифтинг уголков губ", ["дет", "реб", "втор", "нос", "дополнит", "описани", "синус", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["губ"])
	main_func(["нит"], ["подтяж", "лифтин"], "Нитевой лифтинг щек", ["дет", "реб", "втор", "описани", "дополнит", "расшифров", "занятий", "синус", "посещени", "для", "во время", "курс"], ["щек", "щеч"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["shock", "т-шок"], "T-Shock-обертывание", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["целлюлит", "цилюлит", "целюлит", "циллюлит"], "Антицеллюлитное обёртывание", ["дет", "реб", "программ", "массаж", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["винн"], "Винное обертывание", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["водоросл"], "Водорослевое обёртывание", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["горяч"], "Горячее обертывание", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["гряз"], "Грязевое обёртывание", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["мед"], "Медовое обёртывание", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["виски"], "Пеленание виски", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["оберт", "компрессионый бандаж", "бандаж", "пелена"], ["шоколад"], "Шоколадное обёртывание", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радио", "сургитр"], ["циркумцизи", "обрезани"], "Обрезание радиоволновым методом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["озонотерапия"], ["озонотерапия"], "Озонотерапия", ["бед", "сосуд", "ягоди", "ауто", "гемо", "агт", "кров", "вен", "бок", "волос", "живот", "лиц", "руб", "подбор", "сустав", "спин", "тел", "ше", "дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["бедер", "бедрен", "бедра"], "Озонотерапия бедер", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["боков", "боковых", "боковой"], "Озонотерапия боков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["волос"], "Озонотерапия волосистой части головы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["живот"], "Озонотерапия живота", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["лиц"], "Озонотерапия лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["озон", " окс "], ["подбород"], "Озонотерапия подбородка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["озон", " окс "], ["рубц", "рубец"], "Озонотерапия рубцов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["спин"], "Озонотерапия спины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["сустав"], "Озонотерапия сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["озон", " окс "], ["тела", "тело", "грудь", "груди", "декольте", "артикуляр"], "Озонотерапия тела", ["дет", "реб", "втор", "описани", "блок", "расшифров", "занятий", "посещени", "для", "шеи", "шей", "шея", "во время", "курс"])
	main_func(["озон", " окс "], ["шеи", "шейного", "шея"], "Озонотерапия шеи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["смас", "smas"], ["лифтинг", "подтяжк", "lifti"], "SMAS-лифтинг", ["дет", "ше", "ультра", "уз", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["армировани"], ["лиц"], "Армирование лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["биоармирование"], ["биоармирование"], "Биоармирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["буккаль", "букаль"], ["массаж"], "Буккальный массаж лица", ["дет", "реб", "втор", "описани", "расшифров", "без букк", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["микротоки", "микротоками", "микротоков"], ["микроток"], "Микротоки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["термаж"], ["термаж"], "Термаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ультра", "уз", "альтер", "alter"], ["смас", "smas"], "Ультразвуковой SMAS-лифтинг", ["дет", "реб", "от ", "втор", "доп", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["фракц", "фото"], ["термолиз"], "Фракционный фототермолиз", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация акушера-гинеколога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["акушера-гинеколога"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация аллерголога", ["дет", "иммуно", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "иммунолог"], ["аллерголог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация гастроэнтеролога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["гастроэнтеролог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация гематолога", ["дет", "онко", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["гематолог", "гемотолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация гепатолога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["гепатолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация гинеколога", ["дет", "реб", "онко", "уролог", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "акушер", "хирург", "эндокринолог"], ["гинеколог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация дерматолога", ["дет", "онколог", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "онколог"], ["дерматолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация диетолога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["диетолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация инфекциониста", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["инфекционист"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация кардиолога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "интервен", "хирург"], ["кардиолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация ЛОР-врача", ["дет", "реб", "хиру", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "онколог"], ["лор", "оторинолари"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация маммолога-онколога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["маммолог-онколог", "маммолога-онколога"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация нейрохирурга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["нейрохирург"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация нефролога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["нефролог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация онколога", ["дет", "реб", "эндо", "маммолог", "прокто", "втор", "описани", "дермато", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "гинеколог", "лор", "хирург", "оторино", "гемато", "уролог", "кардиоло"] , ["онколог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация офтальмолога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "хирург", "нейро"], ["офтальмолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация педиатра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["педиатр"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация проктолога", ["дет","хирург", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "онколог"], ["проктолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация психиатра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["психиатр"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация психотерапевта", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["психотерапевт"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация пульмонолога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["пульмонолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация ревматолога", ["дет", "реб", "кардио", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ревматолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация семейного врача", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["семейного врача"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация терапевта", ["дет", "рефлекс", "гирудо", "физио", "радио", "химио", "реб", "психо", "натур", "лазер", "гирудо", "баро", "втор", "мануал", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["терапевт"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация уролога", ["дет", "реб", "гинеколог", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "онко", "хирург"], ["уролог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация флеболога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "хирург"], ["флеболог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация флеболога-хирурга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["флеболога-хирурга", "хирурга-флеболога"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация хирурга", ["дет", "реб", "пласти", "офталь", "втор", "гной", "бариатри", "абдомин", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "онко", "уро", "вертебро", "имплан", "онко", "орто", "прокто", "травма", "челю", "эндо", "лор", "рино", "сосуд", "кардио", "торакал", "стомато", "флебо", "гине", "нейро"], ["хирург"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация эндокринолога", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "хирург", "онколог", "гинеколог"], ["эндокринолог"])
	main_func(["онлайн", "он лайн", "он-лайн", "online", "on-line", "on line", "skype", "скайп", "дистанционная", "дистанционный", "телемедицин"], ["консультац", "прием"], "Онлайн консультация невролога", ["дет", "верте", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["невролог"])
	main_func(["дивульси", "девульси"], ["ануса", "анал", "сфинтер"], "Дивульсия ануса", ["дет", "реб", "сеч", "трещин", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["сфинктеролеваторопластика"], ["сфинктеролеваторопластика"], "Сфинктеролеваторопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["баллон", "балон"], ["дилатация аорты"], "Баллонная дилатация аорты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндопротезирование"], ["аорты"], "Эндопротезирование аорты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["илизаров", "иллизаров"], ["бедра"], "Удлинение бедра аппаратом Илизарова", ["дет", "коррег" "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дилатац"], ["подключичн"], "Дилатация подключичной артерии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["дилатац"], ["позвоноч"], "Дилатация позвоночной артерии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["протез"], ["брахиоцефальн"], "Протезирование брахиоцефального ствола", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["сонной"], "Сегментарная резекция внутренней сонной артерии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шунтирован"], ["экстра"], "Экстра-интракраниальное микрошунтирование", ["дет", "или", "либо", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["интра"])
	main_func(["блефарорафи"], ["блефарорафи"], "Блефарорафия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["выворот"], ["век"], "Исправление выворота века", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["заворот"], ["век"], "Исправление заворота века", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["птоз"], ["век"], "Исправление птоза верхнего века", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["верх"])
	main_func(["блефарохалязис"], ["блефарохалязис"], "Иссечение блефарохалязиса", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["опер"], ["трихиаз"], "Оперативное лечение трихиаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["образован"], ["век"], "Удаление новообразований века", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["декомпресс"], ["чревн"], "Декомпрессия чревного ствола", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дилатац"], ["верх"], "Дилатация верхней брыжеечной артерии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["брыж"])
	main_func(["дилатац"], ["чревн"], "Дилатация чревного ствола", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["протез"], ["верх"], "Протезирование верхней брыжеечной артерии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["брыж"])
	main_func(["эндартерэктоми"], ["чревн"], "Эндартерэктомия из чревного ствола", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вестибулэктоми"], ["вестибулэктоми"], "Вестибулэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "резек", "ссеч", "устран"], ["перегород"], "Удаление перегородки влагалища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["влагали"])
	main_func(["шивани"], ["кишечн"], "Ушивание кишечно-влагалищного свища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["свищ"])
	main_func(["шивани"], ["пузыр"], "Ушивание пузырно-влагалищного свища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["свищ"])
	main_func(["шиван"], ["уретро"], "Ушивание уретро-влагалищного свища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["свищ"])
	main_func(["витреолизис"], ["витреолизис"], "Витреолизис", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["иридохрусталиков"], ["диафрагм"], "Восстановление иридохрусталиковой диафрагмы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вскрыт", "дренаж", "дренир"], ["флегмо"], "Вскрытие флегмоны слезного мешка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["слез", "глаз", "век"])
	main_func(["глаз"], ["протез"], "Глазопротезирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["имплант"], ["факичн"], "Имплантация факичных линз", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криопекс"], ["сетчат"], "Криопексия сетчатки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["коагуляция сетчатки", "коагуляции сетчатки"], "Лазерная коагуляция сетчатки", ["дет", "реб", "втор", "описани", "расшифров", "после", "перед", "предоплат", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лакориностоми"], ["лакориностоми"], "Лакориностомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["обтурац", "пломбир"], ["слезной точки", "слезных точек"], "Обтурация слезной точки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["радуж"], "Пластика радужки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радиальн"], ["кератотом"], "Радиальная кератотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["реваскуляр"], ["хориоид"], "Реваскуляризация хориоидеи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["рекон"], ["переднего отрезка"], "Реконструкция переднего отрезка глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["рефрак"], ["замена хрусталика"], "Рефракционная замена хрусталика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["новообраз"], ["слезном мешке", "слезного мешка", "слезных мешков"], "Удаление новообразования слезного мешка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пингвекул"], ["удал", "сечен", "деструк", "резекц"], "Удаление пингвекулы глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экстрасклеральн"], ["пломбир"], "Экстрасклеральное пломбирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["каротид"], ["эндартерэктоми"], "Каротидная эндартерэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["клипир"], ["аневризм"], "Клипирование артериальных аневризм сосудов головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мозг"])
	main_func(["спаек"], ["арахноидит"], "Разъединение спаек при опто-хиазмальном арахноидите", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["рекон", "восстановл"], ["позвоночной артери"], "Реконструкция позвоночной артерии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стереотакс"], ["аспирац"], "Стереотаксическая аспирация внутримозговой гематомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мозговой гематом", "гематомы мозга", "гематомы головного мозга"])
	main_func(["транскрани", "транскарни"], ["гематом"], "Транскраниальное удаление гематомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["абсцесс"], "Удаление абсцесса головного мозга вместе с капсулой", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["с капсулой"])
	main_func(["удал", "иссеч"], ["авм мозга", "авм головного мозга","артерио-венозной мальформации", "артериовенозной мальформации "], "Удаление АВМ головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["аденом"], "Удаление аденомы гипофиза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["гипофиз"])
	main_func(["удал"], ["коллоид"], "Удаление коллоидной кисты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["кист"])
	main_func(["удал"], ["краниофарингиом"], "Удаление краниофарингиом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["менингиомы"], "Удаление менингиомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["неврином"], "Удаление неврином Гассерова узла", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["гассер", "гасер"])
	main_func(["удал"], ["неврином"], "Удаление неврином слухового нерва", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["слух"])
	main_func(["удал"], ["опухоли головного мозга"], "Удаление опухоли головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["субдуральных"], "Удаление хронических субдуральных гематом через фрезевое отверстие", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["фрез"])
	main_func(["арахноидальн"], ["кист"], "Хирургическое лечение арахноидальных кист", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экстра"], ["интра"], "Экстра-интракраниальный анастомоз", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["анастомоз"])
	main_func(["эмболизац"], ["авм мозга", "авм головного мозга", "артерио-венозной мальформации"], "Эмболизация АВМ головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндоваскул"], ["окклю"], "Эндоваскулярная окклюзия аневризм сосудов головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["аневризм"])
	main_func(["эндоскоп"], ["эвакуа"], "Эндоскопическая эвакуация внутримозговой гематомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["гематомы мозга", "внутримозговой гематомы"])
	main_func(["ларингопластик"], ["ларингопластик"], "Ларингопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ларингэктоми"], ["ларингэктоми"], "Ларингэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["трахеостом"], ["трахеостом"], "Трахеостомия", ["дет", "реб", "мен", "обработ", "втор", "описани", "расшифров", "туалет", "очи", "промыв", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["ребр"], "Резекция ребра", ["дет", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["торакотомия"], ["торакотомия"], "Торакотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["кисты"], "Удаление кисты средостения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["средостен"])
	main_func(["рекон", "коррек", "операц", "восстановл"], ["грудн"], "Реконструкция грудной стенки при врожденной деформации", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["врожд"])
	main_func(["рекон", "коррек", "операц", "восстановл"], ["воронко"], "Хирургическая коррекция воронкообразной грудной клетки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["груд"])
	main_func(["рекон", "коррек", "операц", "восстановл"], ["килевид"], "Хирургическая коррекция килевидной грудной клетки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["груд"])
	main_func(["торакопластик"], ["торакопластика"], "Экстраплевральная торакопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["плевролиз", "пневмолиз"], ["плевролиз", "пневмолиз"], "Экстраплевральный пневмолиз", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пласти", "удал", "хирургич", "оператив"], ["диафраг"], "Пластика диафрагмальной грыжи", ["дет", "неоператив", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["грыж"])
	main_func(["ваготомия"], ["ваготомия"], "Ваготомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гастропликация"], ["гастропликация"], "Гастропликация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гастростомия"], ["гастростомия"], "Гастростомия", ["дет", "холе", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гастротомия"], ["гастротомия"], "Гастротомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гастроэнтеростомия"], ["гастроэнтеростомия"], "Гастроэнтеростомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гастрэктоми", "удал"], ["желудка", "гастрэктоми"], "Гастрэктомия / Удаление желудка", ["дет", "эзо", "полип", "без гастрэктомии", "образо", "при", "реб", "инород", "опухол", "образован", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["зондирование"], ["желуд"], "Желудочное зондирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["иссеч", "удале", "резек", "опера"], ["язв"], "Иссечение язвы желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["желудк"])
	main_func(["лапароскоп"], ["фундопликац"], "Лапароскопическая фундопликация", ["дет", "реб", "втор", "ниссен", "нисен", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сужени"], ["привратник"], "Операции при сужениях привратника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["остановк", "купирован"], ["желуд"], "Остановка желудочного кровотечения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["кровоточ", "кровотеч"])
	main_func(["пилоромиотом"], ["пилоромиотом"], "Пилоромиотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["регул"], ["желудочного бандажа", "бандажа желудка"], "Регулировка желудочного бандажа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["реконстру" ,"формирва"], ["гастроэнтероанастомоз"], "Реконструкция гастроэнтероанастомоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "извлеч"], ["баллон"], "Удаление внутрижелудочного баллона", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["желуд"])
	main_func(["удал"], ["опухол"], "Удаление доброкачественной опухоли желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["желудк"])
	main_func(["удал"], ["инородн"], "Удаление инородного тела желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["желудк"])
	main_func(["удаление полип", "полипэктоми"], ["желуд", "жкт"], "Удаление полипов желудка при ЭГДС", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "вне", "под", "для", "во время", "курс"])
	main_func(["уменьше"], ["объем", "желудка"], "Уменьшение объема желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["желудк"])
	main_func(["станов", "имплант", "балонирова", "баллонирован"], ["желуд"], "Установка внутрижелудочного баллона", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["баллон", "балон"])
	main_func(["ушиван"], ["гастростом"], "Ушивание гастростомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ушиван"], ["перфора"], "Ушивание перфоративной язвы желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["язвы желудка", "желудочной язвы"])
	main_func(["фундопликаци"], ["фундопликаци"], "Фундопликация", ["дет", "ниссен", "лапароскоп", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ниссен"], ["ниссен"], "Фундопликация по Ниссену", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндоскоп"], ["лече", "терапи"], "Эндоскопическое лечение язвы желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["язвы желудка", "желудочной язвы"])
	main_func(["дрен"], ["желчного", "желчным", "желчных"], "Дренирование желчного пузыря и протоков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дрен"], ["керр"], "Дренирование желчных протоков по Керру", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дрен"], ["холстед"], "Дренирование желчных протоков по Холстеду", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дрен"], ["рпхг", "ретроградная холангиопанкреатографи", "ретроградной холангиопанкреатографи"], "Назобилиарное дренирование при РПХГ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["при сужени"], ["желчных", "желчного протока"], "Операции при сужениях желчных протоков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удалени", "резекц", "холецистэктоми"], ["желчного", "холецистэктоми"], "Операция по удалению желчного пузыря / Холецистэктомия", ["камн", "допол", "конкремен", "дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["желчного протока", "желчных протоков"], "Пластика желчного протока", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["холедоходуоденостоми"], ["холедоходуоденостоми"], "Холедоходуоденостомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["холедохолитотоми"], ["холедохолитотоми"], "Холедохолитотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["холедохотоми"], ["холедохотоми"], "Холедохотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["холецистостомия"], ["холецистостомия"], "Холецистостомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["холецистотомия"], ["холецистотомия"], "Холецистотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["холецистоэнтеростоми"], ["холецистоэнтеростоми"], "Холецистоэнтеростомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дрен"], ["желчных протоков"], "Чрескожное транспеченочное дренирование желчных протоков", ["дет", "керр", "холстед", "пузыр", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["бужирован"], ["желчных протоков"], "Эндоскопическое бужирование желчных протоков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стент"], ["холедох"], "Эндоскопическое стентирование холедоха", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "извлеч"], ["камн"], "Эндоскопическое удаление камней желчных протоков", ["дет", "реб", "втор", "слюн", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["проток"])
	main_func(["сече"], ["сиблефарон", "симблефарон"], "Рассечение симблефарона", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "вскры"], ["кист"], "Удаление кисты конъюнктивы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["коньюн", "конъюн"])
	main_func(["восстановл"], ["берц"], "Восстановление дистального межберцового синдесмоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["синдесмоз"])
	main_func(["корриг"], ["остеотом"], "Корригирующая высокая остеотомия большеберцовой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["берц"])
	main_func(["пластик"], ["ложных сустав", "ложном суставе"], "Костная пластика при ложных суставах костей голени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["голен"])
	main_func(["илизаров", "иллизаров"], ["голен"], "Удлинение голени аппаратом Илизарова", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["артродез"], ["кист"], "Артродез кистевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пласти", "опер"], ["синдактили"], "Операция при синдактилии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["плантац"], ["кисти"], "Реплантация кисти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["реконс", "восстано"], ["вертлужной"], "Реконструкция вертлужной впадины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["билобэктомия"], ["билобэктомия"], "Билобэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["краев"], ["резекция легкого", "резекция части легкого"], "Краевая резекция легкого", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лобэктомия"], ["лобэктомия"], "Лобэктомия", ["дет", "реб", "втор", "билоб", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["плевропневмонэктомия"], ["плевропневмонэктомия"], "Плевропневмонэктомия по поводу злокачественной мезотелиомы плевры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пневмонэктомия"], ["лимфаденэктоми"], "Пневмонэктомия с медиастинальной лимфаденэктомией", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тораскоп", "торакоскоп"], ["резекция легк"], "Торакоскопическая резекция легкого", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удалени"], ["легкого"], "Удаление легкого", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["энуклеац"], ["опухол", "образован"], "Энуклеация опухоли легкого", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["легк"])
	main_func(["эхинококкэктомия"], ["эхинококкэктомия"], "Эхинококкэктомия", ["дет", "печен", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вальвулопластика"], ["легочной артерии"], "Баллонная вальвулопластика стеноза легочной артерии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["баллон"])
	main_func(["лимфовенозн"], ["шунтировани"], "Лимфовенозное шунтирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вентросуспенз"], ["матки"], "Вентросуспензия матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["конизаци"], ["матк"], "Конизация шейки матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вапоризац"], ["шейки матки"], "Лазерная вапоризация шейки матки", ["дет", "реб", "кист", "обработка", "после вапо", "опухо", "образ", "папил", "кондил", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["лазер"], ["дриллинг"], "Лазерный дриллинг яичников", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["яичник"])
	main_func(["лапароскоп", "лапараскоп"], ["аднексэктоми"], "Лапароскопическая аднексэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лапароскоп", "лапараскоп"], ["клин"], "Лапароскопическая клиновидная резекция яичника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["резекция яичник"])
	main_func(["лапароскоп", "лапараскоп"], ["туботом"], "Лапароскопическая туботомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лапароскоп", "лапараскоп"], ["каутеризац"], "Лапароскопическая электрокаутеризация яичников", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["яичник"])
	main_func(["лапароскоп", "лапараскоп"], ["параовариальной кисты"], "Лапароскопическое удаление параовариальной кисты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лапаротом", "лапаратом"], ["резекция яичника", "резекции яичник"], "Лапаротомная резекция яичника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["манчестер"], ["операци"], "Манчестерская операция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вертгейм"], ["вертгейм"], "Операция Вертгейма", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пункци"], ["яичников", "яичника"], "Пункция яичников", ["дет", "реб", "кист", "образова", "водян", "гидроц", "фолликул", "контрол", "при", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["кист", "ксит"], ["шейки матки", "набот", "ш/матки"], "Удаление кисты шейки матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кисты"], ["яичник"], "Удаление кисты яичника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["удал"])
	main_func(["тубэктомия", "удаление маточных труб"], ["тубэктомия", "удаление маточных труб"], "Удаление маточных труб / Тубэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удалени"], ["шейки матки"], "Удаление шейки матки", ["дет", "реб", "папил", "кондилом", "эктопи", "эрози", "акци", "полип", "кист", "ксит", "опухол", "образован", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["аугментац"], ["цистопластик"], "Аугментационная цистопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дивертикул"], ["мочевого пузыря", "мочевом пузыре"], "Дивертикулэктомия мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сечен"], ["кисты урахус"], "Иссечение кисты урахуса", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кишеч"], ["пласт"], "Кишечная пластика мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мочевого пузыря"])
	main_func(["пластик"], ["мочевого пузыря"], "Пластика мочевого пузыря", ["дет", "реб", "кишеч", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радикальн"], ["цистэктоми"], "Радикальная цистэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["мочевого пузыря"], "Резекция мочевого пузыря", ["дет", "реб", "шей", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["троакар"], ["цистостом"], "Троакарная эпицистостомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["тур", "трансуретраль"], ["опухол"], "ТУР опухоли мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мочевого пузыря"])
	main_func(["тур", "трансуретраль"], ["шейки"], "ТУР шейки мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мочевого пузыря"])
	main_func(["инород"], ["мочевого пузыря", "урологи"], "Удаление инородного тела из мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["цистпростатэктоми"], ["цистпростатэктоми"], "Цистпростатэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эпицистостомия", "эпицистотомия"], ["эпицистостомия", "эпицистотомия"], "Эпицистостомия", ["дет", "троакар", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["баллон", "балон"], ["дилатац"], "Баллонная дилатация мочеточника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мочеточн"])
	main_func(["бужирован"], ["мочеточник"], "Бужирование мочеточника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стриктур"], ["мочеточник"], "Операции при стриктурах мочеточников", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["боари"], ["боари"], "Операция Боари", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["брикер"], ["брикер"], "Операция Брикера", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стент"], ["мочеточник"], "Стентирование мочеточника", ["дет", "реб", "удал", "извлеч", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["уретероуретероанастомоз"], ["уретероуретероанастомоз"], "Уретероуретероанастомоз", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндоскоп"], ["уретероцеле"], "Эндоскопическое рассечение уретероцеле", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сеч"])
	main_func(["эндоскоп"], ["устья мочеточник"], "Эндоскопическое рассечение устья мочеточника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сеч"])
	main_func(["вскрыт", "дренаж", "дренир"], ["абсцесс", "гной"], "Вскрытие и дренирование абсцесса мошонки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мошон"])
	main_func(["лапароскоп", "лапараскоп"], ["варикоцеле"], "Лапароскопическая операция при варикоцеле", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["операц"])
	main_func(["лечени", "терапи", "пункция варикоцеле", "пункции варикоцеле", "пункция при водянке", "пункция водянки", "пункция гидроцеле", "пункции гидроцеле", "пункция при гидроцеле", "операция при гидроцеле", "операция при водянке яич", "операция по поводу гидро", "операция по поводу водянки яи", "винкельман", "бергман"], ["водянки яичка", "пункция водянки", "пункция варикоцеле", "пункции варикоцеле", "пункция при водянке", "пункция гидроцеле", "пункции гидроцеле", "пункция при гидроцеле", "операция при гидроцеле", "операция при водянке яич", "операция по поводу гидро", "операция по поводу водянки яи", "винкельман", "бергман"], "Лечение водянки яичка", ["дет", "реб", "втор", "контрол", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["микрохирургическ"], ["операци"], "Микрохирургическая операция при варикоцеле", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["варикоцел"])
	main_func(["низведен", "орхипекси"], ["яичк", "орхипекси"], "Низведение яичка / Орхипексия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["иванисевич", "иваниссевич"], ["иваниссевич", "иванисевич"], "Операция Иваниссевича при варикоцеле", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["мармар", "marmar"], ["мармар", "marmar"], "Операция Мармара при варикоцеле", ["дет", "реб", "втор", "описани", "расшифров", "рецидив", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["опера"], ["варикоцел"], "Операция при варикоцеле", ["дет", "микро", "иванис", "мармар", "лапар", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пересад"], ["яичка"], "Пересадка яичка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["протез"], ["яичк"], "Протезирование яичка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ревизи"], ["мошонк"], "Ревизия мошонки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекци"], ["яичк"], "Резекция яичка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["атером", "удаление образований половых органов", "удаление образования половых органов"], ["мошонк", "удаление образований половых органов", "удаление образования половых органов"], "Удаление атером мошонки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под ", "для", "во время", "курс"])
	main_func(["удал"], ["кисты"], "Удаление кисты мошонки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мошон"])
	main_func(["удал", "сперматоцелектом", "сперматоцелэктом", "иссечен", "иссечение сперматоцеле"], ["кист", "сперматоцелектом", "сперматоцелэктом", "иссечение сперматоцеле"], "Удаление кисты придатка яичка / Сперматоцелектомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под ", "для", "во время", "курс"], ["придатка", "иссечение сперматоцеле"])
	main_func(["удал", "эпидидимэктоми"], ["придатка яичк", "эпидидимэктоми"], "Удаление придатка яичка / Эпидидимэктомия", ["дет", "кист", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удаление яичка", "орхиэктоми", "орхэктоми", "орхидэктоми"], ["удаление яичка", "орхидэктоми", "орхиэктоми", "орхэктоми"], "Удаление яичка / Орхиэктомия", ["дет", "реб", "кист", "придатк", "опухол", "образован", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["перекрут"], ["яичек", "семенного канатика", "семенных канатиков"], "Устранение перекрута яичек", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["энуклеац"], ["яичек"], "Энуклеация яичек при раке предстательной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сеч"], ["бурс"], "Иссечение бурсы локтевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["локт"])
	main_func(["дюпюитрен"], ["дюпюитрен"], "Операция при контрактуре Дюпюитрена", ["дет", "реб", "втор", "увт", "уда", "терапи", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сеч"], ["кольцевидных связок", "лигаментит", "кольцевидной связки"], "Рассечение кольцевидных связок при стенозирующих лигаментитах", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["акромиально-ключичного сочленения"], "Пластика акромиально-ключичного сочленения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["аутопластик"], ["нервных стволов", "нервного ствола"], "Аутопластика нервных стволов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["декомпресс"], ["нервного ствола", "нервных стволов"], "Декомпрессия нервных стволов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["микроваскул"], ["декомпресс"], "Микроваскулярная декомпрессия корешка лицевого нерва", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лицевого нерва"])
	main_func(["микроваскул"], ["декомпресс"], "Микроваскулярная декомпрессия корешка тройничного нерва", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["тройничного нерва"])
	main_func(["микроваскул"], ["декомпресс"], "Микроваскулярная декомпрессия языкоглоточного-блуждающего нервов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["языко"])
	main_func(["невролиз"], ["невролиз"], "Невролиз периферических нервов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["лицевого нерва"], "Пластика лицевого нерва", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["крио"], ["деструкция носовых раковин", "деструкция нижних носовых раковин", "деструкция нижней носовой раковины", "деструкция носовой раковины"], "Криодеструкция нижних носовых раковины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер", "радио", "диатер"], ["редукц", "коагуля", "томия"], "Лазерная редукция нижних носовых раковин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["носовых раковин", "носовой раковины"])
	main_func(["нижн"], ["конхотоми"], "Нижняя остеоконхотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["резек"], ["заднего конца нижней раковины"], "Резекция заднего конца нижней раковины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["ультразвук"], ["дезинтеграция нижних носовых раковин", "дезинтеграция носовых раковин"], "Ультразвуковая дезинтеграция нижних носовых раковин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["эндопротез"], ["голеностоп"], "Эндопротезирование голеностопного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндопротез"], ["колен"], "Эндопротезирование коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндопротез"], ["локт"], "Эндопротезирование локтевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндопротез"], ["плеч"], "Эндопротезирование плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндопротез"], ["пястно"], "Эндопротезирование пястно-фаланговых суставов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["фаланг"])
	main_func(["эндопротез"], ["тазобедр"], "Эндопротезирование тазобедренного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["elos", "элос"], ["пиляц", "даление волос"], "ЭЛОС-эпиляция", ["дет", "реб", "бикин", "губ", "усов", "усы", "усами", "лица", "лицо" "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["elos", "элос"], ["бикини"], "ЭЛОС-эпиляция бикини", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["elos", "элос"], ["губ"], "ЭЛОС-эпиляция верхней губы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["elos", "элос"], ["лиц"], "ЭЛОС-эпиляция лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электр"], ["эпил", "удаление волос"], "Электроэпиляция", ["дет", "реб", "бикин", "интим", "губ", "усы", "усов", "усами", "лица", "лицо", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])  
	main_func(["электр"], ["эпил", "удаление волос"], "Электроэпиляция бикини", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["бикин"])
	main_func(["электр"], ["эпил", "удаление волос"], "Электроэпиляция верхней губы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["губ"])
	main_func(["электр"], ["эпил", "удаление волос"], "Электроэпиляция лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["электрофорез", "электроферез"], ["электрофорез", "электроферез"], "Электрофорез", ["дет", "белков", "сыворот", "градиен", "иммун", "белка", "мочи", "накож", "ректа", "лидаз", "эуфил", "эндоназал", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электрофорез", "электроферез"], ["кож"], "Накожный электрофорез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электрофорез", "электроферез"], ["ректа"], "Ректальный электрофорез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электрофорез", "электроферез"], ["лида"], "Электрофорез с лидазой", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электрофорез", "электроферез"], ["эуфил"], "Электрофорез с эуфиллином", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электрофорез", "электроферез"], ["эндоназал", "внутриполостн"], "Эндоназальный электрофорез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электролиполиз"], ["электролиполиз"], "Электролиполиз", ["дет", "rf", "рф", "игол", "электрод", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электролиполиз", "липолиз"], ["rf", "рф"], "RF-липолиз", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электролиполиз"], ["игол"], "Электролиполиз игольчатый", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электролиполиз"], ["электрод"], "Электролиполиз электродный", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шугаринг", "сахар"], ["бикин"], "Шугаринг бикини", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шугаринг", "сахар"], ["лиц"], "Шугаринг лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["шугаринг", "сахар"], [" ног", "нижних конечностей", "голен", "до колена"], "Шугаринг ног", ["дет", "ногт", "диабет", "реб", "пальц", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шугаринг", "сахар"], ["подмы"], "Шугаринг подмышечных впадин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["шугаринг", "сахар"], ["рук", "верхних кон", "верхней кон", "до локтя", "до плеча"], "Шугаринг рук", ["дет", "паль", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["чистк"], ["груди"], "Чистка кожи груди", ["дет", "реб", "втор", "лопат", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["чистк"], ["спины"], "Чистка кожи спины", ["дет", "реб", "втор", "описани", "лопат", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["атравматич"], ["чистка"], "Атравматическая чистка лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["вакуум"], ["чистка лица"], "Вакуумная чистка лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дезинкрустация", "гальваническая чистка"], ["дезинкрустация", "гальваническая чистка"], "Дезинкрустация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["комбинирова"], ["чистка лица", "чистки лица", "чистка кожи лица", "чистки кожи лица", "завершающей маской", "дерматологи", "чистка"], "Комбинированная чистка лица", ["дет", "реб", "cпин", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер", "аблятивн"], ["шлиф"], "Лазерная шлифовка лица", ["дет", "реб", "шрам", "руб", "неаблятив", "акци", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["механич"], ["чистк"], "Механическая чистка лица", ["дет", "налет", "зуб", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шлифов"], ["лиц"], "Шлифовка лица", ["дет", "лазер", "реб", "аблятивная", "втор", "описани", "шрам", "руб", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["энзимн", "фермент"], ["пилинг"], "Энзимный пилинг лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ультразвук", "ультрозвук"], ["чистк", "пилинг"], "Ультразвуковая чистка лица", ["дет", "реб", "втор", "челюст", "зуб", "стомато", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "лоп"])
	main_func(["химич"], ["чистк"], "Химическая чистка лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["фракционн"], ["омолож", "шлиф", "термолифтинг"], "Фракционное омоложение", ["дет", "реб", "растя", "руб", "акция", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фракц"], ["лифтинг", "омолож", "шлиф"], "Фракционный термолифтинг", ["дет", "руб", "шрам", "см", "кв", "образ", "операц", "живот", "акция", "лиц", "глаз", "орбит", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фракц"], ["лифтинг", "омолож", "шлиф", "шлиф"], "Фракционный термолифтинг живота", ["дет", "реб", "акция", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["живот"])
	main_func(["фракц"], ["лифтинг", "омолож", "шлиф"], "Фракционный термолифтинг лица", ["дет", "реб", "втор", "акция", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["фракц"], ["ифтинг", "омолож", "шлиф"], "Фракционный термолифтинг области глаз", ["дет", "реб", "акция", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["глаз", "орбит"])
	main_func(["фронтотомия"], ["фронтотомия"], "Фронтотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фото"], ["эпиляц", "удаление волос"], "Фотоэпиляция бикини", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"], ["бикини"])
	main_func(["фото"], ["эпиляц", "удаление волос"], "Фотоэпиляция верхней губы", ["дет", "реб", "пол", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"], ["губ", "усы", "усов"])
	main_func(["фото"], ["эпиляц", "удаление волос"], "Фотоэпиляция лба", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["лба", "лоб"])
	main_func(["фото"], ["эпиляц", "удаление волос"], "Фотоэпиляция лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"], ["лиц"])
	main_func(["фото"], ["эпиляц", "удаление волос"], "Фотоэпиляция ног", ["дет", "реб", "пальц", "более", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"], ["ног", "голен", "нижней конечност", "нижних конечностей"])
	main_func(["фото"], ["эпиляц", "удаление волос"], "Фотоэпиляция подбородка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["подбород"])
	main_func(["фото"], ["акне"], "Фототерапия акне", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "спин", "лба", "лбу", "лоб"], ["акне"])
	main_func(["фото"], ["акне"], "Фототерапия акне на лбу", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лбу", "лоб", "лба"])
	main_func(["фото"], ["акне"], "Фототерапия акне на спине", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["спин"])
	main_func(["фото", "aft"], ["омоложени"], "Фотоомоложение", ["дет", "реб", "втор", "кист", "абонем", "курс", "одной вспышки", "одна вспышка", "1 вспыш", "1 вспы", "рук", "лиц", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["фото", "aft"], ["омоложени"], "Фотоомоложение кистей рук", ["дет", "реб", "втор", "описани", "абонем", "курс", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["кист", "рук"])
	main_func(["фото", "aft"], ["омоложени"], "Фотоомоложение лица", ["дет", "реб", "втор", "описани", "расшифров", "абонем", "курс", "занятий", "посещени", "под", "для", "во время", "курс"], ["лиц"])
	main_func(["фото"], ["коагуляции сосуд", "коагуляция сосуд", "фототерапия сосудистых"], "Фотокоагуляция сосудов в области носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "крыл"], ["нос"])
	main_func(["фото"], ["коагуляции сосуд", "коагуляция сосуд", "фототерапия сосудистых"], "Фотокоагуляция сосудов на крыльях носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["крыльях носа", "крыльев носа"])
	main_func(["фото"], ["коагуляции сосуд", "коагуляция сосуд", "фототерапия сосудистых"], "Фотокоагуляция сосудов на подбородке", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["подбород"])
	main_func(["фото"], ["коагуляции сосуд", "коагуляция сосуд", "фототерапия сосудистых"], "Фотокоагуляция сосудов на щеках", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["щек", "щеч"])
	main_func(["belotero", "белотеро", "belatero"], ["belotero", "белотеро", "belatero"], "Инъекции Belotero", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ellanse"], ["ellanse"], "Инъекции Ellanse", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["juvederm", "ювидерм", "джувидерм", "juviderm"], ["juvederm", "ювидерм", "джувидерм", "juviderm"], "Инъекции Juvederm", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["meso-radiesse", "meso radies", "мезорадиес", "мезо-радиес"], ["meso-radiesse", "мезо-радиес", "мезорадиес", "meso radies"], "Инъекции Meso-Radiesse", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["princess", "принцесс"], ["princess", "принцесс"], "Инъекции Princess", ["дет", "акци", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["radiess", "радиесс"], ["radiess", "радиесс"], "Инъекции Radiesse", ["дет", "meso", "mezo", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["restyline", "restylane", "рестилайн", "restylaine", "restуlаne"], ["restyline", "restуlаne", "restylane", "рестилайн", "restylaine"], "Инъекции Restylane", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["surgiderm", "сургидерм", "сурджидерм"], ["surgiderm", "сургидерм", "сурджидерм"], "Инъекции Surgiderm", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["teosyal", "теосял", "теосиаль", "theosyal", "теасиаль"], ["teosyal", "теосял", "теосиаль", "theosyal", "теасиаль"], "Инъекции Teosyal", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["филлер"], ["губ"], "Уколы филлерами в губы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["филлер"], ["носослез"], "Уколы филлерами в носослезную борозду", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["филлер"], ["глаз", "орбит"], "Уколы филлерами под глаза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["электростимул"], ["электростимул"], "Электростимуляция", ["дет", "волос", "труб", "чрезкож", "чрескож", "мат", "моч", "уретр", "церви", "крани", "предста", "проста", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шотлан"], ["душ"], "Шотландский душ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экстракорп"], ["гемокоррекция"], "Экстракорпоральная гемокоррекция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хвойн"], ["ванн"], "Хвойная ванна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["циркуляр"], ["душ"], "Циркулярный душ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["флюктуоризаци"], ["флюктуоризаци"], "Флюктуоризация", ["дет", "внеполост", "вне полост", "наруж" "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["флюктуоризаци"], ["внеполост", "наружная"], "Флюктуоризация внеполостная", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["солев", "солян"], ["ингал", "пещер"], "Солевая ингаляция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["спелеотерапи"], ["спелеотерапи"], "Спелеотерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тепл"], ["лечени", "терапи"], "Теплолечение", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тракцион"], ["терапи"], "Тракционная терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["подводн"], ["душ", "массаж"], "Подводный душ массаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радон"], ["ванн"], "Радоновая ванна", ["дет", "реб", "парадонт", "пародонт", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сероводород"], ["ванн"], "Сероводородная ванна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["скипидар"], ["ванн"], "Скипидарная ванна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["смв", "сантиметров"], ["терапи"], "СМВ-терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["озокерит"], ["озокерит"], "Озокеритолечение", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["оксигенобаро"], ["терапи"], "Оксигенобаротерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["механотерапия"], ["механотерапия"], "Механотерапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лфк", "лечебная гимнастика", "лечебн"], ["физкультур", "лечебная гимнастика", "лфк"], "Лечебная физкультура / ЛФК", ["онлайн", "лиц", "масс", "скайп", "консультация", "skype", "online", "on line", "on-line", "втор", "описани", "беремен", "расшифров", "занятий", "во время", "курс"])
	main_func(["кислород"], ["ванн"], "Кислородная ванна", ["дет", "реб", "втор", "озон", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кишеч"], ["лаваж"], "Кишечный лаваж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["йодобром"], ["ванн"], "Йодобромная ванна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["каскад"], ["фильтрац"], "Каскадная фильтрация плазмы крови", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дмв"], ["терапи", "лечени"], "ДМВ-терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["душ"], ["виши"], "Душ Виши", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["душ"], ["шарко"], "Душ шарко", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["жемчу"], ["ванн"], "Жемчужная ванна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["иммуносорбц"], ["иммуносорбц"], "Иммуносорбция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["индуктотерми"], ["индуктотерми"], "Индуктотермия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["интерференц"], ["терапи", "лечени", "ток"], "Интерференцтерапия", ["дет", "реб", "доп", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гидромассаж"], ["гидромассаж"], "Гидромассажная ванна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["восходящ"], ["душ"], "Восходящий душ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ароматич"], ["ванн"], "Ароматическая ванна", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["маникюр"], ["аппарат", "апарат"], "Аппаратный маникюр", ["дет", "реб", "лечеб", "медицин",  "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["педикюр"], ["аппарат", "апарат"], "Аппаратный педикюр", ["дет", "реб", "лечеб", "медиц", "гриб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лечебн"], ["маникюр"], "Лечебный маникюр", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["медицин", "лечеб"], ["педикюр"], "Медицинский педикюр", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["педикюр"], ["гриб"], "Педикюр для ногтей с грибком", ["дет", "без гриб", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["укреплени"], ["ibx"], "Укрепление ногтей ibx", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["уродинамическ"], ["исследовани"], "Уродинамические исследования", ["дет", "видео", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ультразвук"], ["лифтинг", "подтяжк"], "Ультразвуковой лифтинг", ["дет", "смас", "гемор", "дезарте", "узл", "узел", "smas", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "допол", "для", "во время", "курс"])
	main_func(["ультразвук"], ["массаж"], "Ультразвуковой массаж", ["дет", "лиц", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ультразвук"], ["массаж лиц"], "Ультразвуковой массаж лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фонофорез", "фоноферез"], ["лиц"], "Ультрафонофорез лица", ["дет", "реб", "втор", "тел", "спин", "деколь", "груд", "ног", "рук", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["уранопластик", "пластик"], ["уранопластик", "расщелины неб"], "Уранопластика", ["дет", "реб", "поватор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["аттикоантротоми"], ["аттикоантротоми"], "Аттикоантротомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кохлеар"], ["имплант"], "Кохлеарная имплантация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мастоидотоми", "мастоидэктоми"], ["мастоидотом", "мастоидэктоми"], "Мастоидотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["протез"], ["слух"], "Слухопротезирование", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стапед"], ["пластик"], "Стапедопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тимпан"], ["пластик"], "Тимпанопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увт", "ударно-волнов"], ["терапи"], "Ударно-волновая терапия", ["дет", "простат", "предста", "шпор", "муж", "перед", "полов", "реб", "втор", "описани", "осмотр", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увт", "ударно-волнов"], ["простат", "предста", "полов", "муж", "эректильной дисфункции"], "Ударно-волновая терапия при заболеваниях мужских половых органов", ["дет", "реб", "втор", "перед", "осмотр", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["экзостоз"], "Удаление экзостоза", ["дет", "реб", "кост", "хрящ", "слух", "уш", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["эпулис"], "Удаление эпулиса", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["яичников"], "Удаление яичников", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лапаротом", "лапаратом"], ["оофорэ"], "Лапаротомная оофорэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "иссеч", "коррекция", "шлиф", "лечение"], ["рубц", "шрам"], "Удаление шрамов и рубцов", ["дет", "реб", "более", "от 5", "от 0.5", "от 0,5", "свыше", "втор", "носа", "лигатур", "с рубцов", "с послеоперационных ру", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экспандер", "эспандер"], ["дермотенз"], "Экспандерная дермотензия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "опера"], ["халязион", "холязион", "халазион"], "Удаление халязиона", ["дет", "лазер", "безоперац", "без операц", "хирурги", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["халязион", "холязион", "халазион"], "Лазерное удаление халязиона", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хирурги", "механич"], ["халязион", "холязион", "халазион"], "Хирургическое удаление халязиона", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["фурункул"], "Удаление фурункула", ["дет", "лазер", "хирург", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хирурги"], ["фурункул"], "Хирургическое удаление фурункула", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "коагул", "деструкци"], ["фибром"], "Удаление фибромы", ["дет", "радио", "сурги", "лазер", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["фибром"], "Радиоволновое удаление фибромы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["фибром"], "Удаление фибромы лазером", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стержн"], ["стержн"], "Удаление стержневой мозоли", ["дет", "механи", "хирурги", "лазер", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер"], ["стержн"], "Удаление стержневой мозоли лазером", ["дет", "реб", "втор", "описани", "более", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мозол"])
	main_func(["хирургичес", "механич"], ["стержн"], "Хирургическое удаление стержневой мозоли", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "лечени"], ["звезд", "сетки", "сеточки", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"], "Удаление сосудистых звездочек", ["дет", "реб", "фото", "свыше", "более", "начиная", "след", "втор", "лазер", "электр", "склеро", "озон", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сосуд", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"])
	main_func(["лазер"], ["звезд", "сетки", "сеточки", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии", "звездчатой гемангиомы"], "Лазерное удаление сосудистых звездочек", ["дет", "реб", "более", "втор", "начиная", "свыше", "след", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["сосуд", "телеангиоэктаз", "телеангиэктазий","ангиэктазии", "звездчатой гемангиомы"])
	main_func(["склер"], ["сосудистых звезд", "сосудистой сет", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"], "Микросклеротерапия сосудистых звездочек", ["дет", "начиная", "после", "свыше", "след", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["озон"], ["сосуд", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"], "Озонотерапия сосудистых звездочек", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "свыше", "начиная", "след", "под", "для", "во время", "курс"], ["сет", "звезд", "телеангиоэктаз", "телеангиэктази", "ангиэктазии"])
	main_func(["удал", "лечени", "устранени"], ["купероз"], "Удаление купероза на лице", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фото"], ["сосуд", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"], "Удаление сосудистых звездочек фотовспышкой", ["дет", "реб", "втор", "начиная", "описани", "расшифров", "более", "занятий", "свыше", "посещени", "след", "под", "для", "во время", "курс"], ["сет", "звезд", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"])
	main_func(["электр"], ["сосуд", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"], "Электрокоагуляция сосудистых звездочек", ["дет", "реб", "начиная", "втор", "описани", "расшифров", "занятий", "более", "посещени", "свыше", "под", "след", "для", "во время", "курс"], ["сет", "звезд", "телеангиоэктаз", "телеангиэктазий", "ангиэктазии"])
	main_func(["удал", "лечение"], ["растяж"], "Удаление растяжек", ["дет", "реб", "втор", "описани", "лазер", "расшифров", "занятий", "более", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["растяж"], "Лазерное удаление растяжек", ["дет", "реб", "втор", "описани", "расшифров", "от 5", "занятий", "более", "посещени", "под", "для", "во время", "курс"], ["удал", "леч", "шлиф"])
	main_func(["удал", "полипотомия носа", "полипотомии носа", "полипэктомия носа"], ["полип", "полипотомия носа", "полипотомии носа"], "Удаление полипов в носу", ["дет", "реб", "хоана", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["носа", "носу", "носовой", "внутриносов", "носовых", "носового"])
	main_func(["удал"], ["хоанальн"], "Удаление хоанального полипа носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "полипэктоми"], ["полип", "полипэктоми"], "Удаление полипов матки", ["дет", "реб", "втор", "радио", "сурги", "гистероскопическое", "гистероскопии", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["матк", "маточны"])
	main_func(["удал"], ["гистероскоп"], "Гистероскопическое удаление полипа матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["полипа матки", "маточного полипа", "полипов матки", "маточных полипов", "полипа шейки матки", "полипов шейки матки"])
	main_func(["удал"], ["радио", "сурги"], "Радиоволновое удаление полипа матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["полипа матки", "маточного полипа", "полипов матки", "маточных полипов", "полипа шейки матки", "полипов шейки матки"])
	main_func(["удал"], ["птериги"], "Удаление птеригиума", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "лечени"], ["подошвенных бородавок", "подошвенной бородавки", "бородавки на подошве"], "Удаление подошвенных бородавок", ["дет", "механич", "более", "лазер", "в рамках", "хирурги", "радио", "след", "сурги", "азот", "крио", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "деструкц"], ["лазер"], "Лазерное удаление подошвенных бородавок", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "более", "в рамках", "посещени", "след", "для", "во время", "курс"], ["подошвенных бородавок", "подошвенной бородавки", "подошвенная борода", "подошвенных вирусных бородавок", "подошвенной вирусной бородавки", "бородавки на подошве"])
	main_func(["удал", "деструкц"], ["радио", "сурги"], "Радиоволновое удаление подошвенных бородавок", ["дет", "реб", "втор", "описани", "расшифров", "более", "в рамках", "занятий", "посещени", "след", "для", "во время", "курс"], ["подошвенных бородавок", "подошвенной бородавки", "подошвенная борода", "бородавки на подошве"])
	main_func(["удал", "деструкц"], ["азот", "крио"], "Удаление подошвенных бородавок жидким азотом", ["дет", "реб", "втор", "описани", "расшифров", "в рамках", "более", "занятий", "посещени", "для", "след", "во время", "курс"], ["подошвенных бородавок", "подошвенной бородавки", "подошвенная борода", "бородавки на подошве"])
	main_func(["удал"], ["хирурги", "механич"], "Хирургическое удаление подошвенных бородавок", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "более", "в рамках", "посещени", "для", "во время", "след", "курс"], ["подошвенных бородавок", "подошвенной бородавки", "подошвенная борода", "бородавки на подошве"])
	main_func(["удал", "леч", "терапи"], ["пигментн"], "Удаление пигментных пятен", ["дет", "более", "лазер", "фото", "хими", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["пигмент"])
	main_func(["удал", "леч", "терап"], ["лазер"], "Лазерное удаление пигментных пятен", ["дет", "реб", "более", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["пигмент"])
	main_func(["фото"], ["пигмент"], "Фототерапия пигментных пятен", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "более", "посещени", "под", "для", "во время", "курс"])
	main_func(["пигмент"], ["пилинг"], "Химический пилинг пигментных пятен", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "более", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["ногт"], "Удаление ногтевой пластины", ["дет", "реб", "гриб", "вросш", "втор", "бород", "описани", "мозол", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "сечени"], ["образован"], "Удаление новообразования полости рта", ["дет", "реб", "гортан", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["рта"])
	main_func(["удал", "сече"], ["образован"], "Удаление новообразований молочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["молочн", "грудной железы"])
	main_func(["радикал"], ["резек", ], "Радикальная резекция молочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["молочной железы", "молочных желез"])
	main_func(["сектор"], ["резек"], "Секторальная резекция молочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["молочной железы", "молочных желез"])
	main_func(["склероз", "склеротерапи"], ["кист"], "Склерозирование кисты молочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["молочной железы", "молочных желез"])
	main_func(["удал"], ["опухол"], "Удаление доброкачественной опухоли молочной железы", ["дет", "рак", "злока", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["молочной железы", "молочных желез"])
	main_func(["удал"], ["кист"], "Удаление кисты молочной железы", ["дет", "склероз", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["молочной железы", "молочных желез"])
	main_func(["удал", "аблация", "абляция", "энуклеац"], ["фиброаденом"], "Удаление фиброаденомы молочной железы", ["дет", "следу", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал"], ["кист"], "Удаление кист влагалища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["влагал", "половых путей"])
	main_func(["удал", "марсупи", "вапоризац", "вскрыти", "вылущивание"], ["кист"], "Удаление кисты бартолиновой железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["бартолин", "бартали", "бактолин", "бортолин"])
	main_func(["удал"], ["образован"], "Удаление новообразований гортани", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["гортан"])
	main_func(["удал"], ["образован"], "Удаление новообразований полости носа", ["дет", "реб", "втор", "пазух", "гаймор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс", "сложност"], [" нос", "внутринос"])
	main_func(["удал", "адреналэктоми"], ["надпочечни", "адреналэктоми"], "Удаление надпочечника / Адреналэктомия", ["дет", "нефр", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "тонзиллэктоми", "тонзилэктоми", "деструкция миндалин"], ["миндалин", "тонзиллэктоми", "тонзилэктоми"], "Удаление миндалин / Тонзиллэктомия", ["лазер", "образ", "пробки", "пробок", "пробка", "абсцес", "сурги", "радио", "хирурги", "дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер"], ["тонзиллэктоми", "удаление миндалин", "удаления миндалин", "деструкция миндалин"], "Лазерная тонзиллэктомия", ["дет", "реб", "втор", "описани", "расшифров", "образ", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["тонзиллэктоми", "удаление миндалин", "удаления миндалин", "деструкция миндалин"], "Радиоволновая тонзиллэктомия", ["дет", "реб", "втор", "описани", "образ", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["хирурги"], ["тонзиллэктоми", "удаление миндалин", "удаления миндалин", "деструкция миндалин"], "Хирургическая тонзиллэктомия", ["дет", "реб", "втор", "описани", "расшифров", "образ", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удален", "сечени", "вскрыт"], ["копчик"], "Удаление копчиковой кисты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["кист"])
	main_func(["удал"], ["кератом"], "Удаление кератомы", ["дет", "радио", "сурги", "более", "свыше", "лаз", "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "след", "механич", "лазер", "азот", "крио", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["кератом"], "Радиоволновое удаление кератомы", ["дет", "более", "свыше", "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "след", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["кератом"], ["азот", "крио"], "Удаление кератомы жидким азотом", ["дет", "реб", "более", "свыше", "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "след", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["кератом"], ["лазер"], "Удаление кератомы лазером", ["дет", "реб", "втор", "описани", "более", "свыше", "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "след", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["хирургиче", "механич"], ["кератом"], "Хирургическое иссечение кератомы", ["дет", "электр", "радио", "реб", "более", "свыше", "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "след", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "извлеч"], ["камн", "конкремен"], "Удаление камней из почек", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "над", "для", "во время", "курс"], ["поч"])
	main_func(["нефролитотоми"], ["нефролитотоми"], "Нефролитотомия", ["дет", "секц", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["нефролитотрипси"], ["нефролитотрипси"], "Нефролитотрипсия", ["дет", "в ходе", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пиелолитотоми"], ["пиелолитотоми"], "Пиелолитотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пиелолитотрипс"], ["пиелолитотрипс"], "Пиелолитотрипсия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "извлеч"], ["камн", "конкремен"], "Удаление камней из желчного пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["желчного пузыря", "желчном пузыре"])
	main_func(["литотрипс"], ["желч"], "Литотрипсия желчных камней", ["дет", "реб", "цисто", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["камн"])
	main_func(["холецистолитотоми"], ["холецистолитотоми"], "Холецистолитотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удалени", "извлеч"], ["камн", "конкремент"], "Удаление камней из мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мочевого пузыря", "мочевом пузыре"])
	main_func(["цистолитотоми"], ["цистолитотоми"], "Цистолитотомия", ["дет", "реб", "холе", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["цистолитотрипси"], ["цистолитотрипси"], "Цистолитотрипсия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["цистолитоэкстракц"], ["цистолитоэкстракц"], "Цистолитоэкстракция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "извлеч"], ["камн", "конкремент"], "Удаление камней из мочеточника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мочеточник"])
	main_func(["уретеролитотрипси"], ["дистанци"], "Дистанционная уретеролитотрипсия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["уретеролитотрипсия"], ["контакт"], "Контактная уретеролитотрипсия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["уретеролитоэкстракци"], ["уретеролитоэкстракц"], "Уретеролитоэкстракция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["инород"], ["верхнечелюстн", "гайморов"], "Удаление инородного тела из верхнечелюстной пазухи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["гигром"], "Удаление гигромы", ["дет", "лазер", "механич", "хирурги", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гигром"], ["лазер"], "Лазерное удаление гигромы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пункци"], ["гигром"], "Пункция гигромы", ["дет", "реб", "втор", "описани", "контрол", "при", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хирургич", "механич"], ["гигром"], "Хирургическое удаление гигромы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удале"], ["жировик", "липом"], "Удаление жировика (липомы)", ["дет", "лазер", "радио", "сурги", "хирурги", "электр", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["жировик", "липом"], "Лазерное удаление жировика / липомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["жировик", "липом"], "Радиоволновое удаление жировика / липомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хирурги"], ["жировик", "липом"], "Хирургическое удаление жировика / липомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["электр"], ["жировик", "липом"], "Электрокоагуляция липомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "леч", "терап", "коагуля"], ["гемангиом"], "Удаление гемангиомы", ["дет", "кри", "свыше", "начиная", "от ", "в рамках", "азот", "след", "лазер", "радио", "сурги", "хирурги", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["крио", "азот"], ["гемангиом"], "Криодеструкция гемангиомы жидким азотом", ["дет", "след", "свыше", "в рамках", "начиная", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["гемангиом"], "Лазерное удаление гемангиомы", ["дет", "реб", "втор", "описани", "в рамках", "свыше", "след", "начиная", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["удал", "коагул", "деструк"])
	main_func(["радио", "сурги"], ["гемангиом"], "Радиоволновое удаление гемангиомы", ["дет", "реб", "втор", "в рамках", "след", "свыше", "начиная", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хирурги"], ["гемангиом"], "Хирургическое удаление гемангиомы", ["дет", "реб", "втор", "описани", "в рамках", "след", "свыше", "начиная", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["удал", "деструк"])
	main_func(["удал", "коагул", "лечение"], ["бородав"], "Удаление бородавок", ["дет", "реб", "подошв", "механич", "след", "более", "от ", "11-20", "6-10", "от 11", "от 10", "6-10", "2-5", "стоп", "в рамках", "лазер", "свыше", "радио", "сурги", "электр", "хирурги", "крио", "азот", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хирургич", "механич", "физическ"], ["бородав"], "Хирургическое удаление бородавок", ["дет", "реб", "подошв", "след", "более", "в рамках", "11-20", "6-10", "от 11", "от 10", "6-10", "2-5", "стоп", "лазер", "свыше", "радио", "сурги", "электр", "крио", "азот", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["крио", "азот"], ["бородав"], "Криодеструкция бородавок жидким азотом", ["дет", "подошв", "след", "более", "в рамках", "реб", "11-20", "6-10", "от 11", "от 10", "6-10", "2-5", "втор", "описани", "свыше", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["бородав"], "Радиоволновое удаление бородавок", ["дет", "подошв", "реб", "втор", "более", "в рамках", "след", "11-20", "6-10", "от 11", "от 10", "6-10", "2-5", "описани", "расшифров", "свыше", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["бородав"], "Удаление бородавок лазером", ["дет", "реб", "втор", "подошв", "описани", "в рамках", "более", "расшифров", "11-20", "6-10", "от 11", "от 10", "6-10", "2-5", "след", "занятий", "посещени", "свыше", "под", "для", "во время", "курс"])
	main_func(["электр"], ["бородав"], "Электрокоагуляция вульгарной бородавки", ["дет", "подошв", "реб", "втор", "в рамках", "более", "описани", "от 21", "11-20", "6-10", "от 11", "от 10", "6-10", "2-5", "расшифров", "след", "занятий", "свыше", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["атером", "кисты сальной железы"], "Удаление атеромы", ["дет", "реб", "радио", "катером", "механич", "сурги", "хирурги", "электр", "лазер", "мошо", "втор", "описани", "расшифров", "в рамках", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["атером", "кисты сальной железы"], "Лазерное удаление атеромы", ["дет", "реб", "мошон" "втор", "уш", "катером", "ух", "описани", "расшифров", "в рамках", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["атером", "кисты сальной железы"], "Радиоволновое удаление атеромы", ["дет", "реб", "втор", "мошон", "катером", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["хирурги", "механич"], ["атером"], "Хирургическое удаление атеромы", ["дет", "реб", "втор", "мошон", "описани", "катером", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["электр"], ["атером", "эпидермальной кисты", "кисты сальной железы"], "Электрокоагуляция атеромы", ["дет", "реб", "втор", "описани", "мошон", "катером", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["удал"], ["базалиом"], "Удаление базалиом", ["дет", "реб", "механич", "крио", "азот", "радио", "свыше", "сурги", "хирурги", "лазер", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "в рамках", "во время", "курс"])
	main_func(["крио", "азот"], ["базалиом"], "Криодеструкция базалиомы жидким азотом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["лазер"], ["базалиом"], "Лазерное удаление базалиом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время" "в рамках", "консультац", "прием", "курс"])
	main_func(["радио", "сурги"], ["базалиом"], "Радиоволновое удаление базалиом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["хирургич", "механич"], ["базалиом"], "Хирургическое удаление базалиом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["удален", "хирургическое лечение"], ["панариц"], "Удаление панариция", ["дет", "реб", "втор", "описани", "расшифров", "врос", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["удалени", "иссеч"], ["мозол"], "Удаление мозоли", ["дет", "механич", "радио", "сурги", "азот", "крио", "лазер", "хирургич", "стерж", "реб", "втор", "описани", "расшифров", "занятий", "в рамках", "посещени", "под", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["мозоли", "мозолей", "мазол"], "Радиоволновое удаление мозоли", ["дет", "реб", "стерж", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "в рамках", "во время", "курс"])
	main_func(["азот", "крио"], ["мозолей", "мозоли", "мазол"], "Удаление мозоли жидким азотом", ["дет", "реб", "стерж", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "в рамках", "во время", "курс"])
	main_func(["лазер"], ["мозоли", "мозолей", "мазол"], "Удаление мозоли лазером", ["дет", "реб", "втор", "стерж", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["хирургич", "механич"], ["мозоли", "мозолей", "мазол"], "Хирургическое удаление мозоли", ["дет", "реб", "втор", "стерж", "описани", "расшифров", "занятий", "посещени", "в рамках", "для", "во время", "курс"])
	main_func(["удал", "коагул"], ["милиум", "миллиум"], "Удаление милиумов", ["дет", "лазер", "радио", "сурги", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["лазер"], ["милиум", "миллиум"], "Лазерное удаление милиумов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "более", "для", "во время", "в рамках", "курс"])
	main_func(["радио", "сурги"], ["милиум", "миллиум"], "Радиоволновое удаление милиумов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["увелич", "увелеч"], ["губ"], "Увеличение губ", ["дет", "гелео", "гиалурон", "полов", "филл", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "в рамках", "курс"])
	main_func(["увел"], ["губ"], "Увеличение губ гиалуроновой кислотой", ["дет", "полов", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["гиалуро", "филл"])
	main_func(["лапароскоп", "лапараскоп"], ["коагул"], "Лапароскопическая коагуляция маточных труб", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["маточных труб"])
	main_func(["лапароскоп", "лапараскоп"], ["стерил"], "Лапароскопическая стерилизация кольцами или клеммами", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["кольц", "клемм"])
	main_func(["стоун", "камн"], ["терапи", "массаж"], "Стоун-терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["парафанго"], ["парафанго"], "Парафанго", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["срар", "cpap", "сипап"], ["терапи"], "СРАР-терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["скелетное вытяжение"], ["большеберц"], "Скелетное вытяжение за бугристость большеберцовой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["скелетное вытяжение"], ["надмы"], "Скелетное вытяжение за надмыщелки бедра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["скелетное вытяжение"], ["пято"], "Скелетное вытяжение за пяточную кость", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["скелетное вытяжение"], ["клапп"], "Скелетное вытяжение по Клаппу", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["синовэктоми"], ["голеностоп"], "Синовэктомия голеностопного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["синовэктоми"], ["колен"], "Синовэктомия коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["синовэктоми"], ["локт"], "Синовэктомия локтевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["синовэктоми"], ["тазобедр"], "Синовэктомия тазобедренного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ринопластика"], ["ринопластика"], "Ринопластика", ["дет", "реб", "втор", "безопер", "без опер", "ронхопат", "уву", "увол", "наполни", "филлер", "филер", "гиалурон", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["коррекц", "пластик"], ["крыльев носа"], "Коррекция крыльев носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик", "коррекц"], ["колумелл"], "Пластика колумеллы носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик", "коррекц"], ["кончика носа"], "Пластика кончика носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["рекон", "восстановл"], ["наружного носа"], "Реконструкция наружного носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["риносептопластик"], ["риносептопластик"], "Риносептопластика носовой перегородки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["септопластик"], ["септопластик"], "Септопластика носовой перегородки", ["дет", "рин", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["горбин"], ["носа", "носу", "носа"], "Удаление горбинки носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["укороч", "умень"], ["носа"], "Укорочение носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["репозиц"], ["внутрен"], "Репозиция внутренней лодыжки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лодыж"])
	main_func(["репозиц"], ["головк"], "Репозиция головки лучевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лучево"])
	main_func(["репозиц"], ["диафиз"], "Репозиция диафиза костей предплечья", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["предплечь"])
	main_func(["репозиц"], ["локт"], "Репозиция диафиза локтевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["репозиц"], ["ключиц"], "Репозиция ключицы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["репозиц"], ["запяст"], "Репозиция костей запястья", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["репозиц"], ["лучево"], "Репозиция лучевой кости", ["дет", "головк", "диафиз", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["репозиц"], ["наружн"], "Репозиция наружной лодыжки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["лодыжк"])
	main_func(["репозиц"], ["плюсн"], "Репозиция плюсневых костей", ["дет", "предплюс", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["репозиц"], ["пястн"], "Репозиция пястной кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лазер"], ["реканализац"], "Лазерная реканализация цервикального канала", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["цервикал"])
	main_func(["резекци"], ["желудк"], "Резекция желудка", ["дет", "дисталь", "продоль", "прокси", "слизист", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["желудк"], "Дистальная резекция желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["дисталь"])
	main_func(["резекц"], ["желудк"], "Продольная резекция желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["продоль"])
	main_func(["резекц"], ["желудк"], "Проксимальная резекция желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["проксималь"])
	main_func(["резекц"], ["слизист"], "Резекция слизистой желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["желудк"])
	main_func(["сечен", "разделени", "разведени", "устранение", "синехиотомия", "удаление"], ["синех", "синехий крайней плоти", "синехи полов", "сенехий крайней", "сенехий полов", "сенехий", "спаек при", "спаек край"], "Рассечение синехий", ["дет", "губ", "нос", "вульв", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["синех", "сенехи"], ["половых губ", "наружных половых органов", "вульв"], "Рассечение синехий половых губ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени",  "для", "во время", "курс"])
	main_func(["синех", "сенехи"], ["носа", "носу", "носов", "полости", " нос"], "Удаление синехий полости носа", ["дет", "мат", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["плевр"], "Плевральная пункция", ["дет", "реб", "биопси", "втор", "описани", "расшифров", "контрол", "при", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["асцитическ"], "Пункция асцитической жидкости", ["дет", "реб", "втор", "описани", "контрол", "при", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["голеностоп"], "Пункция голеностопного сустава", ["дет", "реб", "втор", "описани", "контрол", "при", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["желудочков"], "Пункция желудочков головного мозга", ["дет", "реб", "втор", "описани", "контрол", "при ", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["мозг"])
	main_func(["пункц", "пунктирование"], ["заднего свода", "задний свод"], "Пункция заднего свода влагалища", ["дет", "реб", "втор", "контрол", "при", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["кист"], "Пункция кисты головного мозга", ["дет", "реб", "костн", "втор", "описани", "контрол", "при ", "расшифров", "занятий", "посещени", "для", "во время", "курс", "спинн"], ["мозг"])
	main_func(["пункц", "пунктирование"], ["кист"], "Пункция кисты яичника", ["дет", "реб", "втор", "описани", "расшифров", "контроль", "контроля", "при", "занятий", "посещени", "для", "во время", "курс"], ["яичник"])
	main_func(["пункц", "пунктирование"], ["коленн"], "Пункция коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при ", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["костн"], "Пункция костного мозга", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при ", "занятий", "посещени", "для", "во время", "курс"], ["мозг"])
	main_func(["пункц", "пунктирование"], ["лимф"], "Пункция лимфоузла", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "контроль", "при ", "посещени", "для", "во время", "курс"], ["узл", "узе"])
	main_func(["пункц", "пунктирование"], ["лобных пазух", "лобной пазухи", "верхнечелюст"], "Пункция лобных пазух", ["дет", "реб", "втор", "контрол", "при ", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["локт"], "Пункция локтевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при ", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["лучезап"], "Пункция лучезапястного сустава", ["дет", "реб", "втор", "описани", "контрол", "при ", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["молочн"], "Пункция молочной железы", ["дет", "реб", "навигация", "наведение", "биопсия", "контрол", "при ", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["мошонк"], "Пункция мошонки", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при" , "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование", "перикардиоцентез"], ["перикард", "перикардиоцентез"], "Пункция перикарда", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при ", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["плеч"], "Пункция плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при ", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["подвз", "повзд"], "Пункция подвздошной кости", ["дет", "реб", "втор", "описани", "контрол", "при ", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["предстат", "простат"], "Пункция предстательной железы", ["дет", "реб", "втор", "контрол", "при ", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["спинн"], "Пункция спинного мозга", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при ", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["тазобедр"], "Пункция тазобедренного сустава", ["дет", "реб", "втор", "описани", "контрол", "при ", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["яичник"], "Пункция фолликулов яичников", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при ", "занятий", "посещени", "для", "во время", "курс"], ["фолликул"])
	main_func(["пункц", "пунктирование"], ["стерналь"], "Стернальная пункция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "контрол", "при ", "посещени", "для", "во время", "курс"])
	main_func(["тимпанопункци"], ["тимпанопункци"], "Тимпанопункция", ["дет", "реб", "втор", "описани", "расшифров", "контрол", "при", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["протез"], ["ногт"], "Протезирование ногтевых пластин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пува", "puva"], ["терапи"], "ПУВА-терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["промыва"], ["мочевого пузыр"], "Промывание мочевого пузыря", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["визометрия", "остроты зрения"], ["визометрия", "остроты зрения"], "Визометрия", ["дет", "корре", "испра", "линз", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["подбор"], ["линз", "контактн"], "Подбор контактных линз", ["дет", "реб", "втор", "описани", "очков", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["подбор", "коррекции зрен"], ["очков"], "Подбор очков для зрения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	new_doubler_main(["очков"], ["сложн", "цилиндри", "бифокал", "призм"], ["втор", "дет", "реб", "посеще", "не ц", "не б", "не с", "во время", "курс"], "Подбор очков для зрения", 1132371)
	main_func(["прокол", "прокал", "пирсинг"], ["ушей", "мочек", "мочки", "мочка"], "Прокол ушей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["эрозии", "диатермокоа", "электр"], ["шейки матки"], "Прижигание эрозии шейки матки", ["дет", "реб", "папил", "консерват", "кондил", "образован", "опухол", "солковаг", "лазер", "влаг", "вульв", "крио", "кони", "азот", "радио", "сурги", "медикамент", "набот", "желез", "кист", "желуд", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шейки матки", "шейки мати", "эрозии", "эктопии", "эктопиях", "эктопий", "эндометрийо", "эндометриой"], ["азот", "крио"], "Прижигание эрозии шейки матки жидким азотом", ["дет", "папил", "опухол", "кони", "кондил", "влаг", "вульв", "образован", "деструкцией шейки", "деструкцией матки", "деструкция шейки", "деструкция матки", "деструкции шейки", "деструкции матки", "ампутац", "биопси", "кист", "набот", "желез", "удал", "резекц", "реб", "желуд", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шейки матки", "шейки мати", "эрозии", "эктопии", "эктопиях", "эктопий", "эндометрийо", "эндометриой"], ["лазер"], "Прижигание эрозии шейки матки лазером", ["дет", "этап", "реб", "терап", "папил", "шлиф", "кониз", "кондил", "влаг", "вульв", "опухол", "образован", "втор", "кист", "биопси", "набот", "желез", "желуд", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шейки матки", "шейки мати", "эрозии", "эктопии", "эктопиях", "эктопий", "эндометрийо", "эндометриой"], ["солковаг", "медикамент", "медимикамент", "лекарств", "консервативное лечение", "химическ"], "Прижигание эрозии шейки матки Солковагином", ["дет", "кони", "влаг", "вульв", "папил", "опухол", "кондил", "образован", "реб", "кист", "набот", "желез", "втор", "желуд", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шейки матки", "шейки мати", "эрозии", "эктопии", "эктопиях", "эктопий", "эндометрийо", "эндометриой"], ["радио", "сурги"], "Радиоволновое прижигание эрозии шейки матки", ["дет", "сине", "сеч", "биопси", "реб", "опухол", "кони", "папил", "кондил", "влаг", "вульв", "кист", "набот", "желез", "образован", "втор", "желуд", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кантопекси", "подтяжк"], ["кантопекси", "уголков глаз"], "Кантопексия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["кантопластик"], ["кантопластик"], "Кантопластика", ["дет", "эпикан", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пластик", "пунктопластика", "расширение"], ["слез", "пунктопластика"], "Пластика слезной точки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["точк", "точек", "пунктопластика"])
	main_func(["ампутаци"], ["шейки матки"], "Высокая ампутация шейки матки", ["дет", "конус", "реб", "втор", "описани", "штурмд", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["штурмдорф", "конусовидна"], ["штурмдорф", "ампутация шейки матки", "ампутации шейки матки"], "Конусовидная ампутация шейки матки по Штурмдорфу", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криодеструкци"], ["шейки матки"], "Криодеструкция шейки матки", ["дет", "реб", "эрози", "эктопи", "образ", "опу", "кист", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["шейки матки"], "Пластика шейки матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экстирпац"], ["культ"], "Экстирпация культи шейки матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["шейки матки"])
	main_func(["дезартеризац", "hal"], ["геморр", "гемор", "rar"], "Дезартеризация геморроидальных узлов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["узл", "rar"])
	main_func(["лонго"], ["операц"], "Операция Лонго", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["геморр", "гемор"], "Радиоволновое удаление геморроидального узла", ["дет", "реб", "бахром", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["коагул", "удал", "сечен", "эвакуац"])
	main_func(["склеротерап", "склерозировани"], ["гемор"], "Склеротерапия геморроидальных узлов", ["дет", "реб",])
	main_func(["тромбэктоми", "тромб"], ["гемор"], "Тромбэктомия геморроидального узла", ["дет", "реб", "втор", "1 перевяз", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["фото", "инфра"], ["коагул", "удален"], "Фотокоагуляция геморроидального узла", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["гемор"])
	main_func(["электр"], ["коагул", "удал"], "Электрокоагуляция геморроидального узла", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["гемор"])
	main_func(["антираб"], ["прививк", "вакцин", "сыворотк"], "Антирабическая прививка", ["втор", "книж", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["осмотр", "консультаци", "прием"], ["перед", "для", "предвак", "предприв"], "Осмотр перед вакцинацией", ["втор", "книж", "приема", "внутрь", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"], ["вакцин"])
	main_func(["пневмококк"], ["прививк", "вакцин"], "Пневмококковая прививка", ["дет", "23", "пневмовакс", "книж", "превенар", "реб","втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["пневмо 23", "пневмо-23", "пневмо23"], ["пневмо 23", "пневмо-23", "пневмо23"], "Пневмококковая прививка Пневмо 23", ["втор", "книж", "осмотр ", "прием", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["пневмовакс"], ["пневмовакс"], "Пневмококковая прививка Пневмовакс 23", ["втор", "описани", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["превенар"], ["превенар"], "Пневмококковая прививка Превенар", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"])
	main_func(["аваксим"], ["аваксим"], "Прививка Аваксим", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["адасель"], ["адасель"], "Прививка Адасель", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["адс"], "Прививка АДС", ["дет", "м", "реб","втор", "описани", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["адсм", "адс-м", "адсм  м"], ["адсм", "адс-м", "адсм  м"], "Прививка АДС-М", ["втор", "описани", "книж", "книжк", "доп", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["акдс"], ["акдс"], "Прививка АКДС", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["акт-хиб", "акт хиб"], ["акт-хиб", "акт хиб"], "Прививка АКТ-Хиб", ["втор", "описани", "книж", "расшифров", "занятий", "посещени",  "во время", "курс", "дмс"])
	main_func(["альгавак"], ["альгавак"], "Прививка Альгавак М", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"])
	main_func(["ас-анатокси", "ас анатокси"], ["ас-анатокси", "ас анатокси"], "Прививка АС-анатоксин", ["втор", "книж", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["бивак"], ["полио"], "Прививка БиВак полио", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["бцж", "туберкулез"], ["прививк", "бцж"], "Прививка БЦЖ (от туберкулеза)", ["втор", "описани", "маркер", "инфаркт", "миока", "книж", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["вакта"], ["прививк", "вакцин"], "Прививка Вакта", ["втор", "описани", "расшифров", "занятий", "книж", "посещени",  "во время", "курс", "дмс"])
	main_func(["вианвак"], ["вианвак"], "Прививка Вианвак", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["гардасил"], ["гардасил"], "Прививка Гардасил", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["жкв"], ["прививк", "вакцин", "корь", "кори"], "Прививка ЖКВ", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["инфанрикс", "инфаринкс"], ["инфанрикс", "инфаринкс"], "Прививка Инфанрикс", ["дет", "гекса", "реб","втор", "описани", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["инфанрикс"], ["гекса"], "Прививка Инфанрикс-гекса", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["клещ"], ["э-вак", "э вак", "е-вак", "евак"], "Прививка Клещ-Э-Вак", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["менактра"], ["менактра"], "Прививка Менактра", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["менинго"], ["а+с", "a+c"], "Прививка Менинго А+С", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["менцевакс"], ["менцевакс"], "Прививка Менцевакс", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["опв"], "Прививка ОПВ", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "профилактика"], ["бешенств"], "Прививка от бешенства", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "профилактика"], ["тиф"], "Прививка от брюшного тифа", ["втор", "тифим", "сертифика", "описани", "книж", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["впч", "папиллом"], "Прививка от ВПЧ", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["гемофил"], "Прививка от гемофильной инфекции", ["втор", "описани", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["гепатита а", "гепатита a"], "Прививка от гепатита А", ["втор", "книжк", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "рекомбина", "дрожжев"], ["гепатита б", "гепатита b", "гепатита в"], "Прививка от гепатита В", ["втор", "книж", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "витагерпевак"], ["герпес", "витагерпевак"], "Прививка от герпеса", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "профилактика"], ["дифтери"], "Прививка от дифтерии", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["желт"], "Прививка от желтой лихорадки", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "профилактика"], ["энцефалит"], "Прививка от клещевого энцефалита", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс", "коллектив"])
	main_func(["прививк", "вакцин"], ["коклюш"], "Прививка от коклюша", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "профилактика"], ["кори", "корев"], "Прививка от кори", ["дет", "паротит", "реб","втор", "описани", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "культур", "жив",], ["кори", "корев"], "Прививка от кори и паротита", ["втор", "описани", "краснух", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"], ["паротит"])
	main_func(["краснух"], ["кор"], "Прививка от кори, краснухи, паротита", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"], ["паротит"])
	main_func(["прививк", "вакцин", "профилактика"], ["краснух"], "Прививка от краснухи", ["дет", "кор", "паротит", "реб","втор", "книж", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["менингит"], "Прививка от менингита", ["втор", "описани", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["паротит"], "Прививка от паротита", ["дет", "кор", "краснух" "реб","втор", "книж", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин", "полимилекс"], ["полиомелит", "полиомиелит", "полимилекс"], "Прививка от полиомиелита", ["втор", "орикс", "книж", "имовакс", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["полио"], ["имовакс", "иммовакс"], "Прививка от полиомиелита Имовакс Полио (нет в РФ)", ["втор", "книж", "описани", "расшифров", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["полиорикс"], ["полиорикс"], "Прививка от полиомиелита Полиорикс", ["втор", "описани", "расшифров", "книж", "занятий", "посещени", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["ротавирус"], "Прививка от ротавируса", ["втор", "ротатек", "описани", "расшифров", "книж", "занятий", "посещени",  "во время", "курс", "дмс"])
	main_func(["ротатек", "рота-тек", "рота тек"], ["ротатек", "рота-тек", "рота тек"], "Прививка от ротавируса РотаТек", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "сыворот", "анатоксин", "анатаксин", "профилактика"], ["столбняк", "столбнячн"], "Прививка от столбняка", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"])
	main_func(["пентаксим"], ["пентаксим"], "Прививка Пентаксим", ["втор", "описани", "осмотр ", "прием", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"])
	main_func(["приорикс"], ["приорикс"], "Прививка Приорикс (нет в РФ)", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["менингокок"], "Прививка против менингококковой инфекции", ["втор", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс"], "дмс")
	main_func(["регевак"], ["регевак"], "Прививка Регевак В", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["синфлорикс"], ["синфлорикс"], "Прививка Синфлорикс", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["тифим ви"], ["тифим ви"], "Прививка Тифим ВИ", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["фсме"], ["иммун"], "Прививка ФСМЕ-ИММУН Инжект (нет в РФ)", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["хаврикс"], ["1440", "1 440"], "Прививка Хаврикс 1440", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "во время", "курс", "дмс"])
	main_func(["хаврикс"], ["720"], "Прививка Хаврикс 720", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["хиберикс"], ["хиберикс"], "Прививка Хиберикс", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["церварикс"], ["церварикс"], "Прививка Церварикс", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["шигеллвак"], ["шигеллвак"], "Прививка Шигеллвак", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["энджерикс"], ["энджерикс"], "Прививка Энджерикс В", ["втор", "осмотр ", "прием", "описани", "расшифров", "занятий", "книж", "посещени", "во время", "курс", "дмс"])
	main_func(["энцепур"], ["энцепур"], "Прививка Энцепур", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["эувакс"], ["эувакс"], "Прививка Эувакс", ["втор", "описани", "расшифров", "занятий", "посещени", "во время", "книж", "курс", "дмс"])
	main_func(["грудн", "грудо"], ["массаж"], "Сегментарный массаж грудного отдела позвоночника", ["дет", "реб", "втор", "допол", "все отделы", "всех отделов", "описани", "клетк", "полост", "меда", "медовый", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "шейн"])
	main_func(["пояс", "крестцов", "крестца"], ["массаж"], "Сегментарный массаж пояснично-крестцового отдела позвоночника", ["дет", "реб", "допол", "меда", "медовый", "втор", "все отделы", "всех отделов", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["шейн"], ["массаж"], "Сегментарный массаж шейно-грудного отдела позвоночника", ["дет", "реб", "втор", "воротник", "описани", "меда", "медовый", "расшифров", "допол", "все отделы", "всех отделов", "клет", "полост", "занятий", "посещени", "под", "для", "во время", "курс"], ["груд"])
	main_func(["прививк", "вакцин"], ["грипп"], "Прививка от гриппа", ["втор", "описани", "гриппол", "агрипал", "дому", "последую", "предел", "мкад", "агриппал", "книж", "ваксигрип", "инфлювак", "ультрикс", "расшифров", "занятий", "посещени", "под", "во время", "курс", "дмс"])
	main_func(["ваксигрип"], ["ваксигрип"], "Прививка Ваксигрип", ["втор", "описани", "расшифров", "занятий", "посещени", "под", "книж", "во время", "курс", "дмс"])
	main_func(["гриппол"], ["гриппол"], "Прививка Гриппол", ["втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "книж", "курс", "дмс"])
	main_func(["инфлювак"], ["инфлювак"], "Прививка Инфлювак", ["втор", "описани", "расшифров", "занятий", "посещени", "под", "книж", "во время", "курс", "дмс", "сотрудни"])
	main_func(["ультрикс"], ["ультрикс"], "Прививка Ультрикс", ["описани", "расшифров", "занятий", "посещени", "под", "книж", "во время", "курс", "дмс"])
	main_func(["прививк", "вакцин"], ["ветрян"], "Прививка от ветрянки", ["варилрикс", "окавакс", "втор", "описани", "расшифров", "книж", "занятий", "посещени", "под", "во время", "курс", "дмс"])
	main_func(["варилрикс"], ["варилрикс"], "Прививка от ветрянки Варилрикс", ["втор", "описани", "расшифров", "занятий", "посещени", "книж", "под", "во время", "курс", "дмс"])
	main_func(["окавакс"], ["окавакс"], "Прививка от ветрянки Окавакс", ["втор", "описани", "расшифров", "занятий", "посещени", "под", "книж", "во время", "курс", "дмс"])
	main_func(["прессотерапи", "прессатерап"], ["прессотерапи"], "Прессотерапия", ["дет", "верх", "штан", "сеансов", "рук", "ниж", "низ", "ног", "тела", "тело", "голо", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["прессотерап", "прессатерап"], ["верхни", "рук"], "Прессотерапия верхних конечностей", ["дет", "реб", "втор", "сеансов", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["прессотерап", "прессатерап"], ["всего тела", "тела полностью", "тело полностью", "все тело"], "Прессотерапия всего тела", ["дет", "реб", "костю", "шта", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["прессотерапи", "прессатерап"], ["голов"], "Прессотерапия головы", ["дет", "реб", "втор", "описани", "сеансов", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["прессотерапи", "прессатерап"], ["ниж", "ног"], "Прессотерапия нижних конечностей", ["дет", "реб", "штан", "втор", "сеансов", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["подтяж", "платизмопластик"], ["шеи", "платизмопластик"], "Подтяжка шеи / платизмопластика", ["дет", "нит", "лазер", "dot", "rf", "термо", "безопер", "без опера", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["смас", "smas"], ["шеи"], "SMAS-лифтинг шеи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["нитев", "нитя"], ["подтяж", "лифти"], "Нитевая подтяжка шеи / Soft-Lift", ["дет", "реб", "втор", "синус", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["шеи"])
	main_func(["подтяж"], ["лиц"], "Подтяжка лица", ["дет", "височ", "рубц", "лазер", "dot", "rf", "термо", "коротк", "золот", "нит", "аптос", "aptos", "круг", "глубо", "височ", "эндотин", "круг", "глуб", "реб", "эндоскоп", "средн", "безоперац", "без опера" "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["macs"], ["лифтинг", "подтяжк"], "MACS-лифтинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["space"], ["лифтинг", "подтяж"], "SPACE-лифтинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["безопераци", "без операци"], ["подтяж", "лифтин"], "Безоперационная подтяжка лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лиц"])
	main_func(["височ"], ["подтяж", "лифтин"], "Височная подтяжка лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["глубок"], ["подтяж", "лифтин"], "Глубокая подтяжка лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лиц"])
	main_func(["круг", "ритидэктоми"], ["подтяж", "лифтин", "ритидэктоми"], "Круговая подтяжка лица / Ритидэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лиц", "ритидэктоми"])
	main_func(["подтяж", "лифтин"], ["лба", "лоб"], "Подтяжка лба", ["дет", "лазер", "dot", "rf", "термо", "реб", "нит", "лобка", "лобковой", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж", "лифтин"], ["золот"], "Подтяжка лица золотыми нитями", ["дет", "лазер", "dot", "rf", "термо", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["нит"])
	main_func(["подтяж", "лифтин"], ["лиц"], "Подтяжка лица мезонитями", ["дет", "реб", "втор", "лазер", "dot", "rf", "термо", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["мезони"])
	main_func(["подтяж", "лифтин"], ["лиц"], "Подтяжка лица с коротким рубцом", ["дет", "реб", "лазер", "dot", "rf", "термо", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["корот"])
	main_func(["подтяж", "лифтин"], ["эндотин"], "Подтяжка лица эндотинами", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лиц"])
	main_func(["подтяж", "лифтин", "нити", "коррекц"], ["аптос", "aptos"], "Подтяжка нитями АПТОС", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж", "лифтин"], ["сред"], "Подтяжка средней зоны лица", ["дет", "реб", "втор", "термо", "rf", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лиц"])
	main_func(["подтяж", "лифтин"], ["эндоскоп"], "Эндоскопическая подтяжка лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лиц"])
	main_func(["подтяж", "лифтин", "мастопекси"], ["груди", "мастопекси"], "Подтяжка груди / Мастопексия", ["дет", "класс", "вертикал", "циркуляр", "круг", "ареоляр", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж", "лифтин", "мастопекс"], ["классическ"], "Классическая мастопексия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж", "лифтин", "мастопекс"], ["вертик"], "Мастопексия с вертикальным рубцом", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж", "лифтин", "мастопекс"], ["ареоляр", "циркул", "круговая мастопекс"], "Периареолярная / циркулярная мастопексия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["плацент"], ["терапи", "лечени"], "Плацентарная терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["куразьен", "curacen", "курасен", "curasen"], ["куразьен", "курасен", "curacen", "curasen"], "Инъекции Curacen", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["melsmon", "мелсмон", "мэлсмон"], ["melsmon", "мэлсмон", "мелсмон"], "Инъекции Melsmon", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лаеннек", "laennec"], ["лаеннек", "laennec"], "Лаеннек-терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластика"], ["ягодиц"], "Пластика ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["липомодел", "липоскульптур"], ["ягоди"], "Липомоделирование ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увел"], ["ягоди"], "Увеличение ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["филамен"], ["ягоди"], "Филаментлифтинг ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["подтяжк"], ["ягоди"], "Хирургическая подтяжка ягодиц", ["дет", "нит", "фил", "инъек", "укол", "гиалу", "масс", "терап", "ультра", "безопера", "без опера", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["протез"], ["ягоди"], "Эндопротезирование ягодиц", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["v-line", "v line"], "V-line пластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["димплэктоми"], ["димплэктоми"], "Димплэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["мандибулопластик"], ["мандибулопластик"], "Мандибулопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик", "ментопластик"], ["подбородо", "ментопластик"], "Пластика подбородка / Ментопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["подборо"])
	main_func(["удал"], ["втор"], "Удаление второго подбородка", ["дет", "реб", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["подбород"])
	main_func(["комков", "комочков", "удален"], ["биша"], "Удаление комков биша", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["фронтопластик"], ["фронтопластик"], "Фронтопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["губ"], "Пластика губ", ["дет", "полов", "уздеч", "контур", "v-y", "вестибуло", "ин", "гел", "гиалу", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["v-y"], ["пластика губ"], "V-Y пластика губ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["велофарингопластик"], ["велофарингопластик"], "Велофарингопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["корнер лифт", "подтяж", "лифтин"], ["корнер лифт", "уголков губ"], "Корнер лифт", ["дет", "нит", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["расщелин"], ["верхней губ"], "Коррекция расщелины верхней губы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["булхорн", "буллхорн", "bullhorn"], ["булхорн", "буллхорн", "bullhorn"], "Операция Булхорн", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["кесселринг", "кессельринг"], ["кессельринг", "кесселринг"], "Операция Кессельринга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["умень"], ["губ"], "Уменьшение размера губ", ["дет", "полов", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["плазмолифтинг", "плазмалифтинг"], ["плазмолифтинг", "плазмалифтинг"], "Плазмолифтинг", ["дет", "голов", "дополни", "готов", "акци", "интим", "век", "скид", "цервикал", "полов", "гинекол", "влага", "вульв", "волос", "декольте", "груд", "кист", "лиц", "ше", "десе", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["плазм"], ["век", "параоорбитал"], "Плазмолифтинг век", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "дополни", "готов", "акци", "под", "скид", "для", "во время", "курс"], ["лифт", "терапи"])
	main_func(["плазм"], ["волос", "голов"], "Плазмолифтинг волосистой части головы", ["дет", "реб", "втор", "описани", "расшифров", "готов", "дополни", "занятий", "акци", "скид", "посещени", "под", "для", "во время", "курс"], ["лифт", "терапи"])
	main_func(["плазм"], ["декольт"], "Плазмолифтинг декольте", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "готов", "посещени", "под", "дополни", "акци", "скид", "для", "во время", "курс"], ["лифт", "терапи"])
	main_func(["плазм"], ["рук", "кистей", "кисти", "кисть"], "Плазмолифтинг кистей рук", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "готов", "посещени", "под", "дополни", "акци", "скид", "для", "во время", "курс"], ["лифт", "терапи"])
	main_func(["плазм"], ["лиц"], "Плазмолифтинг лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "готов", "под", "для", "дополни", "акци", "скид", "во время", "курс"], ["лифт", "терап"])
	main_func(["плазм"], ["шеи", "шея", "шейн"], "Плазмолифтинг шеи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "готов", "дополни", "под", "для", "акци", "скид", "во время", "курс"], ["лифт", "терапи"])
	main_func(["плазм"], ["десен"], "Плазмолифтинг десен", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "готов", "под", "акци", "дополни", "скид", "для", "во время", "курс"], ["лифт", "терап"])
	main_func(["плазмотерапия", "плазматерапия", "prp", "аутоплазменная терапия", "введение"], ["плазмы", "плазмотерапия", "плазматерапия", "терапи", "аутоплазменная терапия", "prp"], "Плазмотерапия", ["дет", "век", "мезо", "кист", "после", "костной", "предоплат", "реб", "после", "втор", "описани", "дополни", "расшифров", "офр", "озон", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["подтяж"], ["бедер"], "Подтяжка бёдер", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["увелич", "увелеч"], ["бедер"], "Увеличение бёдер", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик", "увели"], ["голене", "голени"], "Пластика голеней", ["дет", "реб", "лож", "сустав", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["корр"], ["остеотоми"], "Корригирующая остеотомия по Илизарову", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["илизаро", "иллизаров"])
	main_func(["пирсинг", "прокол"], ["бров"], "Пирсинг брови", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пирсинг", "прокол"], ["губ"], "Пирсинг губы", ["дет", "полов", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пирсинг", "прокол"], ["половых губ", "клитор", "вульв", "влага", "интим"], "Пирсинг на интимных местах", ["дет", "реб", "леч", "терап", "кист", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пирсинг", "прокол"], ["нос"], "Пирсинг носа", ["дет", "реб", "втор", "пункци", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пирсинг", "прокол"], ["пупка", "пупок"], "Пирсинг пупка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пирсинг", "прокол"], ["соск"], "Пирсинг сосков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пирсинг", "прокол"], ["язык"], "Пирсинг языка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["arion"], ["arion"], "Грудные импланты Arion", ["дет", "реб", "clarion", "брекет", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["sebbin"], ["sebbin"], "Грудные импланты Sebbin", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["капсулотоми"], ["молочн", "грудн"], "Капсулотомия молочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["коррекц", "пластик"], ["ареол"], "Коррекция ареолы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["коррекц"], ["грудных имплант"], "Коррекция грудных имплантов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["коррекц", "пластик", "рекон"], ["соск"], "Коррекция сосков", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["реконструкци"], ["груди"], "Реконструкция (восстановление) груди", ["дет", "реб", "втор", "после реконстру", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["силикон"], ["грудн", "молочн"], "Силиконовые грудные импланты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["имплант"])
	main_func(["увеличение", "увелеч"], ["груди", "молочных желе"], "Увеличение груди", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удале", "извлеч"], ["грудн", "молоч"], "Удаление грудных имплантов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["имплан", "геля"])
	main_func(["удале"], ["добав", "допол"], "Удаление добавочных молочных желез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["молоч"])
	main_func(["уменьше", "редукц"], ["груд", "молочн", "маммопластик"], "Уменьшение (редукция) груди", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тканев"], ["экспандер"], "Установка тканевого экспандера", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндопротез"], ["груд", "молочн"], "Эндопротезирование груди", ["дет", "реб", "при эндопротез", "втор", "описани", "при", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пилинг"], ["броccаж", "бросаж", "броссаж"], "Пилинг-броссаж лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["срединн", "абр-пилинг", "пилинг абр", "abr пилинг", "пилинг abr"], ["пилинг"], "Срединный пилинг", ["дет", "реб", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["пилинг"], ["тела"], "Пилинг тела", ["дет", "реб", "допол", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["пилинг"], ["спин"], "Пилинг спины", ["дет", "реб", "втор", "описани", "допол", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["пилоропластик"], ["пилоропластик"], "Пилоропластика", ["дет", "микули", "гейнек", "фин", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["микулич", "гейнек"], ["микулич", "гейнек"], "Пилоропластика по Гейнеке Микуличу", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пилинг"], ["голов", "волос"], "Пилинг кожи головы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "во время", "курс"])
	main_func(["уфо крови", "ультрафиолет"], ["уфо крови", "кров"], "УФО крови / Ультрафиолетовое облучение", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["отопластика", "пластика ушной раковины", "устранение деф"], ["отопластика", "ушной раковины"], "Отопластика", ["дет", "лазер", "меато", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["восста", "рекон"], ["ушной раковин"], "Восстановление ушной раковины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик", "коррек", "уменьш", "увел", "восстанов"], ["мочки", "мочек", "мочка"], "Коррекция мочки уха", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер"], ["отопластик"], "Лазерная отопластика", ["дет", "реб", "втор", "меато", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["втор"], ["отопластик"], "Повторная отопластика", ["дет", "реб", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["умень", "реду"], ["ушной ракови", "ушных ракови"], "Уменьшение ушной раковины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["деформац"], ["ушной раковины", "ушных раковин"], "Устранение деформаций ушной раковины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лопоухости"], ["лопоухости"], "Устранение лопоухости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гемосорбци"], ["гемосорбци"], "Гемосорбция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["криоаферез"], ["криоаферез"], "Криоаферез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["плазмаферез", "плазмафарез", "плазмофарез", "плазмоферез", "плазмофорез", "плазмафорез", "плазмоферез"], ["плазмаферез", "плазмафарез", "плазмофарез", "плазмоферез", "плазмофорез", "плазмафорез", "плазмоферез"], "Плазмаферез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вскры", "дренаж", "дренир"], ["парапрокти"], "Вскрытие острого парапроктита", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["сеч", "удал"], ["свищ"], "Иссечение интрасфинктерного свища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["интрасфин"])
	main_func(["сеч", "удал"], ["свищ"], "Иссечение транссфинктерного свища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "интра", "экстра", "курс"], ["сфинктер"])
	main_func(["сеч", "удал"], ["свищ"], "Иссечение экстрасфинктерного свища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["экстрасфин", "параректал"])
	main_func(["гемитиреоидэктоми", "удален"], ["гемитиреоидэктоми", "образований щитовид", "образования щитовид"], "Гемитиреоидэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["узлов", "узла"], ["щитовидно"], "Деструкция узлов щитовидной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["дестр", "разруш", "удал"])
	main_func(["резекц"], ["щитовидной"], "Резекция щитовидной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тиреоидэктоми"], ["лимфодис", "лимфадис"], "Тиреоидэктомия с лимфодиссекцией", ["дет", "геми", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["аденом"], ["щитовидн"], "Удаление аденомы паращитовидной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["удал"])
	main_func(["удал"], ["паращ"], "Удаление паращитовидных желез", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["желез"])
	main_func(["удал"], ["щитовид"], "Удаление щитовидной железы", ["дет", "реб", "пара", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["энуклеац"], ["кист"], "Энуклеация кисты щитовидной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["щитовид"])
	main_func(["удал"], ["образо"], "Удаление образований сигмовидной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сигмовид", "толст"])
	main_func(["удал", "полипэктоми", "полипоктомия"], ["полип", "полипэктоми", "полипоктомия"], "Удаление полипов сигмовидной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сигмовид", "толст", "при колоноскопии", "в ходе колоноскопии", "при диагностической колоноскопии"])
	main_func(["пункц", "пунктирован"], ["сустав"], "Пункция суставов", ["дет", "реб", "втор", "описани", "контрол", "при", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["бахром"], ["бахром"], "Удаление анальных бахромок", ["дет", "реб", "втор", "узл", "гемор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["каловых", "калового", "копролит"], ["камней", "завал", "камня", "копролит"], "Удаление каловых камней", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "полипэктом", "электроэксци", "полипоктомия"], ["полип", "полипэктом", "полипоктомия"], "Удаление полипов прямой кишки", ["дет", "хоанал", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "церви", "церк", "влаг", "матк", "маточ", "во время", "курс"], ["прям", "анал", "толст", "при колоноскопии", "в ходе колоноскопии", "при диагностической колоноскопии", "жкт"])
	main_func(["удаление прямой кишки"], ["удаление прямой кишки"], "Удаление прямой кишки", ["дет", "реб", "образова", "полип", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вазовазостоми"], ["вазовазостоми"], "Вазовазостомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вазоэпидидимостоми"], ["вазоэпидидимостоми"], "Вазоэпидидимостомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гнои"], ["копчик"], "Вскрытие нагноившегося эпителиального копчикового хода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ход"])
	main_func(["анал", "анус", "заднего прохода"], ["трещин"], "Иссечение анальной трещины", ["дет", "реб", "леч", "терапи", "консерв", "без операц", "безопера", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["сеч", "удал", "тер", "опер", "леч", "фотокоагул"])
	main_func(["копчик", "пилонидальной кисты"], ["ход", "пилонидальной кисты"], "Иссечение копчиковых ходов", ["дет", "реб", "гно", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["свищ"], ["прям"], "Иссечение свища прямой кишки", ["дет", "шив", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["киш"])
	main_func(["вапоризац"], ["простат", "предста"], "Лазерная вапоризация простаты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["простатэктоми", "тур предстательной", "тур простаты"], ["простатэктоми", "тур предстательной", "тур простаты"], "Простатэктомия", ["дет", "реб", "цист", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["транусретра", "тур"], ["аденомы простаты", "аденомы предста"], "Трансуретральная резекция (ТУР) аденомы простаты", ["дет", "реб", "вапоризац", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["аденомы простаты", "аденомы предста"], "Удаление аденомы простаты", ["дет", "реб", "тур", "транусретра", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эмболи"], ["арте"], "Эмболизация артерий простаты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["простат", "предста"])
	main_func(["вскры", "дренаж", "дренир"], ["абсцес"], "Вскрытие абсцесса носовой перегородки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["носовой", "носа", "носу"])
	main_func(["вскры", "дренаж", "дренир"], ["фурунку"], "Вскрытие фурункула носа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["нос", "лор-органа", "лор-органов", "лор органа", "лор органов"])
	main_func(["денерв"], ["голов"], "Денервация головки полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["член"])
	main_func(["парауре"], ["кист"], "Иссечение парауретральной кисты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["искривле", "корпоропластика"], ["члена", "корпоропластика"], "Коррекция искривления полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лоскут"], ["корпоропластик"], "Лоскутная корпоропластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["меатотомия", "меатомия", "сечен"], ["меатотомия", "меатомия", "наружного отверстия уретры"], "Меатотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["несбит"], ["несбит"], "Операция Несбита", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["белочной оболочки"], "Пластика белочной оболочки полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["стриктур"], "Пластика стриктур уретры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["уретр"])
	main_func(["уздечк", "плот"], ["пластик", "операция"], "Пластика уздечки полового члена", ["дет", "реб", "втор", "язык", "губ", "тяжа", "слизист", "рот", "рта", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["член", "крайн"])
	main_func(["пластик"], ["уретры"], "Пластика уретры", ["дет", "хольцов", "стриктур", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хольцов"], ["хольцов"], "Пластика уретры по Хольцову", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пликац"], ["белочн"], "Пликация белочной оболочки полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["член"])
	main_func(["сеч", "устра"], ["фимоз", "щемл"], "Рассечение ущемляющего кольца при парафимозе", ["дет", "реб", "без расс", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["кольц"])
	main_func(["реконструк"], ["фаллопластик"], "Реконструктивная фаллопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стент"], ["уретры"], "Стентирование уретры", ["дет", "реб", "втор", "описани", "удал", "извле", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["транспози"], ["уретры"], "Транспозиция уретры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["лигаментотоми", "увелич", "увелеч", "удлинение полового члена"], ["лигаментотоми", "длины полового член", "удлинение полового члена"], "Увеличение длины полового члена / Лигаментотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["увелич", "утолщение полового члена"], ["толщины полового член", "утолщение полового члена"], "Увеличение толщины полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["уретрографи"], ["уретрографи"], "Уретрография", ["дет", "реб", "цисто", "восходящ", "ретроград", "пад", "спирал", "мскт", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["массаж"], ["реб", "дет"], "Детский массаж", ["втор", "описани", "расшифров", "абонемент", "занятий", "посещени", "под", "во время", "допол", "курс"], ["общ", "класс"])
	main_func(["сеч", "опер"], ["дермоид"], "Иссечение дермоидных кист", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["кист"])
	main_func(["кюретаж"], ["подмыше"], "Кюретаж подмышечных впадин", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал"], ["образован"], "Удаление новообразований ободочной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ободочн"])
	main_func(["удал", "полипэктом", "полипоктомия"], ["полип", "полипэктом", "полипоктомия"], "Удаление полипов ободочной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ободоч", "толст", "в ходе колоноскопии", "при колоноскопии", "при диагностической колоноскопии"])
	main_func(["гемисинусотоми"], ["гемисинусотоми"], "Гемисинусотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["инфундибулотоми"], ["инфундибулотоми"], "Инфундибулотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["полисинусотоми"], ["полисинусотоми"], "Полисинусотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сфеноидотоми"], ["сфеноидотоми"], "Сфеноидотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["образован"], "Удаление новообразования околоносовой пазухи", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["носовой пазухи", "носовых пазух", "пазух носа", "носовых пазух", "пазухи в носу"])
	main_func(["этмоидотоми"], ["этмоидотом"], "Этмоидотомия", ["дет", "реб", "втор", "полип", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гепатикоеюностомия"], ["гепатикоеюностомия"], "Гепатикоеюностомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дрен"], ["кист"], "Дренирование кисты печени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["печен"])
	main_func(["краев"], ["резекц"], "Краевая резекция печени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["печен"])
	main_func(["лево"], ["гемигепатэктоми"], "Левосторонняя гемигепатэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["прав"], ["гемигепатэктоми"], "Правосторонняя гемигепатэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пункц", "пунктирование"], ["печен"], "Пункция печени", ["дет", "реб", "втор", "описани", "контрол", "при", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["печен"], "Резекции печени", ["дет", "сегмен", "краев", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["дол"], "Удаление доли печени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["печен"])
	main_func(["удал", "сеч"], ["кист"], "Удаление кисты печени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["печен"])
	main_func(["эхинококкэктоми"], ["печен"], "Эхинококкэктомия из печени", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["баллон", "балон"], ["дилатац"], "Баллонная дилатация пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["пищевод"])
	main_func(["бужирован"], ["пищевод"], "Бужирование пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дивертикул"], ["пищевод"], "Иссечение дивертикула пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пластик"], ["пищевод"], "Пластика пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стриктур"], ["пищевод"], "Рассечение стриктуры пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["стент"], ["пищевод"], "Стентирование пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["доборока", "доброкач"], "Удаление доброкачественных опухолей пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["опухолей пищевода", "опухоли пищевод"])
	main_func(["инородн"], ["пищевод"], "Удаление инородного тела из пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "полипэктоми"], ["полип", "полипэктоми"], "Удаление полипа пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["пищевод"])
	main_func(["эзофагокардиомиотомия"], ["эзофагокардиомиотомия"], "Эзофагокардиомиотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндоскоп"], ["лигир"], "Эндоскопическое лигирование варикозных вен пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["пищевод"])
	main_func(["эндоскоп"], ["склероз"], "Эндоскопическое склерозирование варикозных вен пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["пищевод"])
	main_func(["облитерц", "облитерац"], ["плевр"], "Облитерация плевральной полости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["плеврэктоми"], ["декортац", "декортикац"], "Плеврэктомия с декортикацией легкого", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["плевр"], "Удаление плевры", ["дет", "реб", "втор", "инород", "тел", "жидкости", "образ", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["инородн"], ["член"], "Удаление инородных тел из полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["камн", "конкре"], "Удаление камней уретры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["уретр"])
	main_func(["карункул"], ["карункул"], "Удаление карункула уретры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "сеч"], ["олеогранулем"], "Удаление олеогранулемы полового члена", ["дет", "реб", "молочн", "груд", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "сеч"], ["полип"], "Удаление полипа уретры", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["уретр"])
	main_func(["уретротомия"], ["уретротомия"], "Уретротомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["френулотоми"], ["член", "край", "френулотомия"], "Френулотомия крайней плоти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["сфинктеропластик"], ["сфинктеропластик"], "Сфинктеропластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сфинктеротоми"], ["сфинктеротоми"], "Сфинктеротомия", ["дет", "реб", "папи", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["аденом"], "Удаление аденомы слюнной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["слюнн"])
	main_func(["удал"], ["камн"], "Удаление камня слюнной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["слюнн"])
	main_func(["удал"], ["слюнн"], "Удаление подчелюстной слюнной железы", ["дет", "камн", "кист", "ретен", "образован", "реб", "аден", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["желез"])
	main_func(["удал"], ["кист"], "Удаление ретенционной кисты слюнной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["слюнн"])
	main_func(["удал"], ["беккер", "бейкер", "бейккер"], "Удаление кисты Беккера", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["тимпанотоми"], ["тимпанотоми"], "Ревизионная тимпанотомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["санир"], ["среднего уха", "среднем ухе"], "Санирующая операция на среднем ухе", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["образован"], "Удаление доброкачественных новообразований ушной раковины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ушн", "слух", "ушей", "ушах", "уха"])
	main_func(["свищ"], ["околоуш"], "Удаление околоушного свища", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал"], ["остеом", "экзостоз"], "Удаление остеомы или экзостозов наружного слухового прохода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["слух", "уш", " уха"])
	main_func(["удал", "полипэктом", "полипотомия уха"], ["полип", "полипэктом", "полипотомия уха"], "Удаление полипов среднего уха", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], [" уха", "слухово", "ушей", "ушного"])
	main_func(["вскры", "дренир"], ["мастит"], "Вскрытие и дренирование мастита", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["урологи"], ["массаж"], "Урологический массаж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "резекц", "оператив"], ["ногт"], "Удаление вросшего ногтя", ["дет", "лазер", "бартлет", "радио", "сурги", "хирурги", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["врос"])
	main_func(["лазер"], ["вросшего ногтя", "вросших ногтей", "вросший ноготь"], "Лазерное удаление вросшего ногтя", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "лечен", "курс"])
	main_func(["бартлет"], ["бартлет"], "Операция Бартлетта", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["вросшего ногтя", "вросших ногтей", "вросший ноготь"], "Радиоволновое удаление вросшего ногтя", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["хирурги", "кра"], ["леч", "резек"], "Хирургическое лечение вросшего ногтя", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["вросшего ногтя", "вросших ногт", "ногт"])
	main_func(["хирурги"], ["удал", "резек"], "Хирургическое удаление вросшего ногтя", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "радио", "во время", "курс"], ["ногт", "вросшего ног", "вросших ног"])
	main_func(["кт ", "кт-", "кт -", "кт- ", "компьютерная томограф", "компьютерной томограф", "компьютерно-томографическ" ], ["ангиограф", "томограф"], "КТ-ангиография аорты", ["дет", "реб", "втор", "описани", "с контрастом", "с контрастированием", "расшифров", "мульти", "мскт", "спирал", "занятий", "посещени", "под", "для", "во время", "курс"], ["аорт"])
	main_func(["мрт", "мр-", "мр -", "мр ", "магнито-резонанс", "магниторезонанс", "магнитно-резонанс", "магнитнорезонанс", "магнито резонанс", "магнитно резонанс"], ["ангиограф", "томограф"], "МР-ангиография аорты", ["дет", "спирал", "реб", "мульти", "мскт" "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["аорт"])
	main_func(["мскт", "мультиспиральная томография", "мультисрезовая томография", "спиральная компьютерная томографи", "спиральной компьютерной томограф", "спиральной компьютерной томограф", "мультисрезовая компьютерная томографи", "мультиспиральная кт", "мультисрезовая кт", "мультидетекторная кт", "мультидетекторная томограф", "мультидетекторная компьютерная томо"], ["м"], "МСКТ-ангиография аорты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["аорт"])
	main_func(["кт ", "кт-", "кт -", "кт- ", "компьютерная томограф", "компьютерной томограф", "компьютерно-томографическ"], ["ангиограф", "томограф"], "КТ-ангиография сосудов головного мозга", ["дет", "мскт", "мульти", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сосудов мозга", "сосудов головного мозга", "вен и артерий мозга", "вен и артерий головного мозга", "артерий и вен головного мозга", "артерий и вен мозга"])
	main_func(["мрт", "мр-", "мр -", "мр ", "магнито-резонанс", "магниторезонанс", "магнитно-резонанс", "магнитнорезонанс", "магнито резонанс", "магнитно резонанс"], ["ангиограф"], "МР-ангиография сосудов головного мозга", ["дет", "реб", "наличии", "нашем", "проведенного", "втор", "описани", "спирал", "мскт", "расшифров", "занятий", "посещени", "спин", "под", "для", "кост", "во время", "курс"], ["сосудов мозга", "сосудов головного мозга", "вен и артерий мозга", "вен и артерий головного мозга", "артерий и вен головного мозга", "артерий и вен мозга", "головного мозга", "мозга", "артерий головного"])
	main_func(["мрт", "мр-", "мр -", "мр ", "магнито-резонанс", "магниторезонанс", "магнитно-резонанс", "магнитнорезонанс", "магнито резонанс", "магнитно резонанс"], ["вен ", "вены"], "МР-венография головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "спирал", "мскт", "посещени", "под", "для", "спин", "во время", "курс"], ["мозг"])
	main_func(["мскт", "мультиспиральная томография", "мультисрезовая томография", "спиральная компьютерная томографи", "спиральной компьютерной томограф", "мультисрезовая компьютерная томографи", "мультиспиральная кт", "мультисрезовая кт", "мультидетекторная кт", "мультидетекторная томограф", "мультидетекторная компьютерная томо"], ["м"], "МСКТ-ангиография сосудов головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сосудов мозга", "сосудов головного мозга", "вен и артерий мозга", "вен и артерий головного мозга", "артерий и вен головного мозга", "артерий и вен мозга", "артерий мозга", "артерий головного"])
	main_func(["кт ", "кт-", "кт -", "кт- ", "компьютерная томограф", "компьютерной томограф", "компьютерно-томографическ"], ["ангиограф"], "КТ-ангиография сосудов конечностей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс", "мскт", "мульти"], ["сосудов нижних конечностей", "сосудов нижних конечностей", "сосудов верхних конечностей", "артерий верхних конечностей", "сосудов конечностей", "вен нижних конечностей", "вен верхних конечностей", "артерий нижних конечностей", "артерий конечност"])
	main_func(["мрт", "мр-", "мр -", "мр ", "магнито-резонанс", "магниторезонанс", "магнитно-резонанс", "магнитнорезонанс", "магнито резонанс", "магнитно резонанс"], ["ангиограф"], "МР-ангиография сосудов конечностей", ["дет", "реб", "спирал", "мскт", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сосудов нижних конечностей", "сосудов конечностей", "сосудов нижних конечностей", "сосудов верхних конечностей", "артерий верхних конечностей", "вен нижних конечностей", "вен верхних конечностей", "артерий нижних конечностей"])
	main_func(["мскт", "мультиспиральная томография", "мультисрезовая томография", "спиральная компьютерная томографи", "спиральной компьютерной томограф", "мультисрезовая компьютерная томографи", "мультиспиральная кт", "мультисрезовая кт", "мультидетекторная кт", "мультидетекторная томограф", "мультидетекторная компьютерная томо"], ["м"], "МСКТ-ангиография сосудов конечностей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сосудов нижних конечностей", "сосудов конечностей", "сосудов нижних конечностей", "сосудов верхних конечностей", "артерий верхних конечностей", "вен нижних конечностей", "вен верхних конечностей", "артерий нижних конечностей", "нижних конечностей", "артерий конечност"])
	main_func(["кт ", "кт-", "кт -", "кт- ", "компьютерная томограф", "компьютерной томограф", "компьютерно-томографическ"], ["ангиограф"], "КТ-ангиография сосудов шеи", ["дет", "реб", "мскт", "мульти", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во варемя", "курс"], ["шеи", "шейн"])
	main_func(["мрт", "мр-", "мр -", "мр ", "магнито-резонанс", "магниторезонанс", "магнитно-резонанс", "магнитнорезонанс", "магнито резонанс", "магнитно резонанс"], ["ангиограф", "сосудов ше", "артерий ше"], "МР-ангиография сосудов шеи", ["дет", "спирал", "мскт", "наличии", "нашем", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["шеи", "шейн"])
	main_func(["мскт", "мультиспиральная томография", "мультисрезовая томография", "спиральная компьютерная томографи", "спиральной компьютерной томограф", "мультисрезовая компьютерная томографи", "мультиспиральная кт", "мультисрезовая кт", "мультидетекторная кт", "мультидетекторная томограф", "мультидетекторная компьютерная томо"], ["м"], "МСКТ-ангиография сосудов шеи", ["дет", "реб", "втор", "мяг", "описани", "расшифров", "занятий", "посещени", "под", "для", "позвон", "во время", "курс"], ["шеи", "шейн"])
	main_func(["кт ", "кт-", "кт -", "кт- ", "компьютерная томограф", "компьютерной томограф", "компьютерно-томографическ"], ["ангиограф"], "КТ-ангиография сосудов сердца", ["дет", "реб", "втор", "описани", "спирал", "мскт", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сердц"])
	main_func(["мскт", "мультиспиральная томография", "мультисрезовая томография", "спиральная компьютерная томографи", "спиральной компьютерной томограф", "мультисрезовая компьютерная томографи", "мультиспиральн", "мультисрезов", "мультидетекторн"], ["м"], "МСКТ-ангиография сосудов сердца", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сердц"])
	main_func(["узи", "ультразвук"], ["плеч"], "УЗИ плечевого сплетения", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сплет"])
	main_func(["экг", "электрокардиографи", "электрокардиограм"], ["экг", "электрокардиографи", "электрокардиограмм"], "ЭКГ", ["дет", "акц", "перед", "дом", "запо", "выезд", "после", "проф", "дополн", "неб", "сут", "монитор", "без экг", "ад", "давлен", "холтер", "нагруз", "стресс", "пищевод", "реб", "втор", "занятий", "расшифровка", "с расшифровкой", "описание", "посещени", "под", "для", "во время", "курс"])
	main_func(["ритмокардиографи"], ["ритмокардиографи"], "Ритмокардиография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сут", "смад", "cmaд", "мониторирование", "ад-мониторирование"], ["смад", "cmaд", "артериального давления", "мониторирование ад", "мониторирования ад", "ад-мониторирование"], "СМАД / Суточное мониторирование АД", ["дет", "дом", "заде", "интерпретация данн", "интерпретация резуль", "выезд", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сут", "мониторировани"], ["экг", "холтер", "электрокардиографи", "сердечного ритм", "ритма сердц"], "Суточное мониторирование ЭКГ + АД", ["дет", "реб", "втор", "установка", "описани", "расшифровка", "заде", "занятий", "посещени", "под", "для", "во время", "курс"], ["ад", "давлени"])
	main_func(["сфигмография"], ["сфигмография"], "Сфигмография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["фонокардиограф"], ["фонокардиограф"], "Фонокардиография", ["дет", "плод", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сут", "холтеру", "холтеров", "экг- мониторирование", "экг-мониторирование"], ["экг", "электрокардиография", "холтер", "экг- мониторирование", "экг-мониторирование"], "Холтер / Суточное мониторирование ЭКГ", ["дет", "до 3-х ч", "до 3 час", "до 6 час", "до 16 час", "постан", "оценка результатов", "устан", "ад", "заде", "давле", "интерпретация данн", "интерпретация резуль", "реб", "втор", "описани", "расшифровка", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["холтер", "сут"], ["экг", "холтер"], "Холтер ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "заде", "под", "для", "во время", "курс"], ["реб", "детё"])
	main_func(["пищевод"], ["экг", "электрокардиографи"], "Чреспищеводная электрокардиография", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экг", "электрокардиографи", "электрокардиограмм"], ["дом"], "ЭКГ на дому", ["дет", "реб", "сестр", "инъек", "кров", "втор", "описани", "кардиодом", "расшифровка", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экг", "электрокардиографи", "электрокардиограмм"], ["неб"], "ЭКГ по Небу", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экг", "электрокардиографи", "электрокардиограмм"], ["реб", "дет"], "ЭКГ ребенку", ["втор", "описани", "запо", "выведе", "нарко", "расшифровка", "детокси", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["экг", "электрокардиографи", "электрокардиограмм"], ["нагруз", "упражнениями"], "ЭКГ с нагрузкой", ["дет", "медикаментозной нагрузкой", "реб", "втор", "без нагрузки", "без нагрузок", "описание ", "расшифровка", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["бронхоальвео", "бронхо-альево"], ["лаваж"], "Бронхоальвеолярный лаваж", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["гемиколэктомия"], ["гемиколэктомия"], "Гемиколэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["илеоректальн"], ["анастомоз"], "Наложение илеоректального анастомоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["коло"], ["колоанастомоз"], "Наложение коло-колоанастомоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["колоректально"], ["анастомоз"], "Наложение колоректального анастомоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["трансверзосигмоанастомоз"], ["трансверзосигмоанастомоз"], "Наложение трансверзосигмоанастомоза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["попер"], "Резекция поперечно-ободочной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["ободоч"])
	main_func(["тотал"], ["колэктоми"], "Тотальная колэктомия", ["дет", "реб", "втор", "суб", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["вирсунгодуоденостомия"], ["вирсунгодуоденостомия"], "Вирсунгодуоденостомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["дистал"], ["резекц"], "Дистальная резекция поджелудочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["поджелудоч"])
	main_func(["марс"], ["кист"], "Марсупиализация кисты поджелудочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["поджел"])
	main_func(["некрэктоми"], ["поджелудоч"], "Некрэктомия поджелудочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["опер"], ["фрея"], "Операция Фрея", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["панкреатодуоденальн"], ["резекц"], "Панкреатодуоденальная резекция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["голов"], "Резекция головки поджелудочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["поджел"])
	main_func(["стент"], ["панкреат"], "Стентирование главного панкреатического протока", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["проток"])
	main_func(["трансдуоденальная"], ["папиллэктом"], "Трансдуоденальная папиллэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["цистогастростоми"], ["цистогастростоми"], "Цистогастростомия при кисте поджелудочной железы", ["дет", "холе", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["цистодуоденостоми"], ["цистодуоденостоми"], "Цистодуоденостомия при кисте поджелудочной железы", ["дет", "холе", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["цистоэнтеростоми"], ["цистоэнтеростоми"], "Цистоэнтеростомия при кисте поджелудочной железы", ["дет", "реб", "холе", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["эндоскоп"], ["папиллосфинктеротоми"], "Эндоскопическая папиллосфинктеротомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["эндоскоп"], ["кист"], "Эндоскопическое дренирование кисты поджелудочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["поджелудо"])
	main_func(["энуклеац"], ["опухол"], "Энуклеация опухоли поджелудочной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["поджелуд"])
	main_func(["абсцес"], ["почк", "почеч"], "Вскрытие абсцесса почки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["декапсул"], ["почк"], "Декапсуляция почки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["нефропексия"], ["нефропексия"], "Нефропексия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["нефростомия"], ["нефростомия"], "Нефростомия", ["дет", "рен", "чрезко", "черезкож", "чреско", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["нефротомия"], ["нефротомия"], "Нефротомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["пласти"], ["лохан"], "Пластика лоханочно-мочеточникового сегмента", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["мочеточн"])
	main_func(["резекци", "резакция"], ["почки"], "Резекция почки", ["дет", "реб", "втор", "описани", "образовани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["трансплантац"], ["почк"], "Трансплантация почки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["удал", "замена"], ["нефростомического дренажа", "нефростомы"], "Удаление нефростомического дренажа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "нефрэктомия"], ["почки", "нефрэктомия"], "Удаление почки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["чрескож", "чрезкож"], ["нефростоми"], "Чрескожная пункционная нефростомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["сече"], ["фуникулоцел"], "Иссечение фуникулоцеле", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["орхифуникулэктомия", "орхофуникулоэктоми"], ["орхифуникулэктомия", "орхофуникулоэктоми"], "Орхифуникулэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ликвид", "устран"], ["киш"], "Ликвидация кишечной непроходимости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["непроход"])
	main_func(["проктопластик"], ["проктопластик"], "Проктопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["спаек"], ["кишеч"], "Рассечение спаек при кишечной непроходимости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["проход"])
	main_func(["резекц"], ["кишк"], "Резекция кишки при кишечной непроходимости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["роходимост"])
	main_func(["резекц"], ["прямой кишки"], "Резекция прямой кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["ушиван"], ["свищ"], "Ушивание свища прямой кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["прямо"])
	main_func(["бужирован"], ["слюн"], "Бужирование протока слюнной железы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["паротидэктоми"], ["паротидэктоми"], "Паротидэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["закрыти"], ["колостом"], "Закрытие колостомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["сеч"], ["опухол"], "Иссечение опухоли сигмовидной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"], ["сигмовид"])
	main_func(["колостоми"], ["колостоми"], "Колостомия", ["дет", "реб", "втор", "описани", "расшифров", "энтеро", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["сигмовид"], "Резекция сигмовидной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["рекон", "восстано"], ["колостомы"], "Реконструкция колостомы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["илеостомия"], ["илеостомия"], "Илеостомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "под", "для", "во время", "курс"])
	main_func(["резекц"], ["тонкой"], "Резекция тонкой кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "сегмент", "во время", "курс", "лапароскоп"], ["кишк"])
	main_func(["сегмен"], ["резекц", "иссечение поврежденной тонкой кишки", "иссечение тонкой кишки"], "Сегментарная резекция тонкой кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["тонк"])
	main_func(["инород"], ["тонк"], "Удаление инородного тела тонкого кишечника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["кише", "кишки"])
	main_func(["энтеростом", "энтероэнтеростомия"], ["энтеростом", "энтероэнтеростомия"], "Энтеростомия", ["дет", "гастро", "цисто", "ушиван", "холе", "энтероэнтеростоми", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["делорма"], ["делорма"], "Операция Делорма", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["имплант", "установк"], ["иол", "интраокулярных линз", "интраокулярной линзы"], "Имплантация ИОЛ при афакии", ["дет", "реб", "доп", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["интракапсуляр"], ["экстракц"], "Интракапсулярная экстракция катаракты / ИЭК", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["катаракт"])
	main_func(["лазер"], ["дисциз"], "Лазерная дисцизия задней капсулы хрусталика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["хрусталик"])
	main_func(["лазер"], ["леч"], "Лазерное лечение катаракты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["катарак"])
	main_func(["втор", "реимплант"], ["имплант"], "Повторная имплантация ИОЛ", ["дет", "реб", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["иол", "интраокуля"])
	main_func(["факоаспирац"], ["катарк", "катарак"], "Факоаспирация катаракты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["факоэмульсификаци"], ["катаракт"], "Факоэмульсификация катаракты", ["дет", "глаукома", "реб", "втор", "описани", "после", "предоплат", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["факоэмульсификаци"], ["антиглаукоматозн"], "Факоэмульсификация с антиглаукоматозной операцией", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["экстракапсул"], ["экстракц"], "Экстракапсулярная экстракция катаракты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["катаракт"])
	main_func(["вентрофиксац"], ["матки"], "Вентрофиксация матки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лапароскоп", "лапараскоп"], ["промонтофиксац"], "Лапароскопическая промонтофиксация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["леваторопластик"], ["леваторопластик"], "Леваторопластика", ["дет", "реб", "кольпо", "сфинк", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["регионар"], ["тромболизис"], "Регионарный тромболизис при венозном тромбозе", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["тромбэктоми"], ["полой"], "Тромбэктомия из нижней полой и подвздошных вен", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["подвз"])
	main_func(["тромбэктоми"], ["вен"], "Тромбэктомия при глубоком венозном тромбозе", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["тромбоз"])
	main_func(["удал"], ["кава"], "Удаление кава-фильтра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["фильтр"])
	main_func(["невр"], ["мортон"], "Иссечение невромы Мортона", ["дет", "реб", "блокад", "узи", "ультразвук", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["секвестрэктоми"], ["остеомиелит"], "Секвестрэктомия при остеомиелите", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал"], ["кост", "хрящ"], "Удаление костно-хрящевых экзостозов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["экзостоз"])
	main_func(["лапараскоп", "лапароскоп"], ["сальпингостоми"], "Лапароскопическая сальпингостомия", ["дет", "реб", "хром", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лапараскоп", "лапароскоп"], ["сальпингоовариолизис"], "Лапароскопический сальпингоовариолизис", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["остеосинтез"], ["бугор"], "Остеосинтез большого бугорка плечевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["плеч"])
	main_func(["остеосинтез"], ["вертел"], "Остеосинтез вертельных переломов бедра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "тазо", "курс"], ["бедр", "бедер"])
	main_func(["остеосинтез"], ["верх"], "Остеосинтез верхней челюсти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["челюст"])
	main_func(["остеосинтез"], ["внутр", "медиальн"], "Остеосинтез внутренней лодыжки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лодыж"])
	main_func(["остеосинтез"], ["голов"], "Остеосинтез головки плечевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["плеч"])
	main_func(["остеосинтез"], ["грудины"], "Остеосинтез грудины", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["остеосинтез"], ["диафиз"], "Остеосинтез диафиза большеберцовой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["берц"])
	main_func(["остеосинтез"], ["диафиз"], "Остеосинтез диафиза костей предплечья", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["предплеч"])
	main_func(["остеосинтез"], ["диафиз"], "Остеосинтез диафиза локтевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["локт"])
	main_func(["остеосинтез"], ["диафиз"], "Остеосинтез диафиза лучевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["луч"])
	main_func(["остеосинтез"], ["диафиз"], "Остеосинтез диафиза плечевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс", "пред"], ["плеч"])
	main_func(["остеосинтез"], ["диафиз"], "Остеосинтез диафизарных переломов бедра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["бедр", "бедер"])
	main_func(["остеосинтез"], ["дисталь"], "Остеосинтез дистального метаэпифиза лучевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лучево"])
	main_func(["остеосинтез"], ["зад"], "Остеосинтез заднего края большеберцовой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["берц"])
	main_func(["остеосинтез"], ["ключиц"], "Остеосинтез ключицы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["остеосинтез"], ["ладьевидн", "ладъевид"], "Остеосинтез ладьевидной кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["остеосинтез"], ["локт"], "Остеосинтез локтевого отростка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["отрост"])
	main_func(["остеосинтез"], ["мыщел"], "Остеосинтез мыщелков большеберцовой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["берц"])
	main_func(["остеосинтез"], ["надколен"], "Остеосинтез надколенника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["остеосинтез"], ["наруж", "латераль"], "Остеосинтез наружной лодыжки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["лодыж"])
	main_func(["остеосинтез"], ["нижн"], "Остеосинтез нижней челюсти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["челюст"])
	main_func(["остеосинтез"], ["шейк"], "Остеосинтез шейки бедра", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["бедр"])
	main_func(["остеосинтез"], ["шейк"], "Остеосинтез шейки плечевой кости", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["плеч"])
	main_func(["удал"], ["металлоконструкц"], "Удаление металлоконструкций после остеосинтеза", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["мини"], ["липосакц"], "Минилипосакция щёк", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["щек", "щеч"])
	main_func(["протез"], ["скул"], "Протезирование скуловых дуг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["промыва"], [" нос"], "Промывание носа", ["дет", "кукушк", "в рамках", "прием", "консул", "реб", "соуст", "аттик", "проетц", "проэтц", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["краниаль", "краниосакральн"], ["мануал", "терапи"], "Краниальная мануальная терапия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["терап"])
	main_func(["мануал"], ["терапи"], "Мануальная терапия суставов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["сустав"])
	main_func(["миофасциа"], ["релиз", "релакс", "рилиз"], "Миофасциальный релизинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["остеопат"], ["коррекц"], "Остеопатическая коррекция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["постизо", "пост изо", "пост-изо"], ["релаксац"], "Постизометрическая релаксация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["кинезиотерапи", "кинезитерапия"], ["кинезиотерапи", "кинезитерапия"], "Сеанс кинезиотерапии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["сеанс", "мин", "комплексая мануальная терапия", "комплексная мануальная терапия", "мануальная терапия при "], ["мануальн"], "Сеанс мануальной терапии", ["дет", "реб", "сустав", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["терап"])
	main_func(["сеанс", "мин"], ["остеопат"], "Сеанс остеопатии", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["интравитреал", "стекловид"], ["инъекц"], "Интравитреальная инъекция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["амниоредукц"], ["амниоредукц"], "Амниоредукция", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["стресс"], ["эхокг", "эхокардиограф", "эхо кг", "эхо-кг"], "Стресс-эхокардиография / Стресс ЭхоКГ", ["дет", "реб", "электр", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["инсеминаци"], ["инсеминаци"], "Искусственная инсеминация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пластик"], ["вросш", "ногтевого ложа"], "Пластика вросшего ногтя", ["дет", "реб", "втор", "без пласти", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["b-s", "b/s", "b s", "bs-", "b / s"], ["b-s", "b/s", "b s", "b / s", "пластин"], "Исправление вросшего ногтя B/S-пластиной", ["дет", "сняти", "кобалт", "микро", "лемен", "реб", "микро", "мыш", "токси", "исс", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["ногт"])
	main_func(["скоб", "пласти"], ["зто", "три то"], "Исправление вросшего ногтя скобой ЗТО", ["дет", "реб", "втор", "описани", "кор", "расшифров", "занятий", "сняти", "посещени", "для", "во время", "курс"])
	main_func(["ногт", "пласт", "скоб"], ["фрезер"], "Исправление вросшего ногтя скобой Фрезера", ["дет", "реб", "втор", "описани", "сняти", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["уздечк"], ["губ"], "Пластика уздечки губы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "при пластике", "посещени", "для", "во время", "курс"])
	main_func(["уздечк"], ["язык"], "Пластика уздечки языка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "деструк", "прижига"], ["кондилом", "кандилом"], "Удаление кондилом вульвы", ["дет", "реб", "втор", "описани", "след", "свыше", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["вульв", "наружных полов"])
	main_func(["удал", "деструк", "прижига"], ["папиллом", "папилом"], "Удаление папиллом вульвы", ["дет", "реб", "втор", "описани", "след", "свыше", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["вульв", "наружных полов"])
	main_func(["лазер"], ["образован"], "Лазерное удаление новообразований вульвы", ["дет", "реб", "втор", "описани", "след", "свыше", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["вульв"])
	main_func(["удал", "деструк"], ["кондилом", "кандилом"], "Удаление кондилом во влагалище", ["дет", "реб", "втор", "след", "свыше", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["влагал"])
	main_func(["моллюск", "молюск"], ["моллюск", "молюск"], "Удаление контагиозного моллюска", ["дет", "крио", "от 20", "от 40", "боле", "свыше", "след", "до 5 элементов", "11-20", "6-10", "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "от 15", "от 10", "механич", "в рамках", "азот", "лаз", "радио", "сурги", "хирурги", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["крио", "азот"], ["моллюск", "молюск"], "Криодеструкция контагиозного моллюска жидким азотом", ["дет", "от 20", "от 40", "боле", "свыше", "след", "до 5 элементов", "11-20", "6-10",  "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "от 15", "от 10", "реб", "в рамках", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лаз"], ["моллюск", "молюск"], "Лазерное удаление контагиозного моллюска", ["дет", "реб", "втор", "от 20", "от 40", "боле", "до 5 элементов", "след", "свыше", "11-20", "6-10",  "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "от 15", "от 10", "описани", "расшифров", "в рамках", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["моллюск", "молюск"], "Радиоволновое удаление контагиозного моллюска", ["дет", "от 20", "от 40", "боле", "до 5 элементов", "реб", "след", "свыше", "11-20", "6-10",  "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "от 15", "от 10", "втор", "описани", "расшифров", "в рамках", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["хирургич", "механич"], ["моллюск", "молюск"], "Хирургическое удаление контагиозного моллюска", ["дет", "от 20", "от 40", "боле", "более", "до 5 элементов", "след", "свыше", "11-", "6-",  "от 5", "от 3", "от 6", "от 7", "от 8", "от 9", "от 15", "от 10", "до 5 элементов", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "в рамках", "под", "для", "во время", "курс"])
	main_func(["липофилинг"], ["липофилинг"], "Липофилинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс", "бедер", "бедра", "бедро", "век", "груд", "деколь", "губ", "лба", "лоб", "лиц", "ягоди", "тело", "тела", "щек", "щеч"])
	main_func(["гемиглоссэктоми", "гемиглосэктом"], ["гемиглоссэктоми", "гемиглосэктом"], "Гемиглоссэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["глосэктом", "глоссэктом"], ["глосэктом", "глоссэктом"], "Глоссэктомия", ["дет", "геми", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["расщелины"], ["неба"], "Коррекция расщелины нёба", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["резекц"], ["неба"], "Резекция мягкого неба", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["ринохейлопластика"], ["ринохейлопластика"], "Ринохейлопластика", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "сеч", "опера", "коагул"], ["язы"], "Удаление подъязычной кисты", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["кист"])
	main_func(["резекци"], ["ниж"], "Резекция нижней челюсти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс", "образова"], ["челюст"])
	main_func(["секвестрэктоми", "секвестерэктомия"], ["секвестрэктоми", "секвестерэктомия"], "Секвестрэктомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["инородн"], ["зуб", "корнево", "канала"], "Удаление инородного тела из зубного канала", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "сеч", "опера"], ["кист"], "Удаление кист челюстей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["челюст"])
	main_func(["шинирован"], ["челюсти"], "Шинирование челюсти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["контур"], ["пластик"], "Контурная пластика", ["дет", "интим", "избыточ", "нит", "полов", "гиалу", "кислот", "лба", "лоб", "лиц", "нос", "губ", "слез", "скул", "ше", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал", "механич", "сечен"], ["родин", "невус"], "Удаление родинки / невуса", ["дет", "6", "7", "8", "9", "от 3", "от 4", "от 5", "лазер", "механич", "электр", "радио", "сурги", "крио", "азот", "хирургич", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер"], ["родин", "невус"], "Лазерное удаление родинки / невуса", ["дет", "реб", "7-15", "втор", "6", "7", "8", "от 3", "от 4", "от 5", "9", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["родин", "невус"], "Радиоволновое удаление родинки / невуса", ["дет", "7-15", "реб", "6", "7", "от 3", "от 4", "от 5", "8", "9", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["крио", "азот"], ["родин", "невус"], "Удаление родинки / невуса жидким азотом", ["дет", "реб", "7-15", "6", "7", "8", "от 3", "от 4", "от 5", "9", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["хирургич", "механич"], ["невус", "родин"], "Хирургичекое удаление родинки / невуса", ["дет", "реб", "7-15", "6", "7", "от 3", "от 4", "от 5", "8", "9", "втор", "описани", "электр", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["электр"], ["родин", "невус"], "Электрокоагуляция невуса", ["дет", "реб", "втор", "описани", "расшифров", "и более", "от 1", "7-15", "6", "от 3", "от 4", "от 5", "7", "8", "9", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["парафин"], ["терапи"], "Парафинотерапия", ["дет", "маск", "допол", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["парафин"], ["лиц"], "Парафинотерапия лица", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["парафин"], ["рук"], "Парафинотерапия рук", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пэт"], ["тела", "тело"], "ПЭТ всего тела", ["дет", "кт", "томограф", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пэт"], ["мозга"], "ПЭТ головного мозга", ["дет", "реб", "втор", "кт", "томограф", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пэт"], ["кт"], "ПЭТ-КТ", ["дет", "реб", "тела", "тело", "мозг", "миокард", "серд", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пэт"], ["тела", "тело"], "ПЭТ-КТ всего тела", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пэт"], ["мозг"], "ПЭТ-КТ головного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["пэт"], ["сердц"], "ПЭТ-КТ сердца", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["зимницк"], ["зимницк"], "Проба Зимницкого", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["реберг"], ["реберг"], "Проба Реберга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["сулкович"], ["сулкович"], "Проба Сулковича", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["диаскинтест", "диаскин-тест", "диаскин тест"], ["диаскинтест", "диаскин тест", "диаскин-тест"], "Диаскинтест", ["дет", "реб", "втор", "описани", "расшифров", "после", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["хеликобактер", "helicobacter"], ["фгдс", "гастроскопи", "желудка"], "Анализ на хеликобактер при ФГДС", ["дет", "реб", "втор", "описани", "без", "расшифров", "после", "занятий", "посещени", "для", "курс"])
	main_func(["удал", "сечен", "коагул"], ["кондилом"], "Удаление кондилом в заднем проходе", ["дет", "реб", "втор", "описани", "расшифров", "более", "свыше", "от ",  "после", "занятий", "посещени", "для", "курс"], ["анал", "зад", "анус", "аногенитал", "прямой киш"])
	main_func(["репозици"], ["костей"], "Репозиции костей", ["дет", "реб", "втор", "лодыж", "носа", "носовых", " нос", "носу", "луч", "плеч", "локт", "описани", "ключ", "зап", "плюс", "пяст", "расшифров", "занятий", "посещени", "для", "курс"])
	main_func(["папил"], ["анус", "анал", "задне", "аногенитал"], "Иссечение папиллом ануса", ["дет", "реб", "втор", "более", "свыше", "от ", "описани", "расшифров", "занятий", "посещени", "для", "курс"], ["сеч", "удал", "коагул"])
	main_func(["папил"], ["глот", "гортан", "лор орган", "лор-орган"], "Удаление папиллом глотки / гортани", ["дет", "реб", "втор", "более", "свыше", "от ", "описани", "расшифров", "занятий", "посещени", "для", "курс"], ["сеч", "удал", "коагул"])
	main_func(["образован"], ["влагалищ"], "Удаление новообразований во влагалище", ["дет", "реб", "втор", "более", "свыше", "от ", "кондилом", "папиллом", "лазер", "кист", "описани", "расшифров", "занятий", "посещени", "для", "курс"], ["сеч", "удал", "коагул"])
	main_func(["хромоскопия", "хромоскопии"], ["хромоскопия", "хромоскопии"], "Хромоскопия", ["дет", "реб", "офталь", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["хромоцистоскопия", "хромоцистоскопии"], ["хромоцистоскопия", "хромоцистоскопии"], "Хромоцистоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["цервикоскоп"], ["цервикоскоп"], "Цервикоскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эзофагоскоп", "фэс"], ["эзофагоскоп", "фэс"], "Эзофагоскопия / ФЭС", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эндосонография"], ["12-перстн", "двенадцатипер", "12 перстной", "перстно"], "Эндосонография 12-перстной кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эндосонография"], ["желудк"], "Эндосонография желудка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эндосонография"], ["желчных путей"], "Эндосонография желчных путей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эндосонография"], ["пищевод"], "Эндосонография пищевода", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["энтероскопия"], ["энтероскопия"], "Энтероскопия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["рхпг", "ретроградная холангиопанкреатография"], ["эрхпг", "рхпг", "ретроградная холангиопанкреатография"], "ЭРХПГ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["бронхоскопия", "бронхоскопии"], ["реб", "дет"], "Бронхоскопия ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["вагиноскопия", "вагиноскопии"], ["реб", "дет"], "Вагиноскопия ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["вульвоскопия", "вульвоскопии"], ["реб", "дет"], "Вульвоскопия ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["колоноскопия", "колоноскопии", "фкс"], ["реб", "дет"], "Колоноскопия ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["ларингоскопия"], ["реб", "дет"], "Ларингоскопия ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["риноскопия"], ["реб", "дет"], "Риноскопия ребенку", ["втор", "описани", "расшифров", "занятий", "посещени", "после", "санация"])
	main_func(["фгдс", "гастроскопия"], ["реб", "дет"], "ФГДС ребенку", ["эгдс", "эзо", "втор", "описани", "расшифров", "полип", "в ходе", "удал", "образ", "акци", "занятий", "посещени"])
	main_func(["цистоскопия"], ["реб", "дет"], "Цистоскопия ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эгдс", "эзофагогастродуоденоскоп"], ["реб", "дет"], "ЭГДС ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["энцефалография"], ["энцефалография"], "Энцефалография", ["дет", "рео", "рэо", "электр", "ээг", "реб", "с функц", "с проб", "эхо", "втор", "описани", "расшифров", "занятий", "для", "посещени"])
	main_func(["энцефалография"], ["проб", "функц"], "ЭЭГ с функциональными пробами", ["дет", "рео", "рэо", "электр", "ээг", "реб", "эхо", "втор", "описани", "расшифров", "занятий", "для", "посещени"])
	main_func(["ээг", "электроэнцефалограф", "электроэнцефалограмм"], ["видео", "мониторинг"], "Видео-ЭЭГ мониторинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["рэг", "реоэнцефалография", "реоэнцефалограмм", "рэоэнцефалография"], ["рэг", "рэоэнцефалография", "реоэнцефалография", "реоэнцефалограмм"], "Реоэнцефалография / РЭГ", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эхоэнцефалография", "эхоэг", "эхоэнцефалограмм", "эхо-эг"], ["эхоэнцефалография", "эхоэнцефалограмм", "эхоэг", "эхо-эг"], "Эхоэнцефалография / Эхо-ЭГ", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["эхоэнцефалография", "эхоэг", "эхо-эг", "эхоэнцефалограмм"], ["реб", "дет"], "Эхоэнцефалография / Эхо-ЭГ ребенку", ["втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["ээг", "электроэнцефалографи", "электроэнцефалограмм"], ["ээг", "электроэнцефалографи", "электроэнцефалограмм"], "ЭЭГ", ["дет", "видео", "для", "комис", "води", "монитори", "реб", "втор", "описани", "расшифровка", "занятий", "посещени", "без ээг"])
	main_func(["3д", "3 д", "3-д", "3d", "3-d", "3 d"], ["волюмизаци"], "3D-волюмизация", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени"])
	main_func(["lpg"], ["массаж"], "LPG-массаж", ["дет", "реб", "костюм", "лиц", "тел", "допол", "втор", "описани", "расшифров", "занятий", "посещени", "абонемен"])
	main_func(["lpg"], ["лиц"], "LPG-массаж лица", ["дет", "реб", "втор", "костюм для", "допол", "описани", "расшифров", "занятий", "посещени", "абонемент"])
	main_func(["lpg"], ["тел"], "LPG-массаж тела", ["дет", "реб", "втор", "описани", "допол", "костюм для", "расшифров", "занятий", "посещени", "абонемент"])
	main_func(["lpg"], ["костюм"], "Костюм для LPG-массажа", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "без костю", "с костюм", "включая костюм", "не вклю", "без стоимости кост"])
	main_func(["удал", "извлеч"], ["клещ"], "Удаление клеща", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "курс"])
	main_func(["дуплекс", "уздс", "дс ", "уздг", "допплер", "доплер", "ультразвуковое исследование сосудов полового члена", "узи сосудов полового члена"], ["член", "ультразвуковое исследование сосудов полового члена", "узи сосудов полового члена"], "УЗДГ сосудов полового члена", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "курс"])
	main_func(["дуплекс", "уздс", "дс ", "уздг", "допплер", "доплер"], ["мошон"], "УЗДГ сосудов мошонки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "курс"])
	main_func(["скрин"], ["беремен", "пренатальный", "плода"], "Скрининговое УЗИ 1 триместр", ["дет", "пренатал", "реб", "втор", "описани", "хим", "вне", "не скрин", "расшифров", "занятий", "посещени", "для", "курс", "шейк", "трисоми", "хромос", "аномал", "биохим"], ["1 триместр", "до 13", "до 14", "i триместр", "11-14", "1-го три", "11-13 нед"])
	main_func(["скрин"], ["беремен", "пренатальный", "плода"], "Скрининговое УЗИ 2 триместр", ["дет", "пренатал", "узи 2-го триместра беременности", "реб", "втор", "описани", "хим", "вне", "расшифров", "не скрин", "занятий", "посещени", "для", "курс", "шейк", "трисоми", "хромос", "аномал", "биохим"], ["2 триместр", "до 27", "до 28", "ii триместр", "20-22", "2-го три", "18-21 нед", "ii-iii триместр", "14 по 28"])
	main_func(["скрин"], ["беремен", "пренатальный", "плода"], "Скрининговое УЗИ 3 триместр", ["дет", "реб", "втор", "описани", "хим", "узи 3-го триместра беременности", "вне", "расшифров", "не скрин", "занятий", "посещени", "для", "курс", "шейк", "трисоми", "хромос", "аномал", "биохим"], ["3 триместр", "с 28", "iii триместр", "32-34", "3-го три", "30-34 нед", "ii-iii триместр", "после 28"])
	main_func(["вне скрининг", "не скрининг"], ["узи"], "УЗИ беременным вне скрининга", ["дет", "реб", "втор", "описани", "хим", "расшифров", "занятий", "посещени", "для", "курс"], ["беремен"])
	main_func(uzi, uzi_pe4eni, id_uzi_pe4eni, stop_for_uzi_pe4eni)
	new_doubler_main(["узи", "ультразвук"], ["печени и желчного", "печени с желчным", "желчного и печени", "желчного с печень", "печень и желчный", "желчный и печень", "печень + желчный"], ["дет", "реб", "други", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс", "навед", "пункц", "контрол"], "УЗИ печени", 541726)
	main_func(uzi, uzi_po4ek, id_uzi_po4ek, stop_for_uzi_po4ek)
	new_doubler_main(["узи", "ультразвук"], ["почек и надпочечни", "почек с надпочечникам", "надпочечники и почки", "надпочечников и почек", "почки и надпочечников", "почки + надпочечники", "почек, надпочечников"], ["дет", "реб", "други", "втор", "описани", "расшифров", "пузыр", "занятий", "посещени", "для", "во время", "курс", "навед", "пункц", "контрол"], "УЗИ почек", 540984)
	main_func(["узи диагностика", "ультразвуковая диагностика", "диагностика беременности", "диагностирование беременности"], ["беременность", "диагностика беременности", "диагностирование беременности"], "Определение беременности", ["дет", "триместр", "реб", "втор", "описани", "хим", "расшифров", "занятий", "посещени", "для", "курс"])
	main_func(["амниотест", "тест"], ["амниотест", "подтекани"], "Амниотест", ["дет", "реб", "втор", "описани", "хим", "расшифров", "занятий", "посещени", "курс"], ["вод", "амниотест"])
	main_func(["манту"], ["манту"], "Реакция Манту", ["дет", "реб", "втор", "описани", "без", "расшифров", "после", "занятий", "консульт", "прием", "прив", "справ", "посещени", "для", "курс"])
	main_func(["уреазн", "дыхательный тест", "дыхательного теста", "хелик- тест", "хелик-тест"], ["уреазн", "дыхательный тест", "дыхательного теста", "helicobacter", "хеликобактер", "хелик- тест", "хелик- тест"], "Дыхательный тест на хеликобактер", ["дет", "водород", "реб", "втор", "описани", "без", "расшифров","прием", "консультац", "после", "занятий", "прив", "справ", "посещени", "для", "курс", "плода", "крови матери", "гена", "кровь"])
	main_func(["ведение беременности", "ведению беременности", "ведения беременности"], ["ведени"], "Ведение беременности", ["дет", "реб", "втор", "описани", "без", "расшифров", "после", "занятий", "консульт", "прием", "прив", "справ", "посещени", "для", "курс"])
	main_func(["ээг", "электроэнцефалография", "электроэнцефалограмм"], ["дет", "реб"], "ЭЭГ для детей", ["видео", "монитори", "втор", "описани", "расшифровка", "справ", "занятий", "посещени"])
	main_func(dupleks, dupleks_arteriy_verh, id_dupleks_arteriy_verh, stop_for_dupleks_arteriy_verh, dupleks_arteriy_verh_2)
	main_func(dupleks, dupleks_arteriy_niz, id_dupleks_arteriy_niz, stop_for_dupleks_arteriy_niz, dupleks_arteriy_niz_2)
	main_func(dupleks, dupleks_brahiocefal, id_dupleks_brahiocefal, stop_for_brahiocefal)
	main_func(dupleks, dupleks_ven_verh, id_dupleks_ven_verh, stop_for_dupleks_ven_verh, dupleks_ven_verh_2)
	main_func(dupleks, dupleks_ven_niz, id_dupleks_ven_niz, stop_for_dupleks_ven_niz, dupleks_ven_niz_2)
	new_doubler_main(["дуплекс", "уздс", "дс "], ["артерий и вен верхних конечностей", "сосудов (артерий и вен) верхних конечностей", "артерий и вен рук", "артерий и вен верхней конечности", "артерий и вен на верхн", "вен и артерий рук", "вен и артерий верхн", "сосудов верхн", "сосудов рук"], ["ребен", "перед", "дет", "одного", "бры", "чрев"], "Дуплексное сканирование вен верхних конечностей", 546691)
	new_doubler_main(["дуплекс", "уздс", "дс "], ["артерий и вен нижн", "артерий и вен ног", "вен и артерий ног", "вен и артерий ниж", "сосудов нижн", "сосудов ног", "(артерий и вен) нижних"], ["ребен", "перед", "дет", "одного", "ного", "бры", "чрев"], "Дуплексное сканирование вен нижних конечностей", 2109275)
	new_doubler_main(["дуплекс", "уздс", "дс "], ["артерий и вен верхних конечностей", "артерий и вен рук", "артерий и вен верхней конечности", "артерий и вен на верхн", "сосудов (артерий и вен) верхних конечностей", "вен и артерий рук", "вен и артерий верхн", "сосудов верхн", "сосудов рук"], ["ребен", "перед", "дет", "одного", "бры", "чрев"], "Дуплексное сканирование артерий верхних конечностей", 866862)
	new_doubler_main(["дуплекс", "уздс", "дс "], ["артерий и вен нижн", "артерий и вен ног", "вен и артерий ног", "вен и артерий ниж", "сосудов нижн", "сосудов ног", "(артерий и вен) нижних конечностей"], ["ребен", "перед", "дет", "ного", "одного", "бры", "чрев"], "Дуплексное сканирование артерий нижних конечностей", 2109271)
	main_func(["дуплекс", "триплекс"], ["брюшн"], "УЗДГ брюшной аорты", ["дет", "реб", "втор", "описани", "расшифров", "полост",  "после", "занятий", "консульт", "прием", "прив", "справ", "посещени", "для", "курс"], ["аорт"])
	main_func(["уздг", "доплер", "допплер", "ультразвук", "узи", "уз-диагностика", "уз диагностика", "триплекс"], ["вен", "сосуд", "артери"], "УЗДГ / Допплерография сосудов шеи", ["ребен", "дет", "дуплекс", "дс ", "под", "контрол"], ["шеи", "шейн", "брахиоцефал"])
	main_func(["удаление селезенки", "резекция селезенки", "спленэктомия"], ["спленэктомия", "селез"], "Удаление селезенки", ["ребен", "дет", "во время", "при"])
	main_func(["кава"], ["фильтр"], "Удаление кава-фильтра", ["ребен", "дет", "во время", "при"], ["удал"])
	main_func(["кава"], ["фильтр"], "Установка кава-фильтра", ["ребен", "дет", "во время", "при"], ["станов", "имплантац"])
	new_doubler_main(["вправлени", "устранение"], ["вывиха крупн", "вывихов крупн", "крупных сустав", "крупный сустав", "крупного суста", "крупном суста", "сустав (круп", "большой сустав", "большом суставе", "большого сустава", "больших суст"], ["дет", "кроме", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время"], "Вправление вывиха", 1132290)
	main_func(["сеч", "устра"], ["фимоз"], "Рассечение крайней плоти при фимозе", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["край"])
	main_func(["дивертикул", "девертикул"], ["тонк"], "Дивертикулэктомия тонкой кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["киш"])
	main_func(["диагностическ"], ["лапароскопи"], "Лапароскопия диагностическая", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "при диагностичес", "посещени", "во время", "курс"])
	main_func(["диагностическ"], ["лапаротоми"], "Диагностическая лапаротомия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "при диагностичес", "посещени", "во время", "курс"])
	new_doubler_main(["экг", "электрокардиографи", "электрокардиограм"], ["с расшифровкой", "включая расшифровку", "и расшифровка", "с описанием"], ["дет", "акц", "дом", "выезд", "неб", "сут", "монитор", "ад", "давлен", "холтер", "нагрузка", "нагрузкой", "стресс", "пищевод", "реб", "втор", "занятий", "расшифровка", "описание", "посещени", "под", "для", "во время", "курс"], "ЭКГ", 405)
	main_func(["пласти"], ["крест"], "Артроскопическая пластика задней крестообразной связки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "при", "во время", "курс"], ["зад"])
	main_func(["пласти"], ["крест", "пкс"], "Артроскопическая пластика передней крестообразной связки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "при", "во время", "курс"], ["перед", "пкс"])
	main_func(["артролиз"], ["локт"], "Артроскопический артролиз локтевого сустава", ["дет", "реб", "втор", "описани", "неартроскоп", "не артроскоп", "без артроскоп", "расшифров", "занятий", "при", "посещени", "для", "во время", "курс"], ["сустав"])
	main_func(["шов", "шва"], ["мениск"], "Артроскопический шов мениска", ["дет", "реб", "втор", "описани", "неартроскоп", "не артроскоп", "без артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"])
	perfect_match(["взятие крови из вены", "взятие крови", "забор крови", "забор  крови", "забор крови  из вены", "венопункция (взрослые, дети)", "взятие крови для анализа из вены", "забор крови из вены", "забор крови из периферической вены", "взятие крови из периферической вены", "забор венозной крови", "взятие венозной крови", "забор крови(вена)", "взятие крови (вена)", "взятие крови (из вены)", "забор крови для анализов", "забор крови из вены (пальца)", "взятие крови из периферической вены", "забор крови из периферической вены ( пациенты от 5 лет)"], "Забор крови")
	perfect_match(["взятие мазка", "взятие мазка из носа", "забор мазка/соскоба", "забор мазка из влагалища", "получение цервикального мазка", "забор мазка из носа и зева", "взятие мазка из зева", "забор мазка", "взятие мазков из уретры", "взятие материала ( соскоб)", "взятие (забор) гинекологического мазка", "взятие гинекологического мазка", "получение влагалищного мазка", "забор материала на анализ (мазок) на приеме", "забор мазков для исследования из носа, уха или горла (без стоимости исследования) (1 забор)", "получение урогенитального мазка", "получение влагалищного мазка на флору, онкоцитологию, ПЦР исследование, бакпосев", "взятие мазка из уретры"], "Взятие мазка")
	perfect_match(["Лечение шейки матки, кондилом влагалища солковагином"], "Прижигание эрозии шейки матки Солковагином")
	main_func(["почек", "почеч"], ["триплекс"], "Дуплексное сканирование сосудов почек", ["ребен", "терапия", "воздей", "дет", "под", "контрол"], ["сосуд"])
	main_func(["удал", "лечение одной"], ["папил"], "Удаление папиллом", ["дет", "лазер", "крио", "азот", "от 20", "радио", "сурги", "электро", "дэк", "диатермо", "вульв", "вагин", "влагал", "ротов", "глот", "лор", "анус", "анал", "зад", "более", "свыше", "след", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер"], ["папил"], "Лазерное удаление папиллом", ["дет", "более", "от 20", "свыше", "след", "11-20", "6-", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["папил"], "Радиоволновое удаление папиллом", ["дет", "от 20", "более", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["крио", "азот"], ["папил"], "Удаление папиллом жидким азотом", ["дет", "более", "от 20", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["хирурги", "механич"], ["папил"], "Хирургическое удаление папиллом", ["дет", "ради", "лазер", "магнит", "от 20", "сурги", "коагул", "электр", "более", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["электро", "диатермо", "дэк"], ["папил"], "Электрокоагуляция папиллом", ["дет", "более", "от 20", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["удал"], ["кондил", "кандил"], "Удаление кондилом", ["дет", "лазер", "крио", "азот", "от 20", "радио", "медикамент", "сурги", "электро", "дэк", "диатермо", "солкоде", "лекарств", "вульв", "вагин", "влагал", "ротов", "глот", "лор", "анус", "анал", "зад", "более", "свыше", "след", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["крио", "азот"], ["кондил", "кандил"], "Криодеструкция кондилом жидким азотом", ["дет", "от 20", "более", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["лазер"], ["кондил", "кандил"], "Лазерное удаление кондилом", ["дет", "более", "свыше", "от 20", "след", "11-", "6-", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["радио", "сурги"], ["кондил", "кандил"], "Радиоволновое удаление кондилом", ["дет", "более", "от 20", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["солкодер", "лекарств", "медикамент"], ["кондил", "кандил"], "Удаление кондилом солкодермом", ["дет", "от 20", "более", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["электро", "дэк", "диатермо"], ["кондилом", "кандил"], "Электрокоагуляция кондилом", ["дет", "более", "от 20", "свыше", "след", "11-20", "6-10", "от 5", "от 3", "от 4", "от 6", "от 7", "от 8", "от 9", "от 10", "от 11", "6-10", "2-5", "от 15", "в рамках", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["атрези"], ["хоан"], "Коррекция атрезии хоан", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["хоан"])
	main_func(["пластик"], ["нос"], "Пластика носового клапана", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["клапан"])
	main_func(["опера"], ["при выпадени"], "Операция при выпадении прямой кишки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"], ["киш"])
	main_func(["сеч", "удал"], ["кист"], "Иссечение кисты селезенки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["селезен", "селезён"])
	main_func(["люмбал"], ["дрен"], "Наружное люмбальное дренирование", ["дет", "реб", "внутр", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["дрен"])
	main_func(["удал"], ["интрамедул"], "Удаление интрамедуллярной опухоли", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["опухол"])
	main_func(["удал"], ["менингиом"], "Удаление менингиомы спинного мозга", ["дет", "голов", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["мозг"])
	main_func(["удал"], ["неврином"], "Удаление невриномы спинномозгового корешка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["спинномозг", "спинного мозг", "корешка"])
	main_func(["эмболизаци"], ["аневризм"], "Эмболизация аневризм и АВМ спинного мозга", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["спинного мозга", "спинномозговых"])
	main_func(["артродез"], ["голеностоп"], "Артродез голеностопного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артродез"], ["плеча", "плечев"], "Артродез плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артродез"], ["лисфранк"], "Артродез сустава Лисфранка", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["лисфранк"])
	main_func(["артродез"], ["шопар"], "Артродез сустава Шопара", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["шопар"])
	main_func(["артродез"], ["тазобедренн"], "Артродез тазобедренного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артродез"], ["таранн"], "Артродез таранно-пяточного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артропластик"], ["голеностоп"], "Артропластика голеностопного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артропластик"], ["коленн"], "Артропластика коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артропластик"], ["плеч"], "Артропластика плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артропластик"], ["пястн", "фаланг"], "Артропластика пястно-фаланговых суставов", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артропластик"], ["стопы", "стопе"], "Артропластика суставов стопы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артропластик"], ["тазобедрен"], "Артропластика тазобедренного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["комбин"], ["пластик"], "Артроскопическая комбинированная пластика связок коленного сустава", ["дет", "реб", "неартроскоп", "без артроскоп", "не артроскоп", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["коленного сустава"])
	main_func(["менискэктоми", "удаление мениска"], ["артроскоп"], "Артроскопическая менискэктомия", ["дет", "реб", "втор", "описани", "расшифров", "неартроскоп", "без артроскоп", "не артроскоп", "занятий", "посещени", "при", "для", "во время", "курс"], ["мениск"])
	main_func(["субакромиальн"], ["декомпрес"], "Артроскопическая субакромиальная декомпрессия", ["дет", "реб", "втор", "описани", "неартроскоп", "без артроскоп", "не артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["декомпрес"])
	main_func(["хондропластик"], ["голеностоп"], "Артроскопическая хондропластика голеностопного сустава", ["дет", "реб", "втор", "неартроскоп", "без артроскоп", "не артроскоп", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["хондропластик"], ["локт"], "Артроскопическая хондропластика локтевого сустава", ["дет", "реб", "втор", "описани", "неартроскоп", "без артроскоп", "не артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артролиз"], ["голеностоп"], "Артроскопический артролиз голеностопного сустава", ["дет", "реб", "втор", "описани", "неартроскоп", "без артроскоп", "не артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артролиз"], ["колен"], "Артроскопический артролиз коленного сустава", ["дет", "реб", "втор", "описани", "неартроскоп", "без артроскоп", "не артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["тенодез"], ["сухожил"], "Артроскопический тенодез сухожилия двуглавой мышцы плеча", ["дет", "реб", "втор", "неартроскоп", "без артроскоп", "не артроскоп", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["плеча"])
	main_func(["шов"], ["банкарта"], "Артроскопический шов повреждения Банкарта", ["дет", "реб", "втор", "описани", "неартроскоп", "без артроскоп", "не артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["банкарт"])
	main_func(["шов"], ["ротатор"], "Артроскопический шов повреждения ротаторов плеча", ["дет", "реб", "втор", "неартроскоп", "без артроскоп", "не артроскоп", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["плеч"])
	main_func(["артротомия"], ["голеностоп"], "Артротомия голеностопного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артротомия"], ["колен"], "Артротомия коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["артротомия"], ["плеч"], "Артротомия плечевого сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустав"])
	main_func(["опер"], ["привыч"], "Операция при привычном вывихе надколенника", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["надколенн"])
	main_func(["опер"], ["привыч"], "Операция при привычном вывихе плеча", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["плеч"])
	main_func(["операци"], ["разрыве связок"], "Операция при разрыве связок коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["коленного сустава"])
	main_func(["пластика"], ["пкс", "связок", "связки"], "Пластика связок коленного сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["колен", "пкс"])
	main_func(["сеч"], ["карпал"], "Рассечение карпальной связки", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["связ", "канал"])
	main_func(["резек"], ["молот"], "Резекция молоткообразных деформаций пальцев стопы", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["деформац"])
	main_func(["стабилизаци"], ["сустава"], "Стабилизация сустава", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["сустава"])
	main_func(["удаление"], ["мениск"], "Удаление мениска коленного сустава", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["колен"])
	main_func(["удаление"], ["тел"], "Артроскопическое удаление свободных внутрисуставных тел плечевого сустава", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["плеч"])
	main_func(["шов"], ["сухожил"], "Первичный шов сухожилий пальцев кисти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["кисти", "кисть"])
	main_func(["пластик"], ["ахилл"], "Пластика ахиллова сухожилия", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["ахилл"])
	main_func(["пластик"], ["дистальн"], "Пластика разрыва дистального сухожилия бицепса", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["бицепс"])
	main_func(["пластик"], ["длинной головки"], "Пластика разрыва сухожилия длинной головки бицепса", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["бицепс"])
	main_func(["пластик"], ["сухожил"], "Пластика сухожилий пальцев кисти", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["пальцев кисти"])
	main_func(["тенолиз"], ["ахилл"], "Тенолиз ахиллова сухожилия с бурсэктомией", ["дет", "реб", "втор", "описани", "расшифров", "без бурс", "занятий", "посещени", "при", "для", "во время", "курс"], ["бурс"])
	main_func(["транспозиц"], ["сухожил"], "Транспозиция сухожилий нижних конечностей", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], [" ног", "нижних конеч", "нижней конечно"])
	main_func(["удлинение"], ["сухожил"], "Удлинение сухожилия трицепса", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["трицепс"])
	main_func(["лфк", "лечебная физическая культура"], ["после"], "ЛФК в послеродовом периоде", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["родов"])
	main_func(["лфк", "лечебная физическая культура"], ["беременн"], "ЛФК для беременных", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс"], ["беременн"])
	main_func(["интерпретация результатов", "второе мнение", "консилиум", "совместный осмотр двух врачей"], ["друго", "второе мнение", "консилиум", "профессор", "к. м. н.", "к.м.н.", "совместный осмотр двух врачей"], "Второе мнение врача", ["дет", "привлече", "для участия в консилиуме", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["тест", "проб", "спирометри", "фвд", "внешнего дыхания"], ["бронхолити"], "Тест с бронхолитиком", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["офэкт", "однофотонная эмиссионная компьютерная томография"], ["пече"], "ОФЭКТ печени", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["антероградн", "восходящ"], ["пиелографи"], "Антеградная пиелография", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["рентген"], ["челюсти"], "Рентген челюсти", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["уздг", "доплер", "допплер"], ["орбит"], "УЗДГ сосудов орбиты", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["эндосонография", "эндоузи", "эхоэндоскоп", "эндоскопическое ультразвуковое исследование", "эндоскопическое узи"], ["пищевод", "желудок", "кишеч", "кишк", "эндосонография", "эндоузи"], "Эндосонография / Эндоузи", ["дет", "контролем", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["прием", "консульта"], ["спорт"], "Прием спортивного врача", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"], ["врач"])
	main_func(["лазер"], ["десен", "деснев", "пародонт"], "Лазеротерапия слизистой оболочки десен", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["ультразвук"], ["кавитация"], "Ультразвуковая кавитация", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "гинеколог", "интим", "полов", "посещени", "во время", "курс"])
	main_func(["аппаратн", "аппарате"], ["массаж"], "Аппаратный массаж", ["дет", "реб", "втор", "без", "допол", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["пхо", "первичная хирургическая обработка"], ["головы"], "ПХО ран мягких тканей головы", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "для", "посещени", "во время", "курс"])
	main_func(["орбитотомия"], ["орбитотомия"], "Орбитотомия", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["энуклеаци"], ["глаз"], "Энуклеация глазного яблока", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "для", "во время", "курс"])
	main_func(["вертебропластика"], ["вертебропластика"], "Вертебропластика", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["кифопластика"], ["кифопластика"], "Кифопластика", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["корпорэктомия"], ["корпорэктомия"], "Корпорэктомия", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["ламинэктомия"], ["ламинэктомия"], "Ламинэктомия", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["фикс"], ["кейдж"], "Межпозвонковая фиксации кейджами", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["налож"], ["гало"], "Наложение Гало-аппарата", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["аппарат"])
	main_func(["нуклеопластика"], ["нуклеопластика"], "Нуклеопластика", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["спондилодез", "артродез"], ["спондилодез", "позвон"], "Спондилодез / Артродез позвонков", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["спондилоэктомия", "резекц"], ["спондилоэктомия", "позвонк"], "Спондилоэктомия", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["транспедикуля"], ["фикс"], "Транспедикулярная фиксация позвоночника", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["позвон"])
	main_func(["денервац"], ["поч"], "Денервация почечных артерий", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["дилатац", "дилятац"], ["поч"], "Дилатация почечной артерии", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["торакоско"], ["опух"], "Торакоскопическое удаление доброкачественной опухоли средостения", ["дет", "реб", "втор", "без", "не торако", "неторако", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["средост"])
	main_func(["удал"], ["опух"], "Удаление опухолей средостения", ["дет", "реб", "торако", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["средосте"])
	main_func(["артродез"], ["сустав"], "Артродез (неподвижная фиксация) сустава", ["дет", "реб", "голен", "плеч", "лисфра", "шопар", "тазобед", "таран", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["пластик"], ["дефе"], "Пластика дефектов костей свода черепа", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["свода черепа"])
	main_func(["пластик"], ["базальной ликворе"], "Пластика костей основания черепа при базальной ликворее", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["пластик"], ["костей черепа"], "Пластика костей черепа краниоорбитальной локализации", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["краниоорбитальн"])
	main_func(["удал"], ["опух"], "Удаление опухоли костей черепа", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["костей черепа", "черепных костей", "черепной кости", "кости черепа"])
	main_func(["пластик"], ["мочеточник"], "Пластика мочеточника", ["дет", "реб", "киш", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["киш"], ["пластик"], "Кишечная пластика мочеточника", ["дет", "реб", "втор", "не киш", "некиш", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["мочеточ"])
	main_func(["закры"], ["синус"], "Закрытый синус-лифтинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["лиф"])
	main_func(["откры"], ["синус"], "Открытый синус-лифтинг", ["дет", "реб", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["лиф"])
	main_func(["шугаринг", "пиля"], ["шугаринг", "сахар"], "Шугаринг", ["дет", "реб", "бикин", "лиц", "ног", "подмы", "рук", "конеч", "бедр", "голе", "локт", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["гингивэктомия"], ["гингивэктомия"], "Гингивэктомия", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["гингивопластика"], ["гингивопластика"], "Гингивопластика", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["герме", "закрыти"], ["фиссур"], "Герметизация фиссур", ["втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["сеч"], ["капюшон"], "Иссечение капюшона при перикоронарите", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["каппа", "каппы", "капа", "капы"], ["спорт", "бруксиз", "разгрузоч", "ретенци", "окклюзионн", "лечение", "ортодонти", "релаксац"], "Капы для зубов", ["дет", "реб", "smile", "смайл", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["френкел", "фрекел"], ["френкел", "фрекел"], "Аппарат Френкеля", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["кюретаж"], ["лунк"], "Кюретаж лунки зуба", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["зуб"])
	main_func(["кюретаж"], ["пародонт"], "Кюретаж пародонтального кармана", ["дет", "реб", "втор", "откр", "закр", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["кюретаж"], ["пародонт", "зуб"], "Закрытый кюретаж пародонтального кармана", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["закр"])
	main_func(["кюретаж"], ["пародонт", "зуб"], "Открытый кюретаж пародонтального кармана", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["откр"])
	main_func(["дневной стационар", "дневном стационаре", "суточное пребывание в стационаре"], ["суточное пребывание в стационаре", "дневной стационар", "дневном стационаре"], "Дневной стационар", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["суточный стационар"], ["суточный стационар"], "Круглосуточный стационар", ["дет", "реб", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	function_with_assistant_in_cell_or_and_min_price_main(["026", "медицинское обследование для поступающих в дошкол"], ["детского сада", "детсад", "садик", "детский сад", "медицинское обследование для поступающих в дошкол"], 5999, "Медкарта 026 для детского сада")
	function_with_assistant_in_cell_or_and_min_price_main(["026"], ["школ"], 6000, "Медкарта 026 для школы")
	function_with_assistant_in_cell_or_and_min_price_main(["086"], ["поступлени", "вуз", "учебного", "институт", "справк", "форм"], 4500, "Справка для поступления (086у)")
	function_with_assistant_in_cell_or_and_min_price_main(["справк", "медицинский осмотр для"], [" спорт", "физкультур", "физическ", "атлет"], 899, "Справка для занятий спортом")
	function_with_assistant_in_cell_or_and_min_price_main(["справк"], ["бассейн"], 899, "Справка для бассейна")
	function_with_assistant_in_cell_or_and_min_price_main(["справк"], ["лагер", "079"], 999, "Справка в лагерь")
	function_with_assistant_in_cell_or_and_min_price_main(["справк"], ["состоянии здоровья", "состояние здоровья"], 899, "Справка о состоянии здоровья")
	function_with_assistant_in_cell_or_and_min_price_main(["справк"], ["санатор", "курорт"], 3500, "Санаторно-курортная карта")
	main_func(["медикамент", "лекарств", "мифепрестон", "мифепристон", "мефепрестон" "фармаборт", "мизопристол", "отечественный препарат", "импортный препарат"], ["аборт", "прерывание беременности"], "Медикаментозный аборт", ["дет", "втор", "описани", "расшифров", "занятий", "посещени", "во время", "курс"])
	main_func(["времен"], ["пломб"], "Временная пломба", ["дет", "реб", "имплан", "канал", "снят", "удал", "ортодо", "втор", "без", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["узи", "ультразвук", "ультрозвук", "узд"], ["поч"], "УЗИ почек и мочевого пузыря", ["дет", "реб", "имплан", "канал", "ортодо", "втор", "без", "не ", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["пузыр"])
	main_func(["офтальм", "глаз"], ["тонометр"], "Тонометрия", ["дет", "реб", "имплан", "маклак", "канал", "ортодо", "бескон", "пневм", "втор", "без", "не ", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["чистка", "комплекс", "гигиен", "отложен", "сняти", "гигиен"], ["air", "эйрфлоу", "эйр флоу", "аэрфло"], "Чистка зубов AirFlow", ["дет", "реб", "професси", "комплекс", "втор", "курс", "1 зуб", "скидк", "без", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["flow"])
	main_func(["удал"], ["движ"], "Удаление подвижного зуба", ["дет", "реб", "втор", "времен", "акци", "фрагмент", "молоч", "отбели", "стенки", "стенок", "импла", "описани", "неподвиж", "расшифров", "занятий", "посещени", "во время", "неподвиж", "для", "курс"], ["зуб"])
	main_func(["кор", "щепл", "расшир", "аугмент", "увели", "пластик", "модиф"], ["альвеол"], "Коррекция альвеолярного гребня", ["дет", "реб", "втор", "без", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["греб", "отрост"])
	main_func(["удаление имплант"], ["удаление имплант"], "Удаление зубного имплантата", ["дет", "реб", "контрацеп", "импланон", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["отбел"], ["opales", "опалесенс", "опейлсенс", "опалесценс"], "Отбеливание зубов Opalescence", ["дет", "реб", "курс", "скидк", "втор", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "дом", "курс"])
	main_func(["отбел"], ["beyond", "byond", "бейонд"], "Отбеливание зубов Beyond", ["дет", "реб", "втор", "курс", "скидк", "без", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["отбел"], ["лазер"], "Лазерное отбеливание зубов", ["дет", "реб", "втор", "курс", "скидк", "без", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["зуб"])
	main_func(["ортопантомограмма", "оптг", "ортопантомография", "панорамн"], ["ортопантомограмма", "оптг", "ортопантомография", "рентген"], "Ортопантомограмма", ["дет", "реб", "скидк", "акц", "по результатам","медицинское заключен", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["ортопантомограмма", "оптг", "ортопантомография", "зуб"])
	main_func(["шинир"], ["зуб"], "Шинирование зуба", ["дет", "реб", "втор", "времен", "акци", "реплантация", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс", "след"])
	main_func(["времен"], ["коронк"], "Временные коронки", ["дет", "фиксаци", "реб", "втор", "ампута", "акци", "восстановление куль", "отбели", "цементировка", "временный цемент", "имплант", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["внутри"], ["канал"], "Внутриканальное отбеливание зубов", ["дет", "реб", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "во время", "для", "допол", "курс"], ["отбел"])
	main_func(["отбели"], ["amazing", "амейзинг", "эмейзинг"], "Отбеливание зубов Amazing white", ["дет", "реб", "сеанс", "курс", "скидк", "втор", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["отбели"], ["zoom"], "Отбеливание зубов zoom", ["дет", "реб", "втор", "акци", "описани", "курс", "скидк", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["чист", "проф", "снятие", "удал"], ["зуб", "гигиен", "отложени"], "Чистка зубов ультразвуком", ["дет", "реб", "имплант", "жир", "втор", "имплант", "акци", "лиц", "кож", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["ультразвук", "у/з", "ультрозвук"])
	main_func(["удал"], ["ретенирова", "ретинирова"], "Удаление ретенированного зуба", ["дет", "нерет", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["зуб"])
	main_func(["резе"], ["верх"], "Резекция верхушки корня зуба", ["дет", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["зуб", "корня"])
	main_func(["имплант"], ["adin"], "Импланты Adin", ["дет", "реб", "втор", "формировате", "фдм", "коронк", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["alpha bio", "альфа био", "alfa", "альфа-био", "alpha - bio", "alpha-bio"], "Импланты Alpha Bio", ["дет", "формировате", "аналог", "времен", "фдм", "предложе", "коронк", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["анкилос", "ankilos", "ankylos"], "Импланты ANKYLOS", ["дет", "формировате", "фдм", "реб", "коронк", "аналог", "точка опоры", "времен", "втор", "предложе", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["anyridge"], "Импланты Anyridge", ["дет", "реб", "втор", "акци", "формировате", "фдм", "отбели", "коронк", "аналог", "времен", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["astra", "астр-тек", "астратек", "астра тек"], "Импланты Astra Tech", ["дет", "реб", "полный", "точка опоры", "съемн", "формировате", "фдм", "втор", "коронк", "аналог", "времен", "акци", "заглушк", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["bicon"], "Импланты Bicon", ["дет", "реб", "втор", "акци", "отбели", "формировате", "фдм", "описани", "точка опоры", "коронк", "аналог", "времен", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["biomet"], "Импланты Biomet 3I", ["дет", "реб", "втор", "акци", "отбели", "формировате", "фдм", "коронк", "точка опоры", "описани", "аналог", "времен", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["icx"], "Импланты Icx", ["дет", "реб", "втор", "акци", "отбели", "описани", "формировате", "фдм", "коронк", "точка опоры", "расшифров", "аналог", "времен", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["impro"], "Импланты IMPRO", ["дет", "реб", "втор", "акци", "отбели", "описани", "формировате", "фдм", "коронк", "точка опоры", "расшифров", "аналог", "времен", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["mis"], "Импланты MIS", ["дет", "реб", "втор", "акци", "отбели", "описани", "формировате", "фдм", "расшифров", "коронк", "точка опоры", "занятий", "аналог", "предложе", "времен", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["nobel", "нобел"], "Импланты Nobel Biocare", ["дет", "реб", "втор", "акци", "формировате", "фдм", "отбели", "коронк", "предложе", "описани", "точка опоры", "расшифров", "аналог", "времен", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["осстем", "остем", "osstem", "ostem"], "Импланты Osstem", ["дет", "реб", "формировате", "фдм", "втор", "предложе", "на 4", "на 6", "акци", "коронк", "точка опоры", "отбели", "описани", "аналог", "времен", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["xive"], "Импланты XIVE", ["дет", "реб", "втор", "акци", "отбели", "описани", "формировате", "фдм", "расшифров", "коронк", "занятий", "предложе", "посещени", "точка опоры", "аналог", "времен", "во время", "для", "курс"] )
	main_func(["имплант"], ["zimmer", "циммер", "зиммер"], "Импланты Zimmer", ["дет", "реб", "втор", "формировате", "фдм", "акци", "отбели", "коронк", "описани", "расшифров", "предложе", "точка опоры", "аналог", "времен", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["биогоризонт"], "Импланты Биогоризонт", ["дет", "реб", "втор", "акци", "формировате", "фдм", "отбели", "описани", "коронк", "расшифров", "занятий", "аналог", "времен", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["имплантиум", "implantium"], "Импланты Имплантиум", ["дет", "реб", "втор", "формировате", "фдм", "акци", "отбели", "коронк", "коронк", "описани", "аналог", "времен", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["имплант"], ["strauman", "штрауман"], "Импланты Штрауман", ["дет", "реб", "втор", "акци", "формировате", "фдм", "отбели", "коронк", "описани", "расшифров", "аналог", "времен", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["ороантра", "закрытие соустья с гайморовой"], ["сообще", "закрытие соустья с гайморовой"], "Закрытие ороантрального сообщения", ["дет", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["сня", "удал"], ["зуб"], "Снятие зубных отложений", ["дет", "реб", "втор", "акци", "имплант", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"], ["отлож"] )
	main_func(["удал"], ["дистоп"], "Удаление дистопированного зуба", ["дет", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "недисто", "не дисто", "курс"], ["зуб"] )
	main_func(["vector", "вектор"], ["лечени", "терапи", "снятие", "обработ", "челюст"], "Лечение аппаратом Вектор", ["дет", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "курс"] )
	main_func(["установка", "прямой", "индивидуаль", "alpha", "nobel", "шаровид", "углово", "adin", "alfa", "двойной", "стандартный", "категори"], ["абатмент", "абантмент", "аббатмент"], "Установка абатмента", ["дет", "реб", "времен", "втор", "акци", "отбели", "описани", "расшифров", "без установки абат", "без стоимости абат", "не включая аба", "отдельно", "занятий", "посещени", "во время", "для", "курс"] )
	main_func(["десн", "установ"], ["формировател", "фдм"], "Установка формирователя десны", ["дет", "реб", "втор", "при ", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "без установки форми"], ["фдм", "формировател"])
	main_func(["коронка", "единица", "коронки", "коронкой"], ["на имплант", "с использованием имплантата"], "Коронка на имплант", ["дет", "реб", "втор", "акци", "отбели", "времен", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["пластмас"], ["коронка"], "Пластмассовые коронки", ["дет", "реб", "времен", "втор", "акци", "отбели", "металл", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс"])
	main_func(["культев", "культивая", "культевой"], ["вкладк"], "Культевая вкладка", ["дет", "реб", "втор", "акци", "подготов", "отбели", "времен", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "удал", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["кларити", "clarit"], "Брекеты Clarity", ["дет", "брекета", "этап", "реб", "одного", "использование", "актив", "втор", "акци", "осмотр", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["damon"], "Брекеты Damon Clear", ["дет", "реб", "брекета", "втор", "этап", "акци", "актив", "отбели", "одного", "использование", "описани", "осмотр", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"], ["clear"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["damon"], "Брекеты Damon-Q", ["дет", "реб", "втор", "акци", "брекета", "этап", "одного", "использование", "отбели", "актив", "описани", "расшифров", "осмотр", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"], ["q"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["empower"], "Брекеты Empower", ["дет", "реб", "втор", "акци", "этап", "брекета", "отбели", "актив", "описани", "одного", "использование", "расшифров", "занятий", "осмотр", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["gemini", "gemeni"], "Брекеты Gemini", ["дет", "реб", "втор", "акци", "брекета", "отбели", "этап", "описани", "актив", "расшифров", "одного", "использование", "занятий", "посещени", "осмотр", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["in"], "Брекеты In-Ovation", ["дет", "реб", "металлический", "втор", "одного", "брекета", "использование", "акци", "отбели", "этап", "одного", "использование", "описани", "расшифров", "актив", "занятий", "посещени", "во время", "осмотр", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"], ["ovation"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["ini"], "Брекеты Mini Master", ["дет", "реб", "втор", "акци", "отбели", "этап", "брекета", "описани", "расшифров", "актив", "занятий", "посещени", "одного", "использование", "во время", "для", "осмотр", "курс", "не включа", "привар", "почин", "допол", "снятие"], ["master"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["sprint", "спринт"], "Брекеты Sprint", ["дет", "реб", "втор", "акци", "отбели", "этап", "брекета", "описани", "расшифров", "актив", "занятий", "посещени", "одного", "использование", "во время", "осмотр", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["стб", "stb"], "Брекеты Stb", ["дет", "реб", "втор", "акци", "отбели", "описани", "этап", "брекета", "расшифров", "занятий", "актив", "посещени", "во время", "одного", "использование", "для", "осмотр", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["victory", "виктори"], "Брекеты Victory", ["дет", "реб", "втор", "акци", "отбели", "этап", "брекета", "описани", "расшифров", "занятий", "актив", "посещени", "одного", "использование", "во время", "для", "осмотр", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["блеск"], "Брекеты Блеск", ["дет", "реб", "втор", "акци", "отбели", "описани", "одного", "использование", "брекета", "расшифров", "этап", "занятий", "посещени", "во время", "актив", "для", "курс", "не включа", "осмотр", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["дет", "реб"], "Брекеты для детей", ["втор", "акци", "отбели", "описани", "расшифров", "занятий", "этап", "одного", "брекета", "использование", "посещени", "во время", "для", "актив", "курс", "не включа", "привар", "осмотр", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["инкогнито", "incognito"], "Брекеты Инкогнито", ["дет", "реб", "втор", "акци", "отбели", "описани", "исследование", "диагности", "этап", "расшифров", "одного", "брекета", "использование", "занятий", "посещени", "актив", "во время", "для", "курс", "осмотр", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["ortos", "ортос"], "Брекеты Ортос", ["дет", "реб", "втор", "акци", "отбели", "описани", "одного", "использование", "расшифров", "этап", "брекета", "занятий", "посещени", "во время", "актив", "для", "курс", "не включа", "осмотр", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["сапфир", "inspire"], "Сапфировые брекеты", ["дет", "реб", "втор", "акци", "отбели", "описани", "одного", "использование", "расшифров", "занятий", "брекета", "этап", "посещени", "во время", "для", "актив", "курс", "не включа", "привар", "осмотр", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["самолигир", "безлигатур"], "Самолигирующие / безлигатурные брекеты", ["дет", "реб", "одного", "использование", "при ", "втор", "акци", "брекета", "отбели", "этап", "описани", "расшифров", "занятий", "актив", "посещени", "во время", "осмотр", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["лингвал"], "Лингвальные брекеты", ["дет", "реб", "осмотр", "актив", "втор", "акци", "отбели", "одного", "использование", "описани", "расшифров", "брекета", "этап", "занятий", "посещени", "во время", "для", "курс", "не включа", "осмотр", "привар", "почин", "допол", "снятие"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["снятие"], "Снятие брекет-системы", ["дет", "реб", "элеме", "при ", "кнопк", "замк", "брасл", "втор", "акци", "актив", "отбели", "описани", "расшифров", "брекета", "занятий", "этап", "посещени", "во время", "курс", "не включа", "привар", "почин", "допол"])
	main_func(["гемисекция"], ["гемисекция"], "Гемисекция зуба", ["дет", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "этап", "для", "курс", "не включа", "привар", "почин", "допол"])
	main_func(["удал"], ["молоч", "времен"], "Удаление молочных зубов", ["дет", "реб", "втор", "акци", "пульпит", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол"], ["зуб"])
	function_with_assistant_in_cell_or_and_min_price_main(["invisalign", "инвизилайн"], ["invisalign", "инвизилайн"], 100000, "Элайнеры Invisalign")
	function_with_assistant_in_cell_or_and_min_price_main(["star"], ["smile"], 100000, "Элайнеры Star Smile")
	main_func(["серебр"], ["зуб"], "Серебрение зубов", ["втор", "акци", "отбели", "описани", "расшифров", "занятий", "во время", "курс", "не включа", "привар", "почин", "допол"])
	main_func(["lm", "l-m"], ["активат"], "LM-активатор", ["осмотр", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "контрол"])
	main_func(["угле"], ["ванн"], "Углекислая ванна", ["дет", "реб", "втор", "акци", "отбели", "описани", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "контрол"])
	main_func(["внутри"], ["коронк"], "Внутрикоронковое отбеливание зуба", ["дет", "реб", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "допол", "курс"], ["отбел"])
	main_func(["биометрия"], ["глаз"], "Биометрия глаза", ["дет", "реб", "втор", "описани", "расшифров", "пахиметр", "ультра", "узи", "уз-", "кератометри", "экзо", "занятий", "посещени", "при", "для", "во время", "курс"])
	main_func(["криосауна"], ["криосауна"], "Криосауна", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["альвеол"], ["эктом", "удал"], "Альвеолэктомия", ["дет", "реб", "втор", "кюре", "лунк", "отсро", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["удал", "ампутац"], ["корня "], "Удаление корня зуба", ["дет", "реб", "втор", "молоч", "описани", "кисты", "через лун", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["керами"], ["винир"], "Керамические виниры", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["композит", "прямые", "прямого", "прямым"], ["винир"], "Композитные виниры", ["дет", "реб", "непр", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["люминир"], ["люминир"], "Люминиры", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["опер"], ["опущен"], "Операции при опущении половых органов", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"], ["полов"])
	main_func(["удлинени", "увелич"], ["зуба"], "Удлинение коронковой части зуба", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["квадротти", "кватротти"], ["квадротти", "кватротти"], "Протезы Квадротти", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["дома"], ["opales", "опелс", "опалес", "опалесценс"], "Домашнее отбеливание Opalescence", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["компонир", "кампонир", "капонир", "копонир"], ["компонир", "кампонир", "капонир", "копонир"], "Установка компониров", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["становле"], ["культи"], "Восстановление культи зуба", ["дет", "реб", "втор", "времен", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["циркони"], ["коронк", "единица"], "Циркониевые коронки", ["дет", "реб", "втор", "имплант", "вклад", "alfa", "osstem", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["стимул"], ["овуляц"], "Стимуляция суперовуляции", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["ортопед", "ортезирова", "изготовлен"], ["стелек", "стельки", "формто", "formt", "индивидуальных ортез", "индивидуального ортез"], "Изготовление ортопедических стелек", ["дет", "реб", "втор", "артроза", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["кожн"], ["проба"], "Кожные пробы (аллергопробы)", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["маркиро"], ["молочно"], "Внутритканевая маркировка образований молочной железы", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["ретейнер"], ["ретейнер"], "Ретейнеры", ["дет", "реб", "втор", "сня", "удал", "чис", "чин", "фиксация", "ремон", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["зуб"], ["камн", "камен"], "Удаление зубного камня", ["дет", "реб", "медика", "налет", "мягк", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["налет"], ["зуб"], "Удаление налета с зубов", ["лмк", "книж", "комплекс", "гигиен", "чистка", "профес"])
	main_func(["фтор"], ["зуб", "эмал", "челюст"], "Фторирование эмали", ["дет", "реб", "втор", "описани", "чист", "гиги", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["бели"], ["дома"], "Отбеливание зубов в домашних условиях", ["дет", "реб", "втор", "zoom", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["кисты"], ["зуб"], "Удаление кисты зуба", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["метал", "литые", "литая"], ["коронк", "единица"], "Цельнометаллические коронки", ["дет", "реб", "времен", "молоч", "вклад", "коронковой", "керам", "емакс", "e-max", "сня", "втор", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс", "импла", "удал", "извле"])
	main_func(["метал"], ["керам"], "Металлокерамические коронки", ["дет", "реб", "сня", "втор", "описани", "артроскоп", "расшифров", "вклад", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс", "импла", "удал", "извле"], ["корон", "единица"])
	main_func(["керам", "емакс", "e-max", "emax", "e max"], ["коронк", "единица"], "Керамические коронки", ["дет", "реб", "втор", "вклад", "сня", "метал", "литы", "литой", "литу", "лита", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс", "импла", "удал", "извле"])
	main_func(["ретрак"], ["десн"], "Ретракция десны", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["бабочк", "иммедиа", "имедиа"], ["бабочк", "протез"], "Иммедиат-протез Бабочка", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["штифт", "свш"], ["штифт", "свш"], "Установка зубного штифта", ["дет", "реб", "втор", "удал", "извеч", "описани", "пломбир", "канала", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["icon"], ["кариес"], "Лечение кариеса ICON", ["дет", "реб", "втор", "описани", "артроскоп", "расшифров", "занятий", "посещени", "при", "для", "во время", "курс", "акц", "сотру", "дмс"])
	main_func(["антител"], ["коронавирус", "covid"], "Тест на антитела к коронавирусу", ["лмк", "книж", "профосмо", "профпат", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "вакцин"])
	main_func(["шлиф", "полир"], ["зуб"], "Шлифовка / Полировка зубов", ["дет", "реб", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["реставра", "восстано"], ["зуб"], "Восстановление (реставрация) зубов", ["дет", "пломб", "реставраций", "шлифовка", "витре", "vitre", "реб", "диастем", "корон", "во время", "после", "реж", "клиновид", "культ", "эмал", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["реставра", "восстано"], ["зуб"], "Восстановление коронковой части зуба", ["дет", "реб", "диастем", "молоч", "реж", "культ", "эмал", "во время", "после", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"], ["коронков"])
	main_func(["клиновид"], ["дефект"], "Реставрация клиновидного дефекта", ["дет", "реб", "диастем", "корон", "реж", "культ", "эмал", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["слож"], ["удал"], "Сложное удаление зуба", ["дет", "реб", "втор", "времен", "курс", "постоян", "несложн", "молоч", "не слож", "степен", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"], ["зуб"])
	main_func(["металл"], ["брекет"], "Металлические брекеты", ["дет", "реб", "втор", "снятие", "актива", "брекета", "удале", "извле", "ремонт", "почин", "фиксац", "керам", "1 брекет", "1 зуб", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["композит", "керам"], ["вкладк"], "Композитная вкладка", ["дет", "реб", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"], ["циркон", "композит"])
	main_func(["бюгель"], ["протез"], "Бюгельные протезы", ["дет", "реб", "замен", "матриц", "снятие", "удале", "извле", "под б", "под про", "коронк", "ремонт", "почин", "времен", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["акрил"], ["протез"], "Акриловые протезы", ["дет", "реб", "втор", "времен", "снятие", "удале", "извле", "ремонт", "почин", "замен", "матриц", "безакри", "без акри", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["акри", "acryfree", "acry-free", "acry free"], ["фри", "acryfree", "acry-free", "acry free"], "Протезы Акри-Фри", ["дет", "реб", "втор", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["частич"], ["протез"], "Частичные съемные зубные протезы", ["дет", "реб", "втор", "замен", "ногт", "снятие", "удале", "извле", "ремонт", "почин", "времен", "матриц", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["лоскут"], ["пародонт", "зуб"], "Лоскутная операция при пародонтите", ["дет", "реб", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["трейнер", "треинер"], ["зуб", "выравни", "корре", "исправл", "функционал"], "Трейнеры для зубов", ["миобрейс", "miobrace", "втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "допол", "курс"])
	main_func(["трейнер", "треинер"], ["миобрейс", "miobrace"], "Трейнер Миобрейс", ["втор", "курс", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "допол", "курс"])
	main_func(["брекеты", "брекет-систем", "брекет систем"], ["керам"], "Керамические брекеты", ["дет", "реб", "брекета", "втор", "брекета", "этап", "акци", "актив", "метал", "отбели", "одного", "использование", "описани", "осмотр", "расшифров", "занятий", "посещени", "во время", "для", "курс", "не включа", "привар", "почин", "допол", "снятие"])
	main_func(["нейлон"], ["протез"], "Нейлоновые протезы", ["дет", "реб", "втор", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["десн"], ["рецес"], "Закрытие рецессии десны", ["дет", "реб", "втор", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["дентал"], ["визиограф"], "Дентальная визиография", ["дет", "реб", "втор", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])
	main_func(["дуплекс", "уздс"], ["голов"], "Дуплексное сканирование сосудов головы и шеи / УЗДС", ["дет", "реб", "втор", "позвон", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"], ["шеи", "шея", "шейн"])
	main_func(["допплер", "доплер" "уздг"], ["шеи"], "УЗДГ сосудов шеи ребенку", ["втор", "позвон", "курс", "микро", "голов", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"], ["дет", "реб"])
	main_func(["допплер", "доплер" "уздг"], ["голов"], "УЗДГ сосудов головы ребенку", ["втор", "позвон", "шеи", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"], ["дет", "реб"])
	main_func(["допплер", "доплер" "уздг"], ["головы  и шеи"], "УЗДГ сосудов головы и шеи ребенку", ["втор", "позвон", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"], ["дет", "реб"])
	main_func(["узи", "ультразвук", "ультрозвук", "уз-иссл"], ["лонн"], "УЗИ лонного сочленения", ["дет", "реб", "втор", "позвон", "курс", "микро", "снятие", "удале", "извле", "ремонт", "почин", "мини", "скидк", "акци", "описани", "расшифров", "занятий", "посещени", "во время", "для", "допол", "курс"])



	'''
	Лечение кисты зуба
	'''



	main_func(uzi, uzi_vilo4kovoy_jelezi_2, id_uzi_vilo4kovoy_jelezi, stop_for_uzi_vilo4kovoy_jelezi)
	main_func(uzi, uzi_vn4s_2, id_uzi_vn4s, stop_for_uzi_vn4s)
	main_func(uzi, uzi_kisti_2, id_uzi_kisti, stop_for_uzi_kisti)
	main_func(uzi, uzi_limfouzlov_2, id_uzi_limfouzlov, stop_for_uzi_limfouzlov)
	main_func(uzi, uzi_malogo_taza_2, id_uzi_malogo_taza, stop_for_uzi_malogo_taza)
	main_func(uzi, uzi_malogo_taza__ta_2, id_uzi_malogo_taza_ta, stop_for_uzi_malogo_taza_ta, uzi_malogo_taza__ta_3)
	main_func(uzi, uzi_malogo_taza__tv_2, id_uzi_malogo_taza_tv, stop_for_uzi_malogo_taza_tv, uzi_malogo_taza__tv_3)
	main_func(uzi, uzi_matki_2, id_uzi_matki, stop_for_uzi_matki)
	main_func(uzi, uzi_molo4nih_jelez_2, id_uzi_molo4nih_jelez, stop_for_uzi_molo4nih_jelez, uzi_molo4nih_jelez_3)
	main_func(uzi, uzi_molo4nih_jelez_limfouzel_2, id_uzi_molo4nih_jelez_limfouzel, stop_for_uzi_molo4nih_jelez_limfouzel, uzi_molo4nih_jelez_limfouzel_3)
	main_func(uzi, uzi_mo4evogo_puzirya_2, id_uzi_mo4evogo_puzirya, stop_for_uzi_mo4evogo_puzirya)
	main_func(uzi, uzi_mo4evogo_puzirya_mo4a_2, id_uzi_mo4evogo_puzirya_mo4a, stop_for_uzi_mo4evogo_puzirya_mo4a, uzi_mo4evogo_puzirya_mo4a_3)
	main_func(uzi, uzi_mo4eto4nikov_2, id_uzi_mo4eto4nikov, stop_for_uzi_mo4eto4nikov)
	main_func(uzi, uzi_moshonki_2, id_uzi_moshonki, stop_for_uzi_moshonki)
	main_func(uzi, uzi_4lena_2, id_uzi_4lena, stop_for_uzi_4lena)
	main_func(uzi, uzi_pridat4nih_2, id_uzi_pridat4nih, stop_for_uzi_pridat4nih)
	main_func(uzi, uzi_prostati_2, id_uzi_prostati, stop_for_uzi_prostati)
	main_func(uzi, uzi_truzi_2, id_uzi_truzi, stop_for_uzi_truzi, uzi_truzi_3)
	main_func(uzi, uzi_slunnih, id_uzi_slunnih, stop_for_uzi_slunnih)
	main_func(uzi, uzi_sredosteniya, id_uzi_sredosteniya, stop_for_uzi_sredosteniya)
	main_func(uzi, uzi_stopi, id_uzi_stopi, stop_for_uzi_stopi)
	main_func(uzi, uzi_suhojil, id_uzi_suhojil, stop_for_uzi_suhojil)
	main_func(uzi, uzi_uretri, id_uzi_uretri, stop_for_uzi_uretri)
	main_func(uzi, uzi_parashitovidnih_2, id_uzi_parashitovidnih, stop_for_uzi_parashitovidnih, uzi_parashitovidnih_3)
	main_func(uzi, uzi_shitovid, id_uzi_shitovid, stop_for_uzi_shitovid, uzi_shitovid_2)
	main_func(uzi, uzi_yai4nikov, id_uzi_yai4nikov, stop_for_uzi_yai4nikov)
	main_func(follikulometriya, follikulometriya_2, id_follikulometriya, stop_for_uzi_follikulometriya)
	main_func(uzi, uzi_brushnoy_1, id_uzi_brushnoy,stop_for_uzi_brushnoy, uzi_brushnoy_2)
	main_func(uzi, uzi_jel4nogo, id_uzi_jel4nogo, stop_for_uzi_jel4nogo)
	main_func(uzi, uzi_zabrushinnogo, id_uzi_zabrushinnogo, stop_for_uzi_zabrushinnogo)
	main_func(uzi, uzi_nadpo4e4, id_uzi_nadpo4e4, stop_for_uzi_nadpo4e4)
	main_func(uzi, uzi_podjeludo4, id_uzi_podjeludo4, stop_for_uzi_podjeludo4)
	main_func(uzi, uzi_selezen, id_uzi_selezen, stop_for_uzi_selezen)
	main_func(uzi, uzi_funkcii_jel4nogo_1, id_uzi_funkcii_jel4nogo, stop_for_uzi_funkcii_jel4nogo, uzi_funkcii_jel4nogo_2)
	main_func(uzi, uzi_brushnoy_rebenku_1, id_uzi_brushnoy_rebenku, stop_for_uzi_brushnoy_rebenku, uzi_brushnoy_rebenku_2)
	main_func(uzi, uzi_vilo4kovoy_rebenku_1, id_uzi_vilo4kovoy_rebenku, stop_for_uzi_vilo4kovoy_rebenku, uzi_vilo4kovoy_rebenku_2)
	main_func(uzi, uzi_jel4nogo_rebenku_1, id_uzi_jel4nogo_rebenku, stop_for_uzi_jel4nogo_rebenku, uzi_jel4nogo_rebenku_2)
	main_func(uzi, uzi_kolen_rebenku_1, id_uzi_kolen_rebenku, stop_for_uzi_kolen_rebenku, uzi_kolen_rebenku_2)
	main_func(uzi, uzi_limfouzlov_rebenku_1, id_uzi_limfouzlov_rebenku, stop_for_uzi_limfouzlov_rebenku, uzi_limfouzlov_rebenku_2)
	main_func(uzi, uzi_taza_rebenku_1, id_uzi_taza_rebenku, stop_for_uzi_taza_rebenku, uzi_taza_rebenku_2)
	main_func(uzi, uzi_molo4noy_rebenku, id_uzi_molo4noy_rebenku, stop_for_uzi_molo4noy_rebenku, uzi_molo4noy_rebenku_2)
	main_func(uzi, uzi_mo4evogo_rebenku, id_uzi_mo4evogo_rebenku, stop_for_uzi_mo4evogo_rebenku, uzi_mo4evogo_rebenku_2)
	main_func(uzi, uzi_moshonki_rebenku, id_uzi_moshonki_rebenku, stop_for_uzi_moshonki_rebenku, uzi_moshonki_rebenku_2)
	main_func(uzi, uzi_m9gkih_rebenku, id_uzi_m9gkih_rebenku, stop_for_uzi_m9gkih_rebenku, uzi_m9gkih_rebenku_2)
	main_func(uzi, uzi_nadpo4e4_rebenku, id_uzi_nadpo4e4_rebenku, stop_for_uzi_nadpo4e4_rebenku, uzi_nadpo4e4_rebenku_2)
	main_func(uzi, uzi_pe4eni_rebenku, id_uzi_pe4eni_rebenku, stop_for_uzi_pe4eni_rebenku, uzi_pe4eni_rebenku_2)
	main_func(uzi, uzi_podjelud_rebenku, id_uzi_podjelud_rebenku, stop_for_uzi_podjelud_rebenku, uzi_podjelud_rebenku_2)
	main_func(uzi, uzi_pozvono4_rebenku, id_uzi_pozvono4_rebenku, stop_for_uzi_pozvono4_rebenku, uzi_pozvono4_rebenku_2)
	main_func(uzi, uzi_po4ek_rebenku, id_uzi_po4ek_rebenku, stop_for_uzi_po4ek_rebenku, uzi_po4ek_rebenku_2)
	main_func(uzi, uzi_pazuh_rebenku, id_uzi_pazuh_rebenku, stop_for_uzi_pazuh_rebenku, uzi_pazuh_rebenku_2)
	main_func(uzi, uzi_selezen_rebenku, id_uzi_selezen_rebenku, stop_for_uzi_selezen_rebenku, uzi_selezen_rebenku_2)
	main_func(uzi_serdca_rebenku, uzi_serdca_rebenku_2, id_uzi_serdca_rebenku, stop_for_uzi_serdca_rebenku, uzi_serdca_rebenku_3)
	main_func(uzi, uzi_tazobedr_rebenku, id_uzi_tazobedr_rebenku, stop_for_uzi_tazobedr_rebenku, uzi_tazobedr_rebenku_2)
	main_func(uzi, uzi_shenyogo_rebenku, id_uzi_shenyogo_rebenku, stop_for_uzi_shenyogo_rebenku, uzi_shenyogo_rebenku_2)
	main_func(uzi, uzi_shitovid_rebenku, id_uzi_shitovid_rebenku, stop_for_uzi_shitovid_rebenku, uzi_shitovid_rebenku_2)
	main_func(uzi, uzi_m9gkih_tkaney, id_uzi_m9gkih_tkaney, stop_for_uzi_m9gkih_tkaney)
	main_func(uzi, uzi_m9gkih_tkaney_lica, id_uzi_m9gkih_tkaney_lica, stop_for_uzi_m9gkih_tkaney_lica, id_uzi_m9gkih_tkaney_lica_2)
	main_func(uzi, uzi_m9gkih_tkaney_shei, id_uzi_m9gkih_tkaney_shei, stop_for_uzi_m9gkih_tkaney_shei, id_uzi_m9gkih_tkaney_shei_2)
	main_func(uzi, uzi_m9gkih_plevri, id_uzi_plevri, stop_for_uzi_plevri)
	main_func(uzi, uzi_pozvono4, id_uzi_pozvono4, stop_for_uzi_pozvono4)
	main_func(uzi, uzi_grudnogo_pozvono4, id_uzi_grudnogo_pozvono4, stop_for_uzi_grudnogo_pozvono4, uzi_grudnogo_pozvono4_2)
	main_func(uzi, uzi_kop4ik, id_uzi_kop4ik, stop_for_uzi_kop4ik, uzi_kop4ik_2)
	main_func(uzi, uzi_po9snic, id_uzi_po9snic, stop_for_uzi_po9snic, uzi_po9snic_2)
	main_func(uzi, uzi_sheynogo_pozvon, id_uzi_sheynogo_pozvon, stop_for_uzi_psheynogo_pozvon, uzi_sheynogo_pozvon_2)
	main_func(uzi_serdca, uzi_serdca_2, id_uzi_serdca, stop_for_uzi_serdca)
	main_func(stress_eho, stress_eho_2, id_stress_eho, stop_for_stress_eho, stress_eho_3)
	main_func(uzi, uzi_perikarda, id_uzi_perikarda, stop_for_uzi_perikarda)
	main_func(chrespishevod_eho, chrespishevod_eho_2, id_chrespishevod_eho, stop_for_chrespishevod_eho, chrespishevod_eho_3)
	main_func(uzi, uzi_sustavov, id_uzi_sustavov, stop_for_uzi_sustavov)
	main_func(uzi, uzi_golenostop, id_uzi_golenostop, stop_for_uzi_golenostop)
	main_func(uzi, uzi_kolen, id_uzi_kolen, stop_for_uzi_kolen)
	main_func(uzi, uzi_loktevogo, id_uzi_loktevogo, stop_for_uzi_loktevogo)
	main_func(uzi, uzi_lu4ezap, id_uzi_lu4ezap, stop_for_uzi_lu4ezap)
	main_func(uzi, uzi_ple4, id_uzi_ple4, stop_for_uzi_ple4)
	main_func(uzi, uzi_tazobedr, id_uzi_tazobedr, stop_for_uzi_tazobedr)
	main_func(dupleks, dupleks_aorti, id_dupleks_aorti, stop_for_dupleks_aorti)
	main_func(dupleks, dupleks_uzds, id_dupleks_uzds, stop_for_dupleks_uzds)
	main_func(dupleks, dupleks_po4ek, id_dupleks_po4ek, stop_for_dupleks_po4ek)
	main_func(dupleks, dupleks_shei, id_dupleks_shei, stop_for_dupleks_shei)
	main_func(tripleks, tripleks_arteriy_verh, id_tripleks_arteriy_verh, stop_for_tripleks_arteriy_verh, tripleks_arteriy_verh_2)
	main_func(tripleks, tripleks_arteriy_niz, id_tripleks_arteriy_niz, stop_for_tripleks_arteriy_niz, tripleks_arteriy_niz_2)
	main_func(tripleks, tripleks_ven_verh, id_tripleks_ven_verh, stop_for_tripleks_ven_verh, tripleks_ven_verh_2)
	main_func(tripleks, tripleks_ven_niz, id_tripleks_ven_niz, stop_for_tripleks_ven_niz, tripleks_ven_niz_2)
	main_func(tripleks, tripleks_sosudov_golov_i_shei, id_tripleks_sosudov_golov_i_shei, stop_for_tripleks_sosudov_golov_i_shei, tripleks_sosudov_golov_i_shei_2)
	main_func(uzdg, uzdg_shitovid, id_uzdg_shitovid, stop_for_uzdg_shitovid )
	main_func(tkdg, tkdg_2, id_tkdg, stop_for_tkdg)
	main_func(uzdg, uzdg_sosudov_golovi, id_uzdg_sosudov_golovi, stop_for_uzdg_sosudov_golovi)
	main_func(uzdg, uzdg_sosudov_arteriy_niz, id_uzdg_arteriy_niz, stop_for_uzdg_arteriy_niz, uzdg_arteriy_niz_2)
	main_func(uzdg, uzdg_brushn_aort, id_uzdg_brushn_aort, stop_for_uzdg_brushn_aort, uzdg_brushn_aort_2)
	main_func(uzdg, uzdg_sosudov_ven_verh, id_uzdg_ven_verh, stop_for_ven_verh, uzdg_ven_verh_2)
	main_func(uzdg, uzdg_sosudov_ven_niz, id_uzdg_ven_niz, stop_for_ven_niz, uzdg_ven_niz_2)
	main_func(uzdg, uzdg_glaz_sosud, id_uzdg_glaz_sosud, stop_for_glaz_sosud, uzdg_glaz_sosud_2)
	main_func(uzdg, uzdg_limf_sosud, id_uzdg_limf_sosud, stop_for_limf_sosud)
	main_func(uzdg, uzdg_sosud_brush_polo, id_uzdg_sosud_brush_polo, stop_for_sosud_brush_polo)
	main_func(uzdg, uzdg_sosud_molo4, id_uzdg_molo4, stop_for_molo4)
	main_func(uzdg, uzdg_sosud_po4ek, id_uzdg_po4ek, stop_for_po4ek)
	main_func(uzdg, uzdg_ekstrakranial, id_uzdg_uzdg_ekstrakranial, stop_for_uzdg_ekstrakranial)
	main_func(sonoelastograf, sonoelastograf_2, id_sonoelastograf, stop_for_sonoelastograf)
	main_func(uzgss_mato4nih, mato4nih_2, id_mato4nih, stop_for_mato4nih)
	main_func(elastograf_molo4, elastograf_molo4_2, id_elastograf_molo4, stop_for_elastograf_molo4)
	main_func(elastograf_pe4en, elastograf_pe4en_2, id_elastograf_pe4en, stop_for_elastograf_pe4en)
	main_func(uzi, uzi_glaza, id_uzi_glaza, stop_for_uzi_glaza)
	main_func(neirosonografia, neirosonografia_2, id_neirosonografia, stop_for_neirosonograf)
	main_func(fetometria, fetometria_2, id_fetometria, stop_for_fetometria)
	main_func(kolposkopia, kolposkopia_2, id_kolposkopia, stop_for_kolposkopia)
	main_func(kolposkopia_video, kolposkopia_video_2, id_kolposkopia_video, stop_for_kolposkopia_video, kolposkopia_video_3)
	main_func(anoskop, anoskop_2, id_anoskop, stop_for_anoskop)
	main_func(rektoromanoskop, rektoromanoskop_2, id_rektoromanoskop, stop_for_rektoromanoskop)
	main_func(rektoromanoskop_det, rektoromanoskop_det_2, id_rektoromanoskop_det, stop_for_rektoromanoskop_det, rektoromanoskop_det_3)
	main_func(bronhoskop, bronhoskop_2, id_bronhoskop, stop_for_bronhoskop)
	main_func(vulvoskop, vulvoskop_2, id_vulvoskop, stop_for_vulvoskop)
	main_func(diagnoz_gisteroskop, diagnoz_gisterosko_2, id_diagnoz_gisterosko, stop_for_diagnoz_gisterosko)
	main_func(laringoskop, laringoskop_2, id_laringoskop, stop_for_laringoskop)
	main_func(mediastinoskop, mediastinoskop_2, id_mediastinoskop, stop_for_mediastinoskop)
	main_func(bioimpedansometria, bioimpedansometria_2, id_bioimpedansometria, stop_for_bioimpedansometria)
	main_func(audiogramma, audiogramma_2, id_audiogramma, stop_for_audiogramma)
	main_func(audiogramma_weber, audiogramma_weber_2, id_audiogramma_weber, stop_for_audiogramma_weber, audiogramma_weber_3)
	main_func(audiogramma_porog, audiogramma_porog_2, id_audiogramma_porog, stop_for_audiogramma_porog, audiogramma_porog_3)
	main_func(tamponada_nosa, tamponada_nosa_2, id_tamponada_nosa, stop_for_tamponada_nosa)
	main_func(zaush_blok, zaush_blok_2, id_zaush_blok, stop_for_zaush_blok)
	main_func(mass_pereponki, mass_pereponki_2, id_mass_pereponki, stop_for_mass_pereponki)
	main_func(otoskopia, otoskopia_2, id_otoskopia, stop_for_otoskopia)     
	main_func(kamerton, kamerton_2, id_kamerton, stop_for_kamerton)
	main_func(elektrokohleo, elektrokohleo_2, id_elektrokohleo, stop_for_elektrokohleo)
	main_func(otoakustik, otoakustik_2, id_otoakustik, stop_for_otoakustik)
	main_func(timpanometr, timpanometr_2, id_timpanometr, stop_for_timpanometr)
	main_func(perimetr, perimetr_2, id_perimetr, stop_for_perimetr)
	main_func(skiaskop, skiaskop_2, id_skiaskop, stop_for_skiaskop)
	main_func(maklakov, maklakov_2, id_maklakov, stop_for_maklakov)
	main_func(gonioskop, gonioskop_2, id_gonioskop, stop_for_gonioskop)
	main_func(refraktometr, refraktometr_2, id_refraktometr, stop_for_refraktometr)
	main_func(shirmer, shirmer_2, id_shirmer, stop_for_shirmer)
	main_func(norn, norn_2, id_norn, stop_for_norn)
	main_func(ekzoftalmometr, ekzoftalmometr_2, id_ekzoftalmometr, stop_for_ekzoftalmometr)
	main_func(mass_prostati, mass_prostati_2, id_mass_prostati, stop_for_mass_prostati)
	main_func(ust_pessar, ust_pessar_2, id_ust_pessar, stop_for_ust_pessar)
	main_func(udal_pessar, udal_pessar_2, id_udal_pessar, stop_for_udal_pessar)
	main_func(inorod_vlag, inorod_vlag_2, id_inorod_vlag, stop_for_inorod_vlag)
	main_func(inorod_glaz, inorod_glaz_2, id_inorod_glaz, stop_for_inorod_glaz)
	main_func(inorod_glotki, inorod_glotki_2, id_inorod_glotki, stop_for_inorod_glotki)
	main_func(inorod_gortani, inorod_gortani_2, id_inorod_gortani, stop_for_inorod_gortani)
	main_func(inorod_nosa, inorod_nosa_2, id_inorod_nosa, stop_for_inorod_nosa)
	main_func(inorod_uha, inorod_uha_2, id_inorod_uha, stop_for_inorod_uha)
	main_func(politser, politser_2, id_politser, stop_for_politser)
	main_func(kukushka, kukushka_2, id_kukushka, stop_for_kukushka)
	main_func(bujir_cervik, bujir_cervik_2, id_bujir_cervik, stop_for_bujir_cervik)
	main_func(bujir_uretri, bujir_uretri_2, id_bujir_uretri, stop_for_bujir_uretri)
	main_func(bujir_uretri_jen, bujir_uretri_jen_2, id_bujir_uretri_jen, stop_for_bujir_uretri_jen, bujir_uretri_jen_3)
	main_func(bujir_uretri_muj, bujir_uretri_muj_2, id_bujir_uretri_muj, stop_for_bujir_uretri_muj, bujir_uretri_muj_3)
	main_func(bujir_anal, bujir_anal_2, id_bujir_anal, stop_for_bujir_anal)
	main_func(bujir_kolostom, bujir_kolostom_2, id_bujir_kolostom, stop_for_bujir_kolostom)
	main_func(inorod_kishe4, inorod_kishe4_2, id_inorod_kishe4, stop_for_inorod_kishe4)
	main_func(blokad_grushevid, blokad_grushevid_2, id_blokad_grushevid, stop_for_blokad_grushevid)
	main_func(block_krest_povz, block_krest_povz_2, id_block_krest_povz, stop_for_block_krest_povz, block_krest_povz_3)
	main_func(block_per_nerva, block_per_nerva_2, id_block_per_nerva, stop_for_block_per_nerva)
	main_func(block_p9t_shpor, block_p9t_shpor_2, id_block_p9t_shpor, stop_for_block_p9t_shpor, block_p9t_shpor_2)
	main_func(block_pozvon, block_pozvon_2, id_block_pozvon, stop_for_block_pozvon)
	main_func(block_semen_kanat, block_semen_kanat_2, id_block_semen_kanat, stop_for_block_semen_kanat, block_block_semen_kanat_2)
	main_func(block_sustav, block_sustav_2, id_block_sustav, stop_for_block_sustav)
	main_func(block_to4ek, block_to4ek_2, id_block_to4ek, stop_for_block_to4ek)
	main_func(block_nosa, block_nosa_2, id_block_nosa, stop_for_block_nosa)
	main_func(block_zaushn, block_zaushn_2, id_block_zaushn, stop_for_block_zaushn)
	main_func(block_intratonz, block_intratonz_2, id_block_intratonz, stop_for_block_intratonz)
	main_func(block_koreshkov, block_koreshkov_2, id_block_koreshkov, stop_for_block_koreshkov)
	main_func(block_mezhreber, block_mezhreber_2, id_block_mezhreber, stop_for_block_mezhreber)
	main_func(block_paraprost, block_paraprost_2, id_block_paraprost, stop_for_block_paraprost)
	main_func(block_presakral, block_presakral_2, id_block_presakral, stop_for_block_presakral)
	main_func(block_epidural, block_epidural_2, id_block_epidural, stop_for_block_epidural)
	main_func(piling_karbon, piling_karbon_2, id_piling_karbon, stop_for_piling_karbon)
	main_func(piling_gazojid, piling_gazojid_2, id_piling_gazojid, stop_for_piling_gazojid)
	main_func(piling_korall, piling_korall_2, id_piling_korall, stop_for_piling_korall)
	main_func(piling_mehani4, piling_mehani4_2, id_piling_mehani4, stop_for_piling_mehani4)
	main_func(piling_abr, piling_abr_2, id_piling_abr, stop_for_piling_abr)
	main_func(piling_spa, piling_spa_2, id_piling_spa, stop_for_piling_spa)
	main_func(piling_almaz, piling_almaz_2, id_piling_almaz, stop_for_piling_almaz)
	main_func(piling_glikol, piling_glikol_2, id_piling_glikol, stop_for_piling_glikol)
	main_func(piling_jeltiy, piling_jeltiy_2, id_piling_jeltiy, stop_for_piling_jeltiy)
	main_func(piling_mindal, piling_mindal_2, id_piling_mindal, stop_for_piling_mindal)
	main_func(piling_molo4, piling_molo4_2, id_piling_molo4, stop_for_piling_molo4)
	main_func(piling_jessner, piling_jessner_2, id_piling_jessner, stop_for_piling_jessner)
	main_func(piling_tsa, piling_tsa_2, id_piling_tsa, stop_for_piling_tsa)
	main_func(piling_pirovino, piling_pirovino_2, id_piling_pirovino, stop_for_piling_pirovino)
	main_func(piling_salicil, piling_salicil_2, id_piling_salicil, stop_for_piling_salicil)
	main_func(piling_fenol, piling_fenol_2, id_piling_fenol, stop_for_piling_fenol)
	main_func(piling_ferul, piling_ferul_2, id_piling_ferul, stop_for_piling_ferul)
	main_func(piling_frukt, piling_frukt_2, id_piling_frukt, stop_for_piling_frukt)
	main_func(amnioskopia, amnioskopia_2, id_amnioskopia, stop_for_amnioskopia)
	main_func(amniocentez, amniocentez_2, id_amniocentez, stop_for_amniocentez)
	main_func(kordocentez, kordocentez_2, id_kordocentez, stop_for_kordocentez)
	main_func(placentocentez, placentocentez_2, id_placentocentez, stop_for_placentocentez)
	main_func(placentocentez, placentocentez_2, id_placentocentez, stop_for_placentocentez)
	main_func(prenatal_trisom, prenatal_trisom_2, id_prenatal_trisom, stop_for_prenatal_trisom, prenatal_trisom_3)
	main_func(uzi_vne_skrining, uzi_vne_skrining_2, id_uzi_vne_skrining, stop_for_uzi_vne_skrining)
	main_func(uzi_serdca_ploda,uzi_serdca_ploda_2, id_uzi_serdca_ploda, stop_for_uzi_serdca_ploda, uzi_serdca_ploda_3)
	main_func(cervikometr, cervikometr_2, id_cervikometr, stop_for_cervikometr)
	main_func(diafanoskop, diafanoskop_2, id_diafanoskop, stop_for_diafanoskop)
	main_func(kolonoskop, kolonoskop_2, id_kolonoskop, stop_for_kolonoskop)
	main_func(fks_and_fgds, fks_and_fgds_2, id_fks_and_fgds, stop_for_fks_and_fgds)
	main_func(fks_narkoz, fks_narkoz_2, id_fks_narkoz, stop_for_fks_narkoz)
	main_func(angiograf_set4atk, angiograf_set4atk_2, id_angiograf_set4atk, stop_for_angiograf_set4atk)
	main_func(beskon_tonometr, beskon_tonometr_2, id_beskon_tonometr, stop_for_beskon_tonometr)
	main_func(retinometria, retinometria_2, id_retinometria, stop_for_retinometria)
	main_func(fluor_angio, fluor_angio_2, id_fluor_angio, stop_for_fluor_angio)
	main_func(fluor_proba, fluor_proba_2, id_fluor_proba, stop_for_fluor_proba)
	main_func(cvet_slez_nos, cvet_slez_nos_2, id_cvet_slez_nos, stop_for_cvet_slez_nos, cvet_slez_nos_3)
	main_func(elastotonometr, elastotonometr_2, id_elastotonometr, stop_for_elastotonometr)
	main_func(elektrookulograf, elektrookulograf_2, id_elektrookulograf, stop_for_elektrookulograf)
	main_func(elektroretinograf, elektroretinograf_2, id_elektroretinograf, stop_for_elektroretinograf)
	main_func(biopsia_bronh, biopsia_bronh_2, id_biopsia_bronh, stop_for_biopsia_bronh)
	main_func(biopsia_vulvi, biopsia_vulvi_2, id_biopsia_vulvi, stop_for_biopsia_vulvi)
	main_func(biopsia_mozg, biopsia_mozg_2, id_biopsia_mozg, stop_for_biopsia_mozg)
	main_func(biopsia_gortan, biopsia_gortan_2, id_biopsia_gortan, stop_for_biopsia_gortan)
	main_func(biopsia_jeludka, biopsia_jeludka_2, id_biopsia_jeludka, stop_for_biopsia_jeludka)
	main_func(biopsia_kavern_4lena, biopsia_kavern_4lena_2, id_biopsia_kavern_4lena, stop_for_biopsia_kavern_4lena, biopsia_kavern_4lena_3)
	main_func(biopsia_kishe4, biopsia_kishe4_2, id_biopsia_kishe4, stop_for_biopsia_kishe4)
	main_func(biopsia_koji, biopsia_koji_2, id_biopsia_koji, stop_for_biopsia_koji)
	main_func(biopsia_kost_mozg, biopsia_kost_mozg_2, id_biopsia_kost_mozg, stop_for_biopsia_kost_mozg, biopsia_kost_mozg_3)
	main_func(biopsia_legkih, biopsia_legkih_2, id_biopsia_legkih, stop_for_biopsia_legkih)
	main_func(biopsia_limfouzl, biopsia_limfouzl_2, id_biopsia_limfouzl, stop_for_biopsia_limfouzl)
	main_func(biopsia_molo4, biopsia_molo4_2, id_biopsia_molo4, stop_forbiopsia_molo4, biopsia_molo4_3)
	main_func(biopsia_mo4evo, biopsia_mo4evo_2, id_biopsia_mo4evo, stop_for_biopsia_mo4evo)
	main_func(biopsia_mishc, biopsia_mishc_2, id_biopsia_mishc, stop_for_biopsia_mishc)
	main_func(biopsia_m9gkih, biopsia_m9gkih_2, id_biopsia_m9gkih, stop_for_ebiopsia_m9gkih)
	main_func(biopsia_nakojnih, biopsia_nakojnih_2, id_biopsia_nakojnih, stop_biopsia_nakojnih, biopsia_nakojnih_3)
	main_func(biopsia_obraz_4len, biopsia_obraz_4len_2, id_biopsia_obraz_4len, stop_for_biopsia_obraz_4len, biopsia_obraz_4len_3)
	main_func(biopsia_pe4en, biopsia_pe4en_2, id_biopsia_pe4en, stop_for_biopsia_pe4en)
	main_func(biopsia_pishevod, biopsia_pishevod_2, id_biopsia_pishevod, stop_for_biopsia_pishevod)
	main_func(biopsia_podjelud, biopsia_podjelud_2, id_biopsia_podjelud, stop_for_biopsia_podjelud)
	main_func(biopsia_polip, biopsia_polip_2, id_biopsia_polip, stop_for_biopsia_polip)
	main_func(biopsia_po4ki, biopsia_po4ki_2, id_biopsia_po4ki, stop_for_biopsia_po4ki)
	main_func(biopsia_prostati, biopsia_prostati_2, id_biopsia_prostati, stop_for_biopsia_prostati)
	main_func(biopsia_pr9m_kishki, biopsia_pr9m_kishki_2, id_biopsia_pr9m_kishki, stop_for_biopsia_pr9m_kishki, biopsia_pr9m_kishki_3)
	main_func(biopsia_sustav, biopsia_sustav_2, id_biopsia_sustav, stop_for_biopsia_sustav, biopsia_sustav_3)
	main_func(biopsia_slunnih, biopsia_slunnih_2, id_biopsia_slunnih, stop_for_biopsia_slunnih)
	main_func(biopsia_ton_kishe4, biopsia_ton_kishe4_2, id_biopsia_ton_kishe4, stop_for_biopsia_ton_kishe4, biopsia_ton_kishe4_3)
	main_func(biopsia_sheiki_matki, biopsia_sheiki_matki_2, id_biopsia_sheiki_matki, stop_for_biopsia_sheiki_matki)
	main_func(biopsia_shitovid, biopsia_shitovid_2, id_biopsia_shitovid, stop_for_biopsia_shitovid)
	main_func(laparos_biops_9i4nikov, laparos_biops_9i4nikov_2, id_laparos_biops_9i4nikov, stop_for_laparos_biops_9i4nikov, laparos_biops_9i4nikov_3)
	main_func(nojeva9a_biops_sheiki_matki, nojeva9a_biops_sheiki_matki_2, id_nojeva9a_biops_sheiki_matki, stop_for_nojeva9a_biops_sheiki_matki)
	main_func(biopsia_plevri, biopsia_plevri_2, id_biopsia_plevri, stop_for_biopsia_plevri)
	main_func(biopsia_podmish_limfo, biopsia_podmish_limfo_2, id_biopsia_podmish_limfo, stop_for_biopsia_podmish_limfo, biopsia_podmish_limfo_3)
	main_func(biopsia_pozvonka, biopsia_pozvonka_2, id_biopsia_pozvonka, stop_for_biopsia_pozvonka)
	main_func(biopsi_endometri, biopsi_endometri_2, id_biopsi_endometri, stop_for_biopsi_endometri)
	main_func(aspirac_biop_endometr, aspirac_biop_endometr_2, id_aspirac_biop_endometr, stop_for_aspirac_biop_endometr)
	main_func(paipel, paipel_2, id_paipel, stop_for_paipel, paipel_3)
	main_func(tsug_biopsi, tsug_biopsi_2, id_tsug_biopsi, stop_for_tsug_biopsi, tsug_biopsi_3)
	main_func(fgds_1, fgds_2_1, id_fgds_1, stop_for_fgds_1)
	main_func(fgds_2, fgds_2_2, id_fgds_2, stop_for_fgds_2)
	main_func(fgds_3, fgds_2_3, id_fgds_3, stop_for_fgds_3)
	main_func(fgds_narko, fgds_narko_2, id_fgds_narko_2, stop_for_fgds_narko_2)
	main_func(ezofagogastroduodenoskop, ezofagogastroduodenoskop_2, id_ezofagogastroduodenoskop, stop_for_ezofagogastroduodenoskop)
	main_func(endonazal_fgds, endonazal_fgds_2, id_endonazal_fgds, stop_for_endonazal_fgds)
	main_func(gisteroskop, gisteroskop_2, id_gisteroskop, stop_for_gisteroskop)
	main_func(gisteroskop_rdv, gisteroskop_rdv_2, id_gisteroskop_rdv, stop_for_gisteroskop_rdv)
	main_func(razdel_viskabl, razdel_viskabl_2, id_razdel_viskabl, stop_for_razdel_viskabl)
	main_func(preskalen_biopsia, preskalen_biopsia_2, id_preskalen_biopsia, stop_for_preskalen_biopsia)
	main_func(pricel_biopsia_sheiki_matki, pricel_biopsia_sheiki_matki_2, id_pricel_biopsia_sheiki_matki, stop_for_pricel_biopsia_sheiki_matki, pricel_biopsia_sheiki_matki_3)
	main_func(biopsi_9i4ka, biopsi_9i4ka_2, id_biopsi_9i4ka, stop_for_biopsi_9i4ka)
	main_func(otkrita_biopsi_9i4ka, otkrita_biopsi_9i4ka_2, id_otkrita_biopsi_9i4ka, stop_for_otkrita_biopsi_9i4ka, otkrita_biopsi_9i4ka_3)
	main_func(aromaterapi, aromaterapi_2, id_aromaterapi, stop_for_aromaterapi)
	main_func(barokameri, barokameri_2, id_barokameri, stop_for_barokameri)
	main_func(biorezonans, biorezonans_2, id_biorezonans, stop_for_biorezonans)
	main_func(biotok, biotok_2, id_biotok, stop_for_biotok)
	main_func(bos_terapi, bos_terapi_2, id_bos_terapi, stop_for_bos_terapi)
	main_func(galvanization, galvanization_2, id_galvanization, stop_for_galvanization)
	main_func(vlok, vlok_2, id_vlok, stop_for_vlok, vlok_3)
	main_func(gidrokolonoterap, gidrokolonoterap_2, id_gidrokolonoterap, stop_for_gidrokolonoterap)
	main_func(gr9zele4en, gr9zele4en_2, id_gr9zele4en, stop_for_gr9zele4en)
	main_func(diadinamo, diadinamo_2, id_diadinamo, stop_for_diadinamo)
	main_func(infrakras_terapi, infrakras_terapi_2, id_infrakras_terapi, stop_for_infrakras_terapi)
	main_func(kv4, kv4_2, id_kv4, stop_for_kv4)
	main_func(lazeroterapi, lazeroterapi_2, id_lazeroterapi, stop_for_lazeroterapi)
	main_func(elektromagnit, elektromagnit_2, id_elektromagnit, stop_for_elektromagnit, elektromagnit_3 )
	main_func(magnitolazer, magnitolazer_2, id_magnitolazer, stop_for_magnitolazer, magnitolazer_3)
	main_func(magnitoterap, magnitoterap_2, id_magnitoterap, stop_for_magnitoterap)
	main_func(mikrovolni, mikrovolni_2, id_mikrovolni, stop_for_mikrovolni)
	main_func(okuf, okuf_2, id_okuf, stop_for_okuf)
	main_func(smt_terapi, smt_terapi_2, id_smt_terapi, stop_for_smt_terapi)
	main_func(smt_terapi_vne, smt_terapi_vne_2, id_smt_terapi_vne, stop_for_smt_terapi_vne, smt_terapi_3)
	main_func(transkrani_elektrostim, transkrani_elektrostim_2, id_transkrani_elektrostim, stop_transkrani_elektrostim)
	main_func(transrektal_elektrostim_prostati, transrektal_elektrostim_prostati_2, id_transrektal_elektrostim_prostati, stop_transrektal_elektrostim_prostati, transrektal_elektrostim_prostati_3)
	main_func(transuretral_elektrostim_prostati, transuretral_elektrostim_prostati_2, id_transuretral_elektrostim, stop_transuretral_elektrostim_prostati, transuretral_elektrostim_prostati_3)
	main_func(uv4_terapi, uv4_terapi_2, id_uv4_terapi, stop_uv4_terapi)
	main_func(uz_terapi, uz_terapi_2, id_uz_terapi, stop_uz_terapi)
	main_func(ufo_terapi, ufo_terapi_2, id_ufo_terapi, stop_ufo_terapi)
	main_func(fdt_terapi, fdt_terapi_2, id_fdt_terapi, stop_fdt_terapi)
	main_func(fonoforez, fonoforez_2, id_fonoforez, stop_fonoforez)
	main_func(fotoforez, fotoforez_2, id_fotoforez, stop_fotoforez)
	main_func(chens, chens_2, id_chens, stop_chens)
	main_func(elektromiostimul, elektromiostimul_2, id_elektromiostimul, stop_elektromiostimul)
	main_func(elektroson, eelektroson_2, id_elektroson, stop_elektroson)
	main_func(elektrostimul_mo4evogo, elektrostimul_mo4evogo_2, id_elektrostimul_mo4evogo, stop_elektrostimul_mo4evogo, elektrostimul_mo4evogo_3)
	main_func(elektrostimul_uretri, elektrostimul_uretri_2, id_elektrostimul_uretri, stop_elektrostimul_uretri, elektrostimul_uretri_3)
	main_func(elektrostimul_cervik, elektrostimul_cervik_2, id_elektrostimul_cervik, stop_elektrostimul_cervik, elektrostimul_cervik_3)
	main_func(uretroskopia, uretroskopiak_2, id_uretroskopia, stop_uretroskopia)
	main_func(cistoskopia, cistoskopia_2, id_cistoskopia, stop_cistoskopia)
	main_func(cirkumsicizo, cirkumsicizo_2, id_cirkumsicizo, stop_cirkumsicizo)
	main_func(viskab_cerviko, viskab_cerviko_2, id_viskab_cerviko, stop_viskab_cerviko, viskab_cerviko_3)
	main_func(defloracia, defloracia_2, id_defloracia, stop_defloracia)
	main_func(gineko_massaj, gineko_massaj_2, id_gineko_massaj, stop_gineko_massaj)
	main_func(gsg, gsg_2, id_gsg, stop_gsg)
	main_func(punkc_shitovid, punkc_shitovid_2, id_punkc_shitovid, stop_punkc_shitovid)
	main_func(massaj_veka, massaj_veka_2, id_massaj_veka, stop_massaj_veka)
	main_func(rinoskopia, rinoskopia_2, id_rinoskopia, stop_rinoskopia)
	main_func(reposicia_nosa, reposicia_nosa_2, id_reposicia_nosa, stop_reposicia_nosa)
	main_func(udalenie_probok, udalenie_probok_2, id_udalenie_probok, stop_udalenie_probok)
	main_func(promiv_mndalin, promiv_mndalin_2, id_promiv_mndalin, stop_promiv_mndalin)
	main_func(tonzillor, tonzillor_2, id_tonzillor, stop_tonzillor)
	main_func(vnutriven_ozon, vnutriven_ozon_2, id_vnutriven_ozon, stop_vnutriven_ozon)
	main_func(kateter_sluh, kateter_sluh_2, id_kateter_sluh, stop_kateter_sluh)
	main_func(trihogramm, trihogramm_2, id_trihogramm, stop_trihogramm )
	main_func(trihoskop, trihoskop_2, id_trihoskop, stop_trihoskop)
	main_func(fototrihogramm, fototrihogramm_2, id_fototrihogramm, stop_fototrihogramm)
	main_func(lampa_vuda, lampa_vuda_2, id_lampa_vuda, stop_lampa_vuda)
	main_func(ustanovka_vms, ustanovka_vms_2, id_ustanovka_vms, stop_ustanovka_vms)
	main_func(udalenie_vms, udalenie_vms_2, id_udalenie_vms, stop_udalenie_vms)
	main_func(udal_vros_vms, udal_vros_vms_2, id_udal_vros_vms, stop_udal_vros_vms, udal_vros_vms_3)
	main_func(zondir_polosti_matki, zondir_polosti_matki_2, id_zondir_polosti_matki, stop_zondir_polosti_matki, zondir_polosti_matki_3)
	main_func(kateter_mo4evogo_jen, kateter_mo4evogo_jen_2, id_kateter_mo4evogo_jen, stop_kateter_mo4evogo_jen, kateter_mo4evogo_jen_3)
	main_func(kateter_mo4evogo_muj, kateter_mo4evogo_muj_2, id_kateter_mo4evogo_muj, stop_kateter_mo4evogo_muj, kateter_mo4evogo_muj_3)
	main_func(vnutriven_inekc, vnutriven_inekc_2, id_vnutriven_inekc, stop_vnutriven_inekc)
	main_func(vnutrimish_inekc, vnutrimish_inekc_2, id_vnutrimish_inekc, stop_vnutrimish_inekc)
	main_func(lumbal_punkc, lumbal_punkc_2, id_lumbal_punkc, stop_lumbal_punkc )
	main_func(kt, kt_brushn_aorti_2, id_kt_brushn_aorti, stop_kt_brushn_aorti, kt_brushn_aorti_3)
	main_func(kt, kt_gortani_2, id_kt_gortani, stop_kt_gortani)
	main_func(kt, kt_grud_aorti_2, id_kt_grud_aorti, stop_kt_grud_aorti, kt_grud_aorti_3)
	main_func(kt, kt_kisti_ruki_2, id_kt_kisti_ruki, stop_kt_kisti_ruki)
	main_func(kt, kt_korona_calc_2, id_kt_korona_calc, stop_kt_korona_calc, kt_korona_calc_3)
	main_func(kt, kt_kostey_taza_2, id_kt_kostey_taza, stop_kt_kostey_taza, kt_kostey_taza_3)
	main_func(kt, kt_limfouzlov_2, id_kt_limfouzlov, stop_kt_limfouzlov, kt_limfouzlov_3)
	main_func(kt, kt_licevogo_skelet_2, id_kt_licevogo_skelet, stop_kt_licevogo_skelet, kt_llicevogo_skelet_3 )
	main_func(kt, kt_molo4nih_2, id_kt_molo4nih, stop_kt_molo4nih, kt_molo4nih_3)
	main_func(kt, kt_m9gk_shei_2, id_kt_m9gk_shei, stop_kt_m9gk_shei, kt_m9gk_shei_3)
	main_func(kt, kt_nosoglotki_2, id_kt_nosoglotki, stop_kt_nosoglotki)
	main_func(kt, kt_slun_jele_2, id_kt_slun_jele, stop_kt_slun_jele, kt_slun_jele_3)
	main_func(kt, kt_sredosten_2, id_kt_sredosten, stop_kt_sredosten)
	main_func(kt, kt_stopi_2, id_kt_stopi, stop_kt_stopi)
	main_func(kt, kt_shitovid_2, id_kt_shitovid, stop_kt_shitovid)
	main_func(kt, kt_kolonoskop_2, id_kt_kolonoskop, stop_kt_kolonoskop)
	main_func(kt, kt_brush_polo_2, id_kt_brush_polo, stop_kt_brush_polo, kt_brush_polo_3)
	main_func(kt, kt_jeludok_2, id_kt_jeludok, stop_kt_jeludok)
	main_func(kt, kt_jel4_puz_2, id_kt_kt_jel4_puz, stop_kt_jel4_puz, kt_jel4_puz_3)
	main_func(kt, kt_zabrush_2, id_kt_zabrush, stop_kt_zabrush)
	main_func(kt, kt_kishe4_2, id_kt_kishe4, stop_kt_kishe4)
	main_func(kt, kt_nadpo4e4_2, id_kt_nadpo4e4, stop_kt_nadpo4e4)
	main_func(kt, kt_pe4eni_2, id_kt_pe4eni, stop_kt_pe4eni)
	main_func(kt, kt_podjeludo4_2, id_kt_podjeludo4, stop_kt_podjeludo4)
	main_func(kt, kt_po4ek_2, id_kt_po4ek, stop_kt_po4ek)
	main_func(kt, kt_selezen_2, id_kt_selezen, stop_kt_selezen)
	main_func(kt, kt_golovi_2, id_kt_golovi, stop_kt_golovi)
	main_func(kt, kt_viso4nih_2, id_kt_viso4nih, stop_kt_viso4nih, kt_viso4nih_3)
	main_func(kt, kt_gipofiz_2, id_kt_gipofiz, stop_kt_gipofiz)
	main_func(kt, kt_glaza_2, id_kt_glaza, stop_kt_glaza)
	main_func(kt, kt_golov_mozga_2, id_kt_golov_mozga, stop_kt_golov_mozga)
	main_func(kt, kt_pazuh_2, id_kt_pazuh, stop_kt_pazuh)
	main_func(kt, kt_sedla_2, id_kt_sedla, stop_kt_sedla, kt_sedla_3)
	main_func(kt, kt_4elust_2, id_kt_4elust, stop_kt_4elust)
	main_func(kt, kt_4erep_2, id_kt_4erep, stop_kt_4erep)
	main_func(kt, kt_grudnoy_kletki_2, id_kt_grudnoy_kletki, stop_kt_grudnoy_kletki, kt_grudnoy_kletki_3)
	main_func(kt, kt_legkih_2, id_kt_legkih_2, stop_kt_legkih)
	main_func(kt, kt_serdca_2, id_kt_serdca_2, stop_kt_serdca)
	main_func(kt, kt_bronhoskop_2, id_kt_bronhoskop_2, stop_kt_bronhoskop)
	main_func(kt, kt_malogo_taza_2, id_kt_malogo_taza_2, stop_kt_malogo_taza, kt_malogo_taza_3)
	main_func(kt, kt_matki_2, id_kt_matki_2, stop_kt_matki)
	main_func(kt, kt_mo4evogo_2, id_kt_mo4evogo_2, stop_kt_mo4evogo)
	main_func(kt, kt_prostati_2, id_kt_prostati_2, stop_kt_prostati)
	main_func(kt, kt_9i4nik_2, id_kt_9i4nik_2, stop_kt_9i4nik)
	main_func(kt, kt_pozvono4_2, id_kt_pozvono4, stop_kt_pozvono4)
	main_func(kt, kt_grud_pozvono4_2, id_kt_grud_pozvono4_2, stop_kt_grud_pozvono4, kt_grud_pozvono4_3)
	main_func(kt, kt_kop4ik_2, id_kt_kop4ik_2, stop_kt_kop4ik)
	main_func(kt, kt_po9s_pozvon_2, id_kt_po9s_pozvon_2, stop_kt_po9s_pozvon, kt_po9s_pozvon_3)
	main_func(kt, kt_shei_pozvon_2, id_kt_shei_pozvon_2, stop_kt_shei_pozvon, kt_shei_pozvon_3)
	main_func(kt, kt_vn4s_2, id_kt_vn4s_2, stop_kt_vn4s, kt_vn4s_3)
	main_func(kt, kt_golenostop_2, id_kt_golenostop, stop_kt_golenostop)
	main_func(kt, kt_kolena_2, id_kt_kolena, stop_kt_kolena)
	main_func(kt, kt_lokt9_2, id_kt_lokt9, stop_kt_lokt9)
	main_func(kt, kt_ple4_2, id_kt_ple4, stop_kt_ple4)
	main_func(kt, kt_tazobedr_2, id_kt_tazobedr, stop_kt_tazobedr)
	main_func(duktrograf, duktrograf_2, id_kt_druktograf, stop_kt_duktograf)
	main_func(obzor_mammograf, obzor_mammograf_2, id_kt_obzor_mammograf, stop_obzor_mammograf)
	main_func(pricel_mammograf, pricel_mammograf_2, id_pricel_mammograf, stop_pricel_mammograf)
	main_func(kt, kt_lu4ezap_2, id_kt_lu4ezap, stop_kt_lu4ezap)
	main_func(kt, kt_brush_polo_2_k, id_kt_brush_polo_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_gortani_2_k, id_kt_gortani_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_viskov_2_k, id_kt_viskov_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_gipofiz_2_k, id_kt_gipofiz_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_glaza_2_k, id_kt_glaza_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_mozg_2_k, id_kt_mozg_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_golova_2_k, id_kt_golova_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_grud_pozvon2_k, id_kt_grud_pozvon_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_grud_aorti_k, id_kt_grud_aorti_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_grud_kletki_k, id_kt_grud_kletki_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_zabrush_k, id_kt_zabrush_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_kishe4_k, id_kt_kishe4_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_kostey_taza_k, id_kt_kostey_taza_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_legkih_k, id_kt_legkih_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_molo4nih_2_k, id_kt_molo4nih_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_mo4evoy_2_k, id_kt_mo4evoy_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_shei_2_k, id_kt_shei_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_nadpo4e4_2_k, id_kt_nadpo4e4_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_nosoglot_2_k, id_kt_nosoglot_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_organov_malogo_taza_2_k, id_kt_organov_malogo_taza_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_pe4eni_2_k, id_kt_pe4eni_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_po4ek_2_k, id_kt_po4ek_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_po9s_pozvon_2_k, id_kt_po9s_pozvon_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_pridat_pazuh_2_k, id_kt_pridat_pazuh_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_serdca_2_k, id_kt_serdca_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_slunnih_2_k, id_kt_slunnih_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_stopi_2_k, id_kt_stopi_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_4erepa_2_k, id_kt_4erepa_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_sheynogo_pozvon_2_k, id_kt_sheynogo_pozvon_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_shitovid_2_k, id_kt_shitovid_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(ren, ren_bedra, id_ren_bedra, stop_ren_bedra)
	main_func(ren, ren_viso4, id_ren_viso4, stop_ren_viso4)
	main_func(ren, ren_gortan, id_ren_gortan, stop_ren_gortan)
	main_func(ren, ren_grudini, id_ren_grudini, stop_ren_grudini)
	main_func(ren, ren_kisti, id_ren_kisti, stop_ren_kisti)
	main_func(ren, ren_klu4ic, id_ren_kLu4ic, stop_ren_klu4ic)
	main_func(ren, ren_kostey_nosa, id_ren_kostey_nosa, stop_ren_koste_nosa, ren_kostey_nosa_2)
	main_func(ren, ren_kost_taz, id_ren_kost_taz, stop_ren_kost_taz, ren_kost_taza_2)
	main_func(ren, ren_legkih, id_ren_legkih, stop_ren_legkih, ren_legkih_2)
	main_func(ren, ren_lopatki, id_ren_lopatki, stop_ren_lopatki)
	main_func(ren, ren_nadkolennik, id_ren_nadkolennik, stop_ren_nadkolennik)
	main_func(ren, ren_nosoglotki, id_ren_nosoglotki, stop_ren_nosoglotki)
	main_func(ren, ren_palca, id_ren_palca, stop_ren_palca)
	main_func(ren, ren_predple4, id_ren_predple4, stop_ren_predple4)
	main_func(ren, ren_pazuh_nos, id_ren_pazuh_nos, stop_ren_pazuh_nos, ren_pazuh_nos_2)
	main_func(ren, ren_p9t_kost, id_ren_p9t_kost, stop_ren_p9t_kost, ren_p9t_kost_2)
	main_func(ren, ren_reber, id_ren_reber, stop_ren_reber)
	main_func(ren, ren_skul_kost, id_ren_skul_kost, stop_ren_skul, ren_skul_kost_2)
	main_func(ren, ren_stopi, id_ren_stopi, stop_ren_stopi)
	main_func(ren, ren_stopi, id_ren_stopi, stop_ren_stopi)
	main_func(ren, ren_stopi_nagruz, id_ren_stopi_nagruz, stop_ren_stopi_nagruz, ren_stopi_nagruz_2)
	main_func(ren, ren_taza, id_ren_taza, stop_ren_taza)
	main_func(ren, ren_tur_sed, id_ren_tur_sed, stop_ren_tur_sed, ren_tur_sed_2)
	main_func(ren, ren_4erep, id_ren_4erep, stop_ren_4erep)
	main_func(ren, ren_brushnoy, id_ren_brushnoy, stop_ren_brushnoy, ren_brushnoy_2 )
	main_func(ren, ren_jeludka, id_ren_jeludka, stop_ren_jeludka )
	main_func(ren, ren_pishevod, id_ren_pishevod, stop_ren_pishevod)
	main_func(ren, ren_po4ek, id_ren_po4ek, stop_ren_po4ek)
	main_func(ren, ren_pozvono4_grud, id_ren_pozvono4_grud, stop_ren_pozvono4_grud, ren_pozvono4_grud_2)
	main_func(ren, ren_pozvono4_po9s_krest, id_ren_pozvono4_po9s_krest, stop_ren_pozvono4_po9s_krest, ren_pozvono4_po9s_krest_2)
	main_func(ren, ren_pozvono4_po9s, id_ren_pozvono4_po9s, stop_ren_pozvono4_po9s )
	main_func(ren, ren_pozvono4_shei, id_ren_pozvono4_shei, stop_ren_pozvono4_shei)
	main_func(ren, ren_kop4ik, id_ren_kop4ik, stop_ren_kop4ik)
	main_func(ureteropiel, ureteropiel_2, id_ureteropiel, stop_ureteropiel)
	main_func(fistulograf, fistulograf_2, id_fistulograf, stop_fistulograf)
	main_func(holangiograf, holangiograf_2, id_holangiograf, stop_holangiograf)
	main_func(laterograf_jel, laterograf_jel_2, id_laterograf_jel, stop_laterograf_jel)
	main_func(sialograf, sialograf_2, id_sialograf, stop_sialograf)
	main_func(cistograf, cistograf_2, id_cistograf, stop_cistograf)
	main_func(ren, ren_brush_reb, id_ren_brush_reb, stop_ren_brush_reb, rebenok)
	main_func(ren, ren_viso4_kost_reb, id_ren_viso4_kost_reb, stop_ren_viso4_kost_reb, rebenok)
	main_func(ren, ren_grud_pozvon_reb, id_ren_grud_pozvon_reb, stop_ren_grud_pozvon_reb, rebenok)
	main_func(ren, ren_po9s_pozvon_reb, id_ren_po9s_pozvon_reb, stop_ren_po9s_pozvon_reb, rebenok)
	main_func(ren, ren_shei_pozvon_reb, id_ren_shei_pozvon_reb, stop_ren_shei_pozvon_reb, rebenok)
	main_func(ren, ren_jelud_reb, id_ren_jelud_reb, stop_ren_jelud_reb, rebenok)
	main_func(ren, ren_kisti_reb, id_ren_kisti_reb, stop_ren_kisti_reb, rebenok)
	main_func(ren, ren_kost_taz_reb, id_ren_kost_taz_reb, stop_ren_kost_taz_reb, rebenok)
	main_func(ren, ren_legkih_reb, id_ren_legkih_reb, stop_ren_legkih_reb, rebenok)
	main_func(ren, ren_nosoglot_reb, id_ren_nosoglot_reb, stop_ren_nosoglot_reb, rebenok)
	main_func(ren, ren_orbit_reb, id_ren_orbit_reb, stop_ren_orbit_reb, rebenok)
	main_func(ren, ren_pishevod_reb, id_ren_pishevod_reb, stop_ren_pishevod_reb, rebenok)
	main_func(ren, ren_pridat_reb, id_ren_pridat_reb, stop_ren_pridat_reb, rebenok)
	main_func(ren, ren_stopi_reb, id_ren_stopi_reb, stop_ren_stopi_reb, rebenok)
	main_func(ren, ren_tazobed_reb, id_ren_tazobed_reb, stop_ren_tazobed_reb, rebenok)
	main_func(ren, ren_ton_kish_reb, id_ren_ton_kish_reb, stop_ren_ton_kish_reb, rebenok)
	main_func(ren, ren_trub_kost_reb, id_ren_trub_kost_reb, stop_ren_trub_kost, rebenok)
	main_func(ren, ren_turec_reb, id_ren_turec_reb, stop_ren_turec_kost, rebenok)
	main_func(ren, ren_4elust_reb, id_ren_4elust_reb, stop_ren_4elust_kost , rebenok)
	main_func(ren, ren_4erepa_reb, id_ren_4erepa_reb, stop_ren_4erepa_kost, rebenok)
	main_func(ren, ren_golenostop, id_ren_ren_golenostop, stop_ren_ren_golenostop)
	main_func(ren, ren_kolena, id_ren_ren_kolena, stop_ren_ren_kolena)
	main_func(ren, ren_krest_podvzdo, id_ren_krest_podvzdo, stop_ren_krest_podvzdo, ren_krest_podvzdo_2)
	main_func(ren, ren_lokt9, id_ren_lokt9, stop_ren_lokt9)
	main_func(ren, ren_lu4ezap, id_ren_lu4ezap, stop_ren_lu4ezap)
	main_func(ren, ren_ple4a, id_ren_ple4a, stop_ren_ple4a)
	main_func(ren, ren_tazobedr, id_ren_tazobedr, stop_ren_tazobedr )
	main_func(duodenograf, duodenograf_2, id_duodenograf, stop_duodenograf)
	main_func(mrt, mrt_vse_telo, id_mrt_vse_telo, stop_mrt_vse_telo)
	main_func(mrt, mrt_gortan, id_mrt_gortan, stop_mrt_gortan)
	main_func(mrt, mrt_kist, id_mrt_kist, stop_mrt_kist)
	main_func(mrt, mrt_kraniover, id_mrt_kraniover, stop_mrt_kraniover)
	main_func(mrt, mrt_limfouzl, id_mrt_limfouzl, stop_mrt_limfouzl)
	main_func(mrt, mrt_molo4, id_mrt_molo4, stop_mrt_molo4)
	main_func(mrt, mrt_moshonki, id_mrt_moshonki, stop_mrt_moshonki)
	main_func(mrt, mrt_m9gk_tkan, id_mrt_m9gk_tkan, stop_mrt_m9gk_tkan)
	main_func(mrt, mrt_shei, id_mrt_shei, stop_mrt_shei)
	main_func(mrt, mrt_ple4sple, id_mrt_ple4sple, stop_mrt_ple4sple)
	main_func(mrt, mrt_plod, id_mrt_plod, stop_mrt_plod)
	main_func(mrt, mrt_4len, id_mrt_4len, stop_mrt_4len)
	main_func(mrt, mrt_slun, id_mrt_slun, stop_mrt_slun)
	main_func(mrt, mrt_spinnogo, id_mrt_spinnogo, stop_mrt_spinnogo)
	main_func(mrt, mrt_sredosten, id_mrt_sredosten, stop_mrt_sredoten)
	main_func(mrt, mrt_stopi, id_mrt_stopi, stop_mrt_stopi)
	main_func(mrt, mrt_shitovid, id_mrt_shitovid, stop_mrt_shitovid)
	main_func(mrt, mrt_brush, id_mrt_brush, stop_mrt_brush)
	main_func(mrt, mrt_jelud, id_mrt_jelud, stop_mrt_jelud)
	main_func(mrt, mrt_zabrush, id_mrt_zabrush, stop_mrt_zabrush)
	main_func(mrt, mrt_nadpo4e4, id_mrt_nadpo4e4, stop_mrt_nadpo4e4)
	main_func(mrt, mrt_pe4en, id_mrt_pe4en, stop_mrt_pe4en)
	main_func(mrt, mrt_pishevod, id_mrt_pishevod, stop_mrt_pishevod)
	main_func(mrt, mrt_podjel, id_mrt_podjel, stop_mrt_podjel)
	main_func(mrt, mrt_po4ek, id_mrt_po4ek, stop_mrt_po4ek)
	main_func(mrt, mrt_golova, id_mrt_golova, stop_mrt_golova)
	main_func(mrt, mrt_spektroskop, id_mrt_spektroskop, stop_mrt_spektroskop)
	main_func(mrt, mrt_bn4s, id_mrt_vn4s, stop_mrt_vn4s)
	main_func(mrt, mrt_gipofiz, id_mrt_gipofiz, stop_mrt_gipofiz)
	main_func(mrt, mrt_gippokmap, id_mrt_gippo, stop_mrt_gippo)
	main_func(mrt, mrt_glaz, id_mrt_glaz, stop_mrt_glaz)
	main_func(mrt, mrt_golovnogo, id_mrt_golovnogo, stop_mrt_golovnogo)
	main_func(mrt, mrt_licnerv, id_mrt_licnerv, stop_mrt_licnerv)
	main_func(mrt, mrt_pridatpazuh, id_mrt_pridatpazuh, stop_mrt_pridatpazuh)
	main_func(mrt, mrt_uha, id_mrt_uha, stop_mrt_uha)
	main_func(mrt, mrt_troynerv, id_mrt_troynerv, stop_mrt_troynerv)
	main_func(mrt, mrt_sedlo, id_mrt_sedlo, stop_mrt_sedlo)
	main_func(mrt, mrt_4erep, id_mrt_4erep, stop_mrt_4erep)
	main_func(mrt, mrt_4erepnerv, id_mrt_4erepnerv, stop_mrt_4erepnerv)
	main_func(mrt, mrt_grudkletka, id_mrt_grudkletka, stop_mrt_grudkletka)
	main_func(mrt, mrt_legkih, id_mrt_legkih, stop_mrt_legkih)
	main_func(mrt, mrt_serdca, id_mrt_serdca, stop_mrt_serdca)
	main_func(mrt, mrt_holangio, id_mrt_holangio, stop_mrt_holangio )
	main_func(mrt, mrt_enterograf, id_mrt_enterograf, stop_mrt_enterograf)
	main_func(mrt, mrt_pozvon, id_mrt_pozvon, stop_mrt_pozvon)
	main_func(mrt, mrt_bedrkost, id_mrt_bedrkost, stop_mrt_bedrkost)
	main_func(mrt, mrt_viskost, id_mrt_viskost, stop_mrt_viskost)
	main_func(mrt, mrt_kosttaza, id_mrt_kosttaza, stop_mrt_kosttaza)
	main_func(mrt, mrt_taz, id_mrt_taz, stop_mrt_taz)
	main_func(mrt, mrt_tazwomen, id_mrt_tazwomen, stop_mrt_tazwomen, mrt_tazwomen_2)
	main_func(mrt, mrt_tazmen, id_mrt_tazmen, stop_mrt_tazmen, mrt_tazmen_2)
	main_func(mrt, mrt_matki, id_mrt_matki, stop_mrt_matki)
	main_func(mrt, mrt_mo4evoy, id_mrt_mo4evoy, stop_mrt_mo4evoy)
	main_func(mrt, mrt_prostati, id_mrt_prostati, stop_mrt_prostati)
	main_func(mrt, mrt_9i4nik, id_mrt_9i4nik, stop_mrt_9i4nik)
	main_func(mrt, mrt_golovnogo_open, id_mrt_golovnogo_open, stop_mrt_golovnogo_open, mrt_golovnogo_open_2)
	main_func(mrt, mrt_kolen_open, id_mrt_kolen_open, stop_mrt_kolen_open, mrt_kolen_open_2)
	main_func(mrt, mrt_pozvon_open, id_mrt_pozvon_open, stop_mrt_pozvon_open, mrt_pozvon_open_2)
	main_func(mrt, mrt_grudpozvon, id_mrt_grudpozvon, stop_mrt_grudpozvon, mrt_grudpozvon_2)
	main_func(mrt, mrt_kop4ik, id_mrt_kop4ik, stop_mrt_kop4ik)
	main_func(mrt, mrt_krestpodvz, id_mrt_krestpodzv, stop_mrt_krestpodzv, mrt_krestpodzv_2)
	main_func(mrt, mrt_po9spozvon, id_mrt_po9spozvon, stop_mrt_po9spozvon, mrt_po9spozvon_2)
	main_func(mrt, mrt_sheinogo, id_mrt_sheinogo, stop_mrt_sheinogo, mrt_sheinogo_2)
	main_func(mrt, mrt_golenostop, id_mrt_golenostop, stop_mrt_golenostop)
	main_func(mrt, mrt_kolennogo, id_mrt_kolennogo, stop_mrt_kolennogo)
	main_func(mrt, mrt_lokt9, id_mrt_lokt9, stop_mrt_lokt9)
	main_func(mrt, mrt_lu4ezap, id_mrt_lu4ezap, stop_mrt_lu4ezap)
	main_func(mrt, mrt_ple4evogo, id_mrt_ple4evogo, stop_mrt_ple4evogo)
	main_func(mrt, mrt_tazobedr, id_mrt_tazobedr, stop_mrt_razobedr)
	main_func(mrt, mr_angiograf_mozga_reb, id_mr_angiograf_mozga_reb, stop_mr_angiograf_mozga_reb, rebenok)
	main_func(mrt, mr_brushpol_reb, id_mrt_brupshpol_reb, stop_mrt_brushpol_reb, rebenok)
	main_func(mrt, mr_gipofiz_reb, id_mr_gipofiz_reb, stop_mr_gipofiz_reb, rebenok)
	main_func(mrt, mr_golovnogo_reb, id_mr_golovnogo_reb, stop_mr_golovnogo_reb, rebenok)
	main_func(mrt, mr_grudklet_reb, id_mr_grudklet_reb, stop_mr_grudklet_reb, rebenok)
	main_func(mrt, mr_kolena_reb, id_mr_kolena_reb, stop_mr_kolena_reb, rebenok)
	main_func(mrt, mr_malogo_taza_reb, id_mr_malogotaza_reb, stop_mr_maliyyaz_reb, rebenok)
	main_func(mrt, mr_nadpo4e4nik_reb, id_mr_nadpo4e4nik_reb, stop_mr_nadpo4e4nik_reb, rebenok)
	main_func(mrt, mr_pazuh_reb, id_mr_pazuh_reb, stop_mr_pazuh_reb, rebenok)
	main_func(mrt, mr_po4ek_reb, id_mr_po4ek_reb, stop_mr_po4ek_reb, rebenok)
	main_func(mrt, mr_serdca_reb, id_mr_serdca_reb, stop_mr_serdca_reb, rebenok)
	main_func(mrt, mr_tazobedr_reb, id_mr_tazobedr_reb, stop_mr_tazobedr_reb, rebenok)
	main_func(mrt, mr_ple4esusk, id_ple4esusk, stop_mr_ple4esusk, kt_for_kontrast)
	main_func(mrt, mr_prostatik, id_prostatik, stop_mr_prostatik, kt_for_kontrast)
	main_func(mrt, mr_brushpolk, id_brushpolk, stop_mr_brushpolk, kt_for_kontrast)
	main_func(mrt, mr_vn4sk, id_vn4sk, stop_mr_vn4sk, kt_for_kontrast)
	main_func(mrt, mr_gipofizk, id_gipofizk, stop_mr_gipofizk, kt_for_kontrast)
	main_func(mrt, mr_glazk, id_glazk, stop_mr_glazk, kt_for_kontrast)
	main_func(mrt, mr_golenostopk, id_golenostopk, stop_mr_golenostopk, kt_for_kontrast)
	main_func(mrt, mr_golovnoyk, id_golovnoyk, stop_mr_golovnoyk, kt_for_kontrast)
	main_func(mrt, mr_golovik, id_golovik, stop_mr_golovik, kt_for_kontrast)
	main_func(mrt, mr_grudpozvon, id_mrt_grudpozvonk, stop_mrt_grudpozvonk, kt_for_kontrast)
	main_func(mrt, mr_zabrushk, id_zabrushk, stop_mr_zabrushk, kt_for_kontrast)
	main_func(mrt, mr_kistkik, id_kistkik, stop_mr_kistkik, kt_for_kontrast)
	main_func(mrt, mr_kolenak, id_kolenak, stop_mr_kolenak, kt_for_kontrast)
	main_func(mrt, mr_krestpodvk, id_krestpodvk, stop_mr_krestpodvk, kt_for_kontrast)
	main_func(mrt, mr_maliytazk, id_maliytazk, stop_mr_maliytazk, kt_for_kontrast)
	main_func(mrt, mr_molo4k, id_molo4k, stop_mr_molo4k, kt_for_kontrast)
	main_func(mrt, mr_mo4evoyk, id_mo4evoyk, stop_mr_mo4evoyk, kt_for_kontrast)
	main_func(mrt, mr_moshonkik, id_moshonkik, stop_mr_moshonkik, kt_for_kontrast)
	main_func(mrt, mr_m9gkihk, id_m9gkihk, stop_mr_m9gkihk, kt_for_kontrast)
	main_func(mrt, mr_m9gksheik, id_m9gksheik, stop_mr_m9gksheik, kt_for_kontrast)
	main_func(mrt, mr_nadpo4e4k, id_nadpo4e4k, stop_mr_nadpo4e4k, kt_for_kontrast)
	main_func(mrt, mr_pe4enik, id_pe4enik, stop_mr_pe4enik, kt_for_kontrast)
	main_func(mrt, mr_pozvonkon, id_pozvonkon, stop_mr_pozvonkon, kt_for_kontrast)
	main_func(mrt, mr_4lenk, id_4lenk, stop_mr_4lenk, kt_for_kontrast)
	main_func(mrt, mr_po4ekk, id_po4ekk, stop_mr_po4ekk, kt_for_kontrast)
	main_func(mrt, mr_po9skk, id_po9skk, stop_mr_po9skk, kt_for_kontrast)
	main_func(mrt, mr_pridatkk, id_pridatkk, stop_mr_pridatkk, kt_for_kontrast)
	main_func(mrt, mr_serdcak, id_sercak, stop_mr_sercak, kt_for_kontrast)
	main_func(mrt, mr_spinnoyk, id_spinnoyk, stop_mr_spinnoyk, kt_for_kontrast)
	main_func(mrt, mr_stopik, id_stopik, stop_mr_stopik, kt_for_kontrast)
	main_func(mrt, mr_sustavk, id_sustavk, stop_mr_sustavk, kt_for_kontrast)
	main_func(mrt, mr_tazobedk, id_tazobedk, stop_mr_tazobedk, kt_for_kontrast)
	main_func(mrt, mr_4erepk, id_4erepk, stop_mr_4erepk, kt_for_kontrast)
	main_func(mrt, mr_sheynogok, id_sheynogok, stop_mr_sheynogok, kt_for_kontrast)
	main_func(anorek_manomentr, anorek_manomentr_2, id_anorek_manonemtr, stop_anorek_manometr)
	main_func(keratometr, keratometr_2, id_keratometr, stop_keratometr)
	main_func(komp_kerato, kompkerato_2, id_kompkerato, stop_kompkerato)
	main_func(pahimetr, pahimetr_2, id_pahimetr, stop_pahimetr)
	main_func(uz_biometr_glaz, uz_biometr_glaz_2, id_uz_biometr_glaz, stop_uz_biomtr_glaz)
	main_func(bio_trofekto, biotrofekto_2, id_biotrofekto, stop_biotrofekto)
	main_func(bio_horion, biohorion_2, id_biohorion, stop_biohorion)
	main_func(bio_9zik, bio9zik_2, id_bio9zik, stop_bio9zik)
	main_func(bio_shl9pa, bioshl9pa_2, id_bioshl9pa, stop_bioshl9pa)
	main_func(vidurodin, vidurodin_2, id_vidurodin, stop_vidurodin)
	main_func(viz_pot, viz_pot_2, id_vizpot, stop_vizpot)
	main_func(densitometr, densitometr_2, id_densitometr, stop_densitometr)
	main_func(dermatoskop, dermatoskop_2, id_dermatoskop, stop_dermatoskop)
	main_func(dopplerometr, dopplerometr_2, id_dopplerometr, stop_dopplerometr)
	main_func(invazprenatal, invazprenatal_2, id_invazprenatal, stop_invazprenatal)
	main_func(amniotik, amniotik_2, id_amniotik, stop_amniotik)
	main_func(fono_plod, fonoplod_2, id_fonoplod, stop_fonoplod)
	main_func(ktg, ktg_2, id_ktg, stop_ktg)
	main_func(ninvazprenatal, ninvazprenatal_2, id_ninvazprenatal, stop_ninvazprenatal)
	main_func(kt, kt_pazuhnosr_2, id_kt_pazuhnos, stop_kt_pazuhnosr)
	main_func(kt, kt_krest_2, id_kt_krest, stop_kt_krest)
	main_func(kt, kt_brushpolr_2, id_kt_brushpolr, stop_kt_brushpolr, rebenok)
	main_func(kt, kt_viskosr_2, id_kt_viskosr, stop_kt_viskosr, rebenok)
	main_func(kt, kt_golovnoyr_2, id_kt_golovnoyr, stop_kt_golovnoyr, rebenok)
	main_func(kt, kt_grudkletr_2, id_kt_grudkletr, stop_kt_grudkletr, rebenok)
	main_func(kt, kt_po4ekr_2, id_kt_po4ekr, stop_kt_po4ekr, rebenok)
	main_func(mrt, mr_pozvon_reb, id_mr_pozvon_reb, stop_mr_pozvon_reb, rebenok)
	main_func(mrt, mr_shei_reb, id_mr_shei_reb, stop_mr_shei_reb, rebenok)
	main_func(kt, kt_po4ekr_2_k, id_kt_po4ekr_k, kt_kontrast_stop_reb, kt_for_kontrast)
	main_func(kt, kt_brushaor_2_k, id_kt_brushaor_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(kt, kt_vn4s_2_k, id_kt_vn4s_k, kt_kontrast_stop, kt_for_kontrast)
	main_func(mrt, mrt_cisterno, id_mrt_cisterno, stop_mrt_cisterno)
	main_func(mskt, msktbrushpol_2, id_msktrbrushpol, stop_msktbrushpol)
	main_func(mskt, msktgolovnoy_2, id_golovnoy, stop_golovnoy)
	main_func(mskt, msktgrudklet_2, id_msktgrudklet, stop_msktgrudklet)
	main_func(mskt, msktkorcalc_2, id_msktcorcalc, stop_msktcorcalc)
	main_func(mskt, msktlegkie_2, id_msktlegkie, stop_msktlegkie)
	main_func(mskt, msktnadpo4e4_2, id_msktnadpo4e4, stop_msktnadpo4e4)
	main_func(mskt, msktnosoglot_2, id_msktnosoglot, stop_msktnosoglot)
	main_func(mskt, msktperfuz_2, id_msktperfuz, stop_msktperfuz)
	main_func(mskt, msktpozvon_2, id_msktpozvon, stop_pozvon)
	main_func(mskt, msktpo4ek_2, id_msktpo4ek, stop_msktpo4ek)
	main_func(mskt, msktpazuh_2, id_msktpazuh, stop_msktpazuh)
	main_func(mskt, msktserca_2, id_msktserdca, stop_msktserdca)
	main_func(mskt, msktsustav_2, id_msktsustav, stop_msktsustav)
	main_func(mskt, msktcistouretrograf_2, id_msktcistouretrograf, stop_msktcistouretrograf)
	main_func(mskt, msktcherep_2, id_msktcherep, stop_msktcherep)
	main_func(polplodamater, polplodamater_2, id_polplodamater, stop_polplodamater)
	main_func(rezusmater, resuzmater_2, id_resuzmater, stop_resuzmater, rezusmater_3)
	main_func(oktdiska, oktdiska_2, id_oktdiska, stop_oktdiska)
	main_func(oktmakuli, oktmakuli_2, id_oktmakuli, stop_oktmakuli)
	main_func(oktrogovici, oktrogovici_2, id_oktrogovoci, stop_oktrogovici)
	main_func(oktset, oktset_2, id_oktset, stop_oktset)
	main_func(diafanglaz, diafanglaz_2, id_diafaglaz, stop_diafalglaz)
	main_func(potencizri, potenczri_2, id_potenczri, stop_potenczri)
	main_func(kompkeratotop, kompkeratotop_2, id_kompkeratotop, stop_kompkeratotop)
	main_func(mikrorog, mikrorog_2, id_mikrorog, stop_mikrorog)
	main_func(k4sm, k4sm_2, id_k4sm, stop_k4sm)
	main_func(oftalmoskopop, oftalmoskop_2, id_oftalmoskop, stop_oftalmoskop)
	main_func(suttono, suttono_2, id_suttono, stop_suttono)
	main_func(fotoglaz, fotoglaz_2, id_fotoglaz, stop_fotoglaz)
	main_func(tonoglaz, tonoglaz_2, id_tonoglaz, stop_tonoglaz)
	main_func(polisomno, polisomno_2, id_polisomno, stop_polisomno)
	main_func(ventri, ventri_2, id_ventri, stop_ventri)
	main_func(dakrio, dakrio_2, id_dakrio, stop_dakrio)
	main_func(irrigo, irrigo_2, id_irrigo, stop_irrigo)
	main_func(kavegraf, kavergraf_2, id_kavergraf, stop_kavergraf)
	main_func(mielograf, mielograf_2, id_mielograf, stop_mieograf)
	main_func(rinomano, rinomano_2, id_rinomano, stop_rinomano)
	main_func(tride, tride_2, id_tride, stop_tride)
	main_func(chde, chde_2, id_chde, stop_chde)
	main_func(uzipola, uzipola_2, id_uzipola, stop_uzipola)
	main_func(yzrubc, uzrubc_2, id_uzrubc, stop_uzrubc, uzrubc_3)
	main_func(termografi, termografi_2, id_termografi, stop_termografi)
	main_func(scintrigraf, osteoscin, id_osteoscin, stop_osteoscin)
	main_func(scintrigraf, renosci_2, id_renosci, stop_renosci)
	main_func(scintrigraf, statisci_2, id_statisci, stop_statisci)
	main_func(scintrigraf, scitelo_2, id_scitelo, stop_scitelo)
	main_func(scintrigraf, scikost_2, id_scikost, stop_scikost)
	main_func(scintrigraf, scileg_2, id_scileg, stop_scileg)
	main_func(scintrigraf, scimiokard_2, id_scimiokard, stop_scimiokard)
	main_func(scintrigraf, sciparashit_2, id_sciparashit, stop_sciparashit)
	main_func(scintrigraf, scipe4_2, id_scipe4, stop_scipe4)
	main_func(scintrigraf, scishit_2, id_scishit, stop_scishit)
	main_func(scintrigraf, flebosci_2, id_flebosci, stop_flebosci)
	main_func(scintrigraf, holesci_2, id_holesci, stop_holesci)
	main_func(scintrigraf, scipo4reb_2, id_scipo4reb, stop_scipo4reb)
	main_func(scintrigraf, dinscin_2, id_discin, stop_discin)
	main_func(scintrigraf, scilim_2, id_scilim, stop_scilim)
	main_func(uzperner, uzner_2, id_uzner, stop_uzner)
	main_func(voseretro, vosuretro_2, id_vosuretro, stop_vosuretro)
	main_func(retrouretro, retrouretro_2, id_retruoretro, stop_retrouretro)
	main_func(urograf, urograf_2, id_urograf, stop_ureograf)
	main_func(mrurograf, mrurograf_2, id_mrurograf, stop_mrurograf)
	main_func(obzouro, obzouro_2, id_obzouro, stop_obzouro)
	main_func(uroreb, ureoreb_2, id_uroreb, stop_uroreb)
	main_func(eskuro, eskuro_2, id_eksyro, stop_eksuro)
	main_func(fertiloskop, fertiloskop_2, id_fertiloskop, stop_fertiskolop)
	main_func(fleboskon, flebokon_2, id_flebokon, stop_flebokon)
	main_func(tazozfle, tazofle_2, id_tazofle, stop_tazofle)
	main_func(cheroksi, cheroksi_2, id_cheroksi, stop_cheroksi)
	main_func(fluoro, fluoro_2, id_fluoro, stop_fluoro)
	main_func(aktis, aktis_2, id_aktis, stop_aktis)
	main_func(akurino, akurino_2, id_akurino, stop_akurino)
	main_func(vesti, vesti_2, id_vesti, stop_vesti)
	main_func(phjel, phjel_2, id_phjel, stop_phjel, phjel_3)
	main_func(phpi, phpi_2, id_phpi, stop_phpi, phpi_3)
	main_func(ishetes, ishetest_2, id_ishetest, stop_ishetest)
	main_func(karesmon, karesmon_2, id_karesmon, stop_karesmon)
	main_func(limfograf, limfograf_2, id_limfograf, stop_limfograf)
	main_func(otolito, otolito_2, id_otolito, stop_otolito)
	main_func(olfaktome, olfaktome_2, id_olfaktome, stop_olfaktome)
	main_func(prohosluh, proosluh_2, id_prohosluh, stop_prohosluh)
	main_func(ortoporb, ortoprob_2, id_ortoprob, stop_ortoprob)
	main_func(pikflow, pikflow_2, id_pikflow, stop_pikflow)
	main_func(plantograf, plantograf_2, id_plantograf, stop_plantograf)
	main_func(platismograf, platismograf_2, id_platismograf, stop_platismograf)
	main_func(profilouretri, profilouretri_2, id_profilouretri, stop_profilouretri)
	main_func(pulsoksi, pulsoksi_2, id_pulsoksi, stop_pulsoksi)
	main_func(rvg, rvg_2, id_rvg, stop_rvg)
	main_func(stabilograf, stabilograf_2, id_stabilograf, stop_stabilograf)
	main_func(stroboskop, stroboskoop_2, id_stroboskop, stop_stroboskop)



	for_check_troubles(id_uzi_shitovid, uzi, uzi_shitovid, id_uzi_shitovid, stop_for_uzi_shitovid_2, uzi_shitovid_2)
	for_check_troubles(id_uzi_parashitovidnih, uzi, uzi_parashitovidnih_2, id_uzi_parashitovidnih, stop_for_uzi_parashitovidnih_2, uzi_parashitovidnih_3)
	for_check_troubles(id_uzi_jel4nogo, uzi, uzi_jel4nogo, id_uzi_jel4nogo, stop_for_uzi_jel4nogo_2, uzi_jel4nogo_2)
	for_check_troubles(id_uzi_m9gkih_tkaney, uzi, uzi_m9gkih_tkaney, id_uzi_m9gkih_tkaney, stop_for_uzi_m9gkih_tkaney_2, uzi_m9gkih_tkaney_2)

	# print(array_out_of_service)    
	wb.save('__после скрипта__' + str(filename_strip))

###############################################################################################################################################################




#####################################################ФУНКЦИЯ ПОИСКА ЛАБОРАТОРКИ#################################################################################
def lab_finder():
	wb = openpyxl.load_workbook('__после скрипта__' + str(filename_strip))
	wb.active = 0
	sheet_0 = wb.active
	rows, columns = sheet_0.max_row, sheet_0.max_column

	horionic_gonadotropin, id_horionic_gonadotropin = ["гонадотропин", "хгч"], ["Анализ на ХГЧ"]
	vitamin = ["витамин"] 
	b_12, id_b_12 = ["в12", "в-12", "в 12", "b12", "b-12", "b 12"], ["Анализ на витамин B12"]   
	svobodniy = ["свободн"]


	comma, ints = [","], ["Ц", "ц", "Ф", "ф" "А", "а", "Б", "б", "В", "в", "Г", "г", "Д", "д", "е", "Е", "Ж", "ж", "З", "з", "И", "и", "К", "к", "Л", "л", "М", "м", "Н", "н", "о", "О", "П", "п", "Р", "р", "С", "с", "Т", "т", "у", "У"]
	yo = "ё"
	n = " "
	counter = 0
	for i in range(1, rows+1):
		cell_value_with_comma = str(sheet_0.cell(row = i, column = 3).value)
		if cell_value_with_comma.isprintable() is False:
			cell_value_with_comma = cell_value_with_comma.replace("\xa0", "")
			sheet_0.cell(row = i, column = 3).value = cell_value_with_comma
	
	for i in range(1, rows+1):
		cell_value_with_comma = str(sheet_0.cell(row = i, column = 3).value)
		if cell_value_with_comma.isprintable() is False:
			print("ppc", i)
			cell_value_with_comma = cell_value_with_comma.replace("\xa0", "")
			sheet_0.cell(row = i, column = 3).value = cell_value_with_comma
	
	for i in range(1, rows+1):
		cell_value_with_comma = str(sheet_0.cell(row = i, column = 3).value) 
		cell_with_upper = str(sheet_0.cell(row = i, column = 2).value)
		if sheet_0.cell(row = i, column = 2).value != None:
			cell_without_upper = cell_with_upper.lower()
			sheet_0.cell(row = i, column = 2).value = cell_without_upper
			if yo in sheet_0.cell(row = i, column = 2).value:
				sheet_0.cell(row = i, column = 2).value = sheet_0.cell(row = i, column = 2).value.replace('ё','е')
		if comma[0] in cell_value_with_comma:
			stop_int = re.search('|'.join(ints), cell_value_with_comma) != None
			if stop_int == False:
				q = sheet_0.cell(row = i, column = 3).value.split()
				if len(q) > 1:
					g = str(q[0] + str(q[1]))
					sheet_0.cell(row = i, column = 3).value = g
				if n in sheet_0.cell(row = i, column = 3).value:
					b = str(sheet_0.cell(row = i, column = 3).value)
					d = b.replace(" ", "")
					c = d.replace(",", ".")
					sheet_0.cell(row = i, column = 3).value = int(float((c)))                           
				else:
					e = str(sheet_0.cell(row = i, column = 3).value)
					f = e.replace(",", ".")
					sheet_0.cell(row = i, column = 3).value = int(float((f)))
		elif sheet_0.cell(row = i, column = 3).value != None and sheet_0.cell(row = i, column = 3).value != " ":
			stop_int = re.search('|'.join(ints), cell_value_with_comma) != None
			if stop_int == False:
				try:
					sheet_0.cell(row = i, column = 3).value = int(float(sheet_0.cell(row = i, column = 3).value))
				except ValueError:
					print("в строке " + str(i) + " ошибка, " "проверь ячейку со значением " + sheet_0.cell(row = i, column = 3).value)
					time.sleep(10)
					raise ValueError




	dict_for_rename = {}
	dict_for_yellow_fill = {} # это словарь на случай, если скрипт найдет несколько строкв прайсе клиники, и первая строка будет больше по цене
			####################################################################################################

	def perfect_match_lab(name, cell_id):
			wb.active = 0
			sheet_0 = wb.active
			rows, columns = sheet_0.max_row, sheet_0.max_column
			for i in range(where_is_lab_1, rows + 1):
				cell_value_name_from_price = str(sheet_0.cell(row = i, column = 2).value)
				for j in range(0, len(name)):
					if name[j] == str(sheet_0.cell(row = i, column = 2).value):
						if sheet_0.cell(row = i, column = 3).value != None:
							cell_value_from_price = round(int(sheet_0.cell(row = i, column = 3).value))
							wb.active = 1
							sheet_1 = wb.active
							rows, columns = sheet_1.max_row, sheet_1.max_column
							for k in range(1, rows):
								cell_value_name_from_template = str(sheet_1.cell(row = k, column = 4).value)
								if cell_id == str(sheet_1.cell(row = k, column = 4).value):
									if sheet_1.cell(row = k, column = 5).value == None:             
										sheet_1.cell(row = k, column = 5).value = cell_value_from_price
										sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
										sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
										dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
										link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
										sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
										sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
										dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
										print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
										break
									elif cell_value_from_price < int(sheet_1.cell(row = k, column = 5).value):
										sheet_1.cell(row = k, column = 5).value = cell_value_from_price
										sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
										sheet_1.cell(row = k, column = 6).value = 1 
										sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
										must_rename = dict_for_rename.get(cell_value_name_from_template)
										must_rename.value = ""
										dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
										link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
										sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
										sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
										must_refill = dict_for_yellow_fill.get(cell_value_name_from_template)
										must_refill.fill = PatternFill("solid", fgColor="FFFFFF")
										dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
										print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
										break
									else:
										sheet_1.cell(row = k, column = 6).value = 1
										break

	########################################################################################################
	def new_doubler_lab(name_1, name_2, stopword, what_double, what_double_service):
		wb.active = 1                                            
		sheet_1 = wb.active                                                         
		rows, columns = sheet_1.max_row, sheet_1.max_column
		for i in range(1, rows+1):
			if sheet_1.cell(row = i, column = 4).value == what_double:                  
				sheet_1.insert_rows(i)
				sheet_1.cell(row = i, column = 2).value = what_double_service
				sheet_1.cell(row = i, column = 4).value = what_double
				break   

		wb.active = 0
		sheet_0 = wb.active
		rows, columns = sheet_0.max_row, sheet_0.max_column
		for i in range(where_is_lab_1, rows + 1):
			cell_value_name_from_price = str(sheet_0.cell(row = i, column = 2).value)
			for j in range(0, len(name_1)):     
				if name_1[j] in cell_value_name_from_price:

					for m in range(0, len(name_2)):
						if name_2[m] in cell_value_name_from_price:
							
							stop_word = re.search('|'.join(stopword), cell_value_name_from_price) != None
							if stop_word == False:
								if sheet_0.cell(row = i, column = 3).value != '' and sheet_0.cell(row = i, column = 3).value != None:
									if str(sheet_0.cell(row = i, column = 3).value).isdigit():
										cell_value_from_price = round(int(sheet_0.cell(row = i, column = 3).value)) 
										wb.active = 1
										sheet_1 = wb.active
										rows, columns = sheet_1.max_row, sheet_1.max_column
										for k in range(1, rows):
											cell_value_name_from_template = str(sheet_1.cell(row = k, column = 4).value)                                        
											if what_double == cell_value_name_from_template:

												if cell_value_from_price != sheet_1.cell(row = k, column = 5).value and cell_value_from_price != "":
													if sheet_0.cell(row = i, column = 3).value != None:                             
														if sheet_1.cell(row = k, column = 5).value == None:             
															sheet_1.cell(row = k, column = 5).value = cell_value_from_price
															sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
															sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
															link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
															sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
															sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
															dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
															dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
															print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
															break
														elif int(cell_value_from_price) < int(sheet_1.cell(row = k, column = 5).value):
															sheet_1.cell(row = k, column = 5).value = cell_value_from_price
															sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
															sheet_1.cell(row = k, column = 6).value = 1 
															sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
															must_rename = dict_for_rename.get(cell_value_name_from_template)
															must_rename.value = ""
															dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
															link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
															sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
															sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
															must_refill = dict_for_yellow_fill.get(cell_value_name_from_template)
															must_refill.fill = PatternFill("solid", fgColor="FFFFFF")
															dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
															print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
															break
														else:
															sheet_1.cell(row = k, column = 6).value = 1
													break

	def new_lab_finder(name_1, name_2, cell_id, stopword):
		# print(where_is_lab_1)
		wb.active = 0
		sheet_0 = wb.active
		rows = sheet_0.max_row
		for i in range(where_is_lab_1, rows + 1):
			cell_value_name_from_price = str(sheet_0.cell(row = i, column = 2).value)
			for j in range(0, len(name_1)):     
				if name_1[j] in cell_value_name_from_price:

					for m in range(0, len(name_2)):
						if name_2[m] in cell_value_name_from_price:
							stop_word = re.search('|'.join(stopword), cell_value_name_from_price) != None
							if stop_word == False:
								if sheet_0.cell(row = i, column = 3).value != '' and sheet_0.cell(row = i, column = 3).value != None:
									if str(sheet_0.cell(row = i, column = 3).value).isdigit():
										cell_value_from_price = round(int(sheet_0.cell(row = i, column = 3).value)) 
										wb.active = 1
										sheet_1 = wb.active
										rows, columns = sheet_1.max_row, sheet_1.max_column
										for k in range(1, rows):
											cell_value_name_from_template = str(sheet_1.cell(row = k, column = 4).value)                                    
											if cell_id == cell_value_name_from_template:

												if cell_value_from_price != sheet_1.cell(row = k, column = 5).value and cell_value_from_price != "" and cell_value_from_price != 0:    
													if sheet_1.cell(row = k, column = 5).value == None:          
														sheet_1.cell(row = k, column = 5).value = cell_value_from_price
														sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
														sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
														link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
														sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
														sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
														dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)

														dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
														
														print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
														break
													elif int(cell_value_from_price) < int(sheet_1.cell(row = k, column = 5).value):
														sheet_1.cell(row = k, column = 5).value = cell_value_from_price
														sheet_1.cell(row = k, column = 5).fill = PatternFill("solid", fgColor="008000")
														sheet_1.cell(row = k, column = 6).value = 1 
														sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")

														
														must_rename = dict_for_rename.get(cell_value_name_from_template)

														must_rename.value = ""
														dict_for_rename[cell_value_name_from_template] = sheet_0.cell(row = i, column = 5)
														link_for_sheet_0 = "#" + wb.sheetnames[1] + "!" + str(sheet_1.cell(row = k, column = 4).coordinate)
														sheet_0.cell(row = i, column = 5).value = '=HYPERLINK("{}", "{}")'.format(link_for_sheet_0, cell_value_name_from_template)
														sheet_1.cell(row = k, column = 7).value = cell_value_name_from_price
														must_refill = dict_for_yellow_fill.get(cell_value_name_from_template)
														must_refill.fill = PatternFill("solid", fgColor="FFFFFF")
														dict_for_yellow_fill[cell_value_name_from_template] = sheet_0.cell(row = i, column = 2)
														print(cell_value_name_from_template, cell_value_from_price, cell_value_name_from_price)
														break
													else:
														sheet_1.cell(row = k, column = 6).value = 1
														break
		

	new_lab_finder(["трихомо", "trichomonas", "хламид", "chlamydia"], ["гонор", "gonorrhoeae"], "ПЦР анализ на скрытые инфекции", ["акц", "назна", "схем", "при ", "профила"])
	new_lab_finder(["микопла", "mуcoplasma"], ["гонор", "gonorrhoeae"], "ПЦР анализ на скрытые инфекции", ["акц", "назна", "схем", "при ", "профила"])
	new_lab_finder(["трихомо", "trichomonas", "хламид", "chlamydia"], ["микопла", "mуcoplasma"], "ПЦР анализ на скрытые инфекции", ["акц", "назна", "схем", "при ", "профила"])
	new_lab_finder(["иппп", "зппп", "заболеваний, передающихся половым", "инфекций, передающих", "заболеваний передающихся половым", "инфекций передающих"], ["иппп", "зппп", "заболеваний", "инфекций"], "ПЦР анализ на скрытые инфекции", ["акц", "назна", "схем", "при ", "профила"])
	new_lab_finder(["скрыты"], ["инфекц"], "ПЦР анализ на скрытые инфекции", ["акц", "назна", "схем", "при ", "профила"])
	new_lab_finder(["на скрытые инфекции", "иппп", "chlamydia trachomatis, mycoplasma genitalium, neisseria gonorrhoeae, trichomonas vaginalis"], ["на скрытые инфекции", "иппп", "chlamydia trachomatis, mycoplasma genitalium, neisseria gonorrhoeae, trichomonas vaginalis"], "ПЦР анализ на скрытые инфекции", ["прижиг", "удал", "книж", "назна", "схем", "при ", "профила"])
	new_lab_finder(["впч", "папиллом", "папилом", "hpv", "HPV"], ["впч", "папиллом", "папилом", "hpv"], "Анализ ВПЧ", ["прижиг", "удал", "книж", " высокого"])
	new_doubler_lab(["впч", "папиллом", "папилом", "hpv"], ["высок", "онкоген", "канцеро", "16, 18, 31, 33", "16/18/31/33", "16, 18, 30, 31, 33"], ["прижиг", "удал", "книж", "низ"], "Анализ ВПЧ", 1131878)
	new_lab_finder(["ca", "ca ", "ca-", "са", "са ", "са-", "cа", "cа-", "cа ", "сa", "сa ", "сa-"], ["125"], "Анализ крови на Са-125", ["описани", "книж", "результ", "тест", "125гп"])
	new_lab_finder(["вирус иммунодефицита", "вич", "human immunodeficiency virus", "hiv", " вич"], ["вирус иммунодефицита", "вич", "human immunodeficiency virus", "hiv", " вич"], "Анализ крови на ВИЧ", ["лмк", "книж", "archiv", "профосмо", "профпат", "вакцин", "сулкович", "первич", "вторич", "рабинович"])
	new_lab_finder(["25-oh", "25 oh", "25 он", "25-он", "витамин д", "25-hydroxyvitamin", "кальциферол", "витамин d", "25он", "25oh"], ["25-oh", "25 oh", "25 он", "25-он", "витамин д", "25-hydroxyvitamin", "кальциферол", "витамин d", "25он", "25oh"], "Анализ на витамин Д", ["комплекс"])
	new_lab_finder(["т-спот", "т спот", "t spot", "t-spot"], ["т-спот", "т спот", "t spot", "t-spot"], "Тест (анализ) Т-Спот", ["заглушка"])
	new_lab_finder(["манту"], ["манту"], "Реакция Манту", ["прием", "приём", "консультац"])
	new_lab_finder(["диаскин"], ["диаскин"], "Диаскинтест", ["прием", "приём", "консультац", "лмк"])
	new_lab_finder(["сифилис", "treponema"], ["сифилис", "treponema"], "Анализ на сифилис", ["описани", "книж", "лмк"])
	new_lab_finder(["хламиди", "chlamyd", "chlamid"], ["хламиди", "chlamyd", "chlamid"], "Анализ на хламидии", ["описани", "книж", "типирован", "посев", "лмк"])
	new_lab_finder(["уреаплазм", "ureapla",], ["уреаплазм", "ureapla"], "Анализ на уреаплазму", ["описани", "книж", "типирован", "посев", "лмк"])
	new_lab_finder(["гонор", "gonor"], ["Gonor", "гонор", "gonor", "Гонор"], "Анализ на гонорею", ["описани", "книж", "типирован", "посев", "лмк"])
	new_lab_finder(["трихом", "tricho", "Трихомон"], ["трихом", "tricho"], "Анализ на трихомониаз", ["описани", "книж", "типирован", "посев", "лмк"])
	new_lab_finder(["резус", "rh"], ["крови", "групп"], "Анализ на группу крови и резус фактор", ["плода", "крови матери", "гена", "кровь", "комплекс", "возбу", "респир",])
	new_lab_finder(["уреазн", "дыхательный тест", "дыхательного теста"], ["уреазн", "дыхательный тест", "дыхательного теста", "helicobacter", "хеликобактер", "хиликобактер"], "Дыхательный тест на хеликобактер", ["плода", "водород", "крови матери", "гена", "кровь"])
	new_lab_finder(["гастропанел", "gastropa", "гастротест", "гастро-тест", "гастрокомплекс"], ["гастропанел", "gastropa", "гастротест", "гастро-тест", "гастрокомплекс"], "Гастропанель", ["плода", "крови матери", "гена", "кровь"])
	new_lab_finder(["иммуноглоб", "имуногло", "ig e общий"], ["lge", "lg-e", "lg e", "ige", "ig-e", "ig e", "lgе", "lg-е", "lg е", "igе", "ig-е", "ig е", "ig e общий"], "IgE анализ крови на Иммуноглобулин Е", ["прижиг", "удал", "книж", "iga", "igm", "igg", "профил", "аллерг",])
	new_lab_finder(["педиатриче", "детска", "детей"], ["панел", "аллергокомплекс"], "Педиатрическая панель", ["прижиг", "удал", "книж"])
	new_lab_finder(["скарификац"], ["скарификац"], "Скарификационный тест", ["туширова"])
	new_lab_finder(["деревь"], ["деревь"], "Анализ на аллергию к деревьям", ["туширова"])
	new_lab_finder(["аллерг", "панель"], ["живот", "кошк", "собак", "шерст", "лошад", "коров"], "Анализ на аллергию к животным", ["шерстис", "бухар", "бытов", "дет"])
	new_lab_finder(["аллерг", "панель"], ["пищев"], "Анализ на пищевые аллергены", ["дет", "реб", "токси", "стафил"])
	new_lab_finder(["аллерг", "панель"], ["бытов", "пыль"], "Анализ на бытовые аллергены", ["пыльц"])
	new_lab_finder(["аллерг", "панель"], ["бактери", "плесен", "плесне", "гриб"], "Анализ на бактериальные и грибковые аллергены", ["педиат"])
	new_lab_finder(["аллерг", "панель"], ["пыльц", "цветущих", "трав", "растени"], "Анализ на пыльцу растений", ["дерев"])
	new_lab_finder(["аллерг", "панель"], ["насеком", "таракан"], "Анализ на аллергию к насекомым", ["бытов"])
	new_lab_finder(["prick", "прик тест", "прик-тест"], ["prick", "прик тест", "прик-тест"], "Прик-тест", ["туширова"])
	new_lab_finder(["гомоцистеи", "homocyst"], ["гомоцистеи", "homocyst"], "Анализ крови на гомоцистеин", ["туширова", "комплекс"])
	new_lab_finder(["камня", "камень", "конкремент", "камни", "камней"], ["камня", "камень", "конкремент", "камни", "камней"], "Анализ камня из почки", ["туширова", "камни почечные"])
	new_lab_finder(["токсоплаз", "toxopla"], ["токсоплаз", "toxopla"], "Анализ на токсоплазмоз", ["туширова"])
	new_lab_finder(["torch ", "torch-"], ["torch ", "torch-"], "Анализ на TORCH-инфекции", ["туширова"])
	new_lab_finder(["имуноферментный анализ крови", "иммуноферментный анализ крови"], ["имуноферментный анализ крови", "иммуноферментный анализ крови"], "Иммуноферментный анализ крови", ["туширова"])
	new_lab_finder(["спермограмм", "спермиограмм"], ["спермограмм", "спермиограмм"], "Спермограмма", ["dvd", "заказывается"])
	new_lab_finder(["mar-тест", "мар-тест", "mar test", " мар тест", "mar-", "мар-", "сперм"], ["mar-тест", "мар-тест", "mar test", " мар тест", "mar-", "мар-", "антител"], "MAR-тест", ["туширова", "кров"])
	new_lab_finder(["кальпротектин", "calprotectin"], ["кальпротектин", "calprotectin"], "Анализ кала на кальпротектин", ["групп"])
	new_lab_finder(["ферритин", "ferritin"], ["ферритин", "ferritin"], "Анализ крови на ферритин", ["туширова", "комплекс"])
	new_lab_finder(["мюллер", "амг", "mullerian"], ["мюллер", "амг", "mullerian"], "Анализ на антимюллеров гормон / АМГ", ["туширова"])
	new_lab_finder(["progesterone", "прогестерон"], ["progesterone", "прогестерон"], "Анализ крови на прогестерон", ["туширова"])
	new_lab_finder(["антитела к тг", "антитела к тиреоглобулину", "ат к тиреоглобулину", "thyroglobulin antibodie", "антител к тиреоглобулин", "ат-тг"], ["антитела к тг", "антитела к тиреоглобулину", "ат к тиреоглобулину", "thyroglobulin antibodie", "антител к тиреоглобулин", "ат-тг"], "Анализ крови на антитела к тиреоглобулину", ["туширова", "комплекс"])
	new_lab_finder(["трийодтиронин", "т3", "t3", "т-3", "t-3"], ["трийодтиронин", "т3", "t3", "т-3", "t-3"], "Анализ крови на гормоны Т3 / трийодтиронин", ["туширова"])
	new_lab_finder(["тироксин", "т4", "t4", "t-4", "т-4"], ["тироксин", "т4", "t4", "t-4", "т-4"], "Анализ крови на гормоны Т4 / тироксин", ["туширова"])
	new_lab_finder(["липидограмм", "липидогрaм", "липидный проф", "липидный статус", "липидного проф", "липидного стату"], ["липидограмм", "липидогрaм", "липидный проф", "липидный статус", "липидного проф", "липидного стату"], "Липидограмма", ["туширова"])
	new_lab_finder(["тестостерон", "testosterone"], ["тестостерон", "testosterone"], "Анализ на тестостерон", ["туширова", "Туширова"])
	new_lab_finder(["рака яич", "эпидидимальн", "epididymis", "he4", "he-4", "he 4", "не4", "не-4", " не 4", "he - 4"], ["рака яич", "эпидидимальн", "epididymis", "he4", "he-4", "he 4", "не4", "не-4", " не 4", "he - 4"], "Анализ на онкомаркер HE4", ["туширова"])
	new_lab_finder(["хромогранин", "chromogranin", "хромгранин", "cga", "хга", "нейроэндокринных опухолей"], ["хромогранин", "chromogranin", "хромгранин", "cga", "хга", "нейроэндокринных опухолей"], "Анализ на онкомаркер Хромогранин А", ["туширова"])
	new_lab_finder(["72 4", "72-4", "72.4", "72 - 4"], ["72 4", "72-4", "72.4", "72 - 4"], "Анализ Cа 72-4", ["туширова"])
	new_lab_finder(["енолаз", "nse", "enolase", "нсэ", "нейронспецифическ", "нейро-специфическ"], ["енолаз", "nse", "enolase", "нсэ", "нейронспецифическ", "нейро-специфическ"], "Анализ на нейронспецифическую енолазу", ["туширова", "pratense", "halepense", "перо", "пост"])
	new_lab_finder(["плоскоклеточной карцином", "плоскоклеточную карцином", "scc", "плоскоклеточного", "sсс", "плоскоклеточных"], ["плоскоклеточной карцином", "плоскоклеточную карцином", "scc", "sсс", "плоскоклеточн"], "Анализ на антиген плоскоклеточной карциномы", ["туширова"])
	new_lab_finder(["ca242", "CA-242", "ca-242", "ca 242", "CA242", "CA-242", "CA 242", "са242", "са-242", "са 242", "СА242", "СА-242", "СА 242", "cа242", "cа-242", "cа 242", "CА242", "CА-242", "CА 242", "CА242", "CА-242", "CА 242", "CА242", "CА-242", "CА 242"], ["ca242", "CA-242", "ca-242", "ca 242", "CA242", "CA-242", "CA 242", "са242", "са-242", "са 242", "СА242", "СА-242", "СА 242", "cа242", "cа-242", "cа 242", "CА242", "CА-242", "CА 242",  "CА242", "CА-242", "CА 242", "CА242", "CА-242", "CА 242"], "Анализ на онкомаркер CA-242", ["туширова", "Туширова"])
	new_lab_finder(["s 100", "s-100", "s100"], ["s 100", "s-100", "s100"], "Анализ на онкомаркер белок S100", ["прижиг", "удал", "книж"])
	new_lab_finder(["рэа", "раковый эмбриональ", "раково-эмбриональ", "раково-эмбриональ", "раковый эмбриональ", "ракового эмбриона"], ["рэа", "раковый эмбриональ", "раково-эмбриональ", "раково-эмбриональ", "раковый эмбриональ", "ракового эмбриона"], "Анализ на раково-эмбриональный антиген / РЭА", ["прижиг", "удал", "книж"])
	new_lab_finder(["отцовс"], ["отцовс"], "ДНК тест на отцовство", ["прижиг", "удал", "книж", "выдача"])
	new_lab_finder(["иммунограмм", "имунограмм", "клеточный иммунитет"], ["иммунограмм", "имунограмм", "клеточный иммунитет"], "Иммунограмма", ["прижиг", "удал", "книж"])
	new_lab_finder(["копрограмм", "koprogram", "копрологи", "koprolog"], ["копрограмм", "koprogram", "копрологи", "koprolog"], "Копрограмма", ["прижиг", "удал", "книж"])
	new_lab_finder(["общего анализа кала", "общий анализ кала"], ["общего анализа кала", "общий анализ кала"], "Общий анализ кала", ["прижиг", "удал", "книж"])
	new_lab_finder(["зимниц"], ["зимниц"], "Проба Зимницкого", ["прижиг", "удал", "книж"])
	new_lab_finder(["реберг", "клубочковой фильтрации", "эндогенного креатинина"], ["реберг", "клубочковой фильтрации", "эндогенного креатинина"], "Проба Реберга", ["прижиг", "удал", "книж"])
	new_lab_finder(["сулкович"], ["сулкович"], "Проба Сулковича", ["прижиг", "удал", "книж"])
	new_lab_finder(["нечипор"], ["нечипор"], "Анализ мочи по Нечипоренко", ["прижиг", "удал", "книж"])
	new_lab_finder(["соматомеди", "инсулинозависимый фактор роста", "инсулиноподобный фактор роста", "ипфр", "somatomedin"], ["соматомеди", "инсулинозависимый фактор роста", "инсулиноподобный фактор роста", "ипфр", "somatomedin"], "Анализ на соматомедин-с в крови", ["прижиг", "удал", "книж"])
	new_lab_finder(["фибриноген", "fibrinogen"], ["фибриноген", "fibrinogen"], "Анализ крови на фибриноген", ["прижиг", "удал", "книж"])
	new_lab_finder(["фолликуло", "фоликуло", "фсг", "follicle-stimulating", "follicle stimulating"], ["стим", "фсг", "follicle-stimulating", "follicle stimulating"], "Анализ на ФСГ", ["прижиг", "удал", "книж"])
	new_lab_finder(["катион", "eosinophil", "эозинофил", "катионный протеин эозинофилов"], ["катион", "eosinophil", "эозинофил", "катионный протеин эозинофилов"], "Анализ крови на эозинофильный катионный белок", ["нейтрофилов", "лейкоцитов", "эозинофилы"])
	new_lab_finder(["лютеинизи", "лютенизи", "лютеотропин"], ["лютеинизи", "лютенизи", "лютеотропин"], "Анализ на лютеинизирующий гормон", ["прижиг", "удал", "книж"])
	new_lab_finder(["волос"], ["микроэлемен", "металл", "макроэлемен", "минерал", "элемент", "спектральный ана"], "Анализ волос на микроэлементы", ["прижиг", "удал", "книж", "ногт", "кож"])
	new_lab_finder(["моча", "моче", "мочи", "мочой"], ["наркоти"], "Анализ мочи на наркотики", ["прижиг", "удал", "книж"])
	new_lab_finder(["са", "сa", "ca", "cа"], ["19 9", "19-9", "19 - 9"], "Анализ Са-19-9", ["прижиг", "удал", "книж", "мин"])
	new_lab_finder(["квантиферон", "интерферон", "интерферен", "quantiferon-tb"], ["туберкул", "tubercul", "квантифероновый тест"], "Квантифероновый тест на туберкулез", ["прижиг", "удал", "книж"])
	new_lab_finder(["скрыт"], ["кровь", "крови"], "Анализ кала на скрытую кровь", ["прижиг", "удал", "книж"])
	new_lab_finder(["гемоглобин"], ["гемоглобин"], "Анализ на гемоглобин", ["глик", "кал", "карбокси", "метгемо", "гапто", "fob", "стул", "фекал"])
	new_lab_finder(["гемоглобин"], ["глик"], "Анализ на гликозилированный гемоглобин", ["кал", "карбокси", "метгемо", "гапто", "fob", "стул", "фекал"])
	new_lab_finder(["волчан", "lupus"], ["коагул", "coagula"], "Волчаночный антикоагулянт", ["прижиг", "удал", "книж"])
	new_lab_finder(["антинуклеа", "антиядерные антитела", "антинуклеарные антитела"], ["факт", " тел", "антиядерные антитела", "антинуклеарные антитела"], "Антинуклеарный фактор", ["прижиг", "удал", "книж"])
	new_lab_finder(["аццп", "цитрулин", "ccp", "цитруллин", "сср"], ["аццп", "цитрулин", "ccp", "цитруллин", "сср"], "Анализ АЦЦП", ["вименти", "mcv", "комплекс", "vimentin"])
	new_lab_finder(["calcitonin", "кальцитонин"], ["calcitonin", "кальцитонин"], "Анализ крови на кальцитонин", ["рецептор"])
	new_lab_finder(["тиреотроп", "тсг", "thyroid stimulating", "тиреоидстимул", "ttg", "ттг"], ["тиреотроп", "тсг", "thyroid stimulating", "тиреоидстимул", "ttg", "ттг"], "Анализ крови на гормоны ТТГ", ["антител", "antibodie", "ат ", "at "])
	new_lab_finder(["тиреоглобулин", "thyroglobulin"], ["тиреоглобулин", "thyroglobulin"], "Анализ крови на тиреоглобулин", ["антител", "antibodie", "ат", "at"])
	new_lab_finder(["пролактин", "prolactin"], ["пролактин", "prolactin"], "Анализ на пролактин", ["макропролактин", "macroprolactin"])
	new_lab_finder(["псa", "простатический специфический", "простатический специфич. антиген", "простатический специфический", "простатического специфи", "psa", "prostate-specific", "prostate specific", "пса", "простатспецифический антиген", "простатоспецифический антиген", "простатспецифического антигена", "простатоспецифического антигена"], ["псa", "простатический специфический", "простатический специфический",  "простатический специфич. антиген", "простатического специфи", "psa", "prostate-specific", "prostate specific", "пса", "простатспецифический антиген", "простатоспецифический антиген", "простатспецифического антигена", "простатоспецифического антигена"], "Анализ крови на ПСА / простатический специфический антиген", ["соотнош", "своб"])
	new_lab_finder(["микроглобулин", "microglobulin"], ["микроглобулин", "microglobulin"], "Анализ крови на бета-2-микроглобулин", ["моч"])
	new_lab_finder(["15 3", "15-3", "15 - 3" "карциномы молочн"], ["15 3", "15-3", "15 - 3" "карциномы молочн"], "Анализ крови на Са 15-3", ["мин"])
	new_lab_finder(["ubc", "рака мочевого", "urinary bladder cancer"], ["ubc", "рака мочевого", "urinary bladder cancer"], "Анализ на онкомаркер UBC", ["сyfra", "21"])
	new_lab_finder(["hla"], ["типир", "ген"], "HLA-типирование", ["патоген", "урогени", "андрофлор", "chla"])
	new_doubler_lab(["hla"], ["drb1, dqa1, dqb1", "dqa1, drb1, dqb1", "dqb1, dqa1, drb1", "dqa1, dqb1, drb1", "drb1, dqb1, dqa1", "drb1, dqa1, dqb1", "hla-dqa1, hla-dqb1, hla-drb", "гистосовместимост", "комплекс", "по трем генам"], ["уроген", "андрофлор", "патоген", "chla", "локус dqa1", ": dqa1", "локус dqb1", "локус drb1", "b27", "1 локус", "один локус", "одному локусу", "абакавир", "5701", "чувств"], "HLA-типирование", 1875930)
	new_lab_finder(["гепатит с", "гепатита с", "гепатита c", "гепатит c", "hcv", "нсv", "hсv", "нcv", "hepatitis с", "hepatitis c"], ["гепатит с", "гепатита с", "гепатита c", "гепатит c", "hcv", "нсv", "hсv", "hepatitis с", "hepatitis c"], "Анализ на гепатит C", ["лмк", "книж", "профосмо", "профпат", "вакцин"])
	new_lab_finder(["гепатит а", "гепатита а", "гепатита a", "гепатит a", "hav", "нav", "hаv", "наv" "hepatitis a", "hepatitis а"], ["гепатит а", "гепатита а", "гепатита a", "гепатит a", "hav", "нav", "hаv", "hepatitis a", "hepatitis а"], "Анализ на гепатит A", ["лмк", "книж", "профосмо", "профпат", "вакцин"])
	new_lab_finder(["гепатит b", "гепатита b", "гепатита в", "гепатит в", "hbv", "нbv", "hвv", "нвv", "hepatitis b", "hepatitis в", "hbsag"], ["гепатит b", "гепатита b", "гепатита в", "гепатит в", "hbv", "нbv", "hвv", "нвv", "hepatitis b", "hepatitis в", "hbsag"], "Анализ на гепатит B", ["лмк", "книж", "профосмо", "профпат", "вакцин"])
	new_lab_finder(["гепатит d", "гепатита d", "гепатита д", "гепатит д", "hdv", "нdv", "hdv", "нdv", "hepatitis d", "hepatitis д"], ["гепатит d", "гепатита d", "гепатита д", "гепатит д", "hdv", "нdv", "hdv", "нdv", "hepatitis d", "hepatitis д"], "Анализ на гепатит D", ["лмк", "книж", "профосмо", "профпат", "вакцин"])
	perfect_match_lab(["взятие крови из вены", "взятие крови", "забор крови", "забор крови из вены", "забор крови  из вены", "забор крови из периферической вены", "венопункция (взрослые, дети)", "забор  крови", "взятие крови из периферической вены", "забор венозной крови", "взятие венозной крови", "забор крови(вена)", "взятие крови для анализа из вены", "взятие крови (вена)", "взятие крови (из вены)", "забор крови для анализов", "забор крови из вены (пальца)", "взятие крови из периферической вены", "забор крови из периферической вены ( пациенты от 5 лет)"], "Забор крови")
	perfect_match_lab(["взятие мазка", "забор мазка", "взятие мазка из носа", "забор мазка/соскоба", "забор мазка из влагалища", "получение цервикального мазка", "забор мазка из носа и зева", "взятие мазка из зева", "взятие мазков из уретры", "взятие (забор) гинекологического мазка", "взятие материала ( соскоб)", "забор материала на анализ (мазок) на приеме", "взятие гинекологического мазка", "забор мазков для исследования из носа, уха или горла (без стоимости исследования) (1 забор)", "получение влагалищного мазка", "получение урогенитального мазка", "получение влагалищного мазка на флору, онкоцитологию, ПЦР исследование, бакпосев", "взятие мазка из уретры"], "Взятие мазка")
	new_lab_finder(["тест", "возбудител", "рнк", "забор на коронавирус", "мазок на", "мазка на"], ["коронавирус", "covid", "забор на коронавирус"], "Тест на коронавирус", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "антител", "профпат", "вакцин"])
	new_lab_finder(["антител", "ат", "anti", "анти тел", "анти-тел"], ["c1q"], "Анализ на антитела к C1q", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["антител", "ат", "anti", "анти тел", "анти-тел"], ["ds днк", "днк ds", "двухцепочечной"], "Анализ на антитела к ds ДНК", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["антител", "ат", "anti", "анти тел", "анти-тел"], ["jo-1", "jo - 1", "jo1"], "Анализ на антитела к Jo-1", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["антител", "ат", "anti", "анти тел", "анти-тел"], ["ro/ss-a", "ro / ss-a", "ro/ ss-a", "ro/ssa", "ssa", "ss-a"], "Анализ на антитела к Ro/SS-A", ["лмк", "лист", "гель", "assay", "кал", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["антител", "ат", "anti", "анти тел", "анти-тел"], ["нуклеосомам"], "Анализ на антитела к нуклеосомам", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["антител", "ат", "anti", "анти тел", "анти-тел"], ["цитоплазме нейтрофил"], "Анализ на антитела к цитоплазме нейтрофилов", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["антицентромерные антитела", "антицентромерных антител"], ["антицентромерные антитела", "антицентромерных антител"], "Анализ на антицентромерные антитела", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["анализ крови", "анализ по"], ["форме-50", "форма-50" , "форма 50", "форме 50", "ф-50", "ф 50", "ф50"], "Анализ крови по Форме-50 / Ф-50", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["анализ", "диагностик", "выявление", "возбудител"], ["зппп", "иппп", "инфекции, передающиеся по", "заболевания, передающиеся п", "инфекции передающиеся п", "заболевания передающиес", "заболеваний, передающихся по", "инфекций, передающихся по", "инфекций передающихся п", "заболеваний передающихся п"], "Анализ на ЗППП", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "профпат", "вакцин"])
	new_lab_finder(["волос"], ["наркоти"], "Анализ волос на наркотики", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["крови", "кровь"], ["наркоти"], "Анализ крови на наркотики", ["лмк", "книж", "рино", "адено", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс"])
	new_lab_finder(["кров", "днк", "антиген"], ["хелико", "хилико", "helico", "pylori", "pilori", "pylory"], "Анализ крови на хеликобактер Пилори", ["лмк", "книж", "рино", "адено", "биопт", "желуд", "фгдс", "гастро", "эндо", "слиз", "киш", "профосмо", "парво", "бока", "инфекций", "профпат", "вакцин", "комплекс", "профил", "холес", "витамин", "соэ", "вестер"])



#########################################################################################################################
	
	#для нахождения со вспомогательным not in и максимальной ценой
	def function_with_assistant_not_in_and_max_price(cell_in_our_price, assistant, max_price, cell_id): 
		a = []
		value_for_list = 100000
		q = None
		for i in range(where_is_lab_1, rows + 1):
			cell = sheet_0.cell(row = i, column = 2).value
			for x in range(0, len(cell_in_our_price)):
				if cell_in_our_price[x] in str(cell) and sheet_0.cell(row = i, column = 3).value != "" and sheet_0.cell(row = i, column = 3).value != None and int(float(sheet_0.cell(row = i, column = 3).value)) < int(max_price):
					is_bad = re.search('|'.join(assistant), cell) != None
					if is_bad == False:
						a.append(value_for_list)
						if int(float(sheet_0.cell(row = i, column = 3).value)) < value_for_list:
							value_for_list = int(float(sheet_0.cell(row = i, column = 3).value))
							q = int(float(sheet_0.cell(row = i, column = 3).value))
							p = cell
							sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")
		if q != None:           
			wb.active = 1
			sheet_1 = wb.active
			for m in range(1, 3800):
				cell_sheet_1 = sheet_1.cell(row = m, column = 4).value
				if cell_id[0] == str(cell_sheet_1):
					sheet_1.cell(row = m, column = 5).value = q
					sheet_1.cell(row = m, column = 7).value = p
					if len(set(a)) > 1:
							sheet_1.cell(row = m, column = 6).value = 1

	#для нахождения со вспомогательным  in or и минимальной ценой
	def function_with_assistant_in_cell_or_and_min_price(cell_in_our_price, assistant, min_price, cell_id): 
		a = []
		value_for_list = 100000
		q = None
		for i in range(where_is_lab_1, rows + 1):
			cell = sheet_0.cell(row = i, column = 2).value
			for x in range(0, len(cell_in_our_price)):
				if cell_in_our_price[x] in str(cell) and sheet_0.cell(row = i, column = 3).value != "" and sheet_0.cell(row = i, column = 3).value != None and int(float(sheet_0.cell(row = i, column = 3).value)) > int(min_price):
					for j in range(0, len(assistant)):
						if assistant[j] in str(cell):
							a.append(value_for_list)
							if int(float(sheet_0.cell(row = i, column = 3).value)) < value_for_list:
								value_for_list = int(float(sheet_0.cell(row = i, column = 3).value))
								q = int(float(sheet_0.cell(row = i, column = 3).value))
								p = cell
								sheet_0.cell(row = i, column = 2).fill = PatternFill("solid", fgColor="008000")

		if q != None:           
			wb.active = 1
			sheet_1 = wb.active
			for m in range(1, 3800):
				cell_sheet_1 = sheet_1.cell(row = m, column = 4).value
				if cell_id[0] == str(cell_sheet_1):
					sheet_1.cell(row = m, column = 5).value = q
					sheet_1.cell(row = m, column = 7).value = p
					sheet_1.cell(row = m, column = 5).fill = PatternFill("solid", fgColor="008000")         
					if len(set(a)) > 1:
						sheet_1.cell(row = m, column = 6).value = 1




	function_with_assistant_not_in_and_max_price(horionic_gonadotropin, svobodniy, 1200, id_horionic_gonadotropin)
	function_with_assistant_in_cell_or_and_min_price(b_12, vitamin, 380, id_b_12)
	
					
						
	wb.save('__после скрипта__' + str(filename_strip))

filename_xlsx = glob('*template*.xlsx')
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
path = pathlib.Path(filename_strip)
#print(path.exists()) # Tru
try:
	wbcsv = pd.read_excel(filename_strip)
except FileNotFoundError:
	print("Ошибка! Шаблон xlsx не найден")
	time.sleep(5)
wbcsv = wbcsv.drop('Синоним', 1)
wbcsv = wbcsv.drop('Ссылка', 1)
wbcsv = wbcsv.drop('Просмотры выборки за посл. 30 дней', 1)
header_list = ['Родительские категории', 'service_id', 'Комментарии для контента', 'Название выборки', 'price', 'is_min_price', 'comment', 'clinic_id', 'filial_id', 'excluded_service_ids']
wbcsv = wbcsv.reindex(columns = header_list) 
filename_xlsx = glob('*price.xlsx')
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
wbxlsx = pd.read_excel(filename_strip)
wb = openpyxl.load_workbook(filename_strip)
wb.create_sheet("шаблонАналитикс", 1)

wb.active = 1
sheet_1 = wb.active
for r in dataframe_to_rows(wbcsv, index=True, header=True):
	sheet_1.append(r)
sheet_1.delete_cols(1)
sheet_1.delete_rows(2)
wb.save('__после скрипта__' + str(filename_strip))


# старый код для чтения csv-шаблона цен, может пригодится в будущем
# filename_xlsx = glob('*.csv')
# filename_strip = str(filename_xlsx).rstrip('\']')
# filename_strip = filename_strip.lstrip('[\'')
# path = pathlib.Path(filename_strip)
# try:
# 	wbcsv = pd.read_csv(filename_strip, delimiter=';', encoding='cp1251')
# except FileNotFoundError:
# 	print("Ошибка! Шаблон не найден")
# 	time.sleep(5)
# wbcsv = wbcsv.drop('Синоним', 1)
# wbcsv = wbcsv.drop('Ссылка', 1)
# wbcsv = wbcsv.drop('Просмотры выборки за посл. 30 дней', 1)
# header_list = ['Родительские категории', 'service_id', 'Комментарии для контента', 'Название выборки', 'price', 'is_min_price', 'comment', 'clinic_id', 'filial_id', 'excluded_service_ids']
# wbcsv = wbcsv.reindex(columns = header_list) 
# filename_xlsx = glob('*.xlsx')
# filename_strip = str(filename_xlsx).rstrip('\']')
# filename_strip = filename_strip.lstrip('[\'')
# wbxlsx = pd.read_excel(filename_strip)
# wb = openpyxl.load_workbook(filename_strip)
# wb.create_sheet("шаблонАналитикс", 1)





if checker == 1:
	without_lab_func()
if checker == 2:
	wb.active = 0
	sheet_0 = wb.active
	rows, columns = sheet_0.max_row, sheet_0.max_column
	where_is_lab_1 = 1
	lab_finder()
if checker == 3:
	without_lab_func()
	lab_finder()
wb = openpyxl.load_workbook('__после скрипта__' + str(filename_strip))
add_sostavnie()
wb.save('__после скрипта__' + str(filename_strip))

print(datetime.now() - start_time)


