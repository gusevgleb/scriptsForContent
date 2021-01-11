# Подключаем библиотеки
import httplib2 
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from glob import glob
import time

CREDENTIALS_FILE = 'testproject-299307-145bfdfa1231.json'  # Имя файла с закрытым ключом, вы должны подставить свое

# Читаем ключи из файла
credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive'])

httpAuth = credentials.authorize(httplib2.Http()) # Авторизуемся в системе
service = apiclient.discovery.build('sheets', 'v4', http = httpAuth) # Выбираем работу с таблицами и 4 версию API 

filename_xlsx = glob('*.xlsx')
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
wb.active = 1
sheet_1 = wb.active
rows_0 = sheet_0.max_row + 1
rows_1 = sheet_1.max_row + 1


udaleno = []
dobavleno = []
first_names_dict = {}
second_names_dict = {}
test_old_names = []
test_second_names = []

for i in range(2, rows_0):
  wb.active = 0
  sheet_0 = wb.active
  first_name = sheet_0.cell(row = i, column = 2).value
  first_price = sheet_0.cell(row = i, column = 8).value
  first_comment = sheet_0.cell(row = i, column = 10).value
  if first_name and first_price != None:
    if first_comment is None:
      first_name_id = str(sheet_0.cell(row = i, column = 2).value) + "noneValue"
      first_names_dict[first_name_id] = first_price
      test_old_names.append(first_name_id)
    else:
      first_name_id = str(sheet_0.cell(row = i, column = 2).value) + sheet_0.cell(row = i, column = 10).value
      first_names_dict[first_name_id] = first_price
      test_old_names.append(first_name_id)



for j in range(2, rows_1):
  second_name = sheet_1.cell(row = j, column = 1).value
  second_price = sheet_1.cell(row = j, column = 2).value
  second_comment = sheet_1.cell(row = j, column = 4).value
  if second_name and second_price != None:
    if second_comment is None:
      second_name_id = str(sheet_1.cell(row = j, column = 1).value) + "noneValue"
      second_names_dict[second_name_id] = second_price
      test_second_names.append(second_name_id)
    else:
      second_name_id = str(sheet_1.cell(row = j, column = 1).value) + sheet_1.cell(row = j, column = 4).value
      second_names_dict[second_name_id] = second_price
      test_second_names.append(second_name_id)


clinic_id = sheet_1.cell(row = 3, column = 5).value
if sheet_1.cell(row = 3, column = 6).value != None:
  filial_id = sheet_1.cell(row = 3, column = 6).value
else: filial_id = "0";
clinic_name = filename_strip.rstrip('.xlsx')
izmeneno_dict = {key: first_names_dict[key]-second_names_dict[key] for key in first_names_dict if key in second_names_dict and first_names_dict[key] != second_names_dict[key]}
prejnih_dict = {key: first_names_dict[key]-second_names_dict[key] for key in first_names_dict if key in second_names_dict and first_names_dict[key] == second_names_dict[key]}


for i in range (0, len(test_second_names)):
  if test_second_names[i] not in test_old_names and test_second_names[i] not in dobavleno:
    dobavleno.append(test_second_names[i])

for i in range (0, len(test_old_names)):
  if test_old_names[i] not in test_second_names and test_old_names[i] not in udaleno:
    udaleno.append(test_old_names[i])


print("Было: " + str(len(test_old_names)))
print("Стало: " + str(len(test_second_names)))
print("Прежних: " + str(len(prejnih_dict)))
print("Удалено: " + str(len(udaleno)))
print("Добавлено: " + str(len(dobavleno)))
print("Изменено: " + str(len(izmeneno_dict)))
print("Clinic_id: " + str(clinic_id))
if sheet_1.cell(row = 3, column = 6).value != None:
  print("filial_id : " + str(filial_id))
print("Клиника: " + str(clinic_name))






# Получаем список листов, их Id и название
spreadsheet = service.spreadsheets().get(spreadsheetId = "1DFccIzNTm-hZ058HI49Fjh67_nXU-lpcEOaHPzE4uHo").execute()
sheetList = spreadsheet.get('sheets')
    
sheetId = sheetList[0]['properties']['sheetId']

# получим данные со всей страницы, чтобы найти первую пустую строку
time.sleep(0.5)
ranges = ["Лист1!A" + str(2) + ":I" + str(500)] # 
results = service.spreadsheets().values().batchGet(spreadsheetId = "1DFccIzNTm-hZ058HI49Fjh67_nXU-lpcEOaHPzE4uHo", 
                                     ranges = ranges, 
                                     valueRenderOption = 'FORMATTED_VALUE',  
                                     dateTimeRenderOption = 'FORMATTED_STRING').execute()


sheet_values = results['valueRanges'][0]['values']
print(len(sheet_values))


    
results = service.spreadsheets().values().batchUpdate(spreadsheetId = "1DFccIzNTm-hZ058HI49Fjh67_nXU-lpcEOaHPzE4uHo", body = {
"valueInputOption": "USER_ENTERED", # Данные воспринимаются, как вводимые пользователем (считается значение формул)
"data": [
    {"range": "Лист1!A" + str(len(sheet_values) + 2) + ":I" + str(len(sheet_values) + 2),
     "majorDimension": "ROWS",     # Сначала заполнять строки, затем столбцы
     "values": [
                [clinic_id, filial_id, str(len(test_old_names)), str(len(test_second_names)), str(len(prejnih_dict)), str(len(udaleno)), str(len(dobavleno)), str(len(izmeneno_dict)), clinic_name], # Заполняем первую строку
                 ]}
                ]
    }).execute()


time.sleep(0.5)
ranges = ["Лист1!A" + str(len(sheet_values) + 2) + ":I" + str(len(sheet_values) + 2)] # 
results = service.spreadsheets().values().batchGet(spreadsheetId = "1DFccIzNTm-hZ058HI49Fjh67_nXU-lpcEOaHPzE4uHo", 
                                     ranges = ranges, 
                                     valueRenderOption = 'FORMATTED_VALUE',  
                                     dateTimeRenderOption = 'FORMATTED_STRING').execute()


sheet_values = results['valueRanges'][0]['values']
print("Данные успешно внесены!" + str(sheet_values))
print('https://docs.google.com/spreadsheets/d/' + "1DFccIzNTm-hZ058HI49Fjh67_nXU-lpcEOaHPzE4uHo" + " - ссылка на таблицу")  
time.sleep(20)

    

