import openpyxl
from glob import glob
import time


filename_xlsx = glob('*.xlsx')
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
wb.active = 1
sheet_1 = wb.active
rows_0 = sheet_0.max_row + 1
rows_1 = sheet_1.max_row + 1


udaleno = []
dobavleno = []
first_names_dict = {}
second_names_dict = {}
test_old_names = []
test_second_names = []

for i in range(2, rows_0):
	wb.active = 0
	sheet_0 = wb.active
	first_name = sheet_0.cell(row = i, column = 1).value
	first_price = sheet_0.cell(row = i, column = 2).value
	first_comment = sheet_0.cell(row = i, column = 4).value
	if first_name and first_price != None:
		if first_comment is None:
			first_name_id = sheet_0.cell(row = i, column = 1).value + "noneValue"
			first_names_dict[first_name_id] = first_price
			test_old_names.append(first_name_id)
		else:
			first_name_id = sheet_0.cell(row = i, column = 1).value + sheet_0.cell(row = i, column = 4).value
			first_names_dict[first_name_id] = first_price
			test_old_names.append(first_name_id)



for j in range(2, rows_1):
	second_name = sheet_1.cell(row = j, column = 1).value
	second_price = sheet_1.cell(row = j, column = 2).value
	second_comment = sheet_1.cell(row = j, column = 4).value
	if second_name and second_price != None:
		if second_comment is None:
			second_name_id = sheet_1.cell(row = j, column = 1).value + "noneValue"
			second_names_dict[second_name_id] = second_price
			test_second_names.append(second_name_id)
		else:
			second_name_id = sheet_1.cell(row = j, column = 1).value + sheet_1.cell(row = j, column = 4).value
			second_names_dict[second_name_id] = second_price
			test_second_names.append(second_name_id)


izmeneno_dict = {key: first_names_dict[key]-second_names_dict[key] for key in first_names_dict if key in second_names_dict and first_names_dict[key] != second_names_dict[key]}
prejnih_dict = {key: first_names_dict[key]-second_names_dict[key] for key in first_names_dict if key in second_names_dict and first_names_dict[key] == second_names_dict[key]}


for i in range (0, len(test_second_names)):
	if test_second_names[i] not in test_old_names and test_second_names[i] not in dobavleno:
		dobavleno.append(test_second_names[i])

for i in range (0, len(test_old_names)):
	if test_old_names[i] not in test_second_names and test_old_names[i] not in udaleno:
		udaleno.append(test_old_names[i])


print("Было: " + str(len(test_old_names)))
print("Стало: " + str(len(test_second_names)))
print("Прежних: " + str(len(prejnih_dict)))
print("Удалено: " + str(len(udaleno)))
print("Добавлено: " + str(len(dobavleno)))
print("Изменено: " + str(len(izmeneno_dict)))

time.sleep(30)