import openpyxl
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter
import pathlib 
from glob import glob

filename_xlsx = glob('*.xlsx')
print(filename_xlsx)
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
print(filename_strip)
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column

rows_0 = sheet_0.max_row + 1
##################################################### основная часть #########################################
print("Пошло дело, пошло!")


array_with_values = []

for i in range(1, rows_0):
	counter = 0
	first_value = sheet_0.cell(row = i, column = 2).value
	array_with_values.append(first_value)

print([k for k,v in Counter(array_with_values).items() if v>1])



print("Успех! Белиссимоs!")
wb.save('__после скрипта__' + str(filename_strip))
print("end")
