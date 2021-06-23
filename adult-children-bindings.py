import openpyxl
from openpyxl.styles import PatternFill
import re
import math
import numpy as np
from collections import Counter
import pathlib 
from glob import glob

filename_xlsx = glob('*.xlsx')
print(filename_xlsx)
filename_strip = str(filename_xlsx).rstrip('\']')
filename_strip = filename_strip.lstrip('[\'')
print(filename_strip)
wb = openpyxl.load_workbook(filename_strip)

wb.active = 0
sheet_0 = wb.active
rows, columns = sheet_0.max_row, sheet_0.max_column

rows_0 = sheet_0.max_row + 1
##################################################### основная часть #########################################
print("Пошло дело, пошло!")

for i in range(2, rows_0):
	value_1 = sheet_0.cell(row = i, column = 2).value

	child_analog = 'ребен'
	child_analog_2 = 'детс'
	child_analog_3 = 'детя'

	if child_analog in value_1 or child_analog_2 in value_1 or child_analog_3 in value_1:
		continue

	for m in range(2, rows_0):
		value_2 = sheet_0.cell(row = m, column = 2).value
		if (value_1 in value_2):

			if child_analog in value_2 or child_analog_2 in value_2 or child_analog_3 in value_2:
				print(value_1, value_2)
				sheet_0.cell(row = i, column = 3).value = value_2
				sheet_0.cell(row = i, column = 4).value = sheet_0.cell(row = m, column = 1).value

		

# for i in range(2, rows_0):
# 	value_from_1_list = sheet_0.cell(row = i, column = 4).value
# 	massiv_1 = value_from_1_list.split()
# 	wb.active = 1
# 	sheet_1 = wb.active
# 	rows_1 = sheet_1.max_row
# 	for m in range(2, rows_1):
# 		value_from_2_list = sheet_1.cell(row = m, column = 4).value
# 		massiv_2 = value_from_2_list.split()

# 		if massiv_1[0] == massiv_2[0]:
# 			sheet_1.cell(row = m, column = 4).fill = PatternFill("solid", fgColor="008000")

print("Успех! Белиссимоs!")
wb.save('__после скрипта__' + str(filename_strip))
print("end")
